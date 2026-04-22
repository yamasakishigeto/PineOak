"""
dic_sem_strain_v57.py
====================
実SEM画像に対するDIC + ひずみ場の計算。
GUIウィザード（tkinter）から全パラメータを設定して実行する。

主な機能：
  - 2段階DIC（Stage1粗い探索 → Stage2精密探索）
  - Phase Correlationによるグローバルシフト推定
  - 前段Stage1結果を初期値に使った高速化オプション
  - 2D放物線フィットによるサブピクセル精度
  - joblib並列処理（行単位）
  - NCCマスク処理（信頼性の低い領域をNaN化）
  - nステップ中心差分によるゲージ長さ指定ひずみ計算
  - dic_config.txtへの設定出力・読み込み
"""

import os
import sys
import json
import math
from pathlib import Path

# ── 必須ライブラリの一括チェック ──────────────────────────────────────────
_REQUIRED = {
    'numpy':      'numpy',
    'cv2':        'opencv-python',
    'scipy':      'scipy',
    'matplotlib': 'matplotlib',
    'openpyxl':   'openpyxl',
}
_missing = []
for _mod, _pkg in _REQUIRED.items():
    try:
        __import__(_mod)
    except ImportError:
        _missing.append(_pkg)
if _missing:
    print("=" * 60)
    print("  【エラー】以下のライブラリが見つかりません。")
    print("  以下のコマンドでインストールしてください：")
    print()
    print(f"  pip install {' '.join(_missing)}")
    print("=" * 60)
    sys.exit(1)

import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from numpy.fft import fft2, ifft2, fftshift
from scipy.interpolate import RegularGridInterpolator
import matplotlib.pyplot as plt
try:
    import japanize_matplotlib
    import matplotlib
    matplotlib.rcParams['axes.unicode_minus'] = False
except ImportError:
    pass

# OSに合わせて日本語フォントを直接設定（japanize_matplotlibの有無に関わらず確実に適用）
import matplotlib
import platform as _platform
_jp_fonts = {
    'Windows': ['Meiryo', 'Yu Gothic', 'MS Gothic', 'MS PGothic'],
    'Darwin':  ['Hiragino Sans', 'Hiragino Maru Gothic Pro', 'AppleGothic'],
    'Linux':   ['IPAexGothic', 'IPAPGothic', 'Noto Sans CJK JP'],
}
_candidates = _jp_fonts.get(_platform.system(), [])
from matplotlib import font_manager as _fm
_available = {f.name for f in _fm.fontManager.ttflist}
_found = next((f for f in _candidates if f in _available), None)
if _found:
    matplotlib.rcParams['font.family'] = _found
matplotlib.rcParams['axes.unicode_minus'] = False
import cv2
try:
    from joblib import Parallel, delayed
    JOBLIB_AVAILABLE = True
except ImportError:
    JOBLIB_AVAILABLE = False
    print("joblib が見つかりません。pip install joblib でインストールすると並列処理が高速化されます。")
try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False
    print("tqdm が見つかりません。pip install tqdm でインストールすると進捗バーが表示されます。")

# =============================================================================
# デフォルトパラメータ
# =============================================================================

DEFAULT_ALIGNMENT_JSON = Path("./sem_alignment.json")  # None にすると補正なし
DEFAULT_OUTPUT_DIR     = None   # None にすると REF_DEF の名前から自動生成

TRIM_TOP    = 0         # 上端トリミング [px]
TRIM_BOTTOM = 0         # 下端トリミング [px]
TRIM_LEFT   = 0         # 左端トリミング [px]
TRIM_RIGHT  = 0         # 右端トリミング [px]
SUBSET_SIZE = 31        # サブセットサイズ [px]
NCC_THRESHOLD = 0.2     # NCCマスク閾値（これ以下をNaN化）
PRESCAN_GRID = 3        # 事前スキャンのグリッド数（PRESCAN_GRID × PRESCAN_GRID）
PRESCAN_SEARCH = 50     # 事前スキャンの探索範囲 [px]
SEARCH_MARGIN = 3       # 推奨search範囲のマージン [px]
N_WORKERS = max(1, (os.cpu_count() or 2) - 1)  # 並列ワーカー数（論理スレッド数-1、1スレッドをOS用に残す）

# 二段階探索のパラメータ
STEP_COARSE    = 60   # Stage 1：粗い探索のサブセット間隔 [px]
STAGE1_MARGIN  = 15   # Stage 1：search = グローバルシフト + このマージン [px]
SEARCH_COARSE  = 30   # Stage 1：固定search範囲 [px]（固定モード時）
STAGE1_AUTO    = True # Stage 1：True=グローバルシフト+margin、False=固定値
STEP_FINE      = 15   # Stage 2：細かい探索のサブセット間隔 [px]
SEARCH_FINE    = 5    # Stage 2：探索範囲 [px]

GAUGE_LENGTH   = 1    # ひずみゲージ長さ（サブセット間隔の整数倍）
USE_PREV_STAGE1 = True  # True=2枚目以降のStage1初期値に前段Stage1結果を使用（高速化）
                         # False=毎回グローバルシフトを初期値（独立処理）
                      # 1=隣接サブセット間差分（最高分解能）、2以上=nステップ離れた点の中心差分

# =============================================================================
# GUIウィザード（tkinter）
# =============================================================================

IMAGE_EXTS = {'.bmp', '.png', '.tif', '.tiff'}

def list_images(folder):
    """フォルダ内の画像ファイルを数値順にソートして返す。"""
    import re
    def _sort_key(f):
        # ファイル名中の最初の数値を抽出してソートキーにする
        m = re.search(r'[\d]+(?:\.\d+)?', f.stem)
        return float(m.group()) if m else f.stem
    files = sorted(
        [f for f in folder.iterdir() if f.suffix.lower() in IMAGE_EXTS],
        key=_sort_key,
    )
    return files


def interactive_wizard(initial_dir=None):
    """tkinterによるGUIウィザード。全設定を1ウィンドウで完結。"""
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    result = {}

    root = tk.Tk()
    root.title("DIC解析ウィザード")
    root.resizable(True, True)
    root.attributes('-topmost', True)

    # 画面高さに合わせてウィンドウ高さを制限
    screen_h = root.winfo_screenheight()
    win_h = min(screen_h - 80, 900)
    root.geometry(f"1300x{win_h}")

    # ── フォント・スタイル ──
    fnt      = ('Meiryo UI', 10)
    fnt_bold = ('Meiryo UI', 10, 'bold')
    fnt_sm   = ('Meiryo UI', 9)

    # ── 状態変数 ──
    folder_var   = tk.StringVar()
    json_var     = tk.StringVar(value='（スキップ）')
    images       = []          # Pathのリスト
    ref_var      = tk.IntVar(value=-1)
    def_vars     = []          # BooleanVarのリスト
    s1_step_var   = tk.IntVar(value=STEP_COARSE)
    s1_auto_var   = tk.BooleanVar(value=STAGE1_AUTO)   # True=自動、False=固定
    s1_margin_var = tk.IntVar(value=STAGE1_MARGIN)
    s1_fixed_var  = tk.IntVar(value=SEARCH_COARSE)
    s2_step_var   = tk.IntVar(value=STEP_FINE)
    s2_srch_var   = tk.IntVar(value=SEARCH_FINE)
    subset_var   = tk.IntVar(value=SUBSET_SIZE)
    ncc_thr_var  = tk.DoubleVar(value=NCC_THRESHOLD)
    n_workers_var = tk.IntVar(value=N_WORKERS)
    prescan_srch_var = tk.IntVar(value=PRESCAN_SEARCH)   # 事前スキャン専用search範囲
    prescan_result_var = tk.StringVar(value='（未実行）')
    gauge_length_var = tk.IntVar(value=GAUGE_LENGTH)      # ひずみゲージ長さ（整数倍）
    use_prev_stage1_var = tk.BooleanVar(value=USE_PREV_STAGE1)  # 前段Stage1結果を初期値に使用

    # トリミング設定
    trim_top_var    = tk.IntVar(value=TRIM_TOP)
    trim_bottom_var = tk.IntVar(value=TRIM_BOTTOM)
    trim_left_var   = tk.IntVar(value=TRIM_LEFT)
    trim_right_var  = tk.IntVar(value=TRIM_RIGHT)
    _ref_img_cache  = [None]   # プレビュー用キャッシュ [numpy array or None]

    # ⑤ カラースケール vmin/vmax（空文字=自動）
    SCALE_KEYS = ['u', 'v', 'exx', 'eyy', 'exy', 'e1', 'gamma_max', 'omega_xy']
    SCALE_LABELS = ['u [px]', 'v [px]', 'εxx', 'εyy', 'εxy', 'e1', 'γmax', 'ωxy']
    scale_min_vars = {k: tk.StringVar(value='') for k in SCALE_KEYS}
    scale_max_vars = {k: tk.StringVar(value='') for k in SCALE_KEYS}

    # ── レイアウト用フレーム ──
    pad = dict(padx=6, pady=4)

    root.rowconfigure(0, weight=1)
    root.columnconfigure(0, weight=0)
    root.columnconfigure(1, weight=1)

    # 左列（①②）
    frm_left  = ttk.Frame(root)
    frm_left.grid(row=0, column=0, sticky='nsew', padx=(8,4), pady=8)
    frm_left.rowconfigure(1, weight=1)

    # 右列：Canvasでスクロール可能にする
    _rc = tk.Canvas(root, highlightthickness=0)
    _rc.grid(row=0, column=1, sticky='nsew', padx=(4,0), pady=8)
    _rc_sb = ttk.Scrollbar(root, orient='vertical', command=_rc.yview)
    _rc_sb.grid(row=0, column=2, sticky='ns', pady=8, padx=(0,4))
    _rc.configure(yscrollcommand=_rc_sb.set)

    frm_right = ttk.Frame(_rc)
    _rc_win = _rc.create_window((0, 0), window=frm_right, anchor='nw')

    def _on_right_configure(e):
        _rc.configure(scrollregion=_rc.bbox('all'))
    frm_right.bind('<Configure>', _on_right_configure)

    def _on_rc_resize(e):
        _rc.itemconfig(_rc_win, width=e.width)
    _rc.bind('<Configure>', _on_rc_resize)

    # マウスホイールでスクロール
    def _on_mousewheel(e):
        _rc.yview_scroll(int(-1 * (e.delta / 120)), 'units')
    _rc.bind_all('<MouseWheel>', _on_mousewheel)

    # ===== セクション1: フォルダ・JSON =====
    frm_top = ttk.LabelFrame(frm_left, text=' ① ファイル選択 ', padding=8)
    frm_top.grid(row=0, column=0, sticky='ew', **pad)

    def browse_folder():
        path = filedialog.askdirectory(title='SEM画像フォルダを選択',
                                       initialdir=initial_dir if initial_dir else None)
        if not path:
            return
        folder_var.set(path)
        load_images(Path(path))

    def browse_json():
        folder = folder_var.get()
        path = filedialog.askopenfilename(
            title='アライメントJSONを選択（キャンセルでスキップ）',
            initialdir=folder if folder else None,
            filetypes=[('JSON', '*.json'), ('All', '*.*')]
        )
        json_var.set(path if path else '（スキップ）')

    ttk.Label(frm_top, text='画像フォルダ:', font=fnt).grid(row=0, column=0, sticky='w')
    ttk.Entry(frm_top, textvariable=folder_var, width=45, font=fnt_sm).grid(row=0, column=1, padx=4)
    ttk.Button(frm_top, text='選択…', command=browse_folder).grid(row=0, column=2)

    ttk.Label(frm_top, text='alignment file:', font=fnt).grid(row=1, column=0, sticky='w', pady=(6,0))
    ttk.Entry(frm_top, textvariable=json_var, width=45, font=fnt_sm).grid(row=1, column=1, padx=4, pady=(6,0))
    ttk.Button(frm_top, text='選択…', command=browse_json).grid(row=1, column=2, pady=(6,0))

    # ===== プレビュー用 別ウィンドウ =====
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import matplotlib.figure as _mfig
    _preview_refs = {'win': None, 'fig': None, 'ax': None, 'canvas': None}

    def _ensure_preview_window():
        """プレビューウィンドウが存在しなければ作成して返す。"""
        if _preview_refs['win'] is None or not _preview_refs['win'].winfo_exists():
            win = tk.Toplevel(root)
            win.title('REF 画像プレビュー（赤 = トリミング領域）')
            win.geometry('560x480')
            win.resizable(True, True)
            fig = _mfig.Figure(figsize=(5.6, 4.6), dpi=90, facecolor='#f0f0f0')
            ax  = fig.add_subplot(111)
            ax.axis('off')
            fig.tight_layout(pad=0.5)
            cv  = FigureCanvasTkAgg(fig, master=win)
            cv.get_tk_widget().pack(fill='both', expand=True)
            _preview_refs.update(win=win, fig=fig, ax=ax, canvas=cv)
        return _preview_refs

    def update_ref_preview(*_):
        """REF画像にトリミング境界線を赤く表示してプレビューを更新する。"""
        img = _ref_img_cache[0]
        if img is None:
            return
        pr  = _ensure_preview_window()
        ax, fig, cv = pr['ax'], pr['fig'], pr['canvas']
        h, w = img.shape[:2]
        top    = trim_top_var.get()
        bottom = trim_bottom_var.get()
        left   = trim_left_var.get()
        right  = trim_right_var.get()

        ax.clear()
        ax.imshow(img, cmap='gray', aspect='equal', vmin=0, vmax=255,
                  interpolation='nearest')
        ax.set_xlim(-0.5, w - 0.5)
        ax.set_ylim(h - 0.5, -0.5)

        def _shade(x0, y0, x1, y1):
            from matplotlib.patches import Rectangle
            ax.add_patch(Rectangle((x0, y0), x1 - x0, y1 - y0,
                         linewidth=0, facecolor='red', alpha=0.35))

        if top > 0:
            ax.axhline(y=top - 0.5, color='red', linewidth=1.5, linestyle='--')
            _shade(-0.5, -0.5, w - 0.5, top - 0.5)
        if bottom > 0:
            ax.axhline(y=h - bottom - 0.5, color='red', linewidth=1.5, linestyle='--')
            _shade(-0.5, h - bottom - 0.5, w - 0.5, h - 0.5)
        if left > 0:
            ax.axvline(x=left - 0.5, color='red', linewidth=1.5, linestyle='--')
            _shade(-0.5, -0.5, left - 0.5, h - 0.5)
        if right > 0:
            ax.axvline(x=w - right - 0.5, color='red', linewidth=1.5, linestyle='--')
            _shade(w - right - 0.5, -0.5, w - 0.5, h - 0.5)

        ref_name = images[ref_var.get()].name if ref_var.get() >= 0 and images else ''
        rw = max(0, w - left - right)
        rh = max(0, h - top - bottom)
        ax.set_title(f'{ref_name}\n元サイズ: {w}×{h}px  →  有効: {rw}×{rh}px',
                     fontsize=8, pad=4)
        ax.axis('off')
        fig.tight_layout(pad=0.5)
        cv.draw()

    # トリミング値変更時にプレビューを更新
    for _v in (trim_top_var, trim_bottom_var, trim_left_var, trim_right_var):
        _v.trace_add('write', update_ref_preview)

    # ===== セクション2: REF・DEF選択 =====
    frm_img = ttk.LabelFrame(frm_left, text=' ② REF / DEF 選択 ', padding=8)
    frm_img.grid(row=1, column=0, sticky='nsew', **pad)
    frm_img.rowconfigure(1, weight=1)

    ttk.Label(frm_img, text='REF', font=fnt_bold, foreground='#c0392b').grid(
        row=0, column=0, sticky='w', padx=(0,16))
    ttk.Label(frm_img, text='DEF', font=fnt_bold, foreground='#2980b9').grid(
        row=0, column=1, sticky='w')
    ttk.Label(frm_img, text='ファイル名', font=fnt_bold).grid(
        row=0, column=2, sticky='w', padx=(8,0))

    # 画像リストを表示するスクロール可能フレーム
    canvas_img = tk.Canvas(frm_img, height=220, highlightthickness=0)
    sb = ttk.Scrollbar(frm_img, orient='vertical', command=canvas_img.yview)
    canvas_img.configure(yscrollcommand=sb.set)
    canvas_img.grid(row=1, column=0, columnspan=3, sticky='nsew')
    sb.grid(row=1, column=3, sticky='ns')

    inner = ttk.Frame(canvas_img)
    canvas_img.create_window((0,0), window=inner, anchor='nw')

    def on_frame_configure(e):
        canvas_img.configure(scrollregion=canvas_img.bbox('all'))
    inner.bind('<Configure>', on_frame_configure)

    def load_images(folder):
        nonlocal images, def_vars
        for w in inner.winfo_children():
            w.destroy()
        def_vars.clear()
        ref_var.set(-1)

        try:
            images = list_images(folder)
        except Exception as e:
            messagebox.showerror('エラー', str(e))
            return

        if len(images) < 2:
            messagebox.showwarning('警告', '画像が2枚以上必要です')
            return

        for i, img_path in enumerate(images):
            dv = tk.BooleanVar(value=True)
            def_vars.append(dv)
            ttk.Radiobutton(inner, variable=ref_var, value=i,
                            command=lambda idx=i: on_ref_select(idx)
                            ).grid(row=i, column=0, padx=8)
            ttk.Checkbutton(inner, variable=dv).grid(row=i, column=1, padx=8)
            ttk.Label(inner, text=img_path.name, font=fnt_sm).grid(
                row=i, column=2, sticky='w', padx=(4,20))

        # デフォルト: 最初の画像をREFに
        ref_var.set(0)
        on_ref_select(0)

    def on_ref_select(idx):
        """REF選択時: DEFのチェックを自動調整 + REF画像をキャッシュしてプレビュー更新"""
        for i, dv in enumerate(def_vars):
            if i == idx:
                dv.set(False)
        try:
            _ref_img_cache[0] = imread_safe(images[idx])
        except Exception:
            _ref_img_cache[0] = None
        update_ref_preview()

    # ===== トリミング設定 =====
    frm_trim = ttk.LabelFrame(frm_right, text=' トリミング設定 ', padding=8)
    frm_trim.grid(row=0, column=0, sticky='ew', **pad)

    ttk.Label(frm_trim,
              text='画像の端を除外するピクセル数を設定します。REF選択時にプレビューが別ウィンドウで開きます。',
              font=fnt_sm, foreground='#555').grid(row=0, column=0, columnspan=9, sticky='w', pady=(0,6))

    _trim_defs = [('上 [px]', trim_top_var), ('下 [px]', trim_bottom_var),
                  ('左 [px]', trim_left_var), ('右 [px]', trim_right_var)]
    for _tc, (_tlbl, _tvar) in enumerate(_trim_defs):
        ttk.Label(frm_trim, text=_tlbl, font=fnt).grid(row=1, column=_tc*2, sticky='e', padx=(12 if _tc else 0, 4))
        ttk.Spinbox(frm_trim, textvariable=_tvar, from_=0, to=4096,
                    width=7, font=fnt).grid(row=1, column=_tc*2+1, padx=(0, 4))
    ttk.Button(frm_trim, text='プレビュー表示',
               command=update_ref_preview).grid(row=1, column=8, padx=(12, 0))

    # ===== セクション3: DICパラメータ =====
    frm_param = ttk.LabelFrame(frm_right, text=' ③ DICパラメータ ', padding=8)
    frm_param.grid(row=1, column=0, sticky='ew', **pad)

    def param_row(parent, row, label, var, tooltip=''):
        ttk.Label(parent, text=label, font=fnt).grid(row=row, column=0, sticky='w', pady=2)
        sb = ttk.Spinbox(parent, textvariable=var, from_=1, to=512, width=6, font=fnt)
        sb.grid(row=row, column=1, padx=8, sticky='w')
        if tooltip:
            ttk.Label(parent, text=tooltip, font=fnt_sm, foreground='gray').grid(
                row=row, column=2, sticky='w')

    ttk.Label(frm_param, text='Stage 1（粗い探索）', font=fnt_bold).grid(
        row=0, column=0, columnspan=3, sticky='w', pady=(0,4))
    param_row(frm_param, 1, '  step  [px]', s1_step_var, '← サブセット間隔')

    # 前段Stage1参照チェックボックス
    ttk.Checkbutton(frm_param, text='前段Stage1結果を初期値に使用（2枚目以降を高速化）',
                    variable=use_prev_stage1_var).grid(
        row=2, column=0, columnspan=3, sticky='w', pady=2)

    # search範囲モード切り替え
    s1_mode_frame = ttk.Frame(frm_param)
    s1_mode_frame.grid(row=3, column=0, columnspan=3, sticky='w', pady=2)

    ttk.Label(s1_mode_frame, text='  search範囲', font=fnt).pack(side='left')

    rb_auto  = ttk.Radiobutton(s1_mode_frame, text='グローバルシフト（1枚目のみ）＋',
                                variable=s1_auto_var, value=True)
    rb_auto.pack(side='left', padx=(12,2))
    spn_margin = ttk.Spinbox(s1_mode_frame, textvariable=s1_margin_var,
                              from_=0, to=200, width=5, font=fnt)
    spn_margin.pack(side='left')
    ttk.Label(s1_mode_frame, text='px', font=fnt).pack(side='left', padx=(2,16))

    rb_fixed = ttk.Radiobutton(s1_mode_frame, text='固定値',
                                variable=s1_auto_var, value=False)
    rb_fixed.pack(side='left', padx=(0,2))
    spn_fixed = ttk.Spinbox(s1_mode_frame, textvariable=s1_fixed_var,
                             from_=1, to=512, width=5, font=fnt)
    spn_fixed.pack(side='left')
    ttk.Label(s1_mode_frame, text='px', font=fnt).pack(side='left', padx=(2,0))

    def _update_s1_widgets(*_):
        """モードに応じてspinboxの有効/無効を切り替え。"""
        if s1_auto_var.get():
            spn_margin.config(state='normal')
            spn_fixed.config(state='disabled')
        else:
            spn_margin.config(state='disabled')
            spn_fixed.config(state='normal')
    s1_auto_var.trace_add('write', _update_s1_widgets)
    _update_s1_widgets()  # 初期状態を反映

    ttk.Separator(frm_param, orient='horizontal').grid(
        row=4, column=0, columnspan=3, sticky='ew', pady=6)

    ttk.Label(frm_param, text='Stage 2（精密探索）', font=fnt_bold).grid(
        row=5, column=0, columnspan=3, sticky='w', pady=(0,4))
    param_row(frm_param, 6, '  step  [px]',   s2_step_var, '← サブセット間隔')
    param_row(frm_param, 7, '  search [px]',  s2_srch_var, '← 探索範囲（Stage1初期値使用）')

    ttk.Separator(frm_param, orient='horizontal').grid(
        row=8, column=0, columnspan=3, sticky='ew', pady=6)

    ttk.Label(frm_param, text='共通設定', font=fnt_bold).grid(
        row=9, column=0, columnspan=3, sticky='w', pady=(0,4))
    param_row(frm_param, 10, '  サブセットサイズ [px]', subset_var, '← 奇数推奨（例: 31, 41, 51）')

    ttk.Separator(frm_param, orient='horizontal').grid(
        row=11, column=0, columnspan=3, sticky='ew', pady=6)

    ttk.Label(frm_param, text='品質フィルタ', font=fnt_bold).grid(
        row=12, column=0, columnspan=3, sticky='w', pady=(0,4))
    ttk.Label(frm_param, text='  NCCマスク閾値', font=fnt).grid(row=13, column=0, sticky='w', pady=2)
    ttk.Spinbox(frm_param, textvariable=ncc_thr_var, from_=0.0, to=1.0,
                increment=0.05, width=6, font=fnt, format='%.2f').grid(
        row=13, column=1, padx=8, sticky='w')
    ttk.Label(frm_param, text='← この値未満のNCC点をNaN化（0で無効）',
              font=fnt_sm, foreground='gray').grid(row=13, column=2, sticky='w')

    ttk.Separator(frm_param, orient='horizontal').grid(
        row=14, column=0, columnspan=3, sticky='ew', pady=6)

    ttk.Label(frm_param, text='ひずみ計算', font=fnt_bold).grid(
        row=15, column=0, columnspan=3, sticky='w', pady=(0,4))
    ttk.Label(frm_param, text='  ゲージ長さ [倍]', font=fnt).grid(row=16, column=0, sticky='w', pady=2)
    ttk.Spinbox(frm_param, textvariable=gauge_length_var, from_=1, to=20,
                width=6, font=fnt).grid(row=16, column=1, padx=8, sticky='w')
    ttk.Label(frm_param, text='← Step_fine の整数倍（1=隣接差分、2以上=nステップ中心差分）',
              font=fnt_sm, foreground='gray').grid(row=16, column=2, sticky='w')

    ttk.Separator(frm_param, orient='horizontal').grid(
        row=17, column=0, columnspan=3, sticky='ew', pady=6)

    ttk.Label(frm_param, text='並列処理', font=fnt_bold).grid(
        row=18, column=0, columnspan=3, sticky='w', pady=(0,4))
    ttk.Label(frm_param, text='  ワーカー数', font=fnt).grid(row=19, column=0, sticky='w', pady=2)
    ttk.Spinbox(frm_param, textvariable=n_workers_var, from_=-1, to=32,
                width=6, font=fnt).grid(row=19, column=1, padx=8, sticky='w')
    ttk.Label(frm_param, text='← 起動時に自動設定（論理スレッド数-1）、-1で全スレッド',
              font=fnt_sm, foreground='gray').grid(row=19, column=2, sticky='w')

    # ===== セクション4: 事前スキャン =====
    frm_prescan = ttk.LabelFrame(frm_right, text=' ④ 事前スキャン（search範囲の確認） ', padding=8)
    frm_prescan.grid(row=2, column=0, sticky='ew', **pad)

    ttk.Label(frm_prescan,
              text='最初のDEF（リスト先頭＝最小変形）に対して3×3点だけ粗い探索を行い、\n推奨Stage1 search範囲を表示します。本番実行前に確認してください。',
              font=fnt_sm, foreground='#555').grid(row=0, column=0, columnspan=3, sticky='w', pady=(0,6))

    ttk.Label(frm_prescan, text='探索範囲 [px]', font=fnt).grid(row=1, column=0, sticky='w', pady=2)
    ttk.Spinbox(frm_prescan, textvariable=prescan_srch_var, from_=1, to=512,
                width=6, font=fnt).grid(row=1, column=1, padx=8, sticky='w')
    ttk.Label(frm_prescan, text='← 事前スキャン専用（広めに設定推奨）',
              font=fnt_sm, foreground='gray').grid(row=1, column=2, sticky='w')

    prescan_label = ttk.Label(frm_prescan, textvariable=prescan_result_var,
                              font=fnt_sm, foreground='#2471a3')
    prescan_label.grid(row=2, column=0, columnspan=3, sticky='w', pady=(4,4))

    rec_search_val = tk.IntVar(value=0)  # 推奨search値を保持

    def on_prescan():
        """事前スキャンを実行してsearch推奨値を表示する。"""
        if not folder_var.get():
            messagebox.showwarning('未入力', 'フォルダを選択してください')
            return
        if ref_var.get() < 0:
            messagebox.showwarning('未選択', 'REFを選択してください')
            return
        def_indices = [i for i, dv in enumerate(def_vars) if dv.get()]
        if not def_indices:
            messagebox.showwarning('未選択', 'DEFを1枚以上選択してください')
            return

        ref_idx = ref_var.get()
        def_indices_clean = [i for i in def_indices if i != ref_idx]
        if not def_indices_clean:
            messagebox.showwarning('未選択', 'REF以外のDEFを選択してください')
            return

        ref_path      = images[ref_idx]
        last_def_path = images[def_indices_clean[-1]]  # リスト最後 = 最大変形

        json_path_str = json_var.get()
        json_path = Path(json_path_str) if json_path_str != '（スキップ）' and json_path_str else None

        prescan_result_var.set('事前スキャン実行中...')
        root.update()

        try:
            trim_ps = (trim_top_var.get(), trim_bottom_var.get(),
                       trim_left_var.get(), trim_right_var.get())
            shifts_ps = load_alignment(json_path)
            ref_raw  = load_and_preprocess(ref_path,      trim_ps)
            def_raw  = load_and_preprocess(last_def_path, trim_ps)
            ref_al   = apply_alignment(ref_raw,  ref_path.name,      shifts_ps, ref_raw.shape)
            def_al   = apply_alignment(def_raw,  last_def_path.name, shifts_ps, def_raw.shape)

            if shifts_ps:
                _h, _w = ref_raw.shape[:2]
                roi_ps = calc_valid_roi(shifts_ps, (_h, _w))
                ref_cr = crop_roi(ref_al, roi_ps)
                def_cr = crop_roi(def_al, roi_ps)
            else:
                ref_cr, def_cr = ref_al, def_al

            _, rec, est = prescan(
                ref_cr, def_cr,
                subset_size=subset_var.get(),
                search_range=prescan_srch_var.get(),
                grid=PRESCAN_GRID,
            )
            rec_search_val.set(rec)

            current_s1 = s1_margin_var.get()
            if rec > current_s1:
                status = f'⚠ 推奨 {rec}px ＞ 現在のStage1 search {current_s1}px → 要見直し'
                prescan_label.config(foreground='#c0392b')
            else:
                status = f'✓ 推奨 {rec}px ≤ 現在のStage1 search {current_s1}px → OK'
                prescan_label.config(foreground='#1a7a1a')

            # 参考値テキストを組み立て
            est_lines = ['  変数          min        max']
            lmap = [('u [px]','u'),('v [px]','v'),('exx [-]','exx'),
                    ('eyy [-]','eyy'),('exy [-]','exy'),('e1 [-]','e1'),
                    ('γmax [-]','gamma_max'),('ωxy [-]','omega_xy')]
            for lbl, key in lmap:
                lo, hi = est[key]
                est_lines.append(f'  {lbl:<10}  {lo:>8.4f}  {hi:>8.4f}')
            est_text = '\n'.join(est_lines)

            prescan_result_var.set(
                f'対象: {last_def_path.name}\n{status}\n\n'
                f'【参考：変位・ひずみ推定値】\n{est_text}'
            )

        except Exception as e:
            prescan_result_var.set(f'エラー: {e}')
            prescan_label.config(foreground='red')

    def on_apply_rec():
        """推奨search値をStage1 searchのspinboxに反映する。"""
        rec = rec_search_val.get()
        if rec <= 0:
            messagebox.showinfo('未実行', '先に事前スキャンを実行してください')
            return
        s1_margin_var.set(rec)
        prescan_result_var.set(
            prescan_result_var.get() + f'\n→ Stage1 search を {rec}px に更新しました'
        )

    prescan_btn_frame = ttk.Frame(frm_prescan)
    prescan_btn_frame.grid(row=3, column=0, columnspan=3, sticky='w', pady=(4,0))
    ttk.Button(prescan_btn_frame, text='  事前スキャン実行  ', command=on_prescan).pack(side='left', padx=(0,8))
    ttk.Button(prescan_btn_frame, text='推奨値をStage1 searchに反映', command=on_apply_rec).pack(side='left')

    # ===== ⑤ カラースケール設定 =====
    frm_scale = ttk.LabelFrame(frm_right, text='⑤ カラースケール（空欄=自動）', padding=10)
    frm_scale.grid(row=3, column=0, sticky='ew', **pad)

    ttk.Label(frm_scale, text='事前スキャン後に参考値を確認して入力してください。空欄の場合は各マップの値から自動設定されます。',
              font=fnt_sm, foreground='#555').grid(row=0, column=0, columnspan=7, sticky='w', pady=(0,4))

    def on_load_config():
        """dic_config.txtを読み込んでDICパラメータとカラースケールを復元する。"""
        folder = folder_var.get()
        path = filedialog.askopenfilename(
            title='dic_config.txt を選択',
            initialdir=folder if folder else None,
            filetypes=[('テキスト', '*.txt'), ('All', '*.*')]
        )
        if not path:
            return
        try:
            with open(path, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            cfg = {}
            for line in lines:
                line = line.strip()
                if ':' in line and not line.startswith('[') and not line.startswith('='):
                    key, _, val = line.partition(':')
                    cfg[key.strip()] = val.strip()

            # DICパラメータの復元
            def _get_int(k, var):
                if k in cfg:
                    try: var.set(int(cfg[k].split()[0]))
                    except: pass
            def _get_float(k, var):
                if k in cfg:
                    try: var.set(float(cfg[k].split()[0]))
                    except: pass

            _get_int('subset', subset_var)
            _get_int('gauge', gauge_length_var)
            _get_int('workers', n_workers_var)
            _get_float('NCC閾値', ncc_thr_var)

            # Stage1/Stage2パラメータはsectionで区別
            import re
            in_s1, in_s2 = False, False
            for line in lines:
                line = line.strip()
                if line == '[Stage 1]': in_s1, in_s2 = True, False; continue
                if line == '[Stage 2]': in_s1, in_s2 = False, True; continue
                if line.startswith('['):  in_s1, in_s2 = False, False; continue
                if ':' not in line: continue
                k, _, v = line.partition(':')
                k, v = k.strip(), v.strip()
                if in_s1 and k == 'step':
                    try: s1_step_var.set(int(v.split()[0]))
                    except: pass
                if in_s1 and k == 'search':
                    if 'グローバルシフト' in v:
                        s1_auto_var.set(True)
                        m = re.search(r'\+\s*(\d+)', v)
                        if m: s1_margin_var.set(int(m.group(1)))
                    else:
                        s1_auto_var.set(False)
                        try: s1_fixed_var.set(int(v.split()[0]))
                        except: pass
                if in_s1 and k == '前段参照':
                    use_prev_stage1_var.set('あり' in v)
                if in_s2 and k == 'step':
                    try: s2_step_var.set(int(v.split()[0]))
                    except: pass
                if in_s2 and k == 'search':
                    try: s2_srch_var.set(int(v.split()[0]))
                    except: pass

            # カラースケールの復元
            key_map = {
                'u': 'u', 'v': 'v', 'exx': 'exx', 'eyy': 'eyy', 'exy': 'exy',
                'e1': 'e1', 'gamma_max': 'gamma_max', 'omega_xy': 'omega_xy'
            }
            in_scale = False
            for line in lines:
                line = line.strip()
                if line == '[カラースケール]': in_scale = True; continue
                if line.startswith('['):     in_scale = False; continue
                if not in_scale or ':' not in line: continue
                k, _, rest = line.partition(':')
                k = k.strip()
                if k not in key_map: continue
                # min=xxx  max=yyy
                import re
                m_min = re.search(r'min=([^\s]+)', rest)
                m_max = re.search(r'max=([^\s]+)', rest)
                lo_str = m_min.group(1) if m_min else '自動'
                hi_str = m_max.group(1) if m_max else '自動'
                scale_min_vars[k].set('' if lo_str == '自動' else lo_str)
                scale_max_vars[k].set('' if hi_str == '自動' else hi_str)

            messagebox.showinfo('読み込み完了', f'設定を復元しました:\n{Path(path).name}')

        except Exception as e:
            messagebox.showerror('読み込みエラー', str(e))

    ttk.Button(frm_scale, text='📂 dic_config.txt を読み込む', command=on_load_config).grid(
        row=1, column=0, columnspan=7, sticky='w', pady=(0, 6))

    ttk.Label(frm_scale, text='変数', font=fnt_bold).grid(row=2, column=0, sticky='w', padx=(0,8))
    ttk.Label(frm_scale, text='min', font=fnt_bold).grid(row=2, column=1, sticky='w', padx=4)
    ttk.Label(frm_scale, text='max', font=fnt_bold).grid(row=2, column=2, sticky='w', padx=4)
    ttk.Label(frm_scale, text='', width=4).grid(row=2, column=3)  # spacer
    ttk.Label(frm_scale, text='変数', font=fnt_bold).grid(row=2, column=4, sticky='w', padx=(0,8))
    ttk.Label(frm_scale, text='min', font=fnt_bold).grid(row=2, column=5, sticky='w', padx=4)
    ttk.Label(frm_scale, text='max', font=fnt_bold).grid(row=2, column=6, sticky='w', padx=4)

    for idx, (k, lbl) in enumerate(zip(SCALE_KEYS, SCALE_LABELS)):
        col_base = (idx % 2) * 4  # 左列=0, 右列=4
        row = 3 + idx // 2
        ttk.Label(frm_scale, text=lbl, font=fnt, width=8).grid(
            row=row, column=col_base, sticky='w', padx=(0,8), pady=2)
        ttk.Entry(frm_scale, textvariable=scale_min_vars[k], width=9, font=fnt).grid(
            row=row, column=col_base+1, padx=4, pady=2)
        ttk.Entry(frm_scale, textvariable=scale_max_vars[k], width=9, font=fnt).grid(
            row=row, column=col_base+2, padx=4, pady=2)

    # ===== 実行ボタン =====
    def on_run():
        # バリデーション
        if not folder_var.get():
            messagebox.showwarning('未入力', 'フォルダを選択してください')
            return
        if ref_var.get() < 0:
            messagebox.showwarning('未選択', 'REFを選択してください')
            return
        def_indices = [i for i, dv in enumerate(def_vars) if dv.get()]
        if not def_indices:
            messagebox.showwarning('未選択', 'DEFを1枚以上選択してください')
            return

        ref_idx = ref_var.get()
        if ref_idx in def_indices:
            def_indices = [i for i in def_indices if i != ref_idx]

        json_path_str = json_var.get()
        json_path = Path(json_path_str) if json_path_str != '（スキップ）' and json_path_str else None

        result['ref_path']   = images[ref_idx]
        result['def_paths']  = [images[i] for i in def_indices]
        result['json_path']  = json_path
        result['folder']     = Path(folder_var.get())
        result['s1_step']    = s1_step_var.get()
        result['s1_auto']    = s1_auto_var.get()
        result['s1_margin']  = s1_margin_var.get()
        result['s1_fixed']   = s1_fixed_var.get()
        result['s2_step']    = s2_step_var.get()
        result['s2_search']  = s2_srch_var.get()
        result['subset_size'] = subset_var.get()
        result['ncc_threshold'] = ncc_thr_var.get()
        result['n_workers']     = n_workers_var.get()
        result['gauge_length']  = gauge_length_var.get()
        result['use_prev_stage1'] = use_prev_stage1_var.get()
        result['trim'] = (trim_top_var.get(), trim_bottom_var.get(),
                          trim_left_var.get(), trim_right_var.get())

        # カラースケール（空欄はNoneとして扱う）
        def _parse_scale(v):
            s = v.get().strip()
            try: return float(s) if s else None
            except ValueError: return None
        result['scale'] = {
            k: (_parse_scale(scale_min_vars[k]), _parse_scale(scale_max_vars[k]))
            for k in SCALE_KEYS
        }

        # 確認ダイアログ
        s1_search_desc = (
            f"グローバルシフト（1枚目のみ）＋{result['s1_margin']}px（自動）"
            if result['s1_auto'] else f"固定 {result['s1_fixed']}px"
        )
        _t = result['trim']
        _trim_str = f"上{_t[0]} 下{_t[1]} 左{_t[2]} 右{_t[3]} px"
        msg = (
            f"REF : {result['ref_path'].name}\n"
            f"DEF : {len(result['def_paths'])}枚\n"
            f"JSON: {json_path.name if json_path else 'なし'}\n"
            f"トリミング: {_trim_str}\n\n"
            f"サブセットサイズ: {result['subset_size']}px\n"
            f"Stage1  step={result['s1_step']}px  search={s1_search_desc}\n"
            f"        前段Stage1参照: {'あり（高速化）' if result['use_prev_stage1'] else 'なし（独立処理）'}\n"
            f"Stage2  step={result['s2_step']}px  search={result['s2_search']}px\n"
            f"ゲージ長さ: {result['gauge_length']}px（0=自動）\n"
            f"NCCマスク閾値: {result['ncc_threshold']:.2f}\n"
            f"ワーカー数: {result['n_workers']}（-1=全コア）\n\n"
            "この設定で実行しますか？"
        )
        if messagebox.askokcancel('確認', msg):
            root.destroy()

    btn_frame = ttk.Frame(root)
    btn_frame.grid(row=1, column=0, columnspan=3, pady=10)
    ttk.Button(btn_frame, text='  実行  ', command=on_run).pack(side='left', padx=8)
    ttk.Button(btn_frame, text='キャンセル', command=root.destroy).pack(side='left')

    root.mainloop()

    if not result:
        print("キャンセルされました")
        raise SystemExit(0)

    # 結果をターミナルに表示
    print(f"\n  REF : {result['ref_path'].name}")
    print(f"  DEF : {len(result['def_paths'])}枚 → "
          + ", ".join(p.name for p in result['def_paths']))
    print(f"  JSON: {result['json_path'].name if result['json_path'] else 'なし'}")

    return (result['ref_path'], result['def_paths'], result['json_path'],
            result['folder'],   result['s1_step'],   result['s1_auto'],
            result['s1_margin'], result['s1_fixed'],
            result['s2_step'],  result['s2_search'],  result['subset_size'],
            result['ncc_threshold'], result['n_workers'], result['scale'],
            result['gauge_length'], result['use_prev_stage1'], result['trim'])


# =============================================================================
# アライメント補正
# =============================================================================

def load_alignment(json_path):
    """
    SEM Alignment Tool のJSONを読み込む。

    Returns
    -------
    shifts : dict  {filename: (dx, dy)}
    """
    if json_path is None or not Path(json_path).exists():
        if json_path is not None:
            print(f"  警告: {json_path} が見つかりません。補正なしで続行します。")
        return {}

    with open(json_path, 'r') as f:
        data = json.load(f)

    shifts = {name: (val['dx'], val['dy'])
              for name, val in data.get('shifts', {}).items()}

    print(f"  アライメントJSON読み込み完了: {len(shifts)}件")
    for name, (dx, dy) in shifts.items():
        print(f"    {name}: dx={dx}, dy={dy}")

    return shifts


def apply_alignment(image, filename, shifts, image_shape_hw):
    """
    シフト補正を適用する。

    Parameters
    ----------
    image          : np.ndarray  グレースケール画像
    filename       : str         ファイル名
    shifts         : dict        {filename: (dx, dy)}
    image_shape_hw : tuple       (H, W) 補正前の画像サイズ

    Returns
    -------
    np.ndarray  シフト補正済み画像（クロップ前）
    """
    if filename not in shifts:
        return image

    dx, dy = shifts[filename]
    if dx == 0 and dy == 0:
        return image

    h, w = image_shape_hw
    # ツールでDEFを+dx,+dyドラッグ → DEF画像を+dx,+dyで平行移動して補正
    M = np.float32([[1, 0, dx], [0, 1, dy]])
    corrected = cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_LINEAR,
                               borderMode=cv2.BORDER_CONSTANT, borderValue=0)
    return corrected.astype(np.float64)


def calc_valid_roi(shifts, image_shape_hw):
    """
    全シフト量から有効ROI（全画像で対応が取れる共通領域）を計算する。

    例: shifts = {A:(0,0), B:(30,-10), C:(-15,20)}
      → max_dx=30, max_dy=20
      → ROI = [30:W-30, 20:H-20]

    Returns
    -------
    (x0, y0, x1, y1)
    """
    if not shifts:
        h, w = image_shape_hw
        return (0, 0, w, h)

    dx_vals = [abs(dx) for dx, dy in shifts.values()]
    dy_vals = [abs(dy) for dx, dy in shifts.values()]
    max_dx = max(dx_vals) if dx_vals else 0
    max_dy = max(dy_vals) if dy_vals else 0

    h, w = image_shape_hw
    x0 = int(max_dx)
    y0 = int(max_dy)
    x1 = int(w - max_dx)
    y1 = int(h - max_dy)

    if x0 >= x1 or y0 >= y1:
        raise ValueError(
            f"有効ROIが計算できません（シフトが大きすぎる可能性）: "
            f"max_dx={max_dx}, max_dy={max_dy}, image={w}x{h}"
        )

    print(f"  有効ROI: x=[{x0}, {x1}], y=[{y0}, {y1}]  "
          f"→ サイズ {x1-x0} x {y1-y0} px  "
          f"（シフト最大値: dx=±{max_dx}px, dy=±{max_dy}px）")

    return (x0, y0, x1, y1)


def crop_roi(image, roi):
    """ROIでクロップする。"""
    x0, y0, x1, y1 = roi
    return image[y0:y1, x0:x1]

# =============================================================================
# 画像の読み込み・前処理
# =============================================================================

def imread_safe(path):
    """
    日本語・スペース入りパスに対応した画像読み込み。
    cv2.imreadはWindowsで日本語パスに失敗するため
    np.fromfile + cv2.imdecodeを使用する。
    """
    buf = np.fromfile(str(path), dtype=np.uint8)
    img = cv2.imdecode(buf, cv2.IMREAD_GRAYSCALE)
    return img


def load_and_preprocess(path, trim):
    """trim = (top, bottom, left, right) [px]"""
    img = imread_safe(path)
    if img is None:
        raise FileNotFoundError(f"画像が見つかりません: {path}")
    top, bottom, left, right = trim
    h, w = img.shape[:2]
    y0 = top
    y1 = h - bottom if bottom > 0 else h
    x0 = left
    x1 = w - right if right > 0 else w
    img = img[y0:y1, x0:x1]
    return img.astype(np.float32)


def estimate_global_shift(ref, deformed):
    """Phase Correlationによるグローバルシフト推定。"""
    F1 = fft2(ref)
    F2 = fft2(deformed)
    cross = F1 * np.conj(F2)
    cross /= np.abs(cross) + 1e-10
    pc = np.abs(fftshift(ifft2(cross)))
    peak = np.unravel_index(np.argmax(pc), pc.shape)
    shift_y = peak[0] - ref.shape[0] // 2
    shift_x = peak[1] - ref.shape[1] // 2
    return shift_x, shift_y

# =============================================================================
# NCC整数画素探索・サブピクセル補間
# =============================================================================

def search_integer_displacement(ref, deformed, cx, cy, subset_size,
                                search_range, init_u=0, init_v=0):
    """
    NCC整数画素探索（numpy完全ベクトル化版）。
    探索範囲内の全候補サブセットを一括スタックしてnumpyで一気にNCC計算する。
    Pythonループ版に比べて10〜50倍高速。
    """
    half = subset_size // 2
    subset_ref = ref[cy-half:cy+half+1, cx-half:cx+half+1].astype(np.float32)
    f = subset_ref - subset_ref.mean()
    f_norm = np.sqrt(np.sum(f**2))

    du_vals = np.arange(round(init_u) - search_range, round(init_u) + search_range + 1)
    dv_vals = np.arange(round(init_v) - search_range, round(init_v) + search_range + 1)
    search_size = len(dv_vals)  # = 2*search_range+1

    ncc_map = np.zeros((search_size, search_size), dtype=np.float32)
    best_u, best_v = round(init_u), round(init_v)

    # 有効な候補だけを収集してバッチ処理
    valid_patches = []
    valid_idx = []
    for dv_idx, dv in enumerate(dv_vals):
        for du_idx, du in enumerate(du_vals):
            cy_def = cy + int(dv)
            cx_def = cx + int(du)
            if (cy_def - half < 0 or cy_def + half >= deformed.shape[0] or
                    cx_def - half < 0 or cx_def + half >= deformed.shape[1]):
                continue
            valid_patches.append(deformed[cy_def-half:cy_def+half+1,
                                          cx_def-half:cx_def+half+1])
            valid_idx.append((dv_idx, du_idx))

    if not valid_patches:
        return best_u, best_v, ncc_map

    # (N, subset_size, subset_size) のバッチに積み上げてベクトル化NCC
    batch = np.stack(valid_patches, axis=0).astype(np.float32)       # (N, h, w)
    batch_mean = batch.mean(axis=(1, 2), keepdims=True)               # (N, 1, 1)
    g = batch - batch_mean                                            # (N, h, w)
    g_norm = np.sqrt((g**2).sum(axis=(1, 2)))                        # (N,)

    denom = f_norm * g_norm                                           # (N,)
    valid_mask = denom > 1e-7
    ncc_vals = np.where(valid_mask,
                        (f * g).sum(axis=(1, 2)) / np.where(denom > 1e-7, denom, 1.0),
                        0.0)                                          # (N,)

    for k, (dv_idx, du_idx) in enumerate(valid_idx):
        ncc_map[dv_idx, du_idx] = ncc_vals[k]

    best_k = int(np.argmax(ncc_vals))
    best_dv_idx, best_du_idx = valid_idx[best_k]
    best_u = int(du_vals[best_du_idx])
    best_v = int(dv_vals[best_dv_idx])

    return best_u, best_v, ncc_map


def subpixel_refinement(ncc_map, best_u, best_v, search_range, init_u=0, init_v=0):
    """
    2D放物線フィットによるサブピクセル精度向上。
    ピーク周囲3×3点を使って2次曲面をフィットし、解析的な極大値を求める。
    従来の独立1D放物線より精度が高い。
    """
    peak_v_idx = best_v - round(init_v) + search_range
    peak_u_idx = best_u - round(init_u) + search_range
    if (peak_v_idx <= 0 or peak_v_idx >= ncc_map.shape[0]-1 or
            peak_u_idx <= 0 or peak_u_idx >= ncc_map.shape[1]-1):
        return float(best_u), float(best_v)

    # 3×3近傍を取得
    patch = ncc_map[peak_v_idx-1:peak_v_idx+2, peak_u_idx-1:peak_u_idx+2]

    # 2D放物線フィット: f(u,v) = a*u^2 + b*v^2 + c*u*v + d*u + e*v + f
    # 9点から最小二乗法で係数を求め、極値を解析的に計算
    pts = []
    vals = []
    for dv in [-1, 0, 1]:
        for du in [-1, 0, 1]:
            pts.append([du*du, dv*dv, du*dv, du, dv, 1.0])
            vals.append(patch[dv+1, du+1])

    A = np.array(pts)
    b_vec = np.array(vals)
    try:
        coeffs, _, _, _ = np.linalg.lstsq(A, b_vec, rcond=None)
    except np.linalg.LinAlgError:
        # フィット失敗時は1D放物線にフォールバック
        c_left  = patch[1, 0]; c_mid = patch[1, 1]; c_right = patch[1, 2]
        denom_u = 2.0*(c_left - 2.0*c_mid + c_right)
        delta_u = (c_left - c_right) / denom_u if abs(denom_u) > 1e-10 else 0.0
        c_up    = patch[0, 1]; c_down = patch[2, 1]
        denom_v = 2.0*(c_up - 2.0*c_mid + c_down)
        delta_v = (c_up - c_down) / denom_v if abs(denom_v) > 1e-10 else 0.0
        return best_u + delta_u, best_v + delta_v

    a, b, c, d, e, _ = coeffs
    # 極値: [2a  c][du]   [-d]
    #       [c  2b][dv] = [-e]
    denom = 4*a*b - c*c
    if abs(denom) < 1e-10:
        delta_u, delta_v = 0.0, 0.0
    else:
        delta_u = (c*e - 2*b*d) / denom
        delta_v = (c*d - 2*a*e) / denom

    # 補正量が1px以内に収まらない場合はクリップ
    delta_u = float(np.clip(delta_u, -1.0, 1.0))
    delta_v = float(np.clip(delta_v, -1.0, 1.0))
    return best_u + delta_u, best_v + delta_v


# =============================================================================
# DIC実行（初期値対応版）
# =============================================================================

def _process_row(ref, deformed, cx_row, cy, subset_size, search_range,
                 init_us, init_vs):
    """1行分のサブセットをまとめて処理するワーカー（joblib用）。"""
    row_results = []
    for cx, init_u, init_v in zip(cx_row, init_us, init_vs):
        best_u, best_v, ncc_map = search_integer_displacement(
            ref, deformed, int(cx), int(cy), subset_size, search_range,
            init_u=init_u, init_v=init_v)
        sub_u, sub_v = subpixel_refinement(
            ncc_map, best_u, best_v, search_range,
            init_u=init_u, init_v=init_v)
        row_results.append((int(cx), int(cy), sub_u, sub_v, float(ncc_map.max())))
    return row_results


def run_dic_stage(ref, deformed, subset_size, subset_step, search_range,
                  init_u_func=None, init_v_func=None, label="",
                  ncc_threshold=NCC_THRESHOLD, n_workers=-1):
    """
    DIC探索ステージ。joblib並列処理（行単位）・tqdm・NCCマスク対応。

    Parameters
    ----------
    ncc_threshold : float  NCCピーク値がこれ以下の点をNaN化（0で無効化）
    n_workers     : int    並列ワーカー数（-1で全コア使用）
    """
    half = subset_size // 2
    height, width = ref.shape
    margin = half + search_range + 5
    xs = np.arange(margin, width  - margin, subset_step)
    ys = np.arange(margin, height - margin, subset_step)

    total = len(xs) * len(ys)
    print(f"{label} DIC実行中... サブセット数: {total}  ({len(ys)}行 × {len(xs)}列)")

    # 行ごとに初期値を事前計算
    row_tasks = []
    for cy in ys:
        init_us = [int(round(init_u_func(cx, cy))) if init_u_func else 0 for cx in xs]
        init_vs = [int(round(init_v_func(cx, cy))) if init_v_func else 0 for cx in xs]
        row_tasks.append((xs, int(cy), init_us, init_vs))

    # joblib並列処理（行単位、loky バックエンドでWindows対応）
    if JOBLIB_AVAILABLE and n_workers != 1:
        row_results_list = Parallel(n_jobs=n_workers, backend='loky', verbose=0)(
            delayed(_process_row)(ref, deformed, task[0], task[1],
                                  subset_size, search_range, task[2], task[3])
            for task in (tqdm(row_tasks, desc=label, ncols=80)
                         if TQDM_AVAILABLE else row_tasks)
        )
    else:
        row_results_list = []
        tasks_iter = tqdm(row_tasks, desc=label, ncols=80) if TQDM_AVAILABLE else row_tasks
        for task in tasks_iter:
            row_results_list.append(
                _process_row(ref, deformed, task[0], task[1],
                             subset_size, search_range, task[2], task[3])
            )

    # 結果をフラット化
    results = [item for row in row_results_list for item in row]

    cx_arr = np.array([r[0] for r in results])
    cy_arr = np.array([r[1] for r in results])
    u_arr  = np.array([r[2] for r in results], dtype=float)
    v_arr  = np.array([r[3] for r in results], dtype=float)
    ncc_arr = np.array([r[4] for r in results])

    # NCCマスク処理
    if ncc_threshold > 0:
        mask = ncc_arr < ncc_threshold
        n_masked = mask.sum()
        if n_masked > 0:
            u_arr[mask] = np.nan
            v_arr[mask] = np.nan
            print(f"  NCCマスク: {n_masked}/{total}点をNaN化（NCC < {ncc_threshold}）")

    print(f"  {total}/{total} 完了!")
    return cx_arr, cy_arr, u_arr, v_arr, ncc_arr

# =============================================================================
# 事前スキャン（3×3グリッドで最大変形を推定 → search範囲を推定）
# =============================================================================

def prescan(ref, deformed, subset_size, search_range, grid=PRESCAN_GRID):
    """
    最大変形DEF1枚に対して grid×grid 点だけ粗い探索を行い、
    必要なsearch範囲を推定する。

    Returns
    -------
    results : list of dict  各サブセット点の情報
    recommended_search : int  推奨search範囲（マージン込み）
    """
    half = subset_size // 2
    h, w = ref.shape
    margin = half + search_range + 5

    # grid×gridの均等配置
    xs = np.linspace(margin, w - margin, grid).astype(int)
    ys = np.linspace(margin, h - margin, grid).astype(int)

    results = []
    print(f"\n--- 事前スキャン ({grid}×{grid}点, search={search_range}px) ---")

    for cy in ys:
        for cx in xs:
            best_u, best_v, ncc_map = search_integer_displacement(
                ref, deformed, cx, cy, subset_size, search_range)
            sub_u, sub_v = subpixel_refinement(
                ncc_map, best_u, best_v, search_range)
            ncc_peak = ncc_map.max()
            results.append({
                'cx': cx, 'cy': cy,
                'u': sub_u, 'v': sub_v,
                'ncc': ncc_peak,
                'disp': math.hypot(sub_u, sub_v),
            })

    # 統計表示
    disps = [r['disp'] for r in results]
    us    = [r['u']    for r in results]
    vs    = [r['v']    for r in results]
    nccs  = [r['ncc']  for r in results]

    print(f"\n  【事前スキャン結果】")
    print(f"  {'位置':>14}  {'u [px]':>8}  {'v [px]':>8}  {'変位量':>8}  {'NCC':>6}")
    print(f"  {'-'*56}")
    for r in results:
        flag = " ← 最大" if r['disp'] == max(disps) else ""
        print(f"  ({r['cx']:4d}, {r['cy']:4d})  "
              f"{r['u']:8.2f}  {r['v']:8.2f}  "
              f"{r['disp']:8.2f}  {r['ncc']:6.3f}{flag}")

    max_u = max(abs(r['u']) for r in results)
    max_v = max(abs(r['v']) for r in results)
    max_disp = max(disps)
    rec = int(math.ceil(max(max_u, max_v))) + SEARCH_MARGIN

    # search範囲に引っかかっているかチェック
    clipped = [r for r in results
               if abs(r['u']) >= search_range - 0.5 or
                  abs(r['v']) >= search_range - 0.5]

    print(f"\n  最大変位: u={max_u:.2f}px, v={max_v:.2f}px, 合成={max_disp:.2f}px")
    print(f"  推奨 Stage2 search範囲: {rec} px  （実測最大 + マージン{SEARCH_MARGIN}px）")
    if clipped:
        print(f"  ⚠ {len(clipped)}点が現在のsearch範囲({search_range}px)に張り付いています！")
        print(f"    search範囲を大きくして再スキャンしてください。")
    else:
        print(f"  ✓ 全点がsearch範囲内に収まっています。")

    # ── ひずみ推定（3×3グリッドの変位勾配から） ──────────────────
    u_grid = np.array(us, dtype=float).reshape(grid, grid)
    v_grid = np.array(vs, dtype=float).reshape(grid, grid)
    # グリッド間隔（px）
    dx = (xs[-1] - xs[0]) / max(grid - 1, 1)
    dy = (ys[-1] - ys[0]) / max(grid - 1, 1)
    du_dx = np.gradient(u_grid, dx, axis=1)
    du_dy = np.gradient(u_grid, dy, axis=0)
    dv_dx = np.gradient(v_grid, dx, axis=1)
    dv_dy = np.gradient(v_grid, dy, axis=0)
    exx = du_dx
    eyy = dv_dy
    exy = 0.5 * (du_dy + dv_dx)
    omega_xy = 0.5 * (dv_dx - du_dy)
    # 主ひずみ
    e1 = 0.5*(exx+eyy) + np.sqrt((0.5*(exx-eyy))**2 + exy**2)
    gamma_max = 2 * np.sqrt((0.5*(exx-eyy))**2 + exy**2)

    est = {
        'u':         (float(np.nanmin(u_grid)),    float(np.nanmax(u_grid))),
        'v':         (float(np.nanmin(v_grid)),    float(np.nanmax(v_grid))),
        'exx':       (float(np.nanmin(exx)),       float(np.nanmax(exx))),
        'eyy':       (float(np.nanmin(eyy)),       float(np.nanmax(eyy))),
        'exy':       (float(np.nanmin(exy)),       float(np.nanmax(exy))),
        'e1':        (float(np.nanmin(e1)),        float(np.nanmax(e1))),
        'gamma_max': (float(np.nanmin(gamma_max)), float(np.nanmax(gamma_max))),
        'omega_xy':  (float(np.nanmin(omega_xy)),  float(np.nanmax(omega_xy))),
    }

    print(f"\n  【参考：変位・ひずみ推定値（最大変形DEFの粗い推定）】")
    print(f"  {'変数':<12}  {'min':>10}  {'max':>10}")
    print(f"  {'-'*36}")
    labels = [('u [px]',    'u'), ('v [px]',    'v'),
              ('exx [-]',   'exx'), ('eyy [-]', 'eyy'), ('exy [-]', 'exy'),
              ('e1 [-]',    'e1'), ('γmax [-]', 'gamma_max'), ('ωxy [-]', 'omega_xy')]
    for label, key in labels:
        lo, hi = est[key]
        print(f"  {label:<12}  {lo:>10.4f}  {hi:>10.4f}")

    return results, rec, est


# =============================================================================
# Stage 1の結果から補間関数を作成
# =============================================================================

def make_interpolator(cx_list, cy_list, values):
    xs_unique = np.unique(cx_list)
    ys_unique = np.unique(cy_list)
    x_to_col = {x: i for i, x in enumerate(xs_unique)}
    y_to_row = {y: i for i, y in enumerate(ys_unique)}
    grid = np.zeros((len(ys_unique), len(xs_unique)))
    for cx, cy, val in zip(cx_list, cy_list, values):
        grid[y_to_row[cy], x_to_col[cx]] = val
    interp = RegularGridInterpolator(
        (ys_unique, xs_unique), grid,
        method='linear', bounds_error=False, fill_value=0.0)
    return interp

# =============================================================================
# グリッド変換
# =============================================================================

def make_grid_map(cx_list, cy_list, values, subset_step):
    xs_unique = np.unique(cx_list)
    ys_unique = np.unique(cy_list)
    x_to_col = {x: i for i, x in enumerate(xs_unique)}
    y_to_row = {y: i for i, y in enumerate(ys_unique)}
    grid = np.full((len(ys_unique), len(xs_unique)), np.nan)
    for cx, cy, val in zip(cx_list, cy_list, values):
        grid[y_to_row[cy], x_to_col[cx]] = val
    dx = xs_unique[1] - xs_unique[0] if len(xs_unique) > 1 else subset_step
    dy = ys_unique[1] - ys_unique[0] if len(ys_unique) > 1 else subset_step
    extent = [xs_unique[0]-dx/2, xs_unique[-1]+dx/2,
              ys_unique[-1]+dy/2, ys_unique[0]-dy/2]
    return grid, extent, xs_unique, ys_unique

# =============================================================================
# ひずみ場の計算
# =============================================================================

def calc_strain_field(cx_list, cy_list, u_list, v_list, subset_step, gauge_length=1):
    """
    変位場からひずみテンソルを計算する。

    Parameters
    ----------
    gauge_length : int  ひずみゲージ長さ（サブセット間隔の整数倍、1以上）。
                        1 = 隣接サブセット間の中心差分（最高空間分解能、v33以前と同等）。
                        n = nステップ離れた点の中心差分（ゲージ長さ = n × subset_step）。
                        大きいほどノイズに鈍感になる代わりに空間分解能が落ちる。
                        端点はnステップ前方/後方差分で補完。
    """
    u_grid, extent, xs_unique, ys_unique = make_grid_map(
        cx_list, cy_list, u_list, subset_step)
    v_grid, _, _, _ = make_grid_map(
        cx_list, cy_list, v_list, subset_step)

    dx = xs_unique[1] - xs_unique[0] if len(xs_unique) > 1 else subset_step
    dy = ys_unique[1] - ys_unique[0] if len(ys_unique) > 1 else subset_step

    n = max(1, int(round(gauge_length)))  # 整数倍に強制
    gauge_px = n * dx  # 実際のゲージ長さ [px]（dx≈dyのとき）
    print(f"  ゲージ長さ: {n}倍 = {gauge_px:.1f}px（サブセット間隔 {dx:.1f}px × {n}）")

    def _diff(arr, step, axis, h):
        """
        nステップ中心差分。端点はnステップ片側差分で補完。
        arr  : 2D array（NaNあり可）
        step : nステップ数
        axis : 0=y方向, 1=x方向
        h    : グリッド間隔 [px]
        """
        s = step
        if arr.shape[axis] <= 2 * s:
            # グリッド点数がステップ数の2倍以下ならnp.gradientにフォールバック
            return np.gradient(arr, h, axis=axis)

        result = np.full_like(arr, np.nan)
        if axis == 1:
            # 中心部：中心差分  (u[i+n] - u[i-n]) / (2*n*h)
            result[:, s:-s] = (arr[:, 2*s:] - arr[:, :-2*s]) / (2 * s * h)
            # 左端：前方差分   (u[i+n] - u[i]) / (n*h)
            result[:, :s]   = (arr[:, s:2*s] - arr[:, :s])   / (s * h)
            # 右端：後方差分   (u[i] - u[i-n]) / (n*h)
            result[:, -s:]  = (arr[:, -s:] - arr[:, -2*s:-s]) / (s * h)
        else:
            result[s:-s, :] = (arr[2*s:, :] - arr[:-2*s, :]) / (2 * s * h)
            result[:s,   :] = (arr[s:2*s, :] - arr[:s, :])   / (s * h)
            result[-s:,  :] = (arr[-s:, :] - arr[-2*s:-s, :]) / (s * h)
        return result

    du_dx = _diff(u_grid, n, axis=1, h=dx)
    du_dy = _diff(u_grid, n, axis=0, h=dy)
    dv_dx = _diff(v_grid, n, axis=1, h=dx)
    dv_dy = _diff(v_grid, n, axis=0, h=dy)

    exx = du_dx
    eyy = dv_dy
    exy = 0.5 * (du_dy + dv_dx)

    e_mean    = (exx + eyy) / 2.0
    e_diff    = np.sqrt(((exx - eyy) / 2.0)**2 + exy**2)
    e1        = e_mean + e_diff
    e2        = e_mean - e_diff
    gamma_max = e1 - e2
    omega_xy  = 0.5 * (dv_dx - du_dy)

    return {
        'exx': exx, 'eyy': eyy, 'exy': exy,
        'e1': e1, 'e2': e2,
        'gamma_max': gamma_max,
        'omega_xy': omega_xy,
        'extent': extent,
    }

# =============================================================================
# 可視化
# =============================================================================

def visualize_displacement(cx_list, cy_list, u_corr, v_corr,
                            ncc_peak_list, shift_x, shift_y, subset_step,
                            ncc_threshold=NCC_THRESHOLD, def_stem="",
                            scale_config=None, cmap_sym='RdBu_r', preview=False):
    sc = scale_config or {}
    if not preview:
        OUTPUT_DIR.mkdir(exist_ok=True)

    u_grid, extent, _, _ = make_grid_map(cx_list, cy_list, u_corr, subset_step)
    v_grid, _, _, _      = make_grid_map(cx_list, cy_list, v_corr, subset_step)
    ncc_grid, _, _, _    = make_grid_map(cx_list, cy_list, ncc_peak_list, subset_step)

    img_h = extent[2] - extent[3]
    img_w = extent[1] - extent[0]
    aspect = img_h / img_w

    fig, axes = plt.subplots(1, 3, figsize=(18, 6 * aspect + 1))

    u_abs = max(np.nanpercentile(np.abs(u_grid), 95), 1.0)
    u_lo, u_hi = sc.get('u', (None, None))
    u_vmin = u_lo if u_lo is not None else -u_abs
    u_vmax = u_hi if u_hi is not None else  u_abs
    im0 = axes[0].imshow(u_grid, cmap=cmap_sym, extent=extent,
                          vmin=u_vmin, vmax=u_vmax, aspect='equal')
    plt.colorbar(im0, ax=axes[0], label="u [px]")
    axes[0].set_title(f"u変位マップ\nグローバルシフト: x={shift_x}px")
    axes[0].set_xlabel("X [px]"); axes[0].set_ylabel("Y [px]")

    v_abs = max(np.nanpercentile(np.abs(v_grid), 95), 1.0)
    v_lo, v_hi = sc.get('v', (None, None))
    v_vmin = v_lo if v_lo is not None else -v_abs
    v_vmax = v_hi if v_hi is not None else  v_abs
    im1 = axes[1].imshow(v_grid, cmap=cmap_sym, extent=extent,
                          vmin=v_vmin, vmax=v_vmax, aspect='equal')
    plt.colorbar(im1, ax=axes[1], label="v [px]")
    axes[1].set_title(f"v変位マップ\nグローバルシフト: y={shift_y}px")
    axes[1].set_xlabel("X [px]"); axes[1].set_ylabel("Y [px]")

    ncc_vmin = max(ncc_threshold, 0.0)
    im2 = axes[2].imshow(ncc_grid, cmap='plasma', extent=extent,
                          vmin=ncc_vmin, vmax=1.0, aspect='equal')
    plt.colorbar(im2, ax=axes[2], label="NCC peak")
    axes[2].set_title(f"NCCピーク値マップ\n(下限={ncc_vmin:.2f}=しきい値)")
    axes[2].set_xlabel("X [px]"); axes[2].set_ylabel("Y [px]")

    plt.tight_layout()
    if preview:
        plt.show()
    else:
        suffix = f"_{def_stem}" if def_stem else ""
        plt.savefig(OUTPUT_DIR / f"displacement_map{suffix}.png", dpi=150, bbox_inches='tight')
        plt.close()


def visualize_strain(strain, subset_step, ref_name, def_name, def_stem="",
                     scale_config=None, cmap_sym='RdBu_r', cmap_asym='hot_r',
                     preview=False):
    sc = scale_config or {}
    extent = strain['extent']
    img_h = extent[2] - extent[3]
    img_w = extent[1] - extent[0]
    aspect = img_h / img_w
    col_w = 5.5
    fig, axes = plt.subplots(2, 3, figsize=(col_w * 3, col_w * aspect * 2))

    def plot_strain(ax, data, title, key, cmap='RdBu_r', symmetric=True):
        lo, hi = sc.get(key, (None, None))
        if lo is not None and hi is not None:
            vmin, vmax = lo, hi
        elif symmetric:
            vmax = max(np.nanpercentile(np.abs(data), 98), 1e-6)
            vmin = -vmax
        else:
            vmin = np.nanpercentile(data, 2)
            vmax = np.nanpercentile(data, 98)
        im = ax.imshow(data, cmap=cmap, extent=extent,
                       vmin=vmin, vmax=vmax, aspect='equal')
        plt.colorbar(im, ax=ax, label="[-]", fraction=0.046, pad=0.04)
        ax.set_title(title, pad=4, fontsize=10)
        ax.set_xlabel("X [px]", fontsize=8)
        ax.set_ylabel("Y [px]", fontsize=8)
        ax.tick_params(labelsize=7)

    plot_strain(axes[0,0], strain['exx'],       "εxx（x方向正ひずみ）",   'exx',  cmap=cmap_sym)
    plot_strain(axes[0,1], strain['eyy'],       "εyy（y方向正ひずみ）",   'eyy',  cmap=cmap_sym)
    plot_strain(axes[0,2], strain['exy'],       "εxy（せん断ひずみ）",    'exy',  cmap=cmap_sym)
    plot_strain(axes[1,0], strain['e1'],        "e1（最大主ひずみ）",     'e1',   cmap=cmap_asym, symmetric=False)
    plot_strain(axes[1,1], strain['gamma_max'], "γmax（最大せん断ひずみ）",'gamma_max', cmap=cmap_asym, symmetric=False)
    plot_strain(axes[1,2], strain['omega_xy'],  "ωxy（回転テンソル・参考）",'omega_xy', cmap=cmap_sym)

    fig.suptitle(f"ひずみ場マップ（{ref_name} → {def_name}）",
                 fontsize=12, y=1.005)
    plt.tight_layout()
    if preview:
        plt.show()
    else:
        suffix = f"_{def_stem}" if def_stem else ""
        plt.savefig(OUTPUT_DIR / f"strain_map{suffix}.png", dpi=150, bbox_inches='tight')
        plt.close()

    print("\n========== ひずみ統計 ==========")
    for key in ['exx', 'eyy', 'exy', 'e1', 'gamma_max', 'omega_xy']:
        d = strain[key]
        print(f"  {key:10s}: 平均={np.nanmean(d):.4f}  "
              f"std={np.nanstd(d):.4f}  "
              f"範囲=[{np.nanmin(d):.4f}, {np.nanmax(d):.4f}]")
    print(f"  ※ omega_xyの平均={np.nanmean(strain['omega_xy']):.4f} rad"
          f"（≈{np.degrees(np.nanmean(strain['omega_xy'])):.2f}°）が再マウント回転の残存量")
    print("================================")

# =============================================================================
# Excel出力
# =============================================================================

def export_xlsx(results_list, scale_config, out_path, roi=None):
    """
    DIC結果をExcelファイルに出力する。

    Parameters
    ----------
    results_list : list of dict
        各DEFの結果。キー: 'label', 'cx', 'cy', 'u', 'v', 'strain'
        strain は calc_strain_field の戻り値 dict。
        REFを先頭に含むこと（変位・ひずみ全ゼロ）。
    scale_config : dict
        カラースケール設定 {key: (vmin, vmax)}。平均値の除外判定に使用。
    roi : tuple or None
        (x0, y0, x1, y1) クロップROI。x0, y0 をサブセット座標に加算して
        元のSEM画像座標系に変換する。None の場合はオフセットなし。
    out_path : Path
        出力Excelファイルパス。
    """
    PARAMS = ['u', 'v', 'ncc', 'exx', 'eyy', 'exy', 'e1', 'gamma_max', 'omega_xy']
    PARAM_LABELS = {
        'u': 'u', 'v': 'v', 'ncc': 'ncc',
        'exx': 'exx', 'eyy': 'eyy', 'exy': 'exy',
        'e1': 'e1', 'gamma_max': 'gamma_max', 'omega_xy': 'omega_xy',
    }
    # 平均値シートのヘッダー表示用（特殊文字OK）
    PARAM_HEADERS = {
        'u': 'u [px]', 'v': 'v [px]', 'ncc': 'NCC',
        'exx': 'εxx', 'eyy': 'εyy', 'exy': 'εxy',
        'e1': 'ε1', 'gamma_max': 'γmax', 'omega_xy': 'ωxy',
    }

    wb = openpyxl.Workbook()

    # ヘッダースタイル
    hdr_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
    hdr_font = Font(bold=True)

    def _write_header(ws, row, values):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal='center')

    # =========================================================
    # シート1：平均値
    # =========================================================
    ws_avg = wb.active
    ws_avg.title = '平均値'

    headers = ['段階'] + [PARAM_HEADERS.get(p, p) for p in PARAMS]
    _write_header(ws_avg, 1, headers)

    for res in results_list:
        row_vals = [res['label']]
        for p in PARAMS:
            if p in ('u', 'v', 'ncc'):
                data = np.array(res[p], dtype=float).flatten()
            else:
                d = res['strain'].get(p)
                data = np.array(d, dtype=float).flatten() if d is not None else None

            if data is None:
                row_vals.append('')
                continue

            arr = np.array(data, dtype=float)
            # min/max範囲外をNaN化して平均から除外
            vmin, vmax = scale_config.get(p, (None, None))
            if vmin is not None:
                arr = np.where(arr < vmin, np.nan, arr)
            if vmax is not None:
                arr = np.where(arr > vmax, np.nan, arr)

            valid = arr[~np.isnan(arr)]
            row_vals.append(float(np.mean(valid)) if len(valid) > 0 else '')

        ws_avg.append(row_vals)

    # 列幅調整
    ws_avg.column_dimensions['A'].width = 16
    for col in range(2, len(headers) + 1):
        ws_avg.column_dimensions[
            openpyxl.utils.get_column_letter(col)].width = 14

    # =========================================================
    # シート2以降：各パラメータの生データ
    # =========================================================
    # 座標はREFの座標を基準（全DEFで共通）
    # ROIオフセットを加算して元のSEM画像座標系に変換する
    ref_res = results_list[0]
    cx_base = np.array(ref_res['cx'])
    cy_base = np.array(ref_res['cy'])
    if roi is not None:
        x0_roi, y0_roi = roi[0], roi[1]
        cx_base = cx_base + x0_roi
        cy_base = cy_base + y0_roi
        print(f"  ROIオフセット適用: x+={x0_roi}px, y+={y0_roi}px → 元のSEM画像座標系に変換")

    for p in PARAMS:
        ws = wb.create_sheet(title=PARAM_LABELS.get(p, p))

        col_headers = ['subset_id', 'x [px]', 'y [px]'] + [res['label'] for res in results_list]
        _write_header(ws, 1, col_headers)

        # 各DEFのデータを1Dフラット配列に統一して収集
        arrays = []
        for res in results_list:
            if p in ('u', 'v', 'ncc'):
                arr = np.array(res[p], dtype=float).flatten()
            else:
                d = res['strain'].get(p)
                arr = np.array(d, dtype=float).flatten() if d is not None else None
            arrays.append(arr)

        n_points = len(cx_base)
        for i, (cx, cy) in enumerate(zip(cx_base, cy_base)):
            row = [i + 1, int(cx), int(cy)]  # subset_id は1始まり
            for arr in arrays:
                if arr is None or i >= len(arr):
                    row.append('')
                else:
                    val = arr[i]
                    row.append('' if np.isnan(val) else float(val))
            ws.append(row)

        # 列幅調整
        ws.column_dimensions['A'].width = 12  # subset_id
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        for col in range(4, len(col_headers) + 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = 14

    wb.save(out_path)
    print(f"  → Excel出力完了: {out_path}")


# =============================================================================
# 1ペア分のDIC処理をまとめた関数
# =============================================================================

def run_dic_pair(ref_path, def_path, shifts, roi,
                 ncc_threshold=NCC_THRESHOLD, n_workers=N_WORKERS,
                 prev_cx1=None, prev_cy1=None, prev_u1=None, prev_v1=None,
                 save_png=True):
    """REF・DEF 1ペア分のDIC処理を実行して結果を返す。

    Parameters
    ----------
    prev_cx1, prev_cy1, prev_u1, prev_v1 : array or None
        前段DEFのStage1結果。USE_PREV_STAGE1=Trueかつ2枚目以降で使用。
        Noneの場合はグローバルシフトを初期値とする（1枚目または独立処理）。
    save_png : bool
        Trueの場合は変位・ひずみマップPNGを保存する（デフォルト）。
        Falseの場合は計算のみ行いPNG保存をスキップ（パス1用）。

    Returns
    -------
    cx1, cy1, u1, v1 : array  今回のStage1結果（次のDEFの初期値用）
    """

    def_stem = def_path.stem  # ファイル名（拡張子なし）をsuffixに使用

    print("\n" + "=" * 60)
    print(f"  REF : {ref_path.name}")
    print(f"  DEF : {def_path.name}")
    print(f"  OUT : {OUTPUT_DIR}")
    print("=" * 60)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 画像読み込み + トリミング
    print("\n画像を読み込んでいます...")
    ref_raw      = load_and_preprocess(ref_path, (TRIM_TOP, TRIM_BOTTOM, TRIM_LEFT, TRIM_RIGHT))
    deformed_raw = load_and_preprocess(def_path, (TRIM_TOP, TRIM_BOTTOM, TRIM_LEFT, TRIM_RIGHT))
    print(f"  トリミング後サイズ: {ref_raw.shape[1]} x {ref_raw.shape[0]} px")

    # シフト補正
    ref_aligned      = apply_alignment(ref_raw,      ref_path.name,  shifts, ref_raw.shape)
    deformed_aligned = apply_alignment(deformed_raw, def_path.name,  shifts, deformed_raw.shape)

    # 有効ROIのクロップ
    if roi is not None:
        ref      = crop_roi(ref_aligned,      roi)
        deformed = crop_roi(deformed_aligned, roi)
        print(f"  クロップ後サイズ: {ref.shape[1]} x {ref.shape[0]} px")
    else:
        ref      = ref_aligned
        deformed = deformed_aligned
        print("  アライメント補正なし")

    # グローバルシフト推定（Phase Correlation）
    shift_x, shift_y = estimate_global_shift(ref, deformed)
    print(f"  グローバルシフト推定: x={shift_x}px, y={shift_y}px")

    # Stage1：初期値の決定
    use_prev = (USE_PREV_STAGE1 and
                prev_cx1 is not None and prev_u1 is not None)

    if use_prev:
        # 前段Stage1結果をそのまま初期値に使う（補間不要：同じcoarseグリッド）
        prev_u_dict = {(cx, cy): u for cx, cy, u in zip(prev_cx1, prev_cy1, prev_u1)}
        prev_v_dict = {(cx, cy): v for cx, cy, v in zip(prev_cx1, prev_cy1, prev_v1)}
        _g_init_u = lambda cx, cy: float(prev_u_dict.get((int(cx), int(cy)), shift_x))
        _g_init_v = lambda cx, cy: float(prev_v_dict.get((int(cx), int(cy)), shift_y))

        # search範囲：前段変位を初期値にするので、DEF間の差分分だけ探索すればよい
        stage1_search = STAGE1_MARGIN
        print(f"  Stage1 search = margin({STAGE1_MARGIN}px)  [前段参照]")

        print(f"\n--- Stage 1 (step={STEP_COARSE}px, search={stage1_search}px, "
              f"初期値: 前段Stage1変位場) ---")
    else:
        # グローバルシフトを一律初期値に使う（1枚目または独立処理）
        if STAGE1_AUTO:
            stage1_search = int(max(abs(shift_x), abs(shift_y))) + STAGE1_MARGIN
            print(f"  Stage1 search = |shift|_max({max(abs(shift_x),abs(shift_y))}px)"
                  f" + margin({STAGE1_MARGIN}px) = {stage1_search}px  [自動]")
        else:
            stage1_search = SEARCH_COARSE
            print(f"  Stage1 search = {stage1_search}px  [固定]")

        _g_init_u = lambda cx, cy: float(shift_x)
        _g_init_v = lambda cx, cy: float(shift_y)
        print(f"\n--- Stage 1 (step={STEP_COARSE}px, search={stage1_search}px, "
              f"初期値: x={shift_x}px, y={shift_y}px) ---")
    cx1, cy1, u1, v1, _ = run_dic_stage(
        ref, deformed,
        subset_size=SUBSET_SIZE, subset_step=STEP_COARSE,
        search_range=stage1_search, label="[Stage 1]",
        init_u_func=_g_init_u, init_v_func=_g_init_v,
        ncc_threshold=0,  # Stage1はマスクなし（初期値用なのでNaNにしない）
        n_workers=n_workers
    )
    init_u_func = make_interpolator(cx1, cy1, np.nan_to_num(u1))
    init_v_func = make_interpolator(cx1, cy1, np.nan_to_num(v1))

    # Stage2：Stage1結果を初期値にして精密探索
    print(f"\n--- Stage 2 (step={STEP_FINE}px, search={SEARCH_FINE}px) ---")
    _iu = lambda cx, cy: float(init_u_func([[cy, cx]])[0])
    _iv = lambda cx, cy: float(init_v_func([[cy, cx]])[0])

    cx2, cy2, u2, v2, ncc2 = run_dic_stage(
        ref, deformed,
        subset_size=SUBSET_SIZE, subset_step=STEP_FINE,
        search_range=SEARCH_FINE,
        init_u_func=_iu, init_v_func=_iv,
        label="[Stage 2]",
        ncc_threshold=ncc_threshold, n_workers=n_workers
    )

    # u_corr補正を削除（Phase Correlationによる全体補正は不均一変形に不適切）
    u_corr = u2
    v_corr = v2

    print("\nひずみ場を計算しています...")
    strain = calc_strain_field(cx2, cy2, u_corr, v_corr, STEP_FINE,
                               gauge_length=GAUGE_LENGTH)

    if save_png:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        visualize_displacement(cx2, cy2, u_corr, v_corr,
                               ncc2, shift_x, shift_y, STEP_FINE,
                               ncc_threshold=ncc_threshold, def_stem=def_stem,
                               scale_config=SCALE_CONFIG)
        visualize_strain(strain, STEP_FINE, ref_path.name, def_path.name,
                         def_stem=def_stem, scale_config=SCALE_CONFIG)
        print(f"\n  → 保存完了: displacement_map_{def_stem}.png / strain_map_{def_stem}.png")
    else:
        print(f"  → 計算完了（PNG保存スキップ: {def_stem})")

    return cx1, cy1, u1, v1, {
        'label': def_path.stem,
        'cx': cx2, 'cy': cy2,
        'u': u_corr, 'v': v_corr,
        'ncc': ncc2,
        'strain': strain,
    }


# =============================================================================
# メイン処理
# =============================================================================

if __name__ == "__main__":
    import sys as _sys
    _initial_dir = _sys.argv[1] if len(_sys.argv) > 1 else None

    # GUIウィザードで設定を取得
    REF_PATH, DEF_PATHS, ALIGNMENT_JSON, BASE_FOLDER, \
        STEP_COARSE, STAGE1_AUTO, STAGE1_MARGIN, SEARCH_COARSE, \
        STEP_FINE, SEARCH_FINE, SUBSET_SIZE, \
        NCC_THRESHOLD, N_WORKERS, SCALE_CONFIG, GAUGE_LENGTH, USE_PREV_STAGE1, \
        TRIM = interactive_wizard(_initial_dir)
    # TRIM = (top, bottom, left, right)
    TRIM_TOP, TRIM_BOTTOM, TRIM_LEFT, TRIM_RIGHT = TRIM

    # アライメント情報の読み込み
    print("\nアライメント情報を読み込んでいます...")
    shifts = load_alignment(ALIGNMENT_JSON)

    # 有効ROIの計算（全DEFを通じて共通）
    if shifts:
        # ROI計算用にトリミング後のREFサイズを取得
        _ref_pre = load_and_preprocess(REF_PATH, TRIM)
        _h, _w = _ref_pre.shape[:2]
        del _ref_pre
        roi = calc_valid_roi(shifts, (_h, _w))
    else:
        roi = None

    # DEFをループして順番に処理
    total = len(DEF_PATHS)
    print(f"\n処理開始: {total}ペアのDICを実行します")
    print(f"  REF: {REF_PATH.name}")

    # 出力フォルダを1つに統一（REF名ベース）
    OUTPUT_DIR = BASE_FOLDER / f"dic_{REF_PATH.stem}"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"  出力先: {OUTPUT_DIR}")

    # 前回実行の残留PNGを全削除（今回の結果のみを表示するため）
    _old_pngs = list(OUTPUT_DIR.glob("*.png"))
    if _old_pngs:
        for _p in _old_pngs:
            _p.unlink()
        print(f"  前回の残留PNG {len(_old_pngs)}件を削除しました")

    # ── 事前スキャン ──────────────────────────────────────────────
    # 最後のDEF（最大変形）に対して3×3点で粗い探索を行い
    # Stage1 search範囲の目安を提示する
    # ※最大変形で確認することで search 範囲の過小評価を防ぐ
    prescan_def_path = DEF_PATHS[-1]
    print(f"\n  事前スキャン対象: {prescan_def_path.name}（リスト最後 = 最大変形）")

    _ref_pre      = load_and_preprocess(REF_PATH,           TRIM)
    _def_pre      = load_and_preprocess(prescan_def_path,   TRIM)
    _ref_pre_al   = apply_alignment(_ref_pre,     REF_PATH.name,          shifts, _ref_pre.shape)
    _def_pre_al   = apply_alignment(_def_pre,     prescan_def_path.name,  shifts, _def_pre.shape)
    if roi is not None:
        _ref_pre_cr = crop_roi(_ref_pre_al,  roi)
        _def_pre_cr = crop_roi(_def_pre_al,  roi)
    else:
        _ref_pre_cr = _ref_pre_al
        _def_pre_cr = _def_pre_al

    _, rec_search, _ = prescan(
        _ref_pre_cr, _def_pre_cr,
        subset_size=SUBSET_SIZE,
        search_range=PRESCAN_SEARCH,
        grid=PRESCAN_GRID,
    )
    print(f"\n  ※ 現在の設定: Stage2 search={SEARCH_FINE}px")
    if rec_search > SEARCH_FINE:
        print(f"  ⚠ 推奨値({rec_search}px) > 現在のStage2 search({SEARCH_FINE}px)")
        print(f"    GUIまたはデフォルト値を見直してください。")
    else:
        print(f"  ✓ 現在のStage2 search={SEARCH_FINE}px は推奨値({rec_search}px)以上です。")
    del _ref_pre, _def_pre, _ref_pre_al, _def_pre_al, _ref_pre_cr, _def_pre_cr

    # ── REF基準マップの座標・ゼロデータを準備（PNG保存は後で統一スケールで行う） ──
    _ref_img = load_and_preprocess(REF_PATH, TRIM)
    _ref_al  = apply_alignment(_ref_img, REF_PATH.name, shifts, _ref_img.shape)
    _ref_cr  = crop_roi(_ref_al, roi) if roi is not None else _ref_al
    _h, _w   = _ref_cr.shape
    half     = SUBSET_SIZE // 2
    margin   = half + SEARCH_FINE + 5
    _xs = np.arange(margin, _w - margin, STEP_FINE)
    _ys = np.arange(margin, _h - margin, STEP_FINE)
    _cx = np.array([x for _ in _ys for x in _xs])
    _cy = np.array([y for y in _ys for _ in _xs])
    _zeros = np.zeros(len(_cx))
    _ones  = np.ones(len(_cx))
    ref_stem = REF_PATH.stem
    print(f"  REFグリッド: {len(_cx)}点")
    _ref_strain = calc_strain_field(_cx, _cy, _zeros, _zeros, STEP_FINE,
                                    gauge_length=GAUGE_LENGTH)
    del _ref_img, _ref_al, _ref_cr

    # results_listにREFを先頭として追加
    results_list = [{
        'label': REF_PATH.stem,
        'cx': _cx, 'cy': _cy,
        'u': _zeros.copy(), 'v': _zeros.copy(),
        'ncc': _ones.copy(),
        'strain': _ref_strain,
    }]

    # ── 設定条件を出力 ────────────────────────────────────────────
    config_lines = [
        'DIC解析 設定条件',
        '=' * 40,
        f'REF          : {REF_PATH.name}',
        f'DEF          : {", ".join(p.name for p in DEF_PATHS)}',
        f'alignment file: {ALIGNMENT_JSON.name if ALIGNMENT_JSON else "なし"}',
        f'trim         : 上{TRIM_TOP} 下{TRIM_BOTTOM} 左{TRIM_LEFT} 右{TRIM_RIGHT} px',
        '',
        '[Stage 1]',
        f'  step       : {STEP_COARSE} px',
        f'  search     : {"グローバルシフト + " + str(STAGE1_MARGIN) + " px（自動）" if STAGE1_AUTO else str(SEARCH_COARSE) + " px（固定）"}',
        f'  前段参照   : {"あり（高速化）" if USE_PREV_STAGE1 else "なし（独立処理）"}',
        '',
        '[Stage 2]',
        f'  step       : {STEP_FINE} px',
        f'  search     : {SEARCH_FINE} px',
        '',
        '[共通]',
        f'  subset     : {SUBSET_SIZE} px',
        f'  gauge      : {GAUGE_LENGTH} 倍（Step_fine × {GAUGE_LENGTH} = {GAUGE_LENGTH * STEP_FINE}px）',
        f'  NCC閾値    : {NCC_THRESHOLD}',
        f'  workers    : {N_WORKERS}',
        '',
        '[カラースケール]',
    ]
    for k, (lo, hi) in SCALE_CONFIG.items():
        lo_str = f'{lo}' if lo is not None else '自動'
        hi_str = f'{hi}' if hi is not None else '自動'
        config_lines.append(f'  {k:<12}: min={lo_str}  max={hi_str}')

    config_text = '\n'.join(config_lines)
    config_path = OUTPUT_DIR / 'dic_config.txt'
    with open(config_path, 'w', encoding='utf-8') as f:
        f.write(config_text + '\n')
    print(f"  設定条件を保存しました: {config_path.name}")

    # ── パス1：全DEFのDIC計算（PNG保存なし） ─────────────────────
    print(f"\n{'─' * 60}")
    print(f"  [パス1] 全DEFのDIC計算（PNG保存なし、{total}ペア）")
    print(f"{'─' * 60}")
    prev_cx1 = prev_cy1 = prev_u1 = prev_v1 = None
    for i, def_path in enumerate(DEF_PATHS, 1):
        print(f"\n[{i}/{total}] {def_path.name} を処理中...")
        prev_cx1, prev_cy1, prev_u1, prev_v1, res = run_dic_pair(
            REF_PATH, def_path, shifts, roi,
            ncc_threshold=NCC_THRESHOLD, n_workers=N_WORKERS,
            prev_cx1=prev_cx1, prev_cy1=prev_cy1,
            prev_u1=prev_u1,   prev_v1=prev_v1,
            save_png=False,
        )
        results_list.append(res)

    # ── カラースケール自動統一 ────────────────────────────────────
    # GUI入力済みキー → そのまま使用
    # GUI未入力（None）キー → 全DEF（REF含む）の実データのパーセンタイルで統一
    SCALE_KEYS_DISP   = ['u', 'v']
    SCALE_KEYS_SYM    = ['exx', 'eyy', 'exy', 'omega_xy']   # 対称スケール（±vmax）
    SCALE_KEYS_ASYM   = ['e1', 'gamma_max']                  # 非対称スケール

    def _collect(key):
        """全results_listから指定キーの有効値を結合して返す。"""
        arrays = []
        for res in results_list:
            if key in ('u', 'v', 'ncc'):
                arr = np.array(res[key], dtype=float).flatten()
            else:
                d = res['strain'].get(key)
                arr = np.array(d, dtype=float).flatten() if d is not None else np.array([])
            valid = arr[~np.isnan(arr)]
            if len(valid) > 0:
                arrays.append(valid)
        return np.concatenate(arrays) if arrays else np.array([])

    unified_scale = {}
    for k in SCALE_KEYS_DISP + SCALE_KEYS_SYM + SCALE_KEYS_ASYM:
        gui_lo, gui_hi = SCALE_CONFIG.get(k, (None, None))
        if gui_lo is not None and gui_hi is not None:
            unified_scale[k] = (gui_lo, gui_hi)
            continue
        vals = _collect(k)
        if len(vals) == 0:
            unified_scale[k] = (gui_lo, gui_hi)
            continue
        if k in SCALE_KEYS_DISP:
            # u/v：絶対値95パーセンタイルで対称スケール
            vabs = max(float(np.percentile(np.abs(vals), 95)), 1.0)
            lo = gui_lo if gui_lo is not None else -vabs
            hi = gui_hi if gui_hi is not None else  vabs
        elif k in SCALE_KEYS_SYM:
            # ひずみ対称系：絶対値98パーセンタイルで対称スケール
            vabs = max(float(np.percentile(np.abs(vals), 98)), 1e-6)
            lo = gui_lo if gui_lo is not None else -vabs
            hi = gui_hi if gui_hi is not None else  vabs
        else:
            # 非対称系（e1, gamma_max）：2～98パーセンタイル
            lo = gui_lo if gui_lo is not None else float(np.percentile(vals, 2))
            hi = gui_hi if gui_hi is not None else float(np.percentile(vals, 98))
        unified_scale[k] = (lo, hi)
    unified_scale['ncc'] = SCALE_CONFIG.get('ncc', (None, None))

    print(f"\n{'─' * 60}")
    print("  【カラースケール（全DEF統一）】")
    for k, (lo, hi) in unified_scale.items():
        gui_lo, gui_hi = SCALE_CONFIG.get(k, (None, None))
        src = '(GUI入力)' if (gui_lo is not None and gui_hi is not None) else '(自動集計)'
        lo_s = f'{lo:.5f}' if lo is not None else '自動'
        hi_s = f'{hi:.5f}' if hi is not None else '自動'
        print(f"    {k:<12}: min={lo_s}  max={hi_s}  {src}")
    print(f"{'─' * 60}")

    # ── パス2：統一スケールでPNG保存 ─────────────────────────────
    print(f"\n  [パス2] 統一スケールでPNG保存")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # REF（全ゼロ）
    visualize_displacement(_cx, _cy, _zeros, _zeros,
                           _ones, 0, 0, STEP_FINE,
                           ncc_threshold=NCC_THRESHOLD, def_stem=ref_stem,
                           scale_config=unified_scale)
    visualize_strain(_ref_strain, STEP_FINE, REF_PATH.name, REF_PATH.name,
                     def_stem=ref_stem, scale_config=unified_scale)
    print(f"  → 保存完了: displacement_map_{ref_stem}.png / strain_map_{ref_stem}.png")

    # 各DEF
    for i, (def_path, res) in enumerate(zip(DEF_PATHS, results_list[1:]), 1):
        def_stem_i = def_path.stem
        print(f"  [{i}/{total}] PNG保存: {def_path.name}")
        # shift_x/shift_yはrun_dic_pairの内部変数なので、ここでは0を渡す
        # （グローバルシフトはタイトル表示用のみでスケールに影響しない）
        _shift_x, _shift_y = 0, 0
        visualize_displacement(res['cx'], res['cy'], res['u'], res['v'],
                               res['ncc'], _shift_x, _shift_y, STEP_FINE,
                               ncc_threshold=NCC_THRESHOLD, def_stem=def_stem_i,
                               scale_config=unified_scale)
        visualize_strain(res['strain'], STEP_FINE, REF_PATH.name, def_path.name,
                         def_stem=def_stem_i, scale_config=unified_scale)
        print(f"     → displacement_map_{def_stem_i}.png / strain_map_{def_stem_i}.png")

    # ── Excel出力 ──────────────────────────────────────────────────
    xlsx_path = OUTPUT_DIR / 'dic_results.xlsx'
    print(f"\nExcelファイルを出力しています...")
    export_xlsx(results_list, unified_scale, xlsx_path, roi=roi)

    print(f"\n{'=' * 60}")
    print(f"  全{total}ペアの処理が完了しました！")
    print(f"  出力先: {OUTPUT_DIR}")
    print(f"{'=' * 60}")

    # 全計算完了後にまとめてマップを表示（今回処理した分のみ）
    print("\n  マップを表示しています...")
    # 前回分は削除済みなのでフォルダ内の全PNGを表示
    saved_pngs = sorted(OUTPUT_DIR.glob("*.png"))
    for png in saved_pngs:
        img = plt.imread(str(png))
        fig, ax = plt.subplots(figsize=(min(img.shape[1]/100, 18),
                                        min(img.shape[0]/100, 10)))
        ax.imshow(img)
        ax.axis('off')
        ax.set_title(png.name, fontsize=10)
        plt.tight_layout()
    plt.show()
