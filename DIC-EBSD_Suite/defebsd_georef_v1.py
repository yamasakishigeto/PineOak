"""
defebsd_georef_v1.py
====================
Def_nth EBSD データのジオリファレンス（変形後DIC座標系への位置合わせ）。

処理フロー:
  各変形段階について以下を繰り返す:
  1. dic_results_georef.xlsx から変形後DIC Grain IDマップを生成
  2. Def_nth EBSD ファイル（.mat or .txt）から EBSD Grain IDマップを生成
  3. 左（変形後DIC）→ 右（EBSD）でコントロールポイントを指定
  4. 各EBSDグリッド点に最近傍DICサブセットの subset_id を割り当て
  5. 出力:
     - .mat モード: subset_id + 2D全フィールドを抽出して _georef.mat として保存
     - .txt モード: dic_results_georef.xlsx に新シートを追加

座標系:
  DIC :  SEM画像ピクセル座標 (cx, cy) [px]
  EBSD:  物理座標 (xpos, ypos) [μm]、Grain IDマップ画像座標 (ix, iy) [px]

変換:
  H (2×3 アフィン行列):  変形後DICキャンバスピクセル → EBSDキャンバスピクセル
  H_inv               :  EBSDキャンバスピクセル → 変形後DICキャンバスピクセル
"""

import sys
import os
import json
import numpy as np
from pathlib import Path
from collections import deque

# 非日本語環境でのUnicodeEncodeError対策
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

os.environ['MPLBACKEND'] = 'QtAgg'
import matplotlib
matplotlib.use('QtAgg')

# ── 必須ライブラリチェック ────────────────────────────────────────────────
_REQUIRED = {
    'numpy':      'numpy',
    'scipy':      'scipy',
    'matplotlib': 'matplotlib',
    'openpyxl':   'openpyxl',
    'cv2':        'opencv-python',
}
_missing = []
for _mod, _pkg in _REQUIRED.items():
    try:
        __import__(_mod)
    except ImportError:
        _missing.append(_pkg)
if _missing:
    print("=" * 60)
    print("  [ERROR] Required libraries not found.")
    print(f"  pip install {' '.join(_missing)}")
    print("=" * 60)
    sys.exit(1)

import cv2
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.font_manager as _fm
import platform as _platform
from scipy.spatial import cKDTree

# ── 日本語フォント設定 ────────────────────────────────────────────────────
_JP_FONTS = {
    'Windows': ['Meiryo', 'Yu Gothic', 'MS Gothic'],
    'Darwin':  ['Hiragino Sans', 'Hiragino Maru Gothic Pro', 'AppleGothic'],
    'Linux':   ['IPAexGothic', 'IPAPGothic', 'Noto Sans CJK JP'],
}
_available_fonts = {f.name for f in _fm.fontManager.ttflist}
_jp_candidates   = _JP_FONTS.get(_platform.system(), [])
_jp_font = next((f for f in _jp_candidates if f in _available_fonts), None)
if _jp_font:
    matplotlib.rcParams['font.family'] = _jp_font
matplotlib.rcParams['axes.unicode_minus'] = False


# =============================================================================
# 変数名を MATLAB/scipy.io.savemat 互換の識別子に変換
# =============================================================================
def _mat_safe_name(s):
    """
    文字列を scipy.io.savemat で使える変数名に変換する。
    例: '750MPa' → 's750MPa',  'gamma_max' → 'gamma_max'
    """
    import re
    s = re.sub(r'[^a-zA-Z0-9]', '_', str(s)).strip('_')
    if s and s[0].isdigit():
        s = 's' + s
    return s or 'unnamed'


# =============================================================================
# dic_results_georef.xlsx から変形後DICデータを読み込む
# =============================================================================
def load_deformed_dic_grain_map(georef_xlsx, label):
    """
    dic_results_georef.xlsx から指定ラベルの変形後 DIC Grain ID マップを読み込む。

    Returns
    -------
    cx_def     : ndarray (N,)  変形後X座標 [px]
    cy_def     : ndarray (N,)  変形後Y座標 [px]
    grain_ids  : ndarray (N,)  Grain ID（>0のみ）
    subset_ids : ndarray (N,)  subset_id
    """
    wb = openpyxl.load_workbook(georef_xlsx, read_only=True)

    # ebsd_georef シート: subset_id → grain_id
    rows_eg = list(wb['ebsd_georef'].iter_rows(values_only=True))
    hdr_eg  = rows_eg[0]
    sid_col_eg = next(i for i, h in enumerate(hdr_eg) if h == 'subset_id')
    gid_col_eg = next(i for i, h in enumerate(hdr_eg) if h == 'grain_id')
    sid_to_gid = {int(r[sid_col_eg]): int(r[gid_col_eg])
                  for r in rows_eg[1:] if r[sid_col_eg] is not None}

    # u シート: subset_id, x [px], u[label]
    rows_u  = list(wb['u'].iter_rows(values_only=True))
    hdr_u   = rows_u[0]
    sid_col_u  = next(i for i, h in enumerate(hdr_u) if h == 'subset_id')
    xref_col   = next(i for i, h in enumerate(hdr_u) if h == 'x [px]')
    try:
        u_col = next(i for i, h in enumerate(hdr_u) if h == label)
    except StopIteration:
        wb.close()
        raise ValueError(f"Label '{label}' not found in 'u' sheet of {georef_xlsx}")

    # v シート: subset_id, y [px], v[label]
    rows_v  = list(wb['v'].iter_rows(values_only=True))
    hdr_v   = rows_v[0]
    sid_col_v  = next(i for i, h in enumerate(hdr_v) if h == 'subset_id')
    yref_col   = next(i for i, h in enumerate(hdr_v) if h == 'y [px]')
    try:
        v_col = next(i for i, h in enumerate(hdr_v) if h == label)
    except StopIteration:
        wb.close()
        raise ValueError(f"Label '{label}' not found in 'v' sheet of {georef_xlsx}")

    wb.close()

    # v シートから sid → y_def を構築
    sid_to_ydef = {}
    for r in rows_v[1:]:
        if r[sid_col_v] is None:
            continue
        sid  = int(r[sid_col_v])
        yref = float(r[yref_col]) if r[yref_col] is not None else None
        v    = float(r[v_col])    if r[v_col]    is not None else 0.0
        if yref is not None:
            sid_to_ydef[sid] = yref + v

    # u シートを走査して cx_def, cy_def, grain_id を収集
    subset_ids_list = []
    cx_def_list     = []
    cy_def_list     = []
    grain_ids_list  = []

    for r in rows_u[1:]:
        if r[sid_col_u] is None:
            continue
        sid  = int(r[sid_col_u])
        gid  = sid_to_gid.get(sid, 0)
        if gid <= 0:
            continue
        xref = float(r[xref_col]) if r[xref_col] is not None else None
        u    = float(r[u_col])    if r[u_col]    is not None else 0.0
        if xref is None or sid not in sid_to_ydef:
            continue
        subset_ids_list.append(sid)
        cx_def_list.append(xref + u)
        cy_def_list.append(sid_to_ydef[sid])
        grain_ids_list.append(gid)

    cx_def     = np.array(cx_def_list,    dtype=float)
    cy_def     = np.array(cy_def_list,    dtype=float)
    grain_ids  = np.array(grain_ids_list, dtype=int)
    subset_ids = np.array(subset_ids_list, dtype=int)

    print(f"  Deformed DIC subsets loaded: {len(subset_ids)}  (label='{label}')")
    return cx_def, cy_def, grain_ids, subset_ids


def get_available_labels(georef_xlsx):
    """dic_results_georef.xlsx の u シートから利用可能な変形段階ラベルを返す。"""
    wb   = openpyxl.load_workbook(georef_xlsx, read_only=True)
    hdr  = list(wb['u'].iter_rows(values_only=True, max_row=1))[0]
    wb.close()
    skip = {None, 'subset_id', 'x [px]', 'y [px]'}
    return [str(h) for h in hdr if h not in skip]


# =============================================================================
# 変形後 DIC Grain ID マップを画像（numpy RGB）に描画
# =============================================================================
def render_deformed_grain_image(cx_def, cy_def, grain_ids, step=None):
    """
    変形後DICサブセット座標 (cx_def, cy_def) をGrain IDで色付けし、
    numpy RGB画像（H×W×3, uint8）として返す。

    画像座標 (col, row) と DIC実座標の対応:
        DIC_x = col + x_offset
        DIC_y = row + y_offset

    Returns
    -------
    canvas   : ndarray (H, W, 3) uint8
    x_offset : int  DIC実座標のX最小値（マージン込み）
    y_offset : int  DIC実座標のY最小値（マージン込み）
    step_est : float  推定DICステップ [px]
    """
    if step is None:
        # 点の密度から推定（大変形時でも安定）
        # 変形後の座標は非等間隔になるため、x 差分の中央値は過小推定になりやすい
        x_range = float(cx_def.max() - cx_def.min())
        y_range = float(cy_def.max() - cy_def.min())
        n = len(cx_def)
        if n > 1 and x_range > 0 and y_range > 0:
            step_est = float(np.sqrt(x_range * y_range / n))
        else:
            step_est = 15.0
    else:
        step_est = float(step)

    half   = max(1, int(step_est // 2))
    margin = int(step_est * 3)

    x_offset = max(0, int(cx_def.min()) - margin)
    y_offset = max(0, int(cy_def.min()) - margin)
    W = int(cx_def.max()) + margin - x_offset + 1
    H = int(cy_def.max()) + margin - y_offset + 1

    # Grain IDごとのカラーパレット
    valid_ids = sorted(set(int(g) for g in grain_ids if g > 0))
    pool = []
    for _name in ('tab20', 'tab20b', 'tab20c'):
        _c = matplotlib.colormaps[_name]
        pool.extend([_c(i) for i in range(20)])
    idxs = np.linspace(0, len(pool) - 1, max(len(valid_ids), 1), dtype=int)
    gid_to_color = {
        gid: (np.array(pool[idxs[i]][:3]) * 255).astype(np.uint8)
        for i, gid in enumerate(valid_ids)
    }

    canvas = np.zeros((H, W, 3), dtype=np.uint8)
    for cx, cy, gid in zip(cx_def, cy_def, grain_ids):
        if gid <= 0:
            continue
        color = gid_to_color.get(int(gid))
        if color is None:
            continue
        col = int(round(cx)) - x_offset
        row = int(round(cy)) - y_offset
        r0, r1 = max(0, row - half), min(H, row + half + 1)
        c0, c1 = max(0, col - half), min(W, col + half + 1)
        canvas[r0:r1, c0:c1] = color

    return canvas, x_offset, y_offset, step_est


# =============================================================================
# EBSD Grain File (.txt) 読み込み（ebsd_georef_v68 から流用）
# =============================================================================
def load_grain_file(path):
    """OIM Analysis Grain File（固定11列）を読み込む。"""
    num_rows   = []
    phase_col  = []
    grain_tol_angle = None

    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            line_s = line.strip()
            if not line_s:
                continue
            if line_s.startswith('#'):
                if 'Grain Tolerance Angle' in line_s:
                    try:
                        grain_tol_angle = float(line_s.split(':')[-1].strip())
                    except ValueError:
                        pass
                continue
            parts = line_s.split()
            if len(parts) < 11:
                continue
            try:
                num_rows.append([float(p) for p in parts[:10]])
                phase_col.append(parts[10])
            except ValueError:
                continue

    if not num_rows:
        raise ValueError(f"No data rows found: {path}")

    data        = np.array(num_rows, dtype=np.float64)
    phase_names = np.array(phase_col)

    x_vals  = np.unique(np.round(data[:, 3], 6))
    y_vals  = np.unique(np.round(data[:, 4], 6))
    nx      = len(x_vals)
    ny      = len(y_vals)
    x_step  = float(np.median(np.diff(x_vals))) if nx > 1 else 1.0
    y_step  = float(np.median(np.diff(y_vals))) if ny > 1 else 1.0
    x_min, x_max = float(x_vals.min()), float(x_vals.max())
    y_min, y_max = float(y_vals.min()), float(y_vals.max())

    unique_phases   = sorted(set(phase_col))
    grain_ids       = data[:, 8].astype(int)
    n_valid_grains  = len(np.unique(grain_ids[grain_ids > 0]))

    print(f"  Grain File loaded: {len(data)} pixels  ({nx} x {ny})")
    print(f"  Step: x={x_step:.4f} um, y={y_step:.4f} um")
    print(f"  Valid grains (GrainID>0): {n_valid_grains}")

    meta = dict(
        nx=nx, ny=ny,
        x_step=x_step, y_step=y_step,
        x_min=x_min, y_min=y_min,
        x_max=x_max, y_max=y_max,
        grain_tol_angle=grain_tol_angle,
        phases=unique_phases,
    )
    return data, phase_names, meta


# =============================================================================
# .mat ファイル読み込み（ジオリファレンス用最小フィールド）
# =============================================================================
def load_mat_minimal(mat_path):
    """
    .mat から grain_number / xpos / ypos / euler角 / metadata のみ読み込む。
    大型配列（3次元以上）は読み込まない。
    """
    import scipy.io

    GEOREF_VARS = [
        'grain_number', 'xpos', 'ypos', 'xstep', 'ystep', 'phasetxt',
        'euler_phi1', 'euler_phi', 'euler_phi2',
        'numcols', 'numrows',
    ]
    mat = scipy.io.loadmat(mat_path, variable_names=GEOREF_VARS)

    ny, nx = mat['grain_number'].shape
    xpos   = mat['xpos'].flatten()
    ypos   = mat['ypos'].flatten()
    x_step = float(mat['xstep'][0, 0])
    y_step = float(mat['ystep'][0, 0])
    x_min, x_max = float(xpos.min()), float(xpos.max())
    y_min, y_max = float(ypos.min()), float(ypos.max())

    try:
        phase_name = str(mat['phasetxt'][0, 0][0])
    except Exception:
        phase_name = 'unknown'

    gid_raw = mat['grain_number'].flatten()
    gid     = np.where(np.isnan(gid_raw), 0.0, gid_raw)
    zeros   = np.zeros(len(gid), dtype=np.float64)

    phi1 = np.radians(mat['euler_phi1'].flatten())
    phi  = np.radians(mat['euler_phi'].flatten())
    phi2 = np.radians(mat['euler_phi2'].flatten())

    data = np.column_stack([phi1, phi, phi2, xpos, ypos,
                            zeros, zeros, zeros, gid, zeros])
    phase_names = np.array([phase_name] * len(data))

    print(f"  .mat loaded (minimal): {len(data)} pixels  ({nx} x {ny})")
    print(f"  Step: x={x_step:.4f} um, y={y_step:.4f} um")

    meta = dict(
        nx=nx, ny=ny,
        x_step=x_step, y_step=y_step,
        x_min=x_min, y_min=y_min,
        x_max=x_max, y_max=y_max,
        grain_tol_angle=None,
        phases=[phase_name],
    )
    return data, phase_names, meta


# =============================================================================
# EBSD Grain ID マップ画像生成（ebsd_georef_v68 から流用）
# =============================================================================
def make_grain_id_image(data, meta):
    """Grain ID からカラー RGB 画像（ny × nx × 3, uint8）を生成する。"""
    nx, ny   = meta['nx'], meta['ny']
    x_min    = meta['x_min']
    y_min    = meta['y_min']
    x_step   = meta['x_step']
    y_step   = meta['y_step']

    grain_id_map = np.zeros((ny, nx), dtype=np.int32)
    for row in data:
        ix = int(round((row[3] - x_min) / x_step))
        iy = int(round((row[4] - y_min) / y_step))
        if 0 <= ix < nx and 0 <= iy < ny:
            grain_id_map[iy, ix] = int(row[8])

    valid_ids = sorted(set(int(v) for v in grain_id_map.flat if v > 0))
    pool = []
    for _name in ('tab20', 'tab20b', 'tab20c'):
        _c = matplotlib.colormaps[_name]
        pool.extend([_c(i) for i in range(20)])
    idxs = np.linspace(0, len(pool) - 1, max(len(valid_ids), 1), dtype=int)
    id_to_color = {
        gid: (np.array(pool[idxs[i]][:3]) * 255).astype(np.uint8)
        for i, gid in enumerate(valid_ids)
    }

    rgb = np.zeros((ny, nx, 3), dtype=np.uint8)
    for iy in range(ny):
        for ix in range(nx):
            gid = grain_id_map[iy, ix]
            if gid > 0:
                rgb[iy, ix] = id_to_color[gid]
            elif gid < 0:
                rgb[iy, ix] = [80, 80, 80]

    return rgb, grain_id_map, id_to_color


# =============================================================================
# コントロールポイント指定 GUI（両パネル RGB 対応版）
# =============================================================================
def pick_control_points(left_img, right_img, left_label='', right_label='',
                        ctrl_pts_path=None):
    """
    左（変形後DIC Grain IDマップ）と右（EBSD Grain IDマップ）を並べて表示し、
    左→右の順に交互クリックでコントロールポイントを指定する。

    左パネルは RGB 画像（render_deformed_grain_image の出力）を想定。

    Parameters
    ----------
    left_label     : str        左パネルのデータ識別名（DIC ラベル）
    right_label    : str        右パネルのデータ識別名（EBSD ファイル名など）
    ctrl_pts_path  : str|Path   コントロールポイントの保存/読込先 JSON パス

    右クリック挙動:
      左パネル上: 最近傍ペアをまとめて削除
      右パネル上: 最近傍の右点のみ削除（対応する左点は保留状態で残る）

    Returns
    -------
    pts_left  : ndarray (N, 2)  左パネル画像座標 [px]
    pts_right : ndarray (N, 2)  右パネル画像座標 [px]
    """
    import json as _json
    from matplotlib.widgets import Button

    _C = {
        'bg': '#0a0c0f', 'surface': '#111318', 'surface2': '#1a1d24',
        'border': '#2a2d35', 'accent': '#00d4ff', 'green': '#00ff88',
        'orange': '#ff6b35', 'text': '#e0e4ec', 'dim': '#6b7280',
        'yellow': '#ffd700',
    }

    H_L, W_L = left_img.shape[:2]
    H_R, W_R = right_img.shape[:2]

    plt.style.use('dark_background')
    fig, axes = plt.subplots(1, 2, figsize=(22, 11.5))
    fig.patch.set_facecolor(_C['bg'])
    for ax in axes:
        ax.set_facecolor(_C['surface'])
        for sp in ax.spines.values():
            sp.set_edgecolor(_C['border'])
        ax.tick_params(colors=_C['dim'], labelsize=11)
    fig.subplots_adjust(bottom=0.15, top=0.88, left=0.04, right=0.97, wspace=0.08)

    _win_left  = f'Deformed DIC  [{left_label}]'  if left_label  else 'Deformed DIC Grain ID'
    _win_right = f'EBSD  [{right_label}]'          if right_label else 'EBSD Grain ID'
    fig.canvas.manager.set_window_title(
        f'Def EBSD Georef  |  {_win_left} (L) → {_win_right} (R)'
        '  |  L-Rclick: undo pair  R-Rclick: undo right only  |  q: confirm')

    ax_left, ax_right = axes

    # 両パネルとも RGB 表示
    ax_left.imshow(left_img,  origin='upper')
    ax_right.imshow(right_img, origin='upper')

    _title_left  = f'Deformed DIC Grain ID  —  {left_label}\n[ click here first ]'  if left_label  else 'Deformed DIC Grain ID  [ click here first ]'
    _title_right = f'EBSD Grain ID  —  {right_label}\n[ click here second ]'         if right_label else 'EBSD Grain ID map  [ click here second ]'
    ax_left.set_title(_title_left,  fontsize=12, color=_C['accent'], pad=8)
    ax_right.set_title(_title_right, fontsize=12, color=_C['dim'],   pad=8)

    # オーバーレイ（EBSDをワープして左パネルに半透明表示）
    _overlay_im = ax_left.imshow(
        np.zeros((H_L, W_L, 4), dtype=np.uint8),
        origin='upper', alpha=1.0, zorder=2)
    _overlay_im.set_visible(False)
    _OVERLAY_OPACITY = 0.4

    # ボタン共通スタイル
    def _make_btn(rect, label, color):
        ax_b = fig.add_axes(rect)
        ax_b.set_facecolor(_C['surface2'])
        for sp in ax_b.spines.values():
            sp.set_edgecolor(color)
            sp.set_linewidth(1.2)
        btn = Button(ax_b, label, color=_C['surface2'], hovercolor=_C['surface'])
        btn.label.set_fontsize(11)
        btn.label.set_color(color)
        btn.label.set_fontfamily('monospace')
        return btn

    _btn_load    = _make_btn([0.12, 0.025, 0.14, 0.055], 'Load pts',    _C['orange'])
    _btn_overlay = _make_btn([0.29, 0.025, 0.14, 0.055], 'Overlay: ON', _C['accent'])
    _btn_done    = _make_btn([0.68, 0.025, 0.14, 0.055], 'Confirm',     _C['green'])

    pts_left  = []
    pts_right = []
    artists_left  = []   # 確定ペア分のみ追跡
    artists_right = []
    # 保留中（左クリック済み・右クリック待ち）の左点アーティスト
    state = {'next': 0, 'overlay_visible': True, 'pending_artist': None}

    # ── 保留左点の描画/消去 ──────────────────────────────────────────────────
    def _clear_pending():
        if state['pending_artist'] is not None:
            for a in state['pending_artist']:
                try:
                    a.remove()
                except Exception:
                    pass
            state['pending_artist'] = None

    def _draw_pending():
        """pts_left に未ペアの点がある場合、黄色マーカーで表示する。"""
        _clear_pending()
        if len(pts_left) > len(pts_right):
            i  = len(pts_left)          # このペア番号
            pl = pts_left[-1]
            mk, = ax_left.plot(pl[0], pl[1], '+', color=_C['yellow'],
                               markersize=16, markeredgewidth=2.5, zorder=5)
            lb  = ax_left.annotate(str(i), (pl[0], pl[1]),
                                   color=_C['yellow'], fontsize=14,
                                   fontweight='bold', xytext=(6, 6),
                                   textcoords='offset points', zorder=5)
            state['pending_artist'] = [mk, lb]

    # ── 確定ペアの再描画 ─────────────────────────────────────────────────────
    def _redraw_labels():
        _clear_pending()
        for a_l, a_r in zip(artists_left, artists_right):
            for a in a_l + a_r:
                try:
                    a.remove()
                except Exception:
                    pass
        artists_left.clear()
        artists_right.clear()
        for i, (pl, pr) in enumerate(zip(pts_left, pts_right), start=1):
            mk_l, = ax_left.plot(pl[0],  pl[1],  'r+', markersize=16, markeredgewidth=2.5, zorder=5)
            lb_l  = ax_left.annotate(str(i),  (pl[0], pl[1]),  color='red', fontsize=14,
                                     fontweight='bold', xytext=(6, 6),
                                     textcoords='offset points', zorder=5)
            mk_r, = ax_right.plot(pr[0], pr[1], 'r+', markersize=16, markeredgewidth=2.5, zorder=5)
            lb_r  = ax_right.annotate(str(i), (pr[0], pr[1]), color='red', fontsize=14,
                                      fontweight='bold', xytext=(6, 6),
                                      textcoords='offset points', zorder=5)
            artists_left.append([mk_l, lb_l])
            artists_right.append([mk_r, lb_r])
        _draw_pending()   # 保留左点があれば黄色で表示
        fig.canvas.draw_idle()

    # ── オーバーレイ更新 ─────────────────────────────────────────────────────
    def _update_overlay():
        n = min(len(pts_left), len(pts_right))
        if n < 3 or not state['overlay_visible']:
            _overlay_im.set_visible(False)
            fig.canvas.draw_idle()
            return
        src = np.array(pts_left[:n],  dtype=np.float32)
        dst = np.array(pts_right[:n], dtype=np.float32)
        try:
            if n == 3:
                H_mat = cv2.getAffineTransform(dst[:3], src[:3])
            else:
                H_mat, _ = cv2.estimateAffine2D(dst, src,
                                                method=cv2.RANSAC,
                                                ransacReprojThreshold=10.0)
        except Exception:
            _overlay_im.set_visible(False)
            fig.canvas.draw_idle()
            return
        if H_mat is None:
            _overlay_im.set_visible(False)
            fig.canvas.draw_idle()
            return
        warped = cv2.warpAffine(right_img, H_mat, (W_L, H_L),
                                flags=cv2.INTER_NEAREST,
                                borderMode=cv2.BORDER_CONSTANT, borderValue=(0, 0, 0))
        rgba = np.zeros((H_L, W_L, 4), dtype=np.uint8)
        rgba[:, :, :3] = warped
        is_black = (warped[:, :, 0] == 0) & (warped[:, :, 1] == 0) & (warped[:, :, 2] == 0)
        rgba[:, :, 3] = np.where(is_black, 0, int(_OVERLAY_OPACITY * 255))
        _overlay_im.set_data(rgba)
        _overlay_im.set_visible(True)
        fig.canvas.draw_idle()

    def _update_title():
        n = min(len(pts_left), len(pts_right))
        next_str = 'DIC (left)' if state['next'] == 0 else 'EBSD (right)'
        overlay_str = (f'Overlay: ON (40%)'
                       if n >= 3 and state['overlay_visible']
                       else ('Overlay: OFF' if n >= 3 else 'Overlay: need 3+ pairs'))
        fig.suptitle(
            f'Pairs: {n}  |  Next: click {next_str}  |  4+ pairs recommended  |  {overlay_str}',
            fontsize=12, color=_C['accent'], fontfamily='monospace', y=0.97)
        fig.canvas.draw_idle()

    # ── コントロールポイント 保存/読込 ───────────────────────────────────────
    def _save_ctrl_pts():
        if ctrl_pts_path is None:
            return
        n = min(len(pts_left), len(pts_right))
        data = {
            'label':      left_label,
            'pts_left':   [list(map(float, p)) for p in pts_left[:n]],
            'pts_right':  [list(map(float, p)) for p in pts_right[:n]],
        }
        try:
            with open(ctrl_pts_path, 'w', encoding='utf-8') as f:
                _json.dump(data, f, indent=2)
            print(f"  Ctrl pts saved: {Path(ctrl_pts_path).name}  ({n} pairs)")
        except Exception as e:
            print(f"  Warning: could not save control points: {e}")

    def _load_ctrl_pts(path=None):
        load_path = path or ctrl_pts_path
        if load_path is None or not Path(load_path).exists():
            return False
        try:
            with open(load_path, 'r', encoding='utf-8') as f:
                data = _json.load(f)
            pts_left.clear()
            pts_right.clear()
            pts_left.extend(data['pts_left'])
            pts_right.extend(data['pts_right'])
            state['next'] = 0
            _redraw_labels()
            _update_title()
            _update_overlay()
            print(f"  Ctrl pts loaded: {Path(load_path).name}  ({len(pts_left)} pairs)")
            return True
        except Exception as e:
            print(f"  Warning: could not load control points: {e}")
            return False

    # ── クリックハンドラ ─────────────────────────────────────────────────────
    def on_click(event):
        if event.inaxes is None or event.xdata is None:
            return
        x, y = event.xdata, event.ydata

        if event.button == 3:  # 右クリック
            if event.inaxes is ax_right:
                # ── 右パネル右クリック: 右点のみ削除、左点は保留状態へ ──
                n = len(pts_right)
                if n == 0:
                    return
                dists = [((p[0]-x)**2 + (p[1]-y)**2) for p in pts_right]
                idx   = int(np.argmin(dists))
                saved_left = list(pts_left[idx])   # 対応左点を保存
                pts_left.pop(idx)
                pts_right.pop(idx)
                # 既存の保留左点があれば捨てて、今回の左点を保留に昇格
                if len(pts_left) > len(pts_right):
                    pts_left.pop()
                pts_left.append(saved_left)
                state['next'] = 1
            else:
                # ── 左パネル右クリック: 最近傍ペアをまとめて削除 ──
                n = min(len(pts_left), len(pts_right))
                if n == 0:
                    # 保留左点だけある場合は削除
                    if len(pts_left) > 0:
                        pts_left.pop()
                        state['next'] = 0
                    else:
                        return
                else:
                    dists = [((p[0]-x)**2 + (p[1]-y)**2) for p in pts_left[:n]]
                    idx   = int(np.argmin(dists))
                    pts_left.pop(idx)
                    pts_right.pop(idx)
                    # 保留左点があれば一緒に削除
                    if len(pts_left) > len(pts_right):
                        pts_left.pop()
                    state['next'] = 0
            _redraw_labels()
            _update_title()
            _update_overlay()
            return

        if event.button != 1:
            return

        n_pair = len(pts_right) + 1   # 次のペア番号

        if state['next'] == 0 and event.inaxes is ax_left:
            pts_left.append([x, y])
            state['next'] = 1
            _draw_pending()
            fig.canvas.draw_idle()
        elif state['next'] == 1 and event.inaxes is ax_right:
            pts_right.append([x, y])
            # 保留左点を確定ペアに昇格
            _clear_pending()
            pl = pts_left[-1]
            mk_l, = ax_left.plot(pl[0], pl[1], 'r+', markersize=16, markeredgewidth=2.5, zorder=5)
            lb_l  = ax_left.annotate(str(n_pair), (pl[0], pl[1]), color='red', fontsize=14,
                                     fontweight='bold', xytext=(6, 6),
                                     textcoords='offset points', zorder=5)
            mk_r, = ax_right.plot(x, y, 'r+', markersize=16, markeredgewidth=2.5, zorder=5)
            lb_r  = ax_right.annotate(str(n_pair), (x, y), color='red', fontsize=14,
                                      fontweight='bold', xytext=(6, 6),
                                      textcoords='offset points', zorder=5)
            artists_left.append([mk_l, lb_l])
            artists_right.append([mk_r, lb_r])
            state['next'] = 0
            fig.canvas.draw_idle()
        else:
            return

        _update_title()
        _update_overlay()

    def on_key(event):
        if event.key == 'q':
            _save_ctrl_pts()
            plt.close(fig)
        elif event.key == 'o':
            state['overlay_visible'] = not state['overlay_visible']
            _btn_overlay.label.set_text(
                'Overlay: ON' if state['overlay_visible'] else 'Overlay: OFF')
            _update_title()
            _update_overlay()

    def _on_btn_overlay(_event):
        state['overlay_visible'] = not state['overlay_visible']
        _btn_overlay.label.set_text(
            'Overlay: ON' if state['overlay_visible'] else 'Overlay: OFF')
        _update_title()
        _update_overlay()

    def _on_btn_done(_event):
        _save_ctrl_pts()
        plt.close(fig)

    def _on_btn_load(_event):
        # ファイルダイアログで JSON を選択
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.update()
            init_dir = str(Path(ctrl_pts_path).parent) if ctrl_pts_path else str(Path.cwd())
            chosen = filedialog.askopenfilename(
                title='Load control points',
                initialdir=init_dir,
                filetypes=[('JSON files', '*.json'), ('All files', '*.*')],
            )
            root.destroy()
            if chosen:
                _load_ctrl_pts(path=chosen)
        except Exception as e:
            print(f"  Warning: file dialog failed: {e}")

    fig.canvas.mpl_connect('button_press_event', on_click)
    fig.canvas.mpl_connect('key_press_event',    on_key)
    _btn_overlay.on_clicked(_on_btn_overlay)
    _btn_done.on_clicked(_on_btn_done)
    _btn_load.on_clicked(_on_btn_load)

    # 起動時に保存済みコントロールポイントを自動読込
    if ctrl_pts_path and Path(ctrl_pts_path).exists():
        _load_ctrl_pts()
    else:
        _update_title()

    plt.show()

    n = min(len(pts_left), len(pts_right))
    return (np.array(pts_left[:n],  dtype=np.float32),
            np.array(pts_right[:n], dtype=np.float32))


# =============================================================================
# 変換行列推定（ebsd_georef_v68 から流用）
# =============================================================================
def estimate_transform(pts_left, pts_right):
    """
    左パネル座標 → 右パネル座標 のアフィン変換行列 H (2×3) を推定する。
    """
    n = len(pts_left)
    if n < 3:
        raise ValueError(f"Not enough control points ({n}). At least 3 required.")

    if n == 3:
        H = cv2.getAffineTransform(pts_left[:3], pts_right[:3])
        print(f"  Affine transform estimated (exact, 3 points)")
    else:
        H, inliers = cv2.estimateAffine2D(pts_left, pts_right,
                                          method=cv2.RANSAC,
                                          ransacReprojThreshold=10.0)
        if H is None:
            raise RuntimeError("Affine estimation failed.")
        n_inliers = int(inliers.sum()) if inliers is not None else n
        pts_h  = np.hstack([pts_left, np.ones((n, 1), dtype=np.float32)])
        pts_pred = (H @ pts_h.T).T
        residuals = np.linalg.norm(pts_pred - pts_right, axis=1)
        print(f"  Affine transform estimated (least squares, {n} pts, {n_inliers} inliers)")
        for i, (res, inl) in enumerate(zip(residuals, inliers.flatten())):
            flag = '' if inl else '  [outlier]'
            print(f"    pt{i+1}: residual = {res:.2f} px{flag}")

    return H


# =============================================================================
# EBSDグリッド点への subset_id 割り当て
# =============================================================================
def assign_subset_ids(data, meta, H, cx_def, cy_def, subset_ids,
                      x_offset, y_offset, dic_step):
    """
    各EBSDグリッド点（data の行）に最近傍DICサブセットの subset_id を割り当てる。

    座標変換の流れ:
      EBSD物理座標 (μm)
        → EBSD Grain IDキャンバスピクセル (ix, iy)
        → H_inv → 変形後DICキャンバスピクセル
        → + offset → DIC実座標 (cx_def, cy_def) と比較

    Returns
    -------
    subset_id_map : ndarray (ny × nx, int32)
        各EBSDグリッド点の subset_id。DIC視野外は -1。
    """
    nx, ny   = meta['nx'], meta['ny']
    x_min    = meta['x_min']
    y_min    = meta['y_min']
    x_step   = meta['x_step']
    y_step   = meta['y_step']
    x_max    = meta['x_max']
    y_max    = meta['y_max']

    # EBSD物理座標 → EBSDキャンバスピクセル座標
    xpos = data[:, 3]
    ypos = data[:, 4]
    ebsd_ix = (xpos - x_min) / x_step if x_step > 0 else np.zeros(len(xpos))
    ebsd_iy = (ypos - y_min) / y_step if y_step > 0 else np.zeros(len(ypos))

    # H_inv: EBSDキャンバスピクセル → 変形後DICキャンバスピクセル
    H_full   = np.vstack([H, [0.0, 0.0, 1.0]])
    H_inv    = np.linalg.inv(H_full)[:2]           # (2, 3)

    N = len(ebsd_ix)
    ebsd_pts_h = np.column_stack([ebsd_ix, ebsd_iy, np.ones(N)])  # (N, 3)
    dic_canvas = (H_inv @ ebsd_pts_h.T).T                         # (N, 2)

    # キャンバスピクセル → DIC実座標
    dic_cx_actual = dic_canvas[:, 0] + x_offset
    dic_cy_actual = dic_canvas[:, 1] + y_offset

    # KDTree で最近傍DICサブセットを検索
    max_dist = dic_step * 1.5
    tree = cKDTree(np.column_stack([cx_def, cy_def]))
    dists, idxs = tree.query(np.column_stack([dic_cx_actual, dic_cy_actual]), k=1)

    sid_flat = np.where(dists <= max_dist,
                        np.array(subset_ids)[idxs],
                        -1).astype(np.int32)

    n_assigned = int((sid_flat >= 0).sum())
    n_outside  = N - n_assigned
    print(f"  subset_id assigned: {n_assigned} / {N}")
    if n_outside > 0:
        print(f"  Outside DIC area  : {n_outside}")

    return sid_flat.reshape(ny, nx)


# =============================================================================
# 全ステージ統合 .mat 保存（DIC サブセット単位）
# =============================================================================
def save_integrated_mat(stages_results, georef_xlsx, out_dir):
    """
    全ステージの HR-EBSD データ（CrossCourt4 .mat）と
    DIC データ（dic_results_georef.xlsx）を DIC サブセット単位に統合して
    integrated_georef.mat として保存する。

    出力変数の命名規則:
      静的情報 : subset_id, cx, cy, grain_id, phi1_ref, PHI_ref, phi2_ref, IQ_ref, CI_ref
      DIC      : {sheet}_{safe_stage}   例: exx_s750MPa, u_s0MPa
      HR-EBSD  : {varname}_{safe_stage} 例: map_s11_s750MPa, map_vmstress_s750MPa

    Parameters
    ----------
    stages_results : list of dict
        各ステージの {'label', 'safe_label', 'ebsd_path', 'subset_id_map', 'mode'}
    georef_xlsx    : str   dic_results_georef.xlsx のパス
    out_dir        : Path  出力ディレクトリ
    """
    import scipy.io

    # .mat から除外する変数（大型セル配列・文字列・不要なメタデータ）
    EXCLUDE_VARS = {
        'shiftsx', 'shiftsy', 'peakhgts', 'refloc', 'imagedir',
        'xpos', 'ypos', 'roiloc', 'filters', 'pcfrac',
        'voltage', 'pixelsize', 'stagetilt', 'scanformat',
        'numims', 'numcols', 'numrows', 'xstep', 'ystep',
        'imwid', 'imhig', 'numphase', 'phasetxt',
        'roisize', 'numroi', 'numref',
        'projectname', 'date', 'version',
    }
    NUMERIC_DTYPES = {'double', 'single', 'int32', 'int64', 'uint8', 'uint16', 'uint32'}

    print(f"\n{'='*60}")
    print("  Building integrated_georef.mat ...")
    print(f"{'='*60}")

    # ── [1/3] DIC データ読み込み ─────────────────────────────────────────────
    print("  [1/3] Loading DIC data from xlsx ...")
    wb = openpyxl.load_workbook(georef_xlsx, read_only=True)

    # u シートから全 subset_id と参照座標を取得
    rows_u = list(wb['u'].iter_rows(values_only=True))
    hdr_u  = list(rows_u[0])
    sid_col = hdr_u.index('subset_id')
    x_col   = hdr_u.index('x [px]')
    y_col   = hdr_u.index('y [px]')
    stage_col_labels = [h for h in hdr_u
                        if h not in (None, 'subset_id', 'x [px]', 'y [px]')]

    data_rows      = rows_u[1:]
    N              = len(data_rows)
    all_subset_ids = np.array([int(r[sid_col]) for r in data_rows], dtype=np.int32)
    cx_ref         = np.array([float(r[x_col]) if r[x_col] is not None else np.nan
                               for r in data_rows])
    cy_ref         = np.array([float(r[y_col]) if r[y_col] is not None else np.nan
                               for r in data_rows])
    sid_to_row     = {int(s): i for i, s in enumerate(all_subset_ids)}

    # ebsd_georef シート（参照 EBSD 由来の grain_id 等）
    rows_eg = list(wb['ebsd_georef'].iter_rows(values_only=True))
    hdr_eg  = list(rows_eg[0])
    def _col(hdr, name):
        return hdr.index(name) if name in hdr else None

    grain_id_arr = np.full(N, np.nan)
    phi1_ref_arr = np.full(N, np.nan)
    PHI_ref_arr  = np.full(N, np.nan)
    phi2_ref_arr = np.full(N, np.nan)
    IQ_ref_arr   = np.full(N, np.nan)
    CI_ref_arr   = np.full(N, np.nan)

    c = {k: _col(hdr_eg, k) for k in
         ('subset_id', 'grain_id', 'phi1 [deg]', 'PHI [deg]', 'phi2 [deg]', 'IQ', 'CI')}
    for r in rows_eg[1:]:
        if c['subset_id'] is None or r[c['subset_id']] is None:
            continue
        idx = sid_to_row.get(int(r[c['subset_id']]))
        if idx is None:
            continue
        def _fv(key):
            ci = c.get(key)
            return float(r[ci]) if (ci is not None and r[ci] is not None) else np.nan
        grain_id_arr[idx] = _fv('grain_id')
        phi1_ref_arr[idx] = _fv('phi1 [deg]')
        PHI_ref_arr[idx]  = _fv('PHI [deg]')
        phi2_ref_arr[idx] = _fv('phi2 [deg]')
        IQ_ref_arr[idx]   = _fv('IQ')
        CI_ref_arr[idx]   = _fv('CI')

    # DIC 各シート（ひずみ・変位等）を読み込み
    DIC_SHEETS = ['u', 'v', 'zncc', 'exx', 'eyy', 'exy', 'e1', 'gamma_max', 'omega_xy',
                  'theta_e1', 'theta_gamma']
    out = {
        'subset_id': all_subset_ids.astype(np.float64),
        'cx':        cx_ref,
        'cy':        cy_ref,
        'grain_id':  grain_id_arr,
        'phi1_ref':  phi1_ref_arr,
        'PHI_ref':   PHI_ref_arr,
        'phi2_ref':  phi2_ref_arr,
        'IQ_ref':    IQ_ref_arr,
        'CI_ref':    CI_ref_arr,
    }

    for sh in DIC_SHEETS:
        if sh not in wb.sheetnames:
            continue
        rows_s = list(wb[sh].iter_rows(values_only=True))
        hdr_s  = list(rows_s[0])
        sc_sid = hdr_s.index('subset_id')
        stage_cols_s = [(h, i) for i, h in enumerate(hdr_s)
                        if h not in (None, 'subset_id', 'x [px]', 'y [px]')]
        for stage_lbl, sc_i in stage_cols_s:
            arr = np.full(N, np.nan)
            for r in rows_s[1:]:
                if r[sc_sid] is None:
                    continue
                idx = sid_to_row.get(int(r[sc_sid]))
                if idx is not None and r[sc_i] is not None:
                    arr[idx] = float(r[sc_i])
            out[f'{sh}_{_mat_safe_name(stage_lbl)}'] = arr

    wb.close()
    print(f"    Subsets loaded : {N}")
    print(f"    DIC variables  : {len(out) - 9}")  # 静的9変数を除いた数

    # ── [2/3] HR-EBSD データをサブセット座標にマッピング ─────────────────
    print("  [2/3] Mapping HR-EBSD data to subsets ...")
    for stg in stages_results:
        if stg.get('mode', 'mat') != 'mat':
            continue
        ebsd_path     = stg['ebsd_path']
        safe_label    = stg['safe_label']
        subset_id_map = stg['subset_id_map']
        mat_sfx       = _mat_safe_name(safe_label)

        print(f"    {Path(ebsd_path).name}  ({safe_label}) ...", end='', flush=True)

        # 2D 数値変数のみ抽出（不要変数・セル・文字列を除外）
        mat_info      = scipy.io.whosmat(ebsd_path)
        include_names = [
            name for name, shape, dtype in mat_info
            if len(shape) == 2
            and name not in EXCLUDE_VARS
            and dtype in NUMERIC_DTYPES
        ]

        # EBSDピクセル → サブセット行インデックス の対応表
        sid_flat   = subset_id_map.flatten().astype(np.int64)
        result_idx = np.array([sid_to_row.get(int(s), -1) for s in sid_flat],
                               dtype=np.int64)
        valid_ebsd = result_idx >= 0
        valid_pixel_indices = np.where(valid_ebsd)[0]

        if len(valid_pixel_indices) == 0:
            print(f"  valid EBSDピクセルなし、スキップ")
            continue

        # DIC→EBSD 方向: 各DICサブセットから最近傍EBSDピクセルを検索
        # EBSDの有効ピクセルのDIC座標を取得
        ebsd_dic_x = cx_ref[result_idx[valid_pixel_indices]]
        ebsd_dic_y = cy_ref[result_idx[valid_pixel_indices]]

        # EBSDピクセル座標でKDTree構築
        ebsd_tree = cKDTree(np.column_stack([ebsd_dic_x, ebsd_dic_y]))

        # EBSDステップ推定（最近傍EBSDピクセル間距離の中央値）
        if len(valid_pixel_indices) > 1:
            nn_d, _ = ebsd_tree.query(np.column_stack([ebsd_dic_x, ebsd_dic_y]), k=2)
            ebsd_step_est = float(np.median(nn_d[:, 1]))
        else:
            ebsd_step_est = 1e9
        max_ebsd_dist = ebsd_step_est * 2.0

        # 各DICサブセット → 最近傍EBSDピクセル
        dists_to_ebsd, nearest_in_valid = ebsd_tree.query(
            np.column_stack([cx_ref, cy_ref]), k=1)
        valid_mapping = dists_to_ebsd <= max_ebsd_dist

        # 一括読み込み（失敗時は1変数ずつ）
        try:
            mat_data = scipy.io.loadmat(ebsd_path, variable_names=include_names)
        except Exception:
            mat_data = {}
            for vname in include_names:
                try:
                    tmp = scipy.io.loadmat(ebsd_path, variable_names=[vname])
                    mat_data.update({k: v for k, v in tmp.items()
                                     if not k.startswith('_')})
                except Exception:
                    pass

        n_mapped = 0
        for vname in include_names:
            if vname not in mat_data:
                continue
            var_flat = mat_data[vname].flatten().astype(np.float64)
            if len(var_flat) != len(sid_flat):
                continue

            # 有効EBSDピクセルの値 → 各DICサブセットへ割り当て
            valid_var    = var_flat[valid_pixel_indices]
            mapped_vals  = valid_var[nearest_in_valid]
            out[f'{vname}_{mat_sfx}'] = np.where(
                valid_mapping & ~np.isnan(mapped_vals), mapped_vals, np.nan)
            n_mapped += 1

        print(f"  {n_mapped} vars mapped")

    # ── [3/3] 保存 ───────────────────────────────────────────────────────────
    print("  [3/3] Saving ...")
    out_path = out_dir / 'integrated_georef.mat'
    scipy.io.savemat(str(out_path), out)
    size_mb = out_path.stat().st_size / 1024 / 1024
    print(f"  Saved : {out_path}")
    print(f"  Variables : {len(out)}  /  Subsets : {N}  /  Size : {size_mb:.1f} MB")


# =============================================================================
# dic_results_georef_def.xlsx にシート追加（Grain File モード用）
# =============================================================================
def save_georef_xlsx_sheet(out_dir, sheet_name, data, phase_names, meta,
                           subset_id_map):
    """
    Grain File (.txt) モードの結果を out_dir/dic_results_georef_def.xlsx に書き出す。
    同ファイルが既に存在する場合はシートを追加、なければ新規作成する。
    元の dic_results_georef.xlsx は変更しない。
    """
    out_path = Path(out_dir) / 'dic_results_georef_def.xlsx'
    nx, ny   = meta['nx'], meta['ny']

    # ファイルが既存なら開く、なければ新規作成
    if out_path.exists():
        wb = openpyxl.load_workbook(str(out_path))
        # 同名シートが既にあれば削除して上書き
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    else:
        wb = openpyxl.Workbook()
        # デフォルトシート（"Sheet"）を削除
        wb.remove(wb.active)

    ws = wb.create_sheet(title=sheet_name)

    # ヘッダ行
    header = ['subset_id', 'ebsd_ix', 'ebsd_iy', 'xpos_um', 'ypos_um',
              'grain_id', 'phase',
              'phi1_deg', 'PHI_deg', 'phi2_deg', 'IQ', 'CI']
    ws.append(header)

    # データ行
    for flat_idx, row in enumerate(data):
        iy  = flat_idx // nx
        ix  = flat_idx %  nx
        sid = int(subset_id_map[iy, ix])
        ws.append([
            sid if sid >= 0 else None,
            ix, iy,
            round(float(row[3]), 6), round(float(row[4]), 6),
            int(row[8]),
            phase_names[flat_idx],
            round(float(np.degrees(row[0])), 6),
            round(float(np.degrees(row[1])), 6),
            round(float(np.degrees(row[2])), 6),
            round(float(row[5]), 6) if not np.isnan(row[5]) else None,
            round(float(row[6]), 6) if not np.isnan(row[6]) else None,
        ])

    wb.save(str(out_path))
    print(f"  Sheet '{sheet_name}' saved to {out_path.name}")


# =============================================================================
# オーバーレイ PNG 保存
# =============================================================================
def save_overlay_png(left_img, right_img, pts_left, pts_right, out_path, opacity=0.45):
    """位置合わせ結果のオーバーレイ画像を PNG として保存する。"""
    H_L, W_L = left_img.shape[:2]
    n = len(pts_left)

    src = pts_left.astype(np.float32)
    dst = pts_right.astype(np.float32)
    if n == 3:
        H_mat = cv2.getAffineTransform(dst[:3], src[:3])
    else:
        H_mat, _ = cv2.estimateAffine2D(dst, src,
                                        method=cv2.RANSAC,
                                        ransacReprojThreshold=10.0)
    warped = cv2.warpAffine(right_img, H_mat, (W_L, H_L),
                            flags=cv2.INTER_NEAREST,
                            borderMode=cv2.BORDER_CONSTANT, borderValue=(0, 0, 0))

    left_f  = left_img.astype(np.float32)
    warp_f  = warped.astype(np.float32)
    is_black = ((warped[:,:,0]==0)&(warped[:,:,1]==0)&(warped[:,:,2]==0))
    alpha   = np.where(is_black, 0.0, opacity)[:,:,np.newaxis]
    blended = (left_f*(1-alpha) + warp_f*alpha).clip(0, 255).astype(np.uint8)

    fig, ax = plt.subplots(figsize=(12, 9))
    fig.patch.set_facecolor('#0a0c0f')
    ax.set_facecolor('#111318')
    ax.imshow(blended, origin='upper')
    for i, pl in enumerate(pts_left, start=1):
        ax.plot(pl[0], pl[1], 'r+', markersize=14, markeredgewidth=2, zorder=5)
        ax.annotate(str(i), (pl[0], pl[1]), color='red', fontsize=9,
                    fontweight='bold', xytext=(5, 5),
                    textcoords='offset points', zorder=5)
    ax.set_title(f'EBSD Grain ID overlay on Deformed DIC  (Affine, {n} pts, opacity={int(opacity*100)}%)',
                 fontsize=11)
    plt.tight_layout()
    fig.savefig(str(out_path), dpi=150)
    plt.close(fig)
    print(f"  Saved overlay PNG: {out_path}")


# =============================================================================
# メイン処理
# =============================================================================
if __name__ == '__main__':
    from PyQt6.QtWidgets import QApplication
    import sys as _sys

    _qt_app = QApplication.instance() or QApplication(_sys.argv)

    if len(_sys.argv) < 2:
        print("Usage: python defebsd_georef_v1.py <param_file.json>")
        sys.exit(1)

    with open(_sys.argv[1], encoding='utf-8') as _f:
        params = json.load(_f)

    georef_xlsx = params['georef_xlsx']
    stages      = params['stages']   # list of {label, ebsd_path, mode}
    out_dir     = Path(params.get('out_dir', Path(georef_xlsx).parent / 'defebsd_georef'))
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"  Def EBSD Georef  —  {len(stages)} stage(s)")
    print(f"  Georef xlsx : {georef_xlsx}")
    print(f"  Output dir  : {out_dir}")
    print(f"{'='*60}")

    stages_results = []   # 全ステージのジオリファレンス結果を蓄積

    for stage_idx, stage in enumerate(stages, 1):
        label     = stage['label']
        ebsd_path = stage['ebsd_path']
        mode      = stage['mode']   # 'mat' or 'grain'

        print(f"\n[Stage {stage_idx}/{len(stages)}]  {label}  ({Path(ebsd_path).name})")
        print('-' * 60)

        safe_label    = label.replace('/', '-').replace(' ', '_')
        ctrl_pts_path = out_dir / f'ctrl_pts_{safe_label}.json'

        # [1] 変形後DIC Grain IDマップ読み込み
        print("[1/5] Loading deformed DIC Grain ID map...")
        cx_def, cy_def, grain_ids_dic, subset_ids = \
            load_deformed_dic_grain_map(georef_xlsx, label)

        # [2] 変形後DIC Grain IDマップを画像化
        print("[2/5] Rendering deformed DIC Grain ID image...")
        dic_grain_img, x_offset, y_offset, dic_step = \
            render_deformed_grain_image(cx_def, cy_def, grain_ids_dic)
        print(f"  Canvas size : {dic_grain_img.shape[1]} x {dic_grain_img.shape[0]} px")
        print(f"  DIC step est: {dic_step:.1f} px")

        # [3] EBSD データ読み込みと Grain IDマップ生成
        print("[3/5] Loading EBSD data...")
        if mode == 'mat':
            data, phase_names, meta = load_mat_minimal(ebsd_path)
        else:
            data, phase_names, meta = load_grain_file(ebsd_path)

        print("[3/5] Generating EBSD Grain ID image...")
        ebsd_grain_img, _, _ = make_grain_id_image(data, meta)

        # [4] コントロールポイント指定
        print("[4/5] Please specify control points in the GUI...")
        print("  Left(Deformed DIC) → Right(EBSD) click alternately")
        print("  L-Rclick: undo pair  /  R-Rclick: undo right only  /  q: confirm")
        pts_left, pts_right = pick_control_points(
            dic_grain_img, ebsd_grain_img,
            left_label=label,
            right_label=Path(ebsd_path).name,
            ctrl_pts_path=ctrl_pts_path,
        )

        n_pts = len(pts_left)
        if n_pts < 3:
            print(f"  Skipped: not enough control points ({n_pts} < 3)")
            continue
        print(f"  {n_pts} control point pairs confirmed.")

        H = estimate_transform(pts_left, pts_right)

        # [5] subset_id 割り当て
        print("[5/5] Assigning subset_id to each EBSD point...")
        subset_id_map = assign_subset_ids(
            data, meta, H,
            cx_def, cy_def, subset_ids,
            x_offset, y_offset, dic_step,
        )

        # オーバーレイ PNG 保存
        overlay_png = out_dir / f'overlay_{safe_label}.png'
        save_overlay_png(dic_grain_img, ebsd_grain_img,
                         pts_left, pts_right, overlay_png)

        # ステージ結果を蓄積（matモード・grainモード共通）
        stages_results.append({
            'label':         label,
            'safe_label':    safe_label,
            'ebsd_path':     ebsd_path,
            'subset_id_map': subset_id_map,
            'mode':          mode,
            'data':          data,
            'phase_names':   phase_names,
            'meta':          meta,
        })

        print(f"  Stage '{label}' complete.")

    # ── 全ステージ統合 .mat を1ファイルに出力 ────────────────────────────
    if stages_results:
        mat_stages  = [s for s in stages_results if s['mode'] == 'mat']
        grain_stages = [s for s in stages_results if s['mode'] == 'grain']

        if mat_stages:
            save_integrated_mat(mat_stages, georef_xlsx, out_dir)

        # grain モードは従来通り xlsx にシート追加
        for s in grain_stages:
            sheet_name = f'ebsd_georef_{s["safe_label"]}'
            save_georef_xlsx_sheet(out_dir, sheet_name,
                                   s['data'], s['phase_names'],
                                   s['meta'], s['subset_id_map'])

    print(f"\n{'='*60}")
    print(f"  All {len(stages)} stage(s) complete.")
    print(f"  Output: {out_dir}")
    print(f"{'='*60}")
