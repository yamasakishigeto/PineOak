"""
統合ランチャー main.py  # v12
実行: py -3.13 main.py
"""

import eel
import subprocess
import sys
import os
import threading
from pathlib import Path

# --- パス設定 ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TOOLS_DIR = BASE_DIR
PYTHON = sys.executable

SEM_ALIGN_HTML = os.path.join(TOOLS_DIR, "sem_align_tool_v3.html")

# Eelの初期化（index.html / dic_wizard.html はmain.pyと同じフォルダ）
eel.init(BASE_DIR)

# 起動中のプロセスを追跡
_procs = {}

# 作業フォルダ
_work_dir = None

# DIC解析パラメータ（ウィザードから受け取る）
_dic_params = None


def _run_script(tool_id: str, script_path: str, work_dir: str = None):
    """別スレッドでPythonスクリプトを起動し、完了をHTML側に通知する"""
    try:
        cmd = [PYTHON, script_path]
        if work_dir:
            cmd.append(work_dir)
        proc = subprocess.Popen(cmd, cwd=TOOLS_DIR)
        _procs[tool_id] = proc
        eel.on_tool_started(tool_id)()
        proc.wait()
        returncode = proc.returncode
        _procs.pop(tool_id, None)
        eel.on_tool_finished(tool_id, returncode)()
    except FileNotFoundError:
        eel.on_tool_error(tool_id, f"ファイルが見つかりません: {script_path}")()
    except Exception as e:
        eel.on_tool_error(tool_id, str(e))()


# ================================================================
# ランチャー共通
# ================================================================

@eel.expose
def set_work_dir(path: str):
    global _work_dir
    _work_dir = path if path else None
    return _work_dir


@eel.expose
def get_work_dir():
    return _work_dir or ""


@eel.expose
def browse_work_dir():
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.askdirectory(title='作業フォルダを選択')
    root.destroy()
    if path:
        global _work_dir
        _work_dir = path
        return path
    return ""


@eel.expose
def launch_sem_align():
    """SEM位置合わせ: Eelウィンドウで開く"""
    try:
        eel.on_tool_started("sem_align")()
        eel.start("sem_align_tool_v3.html", size=(1280, 860), block=False)
        eel.on_tool_finished("sem_align", 0)()
    except Exception as e:
        eel.on_tool_error("sem_align", str(e))()


@eel.expose
def launch_dic():
    """DICウィザードHTMLを別ウィンドウで開く"""
    eel.start("dic_wizard.html", size=(1200, 900), block=False)


@eel.expose
def launch_ebsd():
    """EBSDウィザードHTMLを別ウィンドウで開く"""
    eel.start("ebsd_wizard.html", size=(740, 510), block=False)


@eel.expose
def launch_heaviside():
    """Heavisideウィザードを別ウィンドウで開く"""
    eel.start("heaviside_wizard.html", size=(740, 550), block=False)



# ================================================================
# EBSDウィザード用
# ================================================================

@eel.expose
def ebsd_browse_file(title: str, filetypes: list, initialdir: str):
    return _tk_filedialog('file', title,
                          filetypes=[tuple(f) for f in filetypes],
                          initialdir=initialdir or _work_dir)


@eel.expose
def ebsd_browse_dir(title: str, initialdir: str):
    return _tk_filedialog('dir', title, initialdir=initialdir or _work_dir)


@eel.expose
def ebsd_start_analysis(paths: dict):
    threading.Thread(target=_run_ebsd_analysis, args=(paths,), daemon=True).start()


def _run_ebsd_analysis(paths: dict):
    import json, tempfile
    param_file = tempfile.mktemp(suffix='.json')
    with open(param_file, 'w') as f:
        json.dump(paths, f)

    ebsd_script = os.path.join(TOOLS_DIR, "ebsd_georef_v68.py")

    eel.on_tool_started("ebsd")()
    eel.ebsd_on_status("起動中...", "ok")()

    # JSONパスを直接 ebsd_georef_v68.py に渡す（ファイル選択ダイアログをスキップ）
    _env = os.environ.copy()
    _env['MPLBACKEND'] = 'QtAgg'
    _env['PYTHONIOENCODING'] = 'utf-8'
    proc = subprocess.Popen(
        [PYTHON, ebsd_script, param_file],
        cwd=TOOLS_DIR,
        env=_env,
    )
    _procs["ebsd"] = proc
    proc.wait()
    _procs.pop("ebsd", None)

    if proc.returncode == 0:
        eel.on_tool_finished("ebsd", 0)()
        eel.ebsd_on_status("完了", "ok")()
    else:
        eel.on_tool_finished("ebsd", proc.returncode)()
        eel.ebsd_on_status(f"エラー（終了コード {proc.returncode}）", "err")()


# ================================================================
# Heavisideウィザード用
# ================================================================

@eel.expose
def hv_browse_file(title: str, filetypes: list, initialdir: str):
    return _tk_filedialog('file', title,
                          filetypes=[tuple(f) for f in filetypes],
                          initialdir=initialdir or _work_dir)


@eel.expose
def hv_browse_dir(title: str, initialdir: str):
    return _tk_filedialog('dir', title, initialdir=initialdir or _work_dir)


@eel.expose
def hv_start_analysis(paths: dict):
    threading.Thread(target=_run_hv_analysis, args=(paths,), daemon=True).start()


def _run_hv_analysis(paths: dict):
    import json, tempfile
    param_file = tempfile.mktemp(suffix='.json')
    with open(param_file, 'w') as f:
        json.dump(paths, f)

    hv_script = os.path.join(TOOLS_DIR, "heaviside_dic_v81.py")

    eel.on_tool_started("heaviside")()
    eel.hv_on_status("起動中...", "ok")()

    # JSONパスを直接 heaviside_dic_v81.py に渡す（ウィザードをスキップ）
    _env = os.environ.copy()
    _env['MPLBACKEND'] = 'QtAgg'
    _env['PYTHONIOENCODING'] = 'utf-8'
    proc = subprocess.Popen(
        [PYTHON, hv_script, param_file],
        cwd=TOOLS_DIR,
        env=_env,
    )
    _procs["heaviside"] = proc
    proc.wait()
    _procs.pop("heaviside", None)

    if proc.returncode == 0:
        eel.on_tool_finished("heaviside", 0)()
        eel.hv_on_status("完了", "ok")()
    else:
        eel.on_tool_finished("heaviside", proc.returncode)()
        eel.hv_on_status(f"エラー（終了コード {proc.returncode}）", "err")()


@eel.expose
def get_tools_dir():
    return TOOLS_DIR


# ================================================================
# DICウィザード用
# ================================================================

def _tk_filedialog(kind, title, filetypes=None, initialdir=None):
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    if kind == 'dir':
        path = filedialog.askdirectory(title=title, initialdir=initialdir)
    else:
        path = filedialog.askopenfilename(title=title, filetypes=filetypes or [],
                                          initialdir=initialdir)
    root.destroy()
    return path or ""


@eel.expose
def dic_browse_folder():
    """画像フォルダ選択"""
    return _tk_filedialog('dir', 'SEM画像フォルダを選択', initialdir=_work_dir)


@eel.expose
def dic_browse_json(folder: str):
    """アライメントJSON選択"""
    return _tk_filedialog('file', 'アライメントJSONを選択（キャンセルでスキップ）',
                          filetypes=[('JSON', '*.json'), ('All', '*.*')],
                          initialdir=folder or _work_dir)


@eel.expose
def dic_list_images(folder: str):
    """フォルダ内の画像ファイル一覧を返す（自然順ソート）"""
    import re
    exts = {'.bmp', '.png', '.tif', '.tiff', '.jpg', '.jpeg'}

    def natural_key(f):
        # ファイル名中の数字を数値として解釈してソート
        return [int(c) if c.isdigit() else c.lower()
                for c in re.split(r'(\d+)', f.name)]

    try:
        files = sorted(
            [f for f in Path(folder).iterdir()
             if f.is_file() and f.suffix.lower() in exts],
            key=natural_key
        )
        return [{'name': f.name, 'path': str(f)} for f in files]
    except Exception as e:
        return []


@eel.expose
def dic_load_config(folder: str):
    """dic_config.txt を選択して読み込む"""
    import re
    path = _tk_filedialog('file', 'dic_config.txt を選択',
                          filetypes=[('テキスト', '*.txt'), ('All', '*.*')],
                          initialdir=folder or _work_dir)
    if not path:
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        result = {}
        in_s1 = in_s2 = in_scale = in_cmap = False

        for line in lines:
            line = line.strip()
            if line == '[Stage 1]':    in_s1, in_s2, in_scale, in_cmap = True,  False, False, False; continue
            if line == '[Stage 2]':    in_s1, in_s2, in_scale, in_cmap = False, True,  False, False; continue
            if line == '[カラースケール]': in_s1, in_s2, in_scale, in_cmap = False, False, True,  False; continue
            if line == '[カラーマップ]':  in_s1, in_s2, in_scale, in_cmap = False, False, False, True;  continue
            if line.startswith('['):   in_s1, in_s2, in_scale, in_cmap = False, False, False, False; continue
            if ':' not in line: continue

            k, _, v = line.partition(':')
            k, v = k.strip(), v.strip()

            if not in_s1 and not in_s2 and not in_scale and not in_cmap:
                if k == 'subset':   result['subset']   = int(v.split()[0])
                if k == 'gauge':    result['gauge']    = int(v.split()[0])
                if k == 'workers':  result['workers']  = int(v.split()[0])
                if k == 'NCC閾値':  result['ncc_thr']  = float(v.split()[0])

            if in_s1:
                if k == 'step':   result['s1_step'] = int(v.split()[0])
                if k == 'search':
                    if 'グローバルシフト' in v:
                        result['s1_auto'] = True
                        m = re.search(r'\+\s*(\d+)', v)
                        if m: result['s1_margin'] = int(m.group(1))
                    else:
                        result['s1_auto'] = False
                        try: result['s1_fixed'] = int(v.split()[0])
                        except: pass
                if k == '前段参照': result['use_prev_s1'] = 'あり' in v

            if in_s2:
                if k == 'step':   result['s2_step']   = int(v.split()[0])
                if k == 'search': result['s2_search'] = int(v.split()[0])

            if in_scale:
                scale_keys = {'u','v','exx','eyy','exy','e1','gamma_max','omega_xy'}
                if k in scale_keys:
                    m_min = re.search(r'min=([^\s]+)', v)
                    m_max = re.search(r'max=([^\s]+)', v)
                    lo = float(m_min.group(1)) if m_min and m_min.group(1) != '自動' else None
                    hi = float(m_max.group(1)) if m_max and m_max.group(1) != '自動' else None
                    result.setdefault('scale', {})[k] = [lo, hi]

            if in_cmap:
                cmap_keys = {'u','v','exx','eyy','exy','e1','gamma_max','omega_xy'}
                if k in cmap_keys:
                    result.setdefault('cmap', {})[k] = v

        return result
    except Exception as e:
        return {'error': str(e)}


@eel.expose
def dic_run_prescan(params: dict):
    """事前スキャンを実行して結果を返す"""
    try:
        # dic_sem_strain をインポートして prescan 関数を使う
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "dic_sem_strain",
            os.path.join(TOOLS_DIR, "dic_sem_strain_v58.py")
        )
        mod = importlib.util.module_from_spec(spec)

        # サブプロセスで実行する方式（モジュールimportが複雑なため）
        import json, tempfile
        param_file = tempfile.mktemp(suffix='.json')
        result_file = tempfile.mktemp(suffix='.json')
        with open(param_file, 'w') as f:
            json.dump(params, f)

        script = os.path.join(TOOLS_DIR, "_prescan_runner.py")
        _write_prescan_runner(script)

        _env = os.environ.copy()
        _env['PYTHONIOENCODING'] = 'utf-8'
        proc = subprocess.run(
            [PYTHON, script, param_file, result_file,
             os.path.join(TOOLS_DIR, "dic_sem_strain_v58.py")],
            timeout=120,
            env=_env,
        )

        if os.path.exists(result_file):
            with open(result_file, 'r') as f:
                return json.load(f)
        return {'error': '結果ファイルが生成されませんでした'}

    except subprocess.TimeoutExpired:
        return {'error': 'タイムアウト（120秒）'}
    except Exception as e:
        return {'error': str(e)}


def _write_prescan_runner(path: str):
    """事前スキャン用の一時スクリプトを生成する"""
    code = '''
import sys, json
from pathlib import Path

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

param_file  = sys.argv[1]
result_file = sys.argv[2]
dic_module  = sys.argv[3]

with open(param_file) as f:
    p = json.load(f)

import importlib.util
spec = importlib.util.spec_from_file_location("dic", dic_module)
mod  = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

ref_raw  = mod.load_and_preprocess(Path(p["ref_path"]), 0)
def_raw  = mod.load_and_preprocess(Path(p["def_path"]),  0)

json_path = Path(p["json_path"]) if p.get("json_path") else None
shifts = mod.load_alignment(json_path)

ref_al = mod.apply_alignment(ref_raw, Path(p["ref_path"]).name, shifts, ref_raw.shape)
def_al = mod.apply_alignment(def_raw, Path(p["def_path"]).name, shifts, def_raw.shape)

if shifts:
    import cv2, numpy as np
    h, w = ref_al.shape[:2]
    roi = mod.calc_valid_roi(shifts, (h, w))
    ref_cr = mod.crop_roi(ref_al, roi)
    def_cr = mod.crop_roi(def_al, roi)
else:
    ref_cr, def_cr = ref_al, def_al

_, rec, est = mod.prescan(
    ref_cr, def_cr,
    subset_size=p["subset_size"],
    search_range=p["prescan_search"],
)

s1_margin = p.get("s1_margin", 15)
ok = rec <= s1_margin
status = (
    f"✓ 推奨 {rec}px ≤ 現在のStage1 search {s1_margin}px → OK"
    if ok else
    f"⚠ 推奨 {rec}px > 現在のStage1 search {s1_margin}px → 要見直し"
)

lines = [f"対象: {Path(p['def_path']).name}", status, "", "【参考：変位・ひずみ推定値】",
         "  変数          min        max"]
lmap = [("u [px]","u"),("v [px]","v"),("exx [-]","exx"),("eyy [-]","eyy"),
        ("exy [-]","exy"),("e1 [-]","e1"),("γmax [-]","gamma_max"),("ωxy [-]","omega_xy")]
for lbl, key in lmap:
    lo, hi = est[key]
    lines.append(f"  {lbl:<10}  {lo:>8.4f}  {hi:>8.4f}")

result = {
    "rec_search": rec,
    "ok": ok,
    "text": "\\n".join(lines),
    "estimates": {k: list(v) for k, v in est.items()},
}

with open(result_file, "w") as f:
    json.dump(result, f)
'''
    with open(path, 'w', encoding='utf-8') as f:
        f.write(code)


@eel.expose
def dic_start_analysis(params: dict):
    """DIC解析をバックグラウンドで開始する"""
    global _dic_params
    _dic_params = params
    threading.Thread(target=_run_dic_analysis, args=(params,), daemon=True).start()


def _run_dic_analysis(params: dict):
    """DIC解析本体をサブプロセスで実行する"""
    import json, tempfile
    param_file = tempfile.mktemp(suffix='.json')
    with open(param_file, 'w') as f:
        json.dump(params, f, default=str)

    script = os.path.join(TOOLS_DIR, "_dic_runner.py")
    _write_dic_runner(script)

    eel.on_tool_started("dic")()
    eel.dic_on_status("解析中...", "ok")()

    _env = os.environ.copy()
    _env['MPLBACKEND'] = 'QtAgg'
    _env['PYTHONIOENCODING'] = 'utf-8'
    proc = subprocess.Popen(
        [PYTHON, script, param_file, os.path.join(TOOLS_DIR, "dic_sem_strain_v58.py")],
        cwd=TOOLS_DIR,
        env=_env,
    )
    _procs["dic"] = proc
    proc.wait()
    _procs.pop("dic", None)

    if proc.returncode == 0:
        import json as _json
        output_dir = str(Path(params['folder']) / f"dic_{Path(params['ref_path']).stem}")
        eel.on_tool_finished("dic", 0)()
        eel.dic_on_status("解析完了", "ok")()
        eel.dic_on_analysis_done(output_dir)()
        # 解析完了後、最後のDEFをGUI設定で自動プレビュー
        auto_params = {
            'output_dir': output_dir,
            'def_index': -1,
            'scale': params.get('scale', {}),
            'cmap':  params.get('cmap',  {}),
            'ncc_threshold': params.get('ncc_threshold', 0.2),
            'dic_module': os.path.join(TOOLS_DIR, 'dic_sem_strain_v58.py'),
        }
        _env2 = os.environ.copy()
        _env2['PYTHONIOENCODING'] = 'utf-8'
        subprocess.Popen([PYTHON, '-c', _PREVIEW_SCRIPT, _json.dumps(auto_params)], env=_env2)
    else:
        eel.on_tool_finished("dic", proc.returncode)()
        eel.dic_on_status(f"エラー（終了コード {proc.returncode}）", "err")()


def _write_dic_runner(path: str):
    """DIC解析実行用の一時スクリプトを生成する"""
    code = open(__file__.replace('main.py', '_dic_runner_src.py'), 'r', encoding='utf-8').read()
    with open(path, 'w', encoding='utf-8') as f:
        f.write(code)


@eel.expose
def dic_cancel():
    """DIC解析をキャンセルする"""
    proc = _procs.get("dic")
    if proc:
        proc.terminate()


_PREVIEW_SCRIPT = r"""
import sys, json, pickle
from pathlib import Path
import importlib.util
import numpy as np
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt

p            = json.loads(sys.argv[1])
output_dir   = Path(p['output_dir'])
def_index    = int(p.get('def_index', -1))
cmap_map     = p.get('cmap', {})
scale_config = {k: tuple(v) for k, v in p.get('scale', {}).items()}
dic_module   = p['dic_module']
ncc_thr      = float(p.get('ncc_threshold', 0.2))

_CMAP_DEF = {'u':'RdBu_r','v':'RdBu_r','exx':'RdBu_r','eyy':'RdBu_r','exy':'RdBu_r',
             'e1':'hot_r','gamma_max':'hot_r','omega_xy':'RdBu_r'}
def _cm(key): return cmap_map.get(key) or _CMAP_DEF.get(key, 'RdBu_r')

spec = importlib.util.spec_from_file_location('dic', dic_module)
mod  = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

with open(output_dir / 'dic_results.pkl', 'rb') as f:
    data = pickle.load(f)

results_list  = data['results_list']
unified_scale = dict(data['unified_scale'])
step_fine     = data['step_fine']
ref_path      = data['ref_path']
def_paths     = data['def_paths']

for k, (lo, hi) in scale_config.items():
    lo_old, hi_old = unified_scale.get(k, (None, None))
    unified_scale[k] = (lo if lo is not None else lo_old,
                        hi if hi is not None else hi_old)

# 対象DEFを選択（デフォルトは最後のDEF）
idx = def_index if def_index >= 0 else len(results_list) - 1
res      = results_list[idx]
def_name = def_paths[idx - 1] if idx > 0 else ref_path
sc       = unified_scale

def make_grid(arr):
    return mod.make_grid_map(res['cx'], res['cy'], arr, step_fine)[0]

u_grid   = make_grid(res['u'])
v_grid   = make_grid(res['v'])
ncc_grid = make_grid(res['ncc'])
_, extent, _, _ = mod.make_grid_map(res['cx'], res['cy'], res['u'], step_fine)
strain   = res['strain']

img_h = extent[2] - extent[3]
img_w = extent[1] - extent[0]
aspect = img_h / img_w
fig, axes = plt.subplots(3, 3, figsize=(18, 18 * aspect + 2))

def plot_one(ax, data, title, key, cmap, symmetric=True):
    lo, hi = sc.get(key, (None, None))
    if lo is not None and hi is not None:
        vmin, vmax = lo, hi
    elif symmetric:
        vabs = max(float(np.nanpercentile(np.abs(data[~np.isnan(data)]), 98)) if np.any(~np.isnan(data)) else 1e-6, 1e-6)
        vmin, vmax = -vabs, vabs
    else:
        valid = data[~np.isnan(data)]
        vmin = float(np.percentile(valid, 2))  if len(valid) else 0
        vmax = float(np.percentile(valid, 98)) if len(valid) else 1
    im = ax.imshow(data, cmap=cmap, extent=extent, vmin=vmin, vmax=vmax, aspect='equal')
    plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    ax.set_title(title, fontsize=9, pad=3)
    ax.set_xlabel('X [px]', fontsize=7); ax.set_ylabel('Y [px]', fontsize=7)
    ax.tick_params(labelsize=6)

plot_one(axes[0,0], u_grid,              'u変位 [px]',               'u',         _cm('u'))
plot_one(axes[0,1], v_grid,              'v変位 [px]',               'v',         _cm('v'))
ncc_vmin = max(ncc_thr, 0.0)
im = axes[0,2].imshow(ncc_grid, cmap='plasma', extent=extent, vmin=ncc_vmin, vmax=1.0, aspect='equal')
plt.colorbar(im, ax=axes[0,2], fraction=0.046, pad=0.04)
axes[0,2].set_title(f'NCC peak (下限={ncc_vmin:.2f})', fontsize=9, pad=3)
axes[0,2].set_xlabel('X [px]', fontsize=7); axes[0,2].set_ylabel('Y [px]', fontsize=7)
axes[0,2].tick_params(labelsize=6)

plot_one(axes[1,0], strain['exx'],       'εxx（x方向正ひずみ）',      'exx',       _cm('exx'))
plot_one(axes[1,1], strain['eyy'],       'εyy（y方向正ひずみ）',      'eyy',       _cm('eyy'))
plot_one(axes[1,2], strain['exy'],       'εxy（せん断ひずみ）',       'exy',       _cm('exy'))
plot_one(axes[2,0], strain['e1'],        'e1（最大主ひずみ）',        'e1',        _cm('e1'),        symmetric=False)
plot_one(axes[2,1], strain['gamma_max'], 'γmax（最大せん断ひずみ）',  'gamma_max', _cm('gamma_max'), symmetric=False)
plot_one(axes[2,2], strain['omega_xy'],  'ωxy（回転）',              'omega_xy',  _cm('omega_xy'))

fig.suptitle(f'変位・ひずみマップ  {Path(ref_path).name} → {Path(def_name).name}',
             fontsize=11, y=1.002)
plt.tight_layout()
plt.show()
"""


@eel.expose
def dic_preview_maps(params: dict):
    """計算済みpickleを読み込んで1ウィンドウにマップをプレビュー表示する（別プロセス）"""
    import subprocess, sys, json
    params['dic_module'] = os.path.join(TOOLS_DIR, 'dic_sem_strain_v58.py')
    env = os.environ.copy()
    env['PYTHONIOENCODING'] = 'utf-8'
    subprocess.Popen([sys.executable, '-c', _PREVIEW_SCRIPT, json.dumps(params)], env=env)


@eel.expose
def dic_recalc_strain(params: dict):
    """pickleの変位データからひずみのみ再計算してpickleを更新し、プレビューを表示する"""
    import subprocess, sys, json

    _recalc_script = r"""
import sys, json, pickle
from pathlib import Path
import importlib.util
import numpy as np

p            = json.loads(sys.argv[1])
output_dir   = Path(p['output_dir'])
gauge_length = int(p.get('gauge_length', 1))
strain_type  = p.get('strain_type', 'infinitesimal')
dic_module   = p['dic_module']

spec = importlib.util.spec_from_file_location('dic', dic_module)
mod  = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

with open(output_dir / 'dic_results.pkl', 'rb') as f:
    data = pickle.load(f)

results_list  = data['results_list']
step_fine     = data['step_fine']

stype_label = 'グリーン-ラグランジェ' if strain_type == 'green_lagrange' else '微小ひずみ'
print(f"ひずみ再計算: ゲージ長さ={gauge_length}倍 / {stype_label}")

for i, res in enumerate(results_list):
    res['strain'] = mod.calc_strain_field(
        res['cx'], res['cy'], res['u'], res['v'],
        step_fine, gauge_length=gauge_length, strain_type=strain_type)
    print(f"  [{i+1}/{len(results_list)}] {res['label']} 完了")

SCALE_KEYS_SYM  = ['exx', 'eyy', 'exy', 'omega_xy']
SCALE_KEYS_ASYM = ['e1', 'gamma_max']
unified_scale = dict(data['unified_scale'])
for k in SCALE_KEYS_SYM + SCALE_KEYS_ASYM:
    vals = np.concatenate([
        np.array(res['strain'][k], dtype=float).flatten()
        for res in results_list if k in res.get('strain', {})
    ])
    vals = vals[~np.isnan(vals)]
    if len(vals) == 0:
        continue
    if k in SCALE_KEYS_SYM:
        vabs = max(float(np.percentile(np.abs(vals), 98)), 1e-6)
        unified_scale[k] = (-vabs, vabs)
    else:
        unified_scale[k] = (float(np.percentile(vals, 2)),
                            float(np.percentile(vals, 98)))

data['unified_scale'] = unified_scale
data['gauge_length']  = gauge_length
data['strain_type']   = strain_type

with open(output_dir / 'dic_results.pkl', 'wb') as f:
    pickle.dump(data, f)
print("pickle更新完了")

# dic_config.txt の [共通] セクションのgauge・strain_typeを更新
config_path = output_dir / 'dic_config.txt'
if config_path.exists():
    lines = config_path.read_text(encoding='utf-8').splitlines()
    new_lines = []
    in_common = False
    strain_written = False
    for line in lines:
        stripped = line.strip()
        if stripped == '[共通]':
            in_common = True
            new_lines.append(line)
            continue
        if stripped.startswith('[') and stripped != '[共通]':
            # 共通セクションを抜けるとき、strain_typeがまだなら追加
            if in_common and not strain_written:
                new_lines.append(f'  strain_type : {strain_type}')
                strain_written = True
            in_common = False
        if in_common:
            if stripped.startswith('gauge'):
                new_lines.append(f'  gauge      : {gauge_length} 倍')
                continue
            if stripped.startswith('strain_type'):
                new_lines.append(f'  strain_type : {strain_type}')
                strain_written = True
                continue
        new_lines.append(line)
    # ファイル末尾が[共通]で終わる場合
    if in_common and not strain_written:
        new_lines.append(f'  strain_type : {strain_type}')
    config_path.write_text('\n'.join(new_lines) + '\n', encoding='utf-8')
    print("dic_config.txt 更新完了")
"""

    params['dic_module'] = os.path.join(TOOLS_DIR, 'dic_sem_strain_v58.py')
    env = os.environ.copy()
    env['PYTHONIOENCODING'] = 'utf-8'
    try:
        subprocess.run(
            [sys.executable, '-c', _recalc_script, json.dumps(params)],
            env=env, timeout=60
        )
        # 再計算完了後にプレビューを起動
        preview_params = {
            'output_dir':    params['output_dir'],
            'def_index':     -1,
            'scale':         params.get('scale', {}),
            'cmap':          params.get('cmap', {}),
            'ncc_threshold': params.get('ncc_threshold', 0.2),
            'dic_module':    params['dic_module'],
        }
        subprocess.Popen(
            [sys.executable, '-c', _PREVIEW_SCRIPT, json.dumps(preview_params)],
            env=env
        )
        return {'ok': True}
    except Exception as e:
        return {'error': str(e)}


@eel.expose
def dic_save_maps(params: dict):
    """計算済みpickleを読み込んでPNG・Excelを保存する（別プロセス）"""
    import subprocess, sys, json

    _script = r"""
import sys, json, pickle
from pathlib import Path
import importlib.util
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

p            = json.loads(sys.argv[1])
output_dir   = Path(p['output_dir'])
cmap_map     = p.get('cmap', {})
scale_config = {k: tuple(v) for k, v in p.get('scale', {}).items()}
dic_module   = p['dic_module']
ncc_thr      = float(p.get('ncc_threshold', 0.2))

_CMAP_DEF = {'u':'RdBu_r','v':'RdBu_r','exx':'RdBu_r','eyy':'RdBu_r','exy':'RdBu_r',
             'e1':'hot_r','gamma_max':'hot_r','omega_xy':'RdBu_r'}
def _cm(key): return cmap_map.get(key) or _CMAP_DEF.get(key, 'RdBu_r')

spec = importlib.util.spec_from_file_location('dic', dic_module)
mod  = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)
mod.OUTPUT_DIR = output_dir

with open(output_dir / 'dic_results.pkl', 'rb') as f:
    data = pickle.load(f)

results_list  = data['results_list']
unified_scale = dict(data['unified_scale'])
step_fine     = data['step_fine']
ref_path      = data['ref_path']
def_paths     = data['def_paths']
roi           = data.get('roi')

for k, (lo, hi) in scale_config.items():
    lo_old, hi_old = unified_scale.get(k, (None, None))
    unified_scale[k] = (lo if lo is not None else lo_old,
                        hi if hi is not None else hi_old)

sc = unified_scale

def plot_one(ax, data, title, key, cmap, ext, symmetric=True):
    lo, hi = sc.get(key, (None, None))
    if lo is not None and hi is not None:
        vmin, vmax = lo, hi
    elif symmetric:
        valid = data[~np.isnan(data)] if np.any(~np.isnan(data)) else np.array([1e-6])
        vabs = max(float(np.nanpercentile(np.abs(valid), 98)), 1e-6)
        vmin, vmax = -vabs, vabs
    else:
        valid = data[~np.isnan(data)]
        vmin = float(np.percentile(valid, 2))  if len(valid) else 0
        vmax = float(np.percentile(valid, 98)) if len(valid) else 1
    im = ax.imshow(data, cmap=cmap, extent=ext, vmin=vmin, vmax=vmax, aspect='equal')
    plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    ax.set_title(title, fontsize=9, pad=3)
    ax.set_xlabel('X [px]', fontsize=7); ax.set_ylabel('Y [px]', fontsize=7)
    ax.tick_params(labelsize=6)

print(f"マップ保存開始: {output_dir}")
for i, res in enumerate(results_list):
    def_name = def_paths[i-1] if i > 0 else ref_path
    print(f"  [{i+1}/{len(results_list)}] {res['label']}")

    u_grid   = mod.make_grid_map(res['cx'], res['cy'], res['u'],   step_fine)[0]
    v_grid   = mod.make_grid_map(res['cx'], res['cy'], res['v'],   step_fine)[0]
    ncc_grid = mod.make_grid_map(res['cx'], res['cy'], res['ncc'], step_fine)[0]
    _, ext, _, _ = mod.make_grid_map(res['cx'], res['cy'], res['u'], step_fine)
    strain = res['strain']

    img_h = ext[2] - ext[3]
    img_w = ext[1] - ext[0]
    aspect = img_h / img_w
    fig, axes = plt.subplots(3, 3, figsize=(18, 18 * aspect + 2))

    plot_one(axes[0,0], u_grid,              'u変位 [px]',               'u',         _cm('u'),         ext)
    plot_one(axes[0,1], v_grid,              'v変位 [px]',               'v',         _cm('v'),         ext)
    ncc_vmin = max(ncc_thr, 0.0)
    im = axes[0,2].imshow(ncc_grid, cmap='plasma', extent=ext, vmin=ncc_vmin, vmax=1.0, aspect='equal')
    plt.colorbar(im, ax=axes[0,2], fraction=0.046, pad=0.04)
    axes[0,2].set_title(f'NCC peak (下限={ncc_vmin:.2f})', fontsize=9, pad=3)
    axes[0,2].set_xlabel('X [px]', fontsize=7); axes[0,2].set_ylabel('Y [px]', fontsize=7)
    axes[0,2].tick_params(labelsize=6)

    plot_one(axes[1,0], strain['exx'],       'εxx（x方向正ひずみ）',      'exx',       _cm('exx'),       ext)
    plot_one(axes[1,1], strain['eyy'],       'εyy（y方向正ひずみ）',      'eyy',       _cm('eyy'),       ext)
    plot_one(axes[1,2], strain['exy'],       'εxy（せん断ひずみ）',       'exy',       _cm('exy'),       ext)
    plot_one(axes[2,0], strain['e1'],        'e1（最大主ひずみ）',        'e1',        _cm('e1'),        ext, symmetric=False)
    plot_one(axes[2,1], strain['gamma_max'], 'γmax（最大せん断ひずみ）',  'gamma_max', _cm('gamma_max'), ext, symmetric=False)
    plot_one(axes[2,2], strain['omega_xy'],  'ωxy（回転）',              'omega_xy',  _cm('omega_xy'),  ext)

    fig.suptitle(f'変位・ひずみマップ  {Path(ref_path).name} → {Path(def_name).name}',
                 fontsize=11, y=1.002)
    plt.tight_layout()
    suffix = f"_{res['label']}" if res['label'] else ''
    plt.savefig(output_dir / f"map{suffix}.png", dpi=150, bbox_inches='tight')
    plt.close()

xlsx_path = output_dir / 'dic_results.xlsx'
print("Excelファイルを出力しています...")
mod.export_xlsx(results_list, unified_scale, xlsx_path, roi=roi)

config_path = output_dir / 'dic_config.txt'
if config_path.exists():
    lines = config_path.read_text(encoding='utf-8').splitlines()
    cut = next((i for i, l in enumerate(lines) if l.strip() == '[カラースケール]'), len(lines))
    base = '\n'.join(lines[:cut])
    var_keys = ['u','v','exx','eyy','exy','e1','gamma_max','omega_xy']
    scale_sec = ['[カラースケール]']
    for k in var_keys:
        lo, hi = sc.get(k, (None, None))
        scale_sec.append(f'  {k:<12}: min={lo if lo is not None else "自動"}  max={hi if hi is not None else "自動"}')
    cmap_sec = ['[カラーマップ]']
    for k in var_keys:
        cmap_sec.append(f'  {k:<12}: {_cm(k)}')
    config_path.write_text(
        base + '\n' + '\n'.join(scale_sec) + '\n' + '\n'.join(cmap_sec) + '\n',
        encoding='utf-8')
    print('dic_config.txt を更新しました')

print(f"保存完了: {output_dir}")
"""
    params['dic_module'] = os.path.join(TOOLS_DIR, 'dic_sem_strain_v58.py')
    env = os.environ.copy()
    env['PYTHONIOENCODING'] = 'utf-8'
    subprocess.Popen([sys.executable, '-c', _script, json.dumps(params)], env=env)


@eel.expose
def dic_show_trim_preview(params: dict):
    """REF画像にトリミング範囲を重ねて別プロセスのウィンドウで表示する"""
    import subprocess, sys, json

    _script = r"""
import sys, json, os
import numpy as np
import cv2
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle

p      = json.loads(sys.argv[1])
path   = p['ref_path']
top    = int(p.get('trim_top',    0))
bottom = int(p.get('trim_bottom', 0))
left   = int(p.get('trim_left',   0))
right  = int(p.get('trim_right',  0))
subset_size  = int(p.get('subset_size',  31))
subset_step  = int(p.get('subset_step',  15))
search_range = int(p.get('search_range',  5))

buf = np.fromfile(path, dtype=np.uint8)
img = cv2.imdecode(buf, cv2.IMREAD_GRAYSCALE)
if img is None:
    sys.exit(1)
h, w = img.shape

fig, ax = plt.subplots(figsize=(8, 6))
ax.imshow(img, cmap='gray', vmin=0, vmax=255)
ax.set_xlim(-0.5, w - 0.5)
ax.set_ylim(h - 0.5, -0.5)

def shade(x0, y0, bw, bh):
    ax.add_patch(Rectangle((x0, y0), bw, bh, linewidth=0, facecolor='red', alpha=0.35))

if top    > 0: ax.axhline(y=top-0.5,          color='red',lw=1.5,ls='--'); shade(-0.5,-0.5,      w,       top)
if bottom > 0: ax.axhline(y=h-bottom-0.5,     color='red',lw=1.5,ls='--'); shade(-0.5,h-bottom-0.5, w,  bottom)
if left   > 0: ax.axvline(x=left-0.5,         color='red',lw=1.5,ls='--'); shade(-0.5,-0.5,      left,   h)
if right  > 0: ax.axvline(x=w-right-0.5,      color='red',lw=1.5,ls='--'); shade(w-right-0.5,-0.5,right, h)

# サブセットグリッドを描画（トリミング後の有効領域内）
rw = max(0, w - left - right)
rh = max(0, h - top  - bottom)
half   = subset_size // 2
margin = half + search_range + 5
xs_trim = np.arange(margin, rw - margin, subset_step)
ys_trim = np.arange(margin, rh - margin, subset_step)
if len(xs_trim) > 0 and len(ys_trim) > 0:
    xs_full = xs_trim + left
    ys_full = ys_trim + top
    gx = np.tile(xs_full, len(ys_full))
    gy = np.repeat(ys_full, len(xs_full))
    ax.scatter(gx, gy, s=4, c='cyan', marker='.', linewidths=0, alpha=0.7,
               label=f'subsets: {len(gx)} ({len(xs_full)}x{len(ys_full)})')
    ax.legend(loc='lower right', fontsize=7, framealpha=0.6)

ax.set_title(
    f'{os.path.basename(path)}\n'
    f'Original: {w}x{h}px  ->  Valid: {rw}x{rh}px  |  '
    f'step={subset_step}px  subset={subset_size}px',
    fontsize=8)
ax.axis('off')
fig.tight_layout()
plt.show()
"""

    env = os.environ.copy()
    env['PYTHONIOENCODING'] = 'utf-8'
    subprocess.Popen(
        [sys.executable, '-c', _script, json.dumps(params)],
        env=env,
    )


# ================================================================
# Def EBSD Georef ウィザード用
# ================================================================

@eel.expose
def launch_stress_strain_mapper():
    """Stress-Strain Mapper を起動する（matファイルをダイアログで選択）"""
    mat_path = _tk_filedialog(
        'file',
        'integrated_georef.mat を選択',
        filetypes=[('MAT files', '*.mat'), ('All files', '*.*')],
        initialdir=_work_dir,
    )
    if not mat_path:
        return
    script = os.path.join(TOOLS_DIR, 'stress_strain_mapper_v2.py')
    _env = os.environ.copy()
    _env['QT_API'] = 'PyQt6'
    _env['PYTHONIOENCODING'] = 'utf-8'
    threading.Thread(
        target=lambda: subprocess.Popen(
            [PYTHON, script, '--file', mat_path],
            cwd=TOOLS_DIR, env=_env,
        ).wait(),
        daemon=True,
    ).start()


@eel.expose
def launch_defebsd():
    """Def EBSD Georef ウィザードHTMLを別ウィンドウで開く"""
    eel.start("defebsd_wizard.html", size=(820, 800), block=False)


@eel.expose
def defebsd_browse_file(title: str, filetypes: list, initialdir: str):
    return _tk_filedialog('file', title,
                          filetypes=[tuple(f) for f in filetypes],
                          initialdir=initialdir or _work_dir)


@eel.expose
def defebsd_browse_files(title: str, filetypes: list, initialdir: str):
    """複数ファイルを一括選択して、パスのリストを返す"""
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    paths = filedialog.askopenfilenames(
        title=title,
        filetypes=[tuple(f) for f in filetypes],
        initialdir=initialdir or _work_dir,
    )
    root.destroy()
    return list(paths)


@eel.expose
def defebsd_browse_dir(title: str, initialdir: str):
    return _tk_filedialog('dir', title, initialdir=initialdir or _work_dir)


@eel.expose
def defebsd_get_labels(georef_xlsx: str):
    """dic_results_georef.xlsx の 'u' シートからラベル一覧を取得する"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(georef_xlsx, read_only=True)
        if 'u' not in wb.sheetnames:
            wb.close()
            return []
        ws = wb['u']
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        wb.close()
        skip = {'subset_id', 'x [px]', 'y [px]'}
        return [h for h in headers if h and h not in skip]
    except Exception:
        return []


@eel.expose
def defebsd_start_batch(params: dict):
    threading.Thread(target=_run_defebsd_batch, args=(params,), daemon=True).start()


def _run_defebsd_batch(params: dict):
    import json, tempfile, re as _re
    param_file = tempfile.mktemp(suffix='.json')
    with open(param_file, 'w', encoding='utf-8') as f:
        json.dump(params, f, ensure_ascii=False)

    script = os.path.join(TOOLS_DIR, 'defebsd_georef_v1.py')
    stage_labels = [s['label'] for s in params.get('stages', [])]

    _env = os.environ.copy()
    _env['MPLBACKEND'] = 'QtAgg'
    _env['PYTHONIOENCODING'] = 'utf-8'
    _env['PYTHONUNBUFFERED'] = '1'

    proc = subprocess.Popen(
        [PYTHON, script, param_file],
        cwd=TOOLS_DIR,
        env=_env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding='utf-8',
        errors='replace',
    )
    _procs['defebsd'] = proc

    for line in proc.stdout:
        line_s = line.rstrip()
        print(line_s)

        # ステージ開始: [Stage N/M]  label  (file)
        m = _re.match(r'\[Stage \d+/\d+\]\s+(\S+)', line_s)
        if m:
            lbl = m.group(1)
            if lbl in stage_labels:
                try:
                    eel.defebsd_on_stage_status(lbl, '処理中...', 'ok')()
                except Exception:
                    pass

        # ステージ完了: "  Stage 'label' complete."
        m2 = _re.search(r"Stage '(.+?)' complete", line_s)
        if m2:
            lbl = m2.group(1)
            if lbl in stage_labels:
                try:
                    eel.defebsd_on_stage_status(lbl, '完了', 'done')()
                except Exception:
                    pass

    proc.wait()
    _procs.pop('defebsd', None)

    if proc.returncode == 0:
        try:
            eel.defebsd_on_complete(True, '全ステージ完了')()
        except Exception:
            pass
    else:
        try:
            eel.defebsd_on_complete(False, f'エラー（終了コード {proc.returncode}）')()
        except Exception:
            pass


# ================================================================
# EBSD PatRep ウィザード用
# ================================================================

PATREP_DIR = BASE_DIR


@eel.expose
def launch_patrep():
    eel.start("patrep_wizard.html", size=(700, 780), block=False)


@eel.expose
def patrep_browse_dir(title: str, initialdir: str):
    return _tk_filedialog('dir', title, initialdir=initialdir or _work_dir)


@eel.expose
def patrep_get_info(parent_folder: str):
    """親フォルダを検査して nth 名一覧・Phase 情報・tif フォルダ一覧を返す"""
    try:
        import numpy as _np
        from pathlib import Path as _Path
        sys.path.insert(0, TOOLS_DIR)
        from preprocessed_loader import smart_loadmat

        parent = _Path(parent_folder)

        # pre-processed {name}.mat が存在するものを nth として列挙
        mats = sorted(parent.glob("pre-processed *.mat"))
        nth_names = []
        for m in mats:
            name = m.stem.replace("pre-processed ", "")
            if name.lower() == "0th":
                continue
            xlsx = parent / f"pre-processed {name}.xlsx"
            if xlsx.exists():
                nth_names.append(name)

        # 0th .mat から phase 情報を読む
        mat0_cands = sorted(parent.glob("pre-processed 0th*.mat"))
        if not mat0_cands:
            return {"error": "pre-processed 0th*.mat が見つかりません"}

        mat0 = smart_loadmat(str(mat0_cands[0]), variable_names=["phase_index", "phasetxt"])
        phase_idx_map = mat0["phase_index"]
        phase_names_raw = [str(n) for n in mat0["phasetxt"][0]]
        idxs = sorted(set(
            int(v) for v in phase_idx_map.flatten()
            if not (isinstance(v, float) and _np.isnan(v))
        ))
        phases = [
            {"index": i,
             "name": phase_names_raw[i] if i < len(phase_names_raw) else f"Phase{i}"}
            for i in idxs
        ]

        # tif ファイルを含むサブフォルダを列挙
        tif_folders = sorted(
            d.name for d in parent.iterdir()
            if d.is_dir() and any(d.glob("*.tif"))
        )

        return {"nth_names": nth_names, "phases": phases, "tif_folders": tif_folders}

    except Exception as e:
        return {"error": str(e)}


@eel.expose
def patrep_start_batch(params: dict):
    threading.Thread(target=_run_patrep_batch, args=(params,), daemon=True).start()


def _run_patrep_batch(params: dict):
    import json, tempfile, re as _re
    params["patrep_dir"] = PATREP_DIR

    param_file = tempfile.mktemp(suffix=".json")
    with open(param_file, "w", encoding="utf-8") as f:
        json.dump(params, f, ensure_ascii=False)

    script = os.path.join(TOOLS_DIR, "_patrep_runner.py")
    nth_names = params.get("nth_names", [])

    _env = os.environ.copy()
    _env["PYTHONIOENCODING"] = "utf-8"
    _env["PYTHONUNBUFFERED"] = "1"

    proc = subprocess.Popen(
        [PYTHON, script, param_file],
        cwd=TOOLS_DIR,
        env=_env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    _procs["patrep"] = proc

    for line in proc.stdout:
        line_s = line.rstrip()
        print(line_s)

        # "Processing: name" → RUNNING
        m = _re.search(r"Processing:\s+(.+)", line_s)
        if m:
            name = m.group(1).strip()
            if name in nth_names:
                try: eel.patrep_on_nth_status(name, "running")()
                except Exception: pass

        # "name: 完了" → DONE
        m2 = _re.search(r"^\s+(.+?):\s+完了", line_s)
        if m2:
            name = m2.group(1).strip()
            if name in nth_names:
                try: eel.patrep_on_nth_status(name, "done")()
                except Exception: pass

        # "ERROR [name]" → ERROR
        m3 = _re.search(r"ERROR \[(.+?)\]", line_s)
        if m3:
            name = m3.group(1).strip()
            if name in nth_names:
                try: eel.patrep_on_nth_status(name, "error")()
                except Exception: pass

    proc.wait()
    _procs.pop("patrep", None)

    if proc.returncode == 0:
        try: eel.patrep_on_complete(True, "全処理完了")()
        except Exception: pass
    else:
        try: eel.patrep_on_complete(False, f"エラー（終了コード {proc.returncode}）")()
        except Exception: pass


# ================================================================
# メイン
# ================================================================

if __name__ == "__main__":
    print("=" * 50)
    print("  PineOak v1.0  —  DIC/EBSD Suite")
    print(f"  Tools dir: {TOOLS_DIR}")
    print("=" * 50)
    eel.start(
        "index.html",
        size=(820, 1020),
        port=8765,
        host="localhost",
    )
