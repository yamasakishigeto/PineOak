"""
heaviside_dic_v80.py
===================
v80での修正点:
  キーボードヒントを「←/→ しきい値微調整」のみに変更。
  「a」キーによる結晶粒追加機能を削除。

v79での修正点:
  キーボードショートカットの左右矢印をθ調整からしきい値調整に変更。
  1ステップ = (thr_max - thr_min) / 1000。Shift+矢印は廃止。

v78での修正点:
  ひずみマップの vmin/vmax 範囲外ピクセルを描画しない仕様を削除。

v77での修正点:
  scatter マーカーサイズを ax_main の物理サイズから動的計算。
  リサイズのたびに _redraw を呼び直してマーカーサイズを更新。

v76での修正点:
  不連続線の描画を imshow から scatter に変更。
  ひずみマップと完全に同じ座標系・描画方式になりリサイズでずれない。

v75での修正点:
  不連続線マップの座標を通常DIC変位 (u0,v0) ベースに統一。
  ひずみマップ（scatter）と同じ座標系になり位置ずれが解消。

v74での修正点:
  imshow と scatter の座標系ずれを修正。
  ax_main に set_aspect("equal") を設定し常に1:1スケールを維持。

v73での修正点:
  scatter の alpha を 0.85 -> 1.0 に変更（ベース画像の半透明を解消）。

v72での修正点:
  ベースカラーマップ選択が反映されない問題を再修正。
  trace_add をやめ ComboboxSelected + root.focus_set() + _redraw() に変更。
  base_cmap_cb.get() で確実に選択値を取得する。

v71での修正点:
  ベースカラーマップ選択が反映されない問題を修正。
  Combobox の bind から StringVar の trace_add に変更。
  選択直後に確実に _redraw が呼ばれる。

v70での修正点:
  右パネルが表示されない問題を修正。
  canvas より先に ctrl_outer を pack することで領域を確保。
  ctrl_outer に pack_propagate(False) と固定幅を設定。

v69での修正点:
  【1】リサイズ時に imshow と scatter がずれる問題を修正
    resize_event に set_xlim/ylim の再設定を登録
  【2】ベースカラーマップ選択が効かない問題を修正
    重複していた BASE_CMAPS/DISC_CMAPS 定義を削除
    Combobox に FocusOut バインドを追加
  【3】右パネルをウィンドウ上下中央に配置
    ctrl_outer に上下スペーサーを追加

v68での変更点:
  ベース画像のカラーマップをGUIで選択可能に。
    候補: RdBu_r / hot / cividis / magma / viridis
    ドロップダウン（Combobox）で切り替え、即座に再描画。

v67での修正点:
  結果確認GUIを Tk ベースに全面書き直し。
    - matplotlib figure を左側、Tk Frame を右側に配置
    - 全コントロール（ボタン・入力欄）を tkinter ウィジェットで実装
    - ウィンドウリサイズに入力欄が追従
    - 入力時の重さを解消（tkinter.Entry はキー入力で再描画しない）
    - ボタンが確実に動作

v66での修正点:
  入力フィールドの重さを解消: matplotlib TextBox を廃止し
  tkinter.Entry を canvas に直接 place で配置する方式に変更。
  キー入力のたびに canvas.draw が発火しないため入力が軽快になる。

v65での修正点:
  【1】ひずみベース画像の極端値表示を修正
    vmin/vmax 範囲外の点を scatter から除外して描画
  【2】不連続線をひずみマップの最上面に表示
    imshow に zorder=2 を設定し scatter（zorder=1）より前面に

v64での修正点:
  【1】scatter の範囲外点を除外（端部外れ値対策）
    x_def/y_def が画像範囲 [0,W]x[0,H] 外の点を除いてプロット
  【2】base_vmin/vmax の初期値を _strain_vlim（2〜98パーセンタイル）で設定
  【3】ベース画像の alpha を 0.85->0.5 に下げて不連続線を見えやすく
  【4】レイアウト再設計（ax_main 幅縮小、カラーバー2本を重ならない位置に）
  【5】カラーバーの tick 方向（ベース用=左、不連続線用=右）
  【6】TextBox の on_text_change を無効化して入力時の重さを解消

v63での修正点:
  【1】周囲の極端値 -> ひずみベース画像の vmin/vmax を state で管理・編集可能に
  【2】不連続線が水平になる -> scatter 後に set_xlim/ylim を明示的に再設定
  【3】カラーバーを2本に -> ax_cb_base（ひずみベース用）と ax_cb_disc（不連続線用）
  【4】TextBox 入力が重い -> 「適用」ボタン押下時のみ _redraw（リアルタイム更新なし）
  【5】カラーバーのレイアウト -> マップ左・カラーバー2本中央・コントロール右に再設計

v62での修正点:
  ひずみマップのベース画像を全サブセットの変位補正済み座標で描画。
    u_grid/v_grid で (cx+u0, cy+v0) を計算し scatter でプロット。
    変形後SEM像と座標が一致する。

v61での修正点:
  ひずみマップのベース画像を all_sheet_data のグリッドを imshow で表示するシンプルな実装に変更。
  （変位補正による scatter プロットをやめ、通常DICの結果グリッドをそのまま描画）

v60での変更点:
  【1】結果確認GUIのベース画像（ひずみシート）を変位補正済み座標で描画
    - subset_list の (cx+u0, cy+v0) を変形後座標として scatter でプロット
    - run_heaviside_dic の戻り値に subset_list を追加
    - _show_result_gui の引数を extra_imgs -> strain_data に変更
  【2】マップが再描画のたびに縮小する問題を修正
    - カラーバー軸 ax_cb を固定位置で一度だけ作成し cax= で更新
    - ax_main の位置・サイズが変化しないよう fig.add_axes で固定

v59での変更点:
  【機能追加】計算後に結果確認GUIを起動（PNG自動保存を廃止）
    - 表示データを delta_s / delta_n から選択可能
    - ベース画像を SEM像 / 各ひずみシート から選択可能
    - vmin / vmax を TextBox で直接入力して「適用」
    - 「自動設定」で 2〜98 パーセンタイルに戻す
    - 「PNG 保存」で現在の設定通りに保存してウィンドウを閉じる
    - run_heaviside_dic の戻り値を (results, deformed) に変更

v58での修正点:
  【バグ修正】あるしきい値を変えると他の粒のサブセット数も変わる問題を修正。
    _draw_list と hdic_config.txt のサブセット数集計で self.disc_mask
    （共通マスク）を使っていたため、どの粒も現在のスライダー値で
    カウントされていた。grain_thr_map の粒ごとのしきい値で
    self.map_grid を直接参照して個別に集計するよう修正。

v57での修正点:
  【バグ修正】しきい値が全結晶粒で同一値になっていた問題を修正。
    run_heaviside_dic に grain_thr_map 引数を追加し、
    サブセット抽出を粒ごとのしきい値で個別に行うよう変更。
    ターミナルに各粒のしきい値とサブセット数を表示。

v56での修正点:
  しきい値スライダーの範囲を 0〜99パーセンタイルに変更。
    最小値: 0パーセンタイル（実データの最小値と同等）
    最大値: 99パーセンタイル（上位1%の外れ値を除外）

v55での修正点:
  【1】しきい値スライダーの範囲をパーセンタイルから実データの最小〜最大値に変更
    上限が98パーセンタイルで頭打ちになっていたため高ひずみ側の
    外れ値除外ができなかった問題を修正。
  【2】プレビューマップに水色の線が2本表示される問題を修正
    _update_hough と _on_slider の両方で線を描いていたため2本になっていた。
    _update_hough 側の線描画を削除し _on_slider のみに統一。
    （slider.set_val 呼び出しで _on_slider が発火するため線は1本描かれる）

v54での修正点:
  【バグ修正】_on_slider 内のリスト内包表記で NameError が発生する問題を修正。
    xs[c], ys[r] -> xs[c_idx], ys[r_idx] に修正。

v53での修正点:
  【1】しきい値スライダーの valtext がウィンドウ右端にはみ出る問題を修正
    スライダー軸幅を RW -> RW*0.82 に縮小して右側に余白を確保
    （θスライダーも同様に修正）
  【2】gamma_max / omega_xy ボタンの文字がはみ出る問題を修正
    マップ選択ボタンのフォントサイズを 9.5 -> 7.5 に縮小
  【3】使い方テキストの誤記修正
    「走向角」-> 「不連続線の角度」
    「粒を追加」-> 「結晶粒を追加」

v52での修正点:
  マップ切り替え時のしきい値スライダー範囲が正しく更新されない問題を修正。
    matplotlib の Slider は valmin/valmax を直接書き換えても
    トラック（バー）の描画範囲（ax.xlim）が追従しないため、
    _on_map_btn で ax.set_xlim(thr_min, thr_max) も同時に更新するよう修正。

v51での修正点:
  gamma_max に依存した変数名・引数名を中立的な名前に変更。
    gmax_threshold -> disc_threshold（run_heaviside_dic 引数）
    gmax_cols      -> init_cols
    gmax_grid      -> init_grid
    GMAX_THR_INIT  -> THR_INIT
    self.gmax_thr  -> self.disc_thr
    gui.gmax_thr   -> gui.disc_thr

v50での修正点:
  【バグ修正】disc_mask が常に gamma_max シートで生成されていた問題を修正。
    GUI でどのマップシートを選択していても、run_heaviside_dic 内で
    gamma_max を固定読み込みしていたため、exx 等に切り替えても
    実計算では gamma_max のしきい値判定が使われていた。
    - run_heaviside_dic に disc_sheet 引数を追加
    - 呼び出し時に gui.map_sheet を渡すよう修正
    - シートが存在しない場合は gamma_max にフォールバック
    - ターミナル出力のシート名も実際のシート名を表示するよう修正

v49での修正点:
  hdic_config.txt の [HoughGUI設定] を GUI登録リストと同じ形式に変更。
    - 共通しきい値の記載を廃止
    - 登録粒ごとに theta [deg] / threshold / n_subsets を記載

v48での修正点:
  v47で上マップに追加したtheta線描画を削除。
  スライダーと連動して線が動くのは一番下のプレビューマップのみ（元の仕様に戻す）。

v47での修正点:
  【1】θスライダーを動かすと上マップの水色直線が更新されない問題を修正
    - _draw_strain_map 内に選択粒のtheta線描画を追加
    - _on_slider から _draw_strain_map を呼び出すよう修正
  【2】登録リストのしきい値を「追加時点の値」として粒ごとに記録・表示
    - grain_thr_map {grain_id: thr} を追加
    - _on_add でしきい値を保存、_on_delete で削除
    - リスト上部の共通しきい値表示を廃止し粒ごとに表示
  【3】リスト中の角度表記を「theta=+75.0 [deg]」形式に変更
  【4】GUI文字サイズをさらに2段階拡大（+4pt）

v46での修正点:
  ターミナル出力の文字化け対策
    print文内の特殊文字を ASCII に置換:
      θ -> theta, × -> x, ° -> deg, γ -> gamma, → -> ->  ! -> !

v45での修正点:
  【1】GUI文字サイズを全体的に拡大（fontsize 8-9 → 10-11）
  【2】「閾値スライダー」→「しきい値」に変更
  【3】「走向角 θ [deg]」→「不連続線の角度 θ [deg]」に変更
  【4】登録リストの文字化け修正（family='monospace' を除去）
  【5】登録リストにしきい値・対象サブセット数を追記
  【6】「＋ 粒を追加」→「＋ 結晶粒を追加」に変更

v44での変更点:
  GUI内のバージョン番号表記をすべて削除。
  ウィンドウタイトル・ラベル・設定ファイル・PNGタイトルからバージョン文字列を除去。
  以降の修正はすべてバージョン番号をインクリメントして管理する。

v42での改良点:
  【1】プレビュー直線を選択粒のdisc点群の重心・広がりを基準に描画
    - θスライダー操作時の不連続線プレビューがマップ全体中央ではなく
      選択粒のdisc点の重心を中心とした適切な長さで描かれるよう修正
  【2】「削除」ボタンの挙動改善
    - 常に最後の粒を削除していたのを「現在選択中の粒を優先削除」に変更
    - 未選択時は従来どおり最後の粒を削除
  【3】キーボードショートカット追加
    - a: 粒を追加（Addボタンと同等）
    - d: 選択粒を削除（Deleteボタンと同等）
    - ←/→: θ を ±0.5° 微調整
    - Shift+←/→: θ を ±5° 大きく調整
  【4】シリアル実行時のマスクキャッシュ最適化
    - 粒ごとに build_mask_cache() を事前生成してサブセット間で共有
    - process_one_subset に _mask_cache 引数を追加（joblib 並列時は使わない）
    - 粒数が少ない場合に重複マスク生成を排除して高速化

v41での改良点（GUI・可視化のさらなるブラッシュアップ）:
  【1】HoughGUI マップ上の粒ハイライト改善
    - 登録済み粒の重心位置に grain_id 番号を黄色ラベルで表示
    - 現在選択中の粒をシアン色でハイライト（未登録でも視認可能）
    - タイトルバーに「登録粒:N件」をリアルタイム表示
  【2】計算進捗ウィンドウ追加（シリアル実行時）
    - joblib 非使用時に Tkinter プログレスバー付きウィンドウを表示
    - 5サブセットごとに更新、完了後0.8秒で自動クローズ
  【3】結果 PNG の大幅改良
    - 1枚に delta_s（走向すべり）と delta_n（法線方向変位）の2パネル構成
    - ダークテーマ背景（#1c1c2e）で視認性向上
    - 各粒の重心位置に grain_id / θ ラベルを重ねて表示
    - 図下部に n・mean・std の統計情報テキストを自動付与

v40での改良点（GUI全面ブラッシュアップ）:
  【1】ファイル選択ランチャー（Tkinter）をウィザード形式に刷新
    - 全ステップ（xlsx / config / alignment / 出力先）を1つのダイアログにまとめた
    - 進行バー（ttk.Progressbar）で現在ステップを表示
    - 選択済みパスをラベルでリアルタイム表示
  【2】応力段階選択ダイアログの強化
    - ラジオボタン一覧をスクロール可能リストボックスに変更
    - フォントサイズ・パディングを拡大し視認性向上
  【3】HoughGUI レイアウト全面改修
    - GridSpec を 3×2 に再設計（左列: マップ/Hough/プレビュー、右列: コントロール）
    - 右コラム内のウィジェット間隔を均等化
    - OFFSET_MAX / MIN_SIDE_PX の入力欄をラベル付きで見やすく配置
    - マップ選択ボタンを見やすいサイズ・配色に調整
    - 「粒を追加」「削除」「H-DIC実行」ボタンの色分け（緑/灰/橙）
    - 登録リスト表示を ax_list で更新時のちらつきを軽減
  【4】PNG タイトルのバージョン番号を v39 → v40 に修正

v38での修正点（HoughGUI ウィジェットをmatplotlibネイティブに統一）:
  Tkinter の Combobox/Spinbox が Windows 環境で正しく表示されなかった問題を修正。
  全ウィジェットを matplotlib ネイティブに置き換え:
    マップ選択: シート名ごとのトグルボタンを横並び配置（選択中は濃いグレー）
    OFFSET_MAX / MIN_SIDE_PX: TextBox で直接数値入力
  「H-DIC実行」ボタン押下時に TextBox の値を読み取って確定。

v37での修正点:
  【1】スキップされたサブセットも Excel に出力
    process_one_subset が None を返したサブセット（全組み合わせでスキップされたもの）
    を 'skipped' 列で識別できるよう Excel に記載。offset 等は空欄。
  【2】MIN_SIDE_PX のデフォルト値を SUBSET_SIZE^2 * 0.2 に変更
    31×31×0.2 ≈ 192px。Side A/B のどちらかがこれ未満ならスキップ。
  【3】パラメータ設定ダイアログの推奨値・連動表示を削除
    シンプルな OFFSET_MAX / MIN_SIDE_PX スピンボックスのみに。
  【4】OFFSET_MAX / MIN_SIDE_PX の設定を HoughGUI 内に統合
    別ダイアログ（ステップ3.5）を廃止し HoughGUI の右下パネルに組み込み。
  【5】HoughGUI で設定した全パラメータをテキストファイルに保存
    出力フォルダに hdic_config_{label}.txt を出力。
  【6】HoughGUI のマップ表示をプルダウンで選択可能に
    gamma_max 固定から u/v/ncc/exx/eyy/exy/e1/gamma_max/omega_xy の
    任意シート×任意応力段階をプルダウンで選択できるよう変更。
    閾値スライダーの範囲はシート切り替え時に自動更新。
  【7】filter_by_neighbors の内積符号チェック追加（v37改）
    内積 ≤ 0 の候補（逆方向・直交）をスキップ。
    正方向に隣がなければ「隣なし→保持」と明示的に判断。

v35での修正点（filter_by_neighbors の隣座標計算バグ修正）:
  旧実装では raw座標を STEP で割って丸める方式だったが、サブセット座標が
  STEP の倍数とは限らないため result_map に全くヒットせず、フィルタが
  v31以降ずっと機能していなかった。
  → 上下左右4候補（±STEP）のうち法線方向への内積が最大のものを選ぶ方式に変更。
     サブセット座標に依存せず確実に隣を検索できる。

v34での修正点（パラメータ設定GUIの追加、joblib対応）:
  HoughGUI終了後にパラメータ設定ダイアログを追加。
    - dic_config.txt から読み取った SUBSET_SIZE・STEP を参照表示
    - OFFSET_MAX スピンボックス → 変更すると Side B 最小px と
      MIN_SIDE_PX 推奨値を自動計算して連動表示
    - MIN_SIDE_PX スピンボックス → 変更すると有効な最大 offset を自動計算
  process_one_subset / ncc_masked_search に offset_max・min_side_px 引数を追加。
  run_heaviside_dic にも同引数を追加し joblib サブプロセスに正しく渡す。
  これにより GUI で設定した値がサブプロセスにも確実に反映される。

v33での修正点（一時表示GUIを削除、PNG自動保存に戻す）:
  show_result_gui を削除。計算完了後に _save_png で自動保存する形に復帰。
  _save_png は v32 の変形後座標描画（_build_disc_line_map）を維持。

v32での修正点（描画座標を変形後座標に統一）:
  _build_disc_line_map / show_result_gui の描画位置を変形後座標に修正。
    不連続線: 参照座標 + (u_A + u_B) / 2
    全対象サブセット: 参照座標 + 通常DIC変位 (u0, v0)
    境界到達サブセット: 参照座標 + (u_A + u_B) / 2
  変形後SEM画像と正しく位置が一致するようになった。

v31での修正点（dic_config.txt 読み込みによるパラメータ自動設定）:
  通常DICの設定ファイル dic_config.txt を読み込み、以下のパラメータを自動上書き:
    SUBSET_SIZE  ← [共通] subset
    STEP         ← [Stage 2] step
    TRIM_BOTTOM  ← trim_bottom
    NCC_THRESHOLD← [共通] NCC閾値
    OFFSET_MAX   ← SUBSET_SIZE // 2 - 1  （サブセット端まで探索、MIN_SIDE_PX制約内）
  dic_config.txt はステップ1でスキップ可能（スキップ時はハードコード値を使用）。
  ファイル読み込み後、適用したパラメータをコンソールに表示。
  offset境界到達フィルタを追加: |offset| == offset_max_int のサブセットを
  後処理で除外し、境界張り付き誤検出を排除する。

v30での修正点（バグ修正・非効率コード改善）:
  【Bug 1】estimate_theta() 戻り値の個数不一致（クラッシュ確定）
    hspace.max()==0 のフォールバック時に4値を返していたが、
    呼び出し元は5値でアンパックするため ValueError でクラッシュ。
    → return 0.0, 0.0, hspace, theta_angles, d の5値に統一。

  【Bug 2】vlim 計算が非対称（カラーバーずれ）
    np.max(np.abs(np.percentile(valid_ds, [2, 98]))) は
    正側・負側の絶対値の大きい方を採用するだけで、コメントの「対称・中央=0」
    と実質的に整合しない場合がある。
    → p2, p98 を個別に取得し max(abs(p2), abs(p98)) に修正。

  【Bug 3】_on_slider のcyan線削除が色文字列に依存（脆弱）
    line.get_color() == 'cyan' はmatplotlib内部表現に依存し
    '#00ffff' 等で不一致になる場合がある。
    → axvline の戻り値を self._hough_vline / self._preview_line に保持し
      直接 remove() する方式に変更。

  【Bug 4】nx, ny 変数が process_one_subset で未使用（デッドコード）
    → tx/ty/nx/ny の個別変数をやめ、np.array のベクトル演算に整理。

  【非効率 1】_save_png の二重Pythonループを NumPy 化
    on_line マスクの行列演算で一括処理するよう置き換え。
    サブセット数×(31×31)回のループが 1回の NumPy 演算になり高速化。

  【非効率 2】make_heaviside_mask をキャッシュ化
    mask は (theta, offset) の関数のみで画像と無関係なため
    functools.lru_cache でメモ化。165回×サブセット数分の配列生成を削減。

  【非効率 3】_on_slider の未使用変数 y0 を削除

v9での修正点（根本的なファイル選択フローの再設計）:
  - __main__ を全面書き直し。任意のdic_results.xlsxに対応。
    ① dic_results.xlsx を選択
    ② Excelの列名を自動読み取り → チェックボックスで処理段階を複数選択
    ③ 参照SEM画像 (Ref) を選択
    ④ 選択した各段階のSEM画像を順番に選択
    ⑤ 出力先フォルダを選択
    ⑥ 全段階を順番に処理 → heaviside_results_{label}.xlsx + _delta_s.png を出力
  - label を「SEM画像ファイル名のstem」から「Excelの列名」に変更
    （旧: Path(def_path).stem → 新: チェックボックスで選択した列名）
  - _save_png() のタイトルにlabelを表示するよう修正

v2からの修正点（v3で実装、v4でGUIファイル選択に対応、v5でNCC_THRESHOLD追加）:
  - offsetを「y方向」定義から「法線方向」定義に変更
    旧: rr - cc*tan(θ) = offset_y  （y方向距離）
    新: cc*(-sin(θ)) + rr*cos(θ) = offset_n  （法線方向距離）
  - OFFSET_MAXをSTEP/√2 ≈ 10.6pxに変更（θ非依存・格子間隔の半分/√2）
  - NCC_THRESHOLD = 0.2 追加

後処理フィルタ: filter_by_neighbors()
  全サブセット処理後に、隣接サブセット比較によるフィルタリングを実施する。
  不連続線のoffset方向（法線ベクトル n = (-sinθ, cosθ) 方向）に存在する
  隣接サブセットのみを評価対象とする（v31改善）。
  その隣が解析対象（result_mapに存在）かつ以下の条件を両方満たす場合、
  当該サブセットを除外する（不連続線から外れた誤検出と判断）:
    条件1: |offset_neighbor| < |offset_self|   （隣の方がoffsetが小さい）
    条件2: score_neighbor    > score_self       （隣の方がNCCスコアが高い）
  隣が解析対象でない場合はフィルタを適用しない（端部・γmax低い場合も保持）。
  注意: 不連続線の端部サブセットは隣が存在しないため除外されない。

座標系定義（変更禁止）:
  画像座標系: x右向き、y下向き
  走向ベクトル: t = (cos(θ), sin(θ))
  法線ベクトル: n = (-sin(θ), cos(θ))  ← 走向を反時計回り90°
  不連続線の式: cc*(-sin(θ)) + rr*cos(θ) = offset_n
  Side A: cc*(-sin(θ)) + rr*cos(θ) < offset_n
  Side B: cc*(-sin(θ)) + rr*cos(θ) >= offset_n
  on_line: |cc*(-sin(θ)) + rr*cos(θ) - offset_n| < 1.5
"""

import os
import numpy as np
import cv2
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.font_manager as _fm
import platform as _platform
from pathlib import Path
import openpyxl
from skimage.transform import hough_line, hough_line_peaks

# 日本語フォント設定
_JP_FONTS = {
    'Windows': ['Meiryo', 'Yu Gothic', 'MS Gothic'],
    'Darwin':  ['Hiragino Sans', 'Hiragino Maru Gothic Pro'],
    'Linux':   ['IPAexGothic', 'Noto Sans CJK JP'],
}
_avail = {f.name for f in _fm.fontManager.ttflist}
_jp = next((f for f in _JP_FONTS.get(_platform.system(), []) if f in _avail), None)
if _jp:
    matplotlib.rcParams['font.family'] = _jp
matplotlib.rcParams['axes.unicode_minus'] = False

try:
    from joblib import Parallel, delayed
    JOBLIB_AVAILABLE = True
except ImportError:
    JOBLIB_AVAILABLE = False

# =============================================================================
# パラメータ
# =============================================================================
SUBSET_SIZE   = 31
STEP          = 15
MAD_K         = 5.0
THETA_RANGE   = 5.0
THETA_STEP    = 1.0
OFFSET_MAX    = float(STEP)          # = 15px。dic_config.txt読込後に上書きされる
OFFSET_STEP   = 1
MIN_SIDE_PX   = int(SUBSET_SIZE * SUBSET_SIZE * 0.2)  # ≈192px（20%）
NCC_THRESHOLD = 0.2
SEARCH_RANGE  = 3      # 通常DIC初期値からの探索範囲 [±px]
TRIM_BOTTOM   = 128    # SEM情報バーのトリミング [px] ← dic_sem_strain_v57と合わせる

# =============================================================================
# dic_config.txt パーサー
# =============================================================================
def load_dic_config(config_path):
    """
    通常DICの設定ファイル dic_config.txt を読み込み、グローバルパラメータを更新する。

    読み取り項目:
      trim_bottom      → TRIM_BOTTOM
      [Stage 2] step   → STEP
      [共通]   subset  → SUBSET_SIZE
      [共通]   NCC閾値 → NCC_THRESHOLD
    OFFSET_MAX は読み込み後に float(STEP) で再計算する。

    戻り値: dict（適用した値のサマリー）
    """
    global SUBSET_SIZE, STEP, TRIM_BOTTOM, NCC_THRESHOLD, OFFSET_MAX

    current_section = None
    applied = {}

    with open(config_path, encoding='utf-8') as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith('=') or line.startswith('DIC'):
                continue

            # セクション行
            if line.startswith('['):
                current_section = line.strip('[]').strip()
                continue

            # "キー : 値 単位" の形式を分割
            if ':' not in line:
                continue
            key, _, rest = line.partition(':')
            key  = key.strip()
            val_str = rest.strip().split()[0]  # 先頭トークン（数値部分）

            try:
                val = float(val_str)
            except ValueError:
                continue

            # セクション外: trim_bottom
            if current_section is None and key == 'trim_bottom':
                TRIM_BOTTOM = int(val)
                applied['TRIM_BOTTOM'] = TRIM_BOTTOM

            # [Stage 2]: step → STEP
            elif current_section == 'Stage 2' and key == 'step':
                STEP = int(val)
                applied['STEP'] = STEP

            # [共通]: subset → SUBSET_SIZE
            elif current_section == '共通' and key == 'subset':
                SUBSET_SIZE = int(val)
                applied['SUBSET_SIZE'] = SUBSET_SIZE

            # [共通]: NCC閾値 → NCC_THRESHOLD
            elif current_section == '共通' and key == 'NCC閾値':
                NCC_THRESHOLD = float(val)
                applied['NCC_THRESHOLD'] = NCC_THRESHOLD

    # OFFSET_MAX を STEP で再計算
    OFFSET_MAX = float(STEP)
    applied['OFFSET_MAX'] = OFFSET_MAX

    return applied


# =============================================================================
# データ読み込み
# =============================================================================
def load_sheet(ws):
    """
    u/v シートを読み込む。新旧2形式に自動対応。
      新形式: subset_id | x [px] | y [px] | Ref | Def_1st | ...
      旧形式: x [px]    | y [px] | Ref | Def_1st | ...
    ヘッダ先頭列が 'subset_id' かどうかで判別する。
    """
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1:]
    if header[0] == 'subset_id':
        # 新形式: 列0=subset_id, 列1=x, 列2=y, 列3〜=データ
        x = np.array([r[1] for r in data], dtype=float)
        y = np.array([r[2] for r in data], dtype=float)
        cols = {header[i]: np.array([r[i] if r[i] is not None else np.nan
                                      for r in data], dtype=float)
                for i in range(3, len(header)) if header[i] is not None}
    else:
        # 旧形式: 列0=x, 列1=y, 列2〜=データ
        x = np.array([r[0] for r in data], dtype=float)
        y = np.array([r[1] for r in data], dtype=float)
        cols = {header[i]: np.array([r[i] if r[i] is not None else np.nan
                                      for r in data], dtype=float)
                for i in range(2, len(header)) if header[i] is not None}
    return x, y, cols

def to_grid(x, y, vals):
    xs = np.unique(x)
    ys = np.unique(y)
    xi = {v: i for i, v in enumerate(xs)}
    yi = {v: i for i, v in enumerate(ys)}
    grid = np.full((len(ys), len(xs)), np.nan)
    for cx, cy, val in zip(x, y, vals):
        grid[yi[cy], xi[cx]] = val
    return grid, xs, ys

# =============================================================================
# 変位差マップ + MAD閾値
# =============================================================================
def calc_disc(u_grid, v_grid):
    du_x = np.full_like(u_grid, np.nan)
    du_x[:, 1:] = np.abs(u_grid[:, 1:] - u_grid[:, :-1])
    du_y = np.full_like(u_grid, np.nan)
    du_y[1:, :] = np.abs(u_grid[1:, :] - u_grid[:-1, :])
    dv_x = np.full_like(v_grid, np.nan)
    dv_x[:, 1:] = np.abs(v_grid[:, 1:] - v_grid[:, :-1])
    dv_y = np.full_like(v_grid, np.nan)
    dv_y[1:, :] = np.abs(v_grid[1:, :] - v_grid[:-1, :])
    return np.nanmax(np.stack([du_x, du_y, dv_x, dv_y]), axis=0)

def mad_threshold(disc, k=MAD_K):
    valid = disc[~np.isnan(disc)]
    med = np.median(valid)
    mad = np.median(np.abs(valid - med))
    return med + k * mad

# =============================================================================
# Hough変換
# =============================================================================
def estimate_theta(disc_mask):
    """
    disc_maskにHough変換を適用して走向角theta [deg] を推定する。
    ピーク0件時のフォールバック付き（IndexError修正）。
    戻り値: (theta_deg, raw_hough_deg, hspace, theta_angles, d)
      theta_deg    : 推定走向角 [-90, 90]（H-DICのtheta_centerとして使用）
      raw_hough_deg: Hough空間x軸座標（axvline用）
      hspace       : Houghアキュムレータ（可視化用）
      theta_angles, d: Hough変換の角度・距離配列（可視化用）
    """
    binary = disc_mask.astype(np.uint8)
    tested_angles = np.linspace(-np.pi/2, np.pi/2, 360, endpoint=False)
    hspace, theta_angles, d = hough_line(binary, theta=tested_angles)

    if hspace.max() == 0:
        # Bug 1修正: 呼び出し元は5値でアンパックするため5値を返す
        return 0.0, 0.0, hspace, theta_angles, d

    # Step1: 通常閾値
    _, angles, _ = hough_line_peaks(hspace, theta_angles, d, num_peaks=1,
                                    min_distance=20, min_angle=10,
                                    threshold=0.3 * hspace.max())
    if len(angles) == 0:
        # Step2: 閾値を緩めて再試行
        _, angles, _ = hough_line_peaks(hspace, theta_angles, d, num_peaks=1,
                                        min_distance=1, min_angle=1,
                                        threshold=0.1 * hspace.max())
    if len(angles) == 0:
        # Step3: アキュムレータ最大値位置から直接取得
        peak_idx = np.unravel_index(hspace.argmax(), hspace.shape)
        angles = [theta_angles[peak_idx[1]]]
        print("  estimate_theta: フォールバック使用（アキュムレータ最大値）")

    # raw_hough_deg: Hough空間のx軸座標（-90°〜+90°）= axvline用
    raw_hough_deg = float(np.degrees(angles[0]))

    # theta_deg: 走向角（-90°〜+90°）= H-DIC の theta_center として使う値
    theta_deg = raw_hough_deg + 90
    if theta_deg > 90:
        theta_deg -= 180
    return theta_deg, raw_hough_deg, hspace, theta_angles, d

# =============================================================================
# Heavisideマスク
# =============================================================================
def make_heaviside_mask(subset_size, theta_deg, offset):
    """
    座標系定義（変更禁止）
    画像座標系: x右向き、y下向き
    cc: サブセット中心からのx距離、rr: y距離

    不連続線の式（法線方向offset定義）:
      cc*(-sin(θ)) + rr*cos(θ) = offset
      offset = 法線方向 n=(-sin(θ), cos(θ)) への射影距離 [px]

    Side A: cc*(-sin(θ)) + rr*cos(θ) < offset
    Side B: cc*(-sin(θ)) + rr*cos(θ) >= offset
    """
    half = subset_size // 2
    cols = np.arange(-half, half + 1)
    rows = np.arange(-half, half + 1)
    cc, rr = np.meshgrid(cols, rows)
    theta_rad = np.radians(theta_deg)
    proj = cc * (-np.sin(theta_rad)) + rr * np.cos(theta_rad)
    mask_A = proj < offset
    return mask_A, ~mask_A


def build_mask_cache(theta_center):
    """
    非効率2修正（v30改）: process_one_subset の呼び出し前に
    (theta, offset) の全組み合わせのマスクを辞書として事前生成して返す。
    lru_cache はjoblibのLokyバックエンド（サブプロセス）でpickle不可のため、
    サブプロセス内ローカルの辞書キャッシュ方式に変更。
    """
    thetas = np.arange(theta_center - THETA_RANGE,
                       theta_center + THETA_RANGE + THETA_STEP * 0.5,
                       THETA_STEP)
    offset_max_int = int(np.floor(OFFSET_MAX))
    offsets = np.arange(-offset_max_int, offset_max_int + 1, OFFSET_STEP)
    cache = {}
    for theta in thetas:
        for offset in offsets:
            key = (float(theta), int(offset))
            cache[key] = make_heaviside_mask(SUBSET_SIZE, theta, offset)
    return cache

# =============================================================================
# マスク付きNCC + 整数ピクセル探索
# =============================================================================
def ncc_masked_search(ref, deformed, cx, cy, subset_size,
                      u_init, v_init, mask, min_side_px=None):
    """
    mask内のピクセルのみ使ってNCC計算し、
    u_init±SEARCH_RANGE の範囲で最良変位を探索。
    戻り値: (best_ncc, best_du, best_dv)
    """
    if min_side_px is None:
        min_side_px = MIN_SIDE_PX
    half = subset_size // 2
    if int(np.sum(mask)) < min_side_px:
        return np.nan, np.nan, np.nan

    # 参照パッチ（マスク適用）
    ref_patch = ref[cy-half:cy+half+1, cx-half:cx+half+1].astype(np.float32)
    ref_masked = ref_patch[mask]
    ref_mean = ref_masked.mean()
    f = ref_masked - ref_mean
    f_norm = np.sqrt(np.sum(f**2))
    if f_norm < 1e-7:
        return np.nan, np.nan, np.nan

    du_init_int = int(round(u_init))
    dv_init_int = int(round(v_init))

    best_ncc = -np.inf
    best_du  = du_init_int
    best_dv  = dv_init_int

    for dv in range(dv_init_int - SEARCH_RANGE, dv_init_int + SEARCH_RANGE + 1):
        for du in range(du_init_int - SEARCH_RANGE, du_init_int + SEARCH_RANGE + 1):
            cy_def = cy + dv
            cx_def = cx + du
            if (cy_def - half < 0 or cy_def + half >= deformed.shape[0] or
                    cx_def - half < 0 or cx_def + half >= deformed.shape[1]):
                continue
            def_patch = deformed[cy_def-half:cy_def+half+1,
                                  cx_def-half:cx_def+half+1].astype(np.float32)
            def_masked = def_patch[mask]
            def_mean = def_masked.mean()
            g = def_masked - def_mean
            g_norm = np.sqrt(np.sum(g**2))
            denom = f_norm * g_norm
            if denom < 1e-7:
                continue
            ncc_val = float(np.sum(f * g) / denom)
            if ncc_val > best_ncc:
                best_ncc = ncc_val
                best_du  = du
                best_dv  = dv

    if best_ncc < NCC_THRESHOLD:
        return np.nan, np.nan, np.nan

    # サブピクセル精度（2D放物線フィット）
    du_sub, dv_sub = subpixel_fit(
        ref, deformed, cx, cy, subset_size, mask,
        f, f_norm, best_du, best_dv)

    return best_ncc, du_sub, dv_sub

def subpixel_fit(ref, deformed, cx, cy, subset_size, mask,
                 f, f_norm, best_du, best_dv):
    """3×3近傍のNCC値から2D放物線フィットでサブピクセル精度向上"""
    half = subset_size // 2
    ncc_3x3 = np.zeros((3, 3))
    for dv_off, dv in enumerate([best_dv-1, best_dv, best_dv+1]):
        for du_off, du in enumerate([best_du-1, best_du, best_du+1]):
            cy_def = cy + dv
            cx_def = cx + du
            if (cy_def - half < 0 or cy_def + half >= deformed.shape[0] or
                    cx_def - half < 0 or cx_def + half >= deformed.shape[1]):
                ncc_3x3[dv_off, du_off] = 0.0
                continue
            def_patch = deformed[cy_def-half:cy_def+half+1,
                                  cx_def-half:cx_def+half+1].astype(np.float32)
            g = def_patch[mask] - def_patch[mask].mean()
            g_norm = np.sqrt(np.sum(g**2))
            denom = f_norm * g_norm
            ncc_3x3[dv_off, du_off] = (float(np.sum(f * g) / denom)
                                        if denom > 1e-7 else 0.0)

    # 2D放物線フィット
    pts, vals = [], []
    for dv in [-1, 0, 1]:
        for du in [-1, 0, 1]:
            pts.append([du*du, dv*dv, du*dv, du, dv, 1.0])
            vals.append(ncc_3x3[dv+1, du+1])
    A = np.array(pts)
    b_vec = np.array(vals)
    try:
        coeffs, _, _, _ = np.linalg.lstsq(A, b_vec, rcond=None)
        a, b, c, d, e, _ = coeffs
        denom = 4*a*b - c*c
        if abs(denom) > 1e-10:
            delta_u = float(np.clip((c*e - 2*b*d) / denom, -1.0, 1.0))
            delta_v = float(np.clip((c*d - 2*a*e) / denom, -1.0, 1.0))
            return best_du + delta_u, best_dv + delta_v
    except Exception:
        pass
    return float(best_du), float(best_dv)

# =============================================================================
# 1サブセットのHeaviside DIC
# =============================================================================
def process_one_subset(ref, deformed, cx, cy, u_init, v_init, theta_center,
                       offset_max=None, min_side_px=None, _mask_cache=None):
    if offset_max   is None: offset_max   = OFFSET_MAX
    if min_side_px  is None: min_side_px  = MIN_SIDE_PX
    half = SUBSET_SIZE // 2
    if (cy - half < 0 or cy + half >= ref.shape[0] or
            cx - half < 0 or cx + half >= ref.shape[1]):
        return None

    thetas  = np.arange(theta_center - THETA_RANGE,
                        theta_center + THETA_RANGE + THETA_STEP * 0.5,
                        THETA_STEP)
    offset_max_int = int(np.floor(offset_max))
    offsets = np.arange(-offset_max_int, offset_max_int + 1, OFFSET_STEP)

    # 外部キャッシュがあればそれを使う（シリアル実行で粒ごとに再利用）
    # なければサブプロセス内ローカルキャッシュを生成
    mask_cache = _mask_cache if _mask_cache is not None else {}

    best_score  = -np.inf
    best_result = None

    for theta in thetas:
        for offset in offsets:
            key = (float(theta), int(offset))
            if key not in mask_cache:
                mask_cache[key] = make_heaviside_mask(SUBSET_SIZE, theta, offset)
            mask_A, mask_B = mask_cache[key]
            if mask_A.sum() < min_side_px or mask_B.sum() < min_side_px:
                continue

            ncc_A, u_A, v_A = ncc_masked_search(
                ref, deformed, cx, cy, SUBSET_SIZE, u_init, v_init, mask_A,
                min_side_px=min_side_px)
            ncc_B, u_B, v_B = ncc_masked_search(
                ref, deformed, cx, cy, SUBSET_SIZE, u_init, v_init, mask_B,
                min_side_px=min_side_px)

            if np.isnan(ncc_A) or np.isnan(ncc_B):
                continue

            score = ncc_A + ncc_B
            if score > best_score:
                best_score = score
                best_result = dict(cx=cx, cy=cy, theta=theta, offset=offset,
                                   u_A=u_A, v_A=v_A, u_B=u_B, v_B=v_B,
                                   ncc_A=ncc_A, ncc_B=ncc_B)

    if best_result is None:
        return None

    theta_rad = np.radians(best_result['theta'])
    du = best_result['u_A'] - best_result['u_B']
    dv = best_result['v_A'] - best_result['v_B']

    cos_t = np.cos(theta_rad)
    sin_t = np.sin(theta_rad)
    best_result['delta_u'] = du
    best_result['delta_v'] = dv
    best_result['delta_s'] = du * cos_t  + dv * sin_t
    best_result['delta_n'] = du * (-sin_t) + dv * cos_t
    return best_result

# =============================================================================
# 後処理フィルタ: 隣接サブセット比較
# =============================================================================
def filter_by_neighbors(results):
    """
    不連続線がずれている方向の隣接サブセットのみと比較し、
    その隣が解析対象かつ「より中心に近い（|offset|が小さく、scoreも高い）」
    場合に自サブセットを除外する。

    v35修正: 隣座標の計算方法を修正。
      旧（バグあり）: raw座標を STEP で割って丸める方式
        → サブセット座標が STEP の倍数とは限らないため result_map にヒットせず
           フィルタが全く機能していなかった
      新（修正後）: 上下左右4候補（±STEP）のうち、法線方向への内積が
        最大（最も法線方向に近い）ものを選ぶ方式
        → サブセット座標に依存しないため確実にヒットする

    除外条件（変更なし）:
      |offset_neighbor| < |offset_self|  かつ  score_neighbor > score_self
    """
    result_map = {(r['cx'], r['cy']): r for r in results}

    kept = []
    n_removed = 0
    for r in results:
        cx, cy       = r['cx'], r['cy']
        score_self   = r['ncc_A'] + r['ncc_B']
        offset_self  = r['offset']
        theta_rad    = np.radians(r['theta'])

        sign = 1 if offset_self >= 0 else -1
        nx = -np.sin(theta_rad)
        ny =  np.cos(theta_rad)
        n_vec = np.array([nx, ny]) * sign  # 不連続線が外れている方向

        # 上下左右4候補のうち n_vec への内積が正かつ最大のものを選ぶ
        # （内積 ≤ 0 = 逆方向・直交なのでスキップ）
        candidates = [(cx+STEP, cy), (cx-STEP, cy), (cx, cy+STEP), (cx, cy-STEP)]
        best_nb  = None
        best_dot = -np.inf
        for (nbx, nby) in candidates:
            d = np.array([nbx - cx, nby - cy], dtype=float)
            d_norm = d / (np.linalg.norm(d) + 1e-9)
            dot = float(np.dot(d_norm, n_vec))
            if dot <= 0:
                continue
            nb_candidate = result_map.get((nbx, nby))
            if nb_candidate is not None and dot > best_dot:
                best_dot = dot
                best_nb  = nb_candidate

        # 隣が解析対象でない場合はフィルタを適用しない
        if best_nb is None:
            kept.append(r)
            continue

        score_nb  = best_nb['ncc_A'] + best_nb['ncc_B']
        offset_nb = abs(best_nb['offset'])

        if abs(offset_nb) < abs(offset_self) and score_nb > score_self:
            n_removed += 1
        else:
            kept.append(r)

    print(f"filter_by_neighbors: {len(results)} -> {len(kept)} "
          f"({n_removed} removed)")
    return kept


# =============================================================================
# メイン処理
# =============================================================================

def apply_alignment(image, filename, shifts):
    """
    dic_sem_strain_v57.py と同じシフト補正を変形後画像に適用する。
    shifts: {filename: (dx, dy)}  ← sem_alignment.json の shifts から生成
    """
    import cv2 as _cv2
    if filename not in shifts:
        return image
    dx, dy = shifts[filename]
    if dx == 0 and dy == 0:
        return image
    h, w = image.shape
    M = np.float32([[1, 0, dx], [0, 1, dy]])
    corrected = _cv2.warpAffine(image, M, (w, h),
                                flags=_cv2.INTER_LINEAR,
                                borderMode=_cv2.BORDER_CONSTANT,
                                borderValue=0)
    return corrected.astype(np.uint8)



def load_grain_assignment(georef_xlsx_path):
    """
    dic_results_georef.xlsx の ebsd_georef シートを読み込み、
    (cx, cy) -> grain_id の辞書を返す。grain_id <= 0 は除外。
    """
    wb = openpyxl.load_workbook(str(georef_xlsx_path), read_only=True)
    rows = list(wb['ebsd_georef'].iter_rows(values_only=True))
    wb.close()
    hdr = rows[0]
    cx_col  = next(i for i, h in enumerate(hdr) if h == 'cx [px]')
    cy_col  = next(i for i, h in enumerate(hdr) if h == 'cy [px]')
    gid_col = next(i for i, h in enumerate(hdr) if h == 'grain_id')
    coord_to_grain = {}
    for r in rows[1:]:
        if r[cx_col] is None or r[gid_col] is None:
            continue
        gid = int(r[gid_col])
        if gid <= 0:
            continue
        coord_to_grain[(int(round(float(r[cx_col]))),
                        int(round(float(r[cy_col]))))] = gid
    return coord_to_grain


def run_heaviside_dic(xlsx_path, ref_path, def_path, label,
                      grain_theta_map, coord_to_grain,
                      out_path='heaviside_results.xlsx', n_jobs=max(1, (os.cpu_count() or 2) - 1),
                      alignment_json_path=None,
                      disc_threshold=0.02,
                      offset_max=None, min_side_px=None,
                      disc_sheet='gamma_max',
                      grain_thr_map=None):
    """
    grain_theta_map    : {grain_id: theta_center_deg}
    coord_to_grain     : {(cx,cy): grain_id}
    alignment_json_path: sem_alignment.json のパス（省略可）
    offset_max         : offsetの探索上限 [px]（省略時はグローバルOFFSET_MAX）
    min_side_px        : Side A/B の最小ピクセル数（省略時はグローバルMIN_SIDE_PX）
    """
    if offset_max  is None: offset_max  = OFFSET_MAX
    if min_side_px is None: min_side_px = MIN_SIDE_PX
    print(f"Loading DIC results: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    x, y, u_cols     = load_sheet(wb['u'])
    _, _, v_cols     = load_sheet(wb['v'])
    _disc_sheet = disc_sheet if disc_sheet in wb.sheetnames else 'gamma_max'
    if _disc_sheet != disc_sheet:
        print(f"  Warning: sheet '{disc_sheet}' not found, falling back to 'gamma_max'")
    _, _, disc_cols_run = load_sheet(wb[_disc_sheet])
    wb.close()

    print(f"Loading images...")
    from PIL import Image as _PIL_Image
    ref_raw      = np.array(_PIL_Image.open(str(ref_path)).convert('L'))
    deformed_raw = np.array(_PIL_Image.open(str(def_path)).convert('L'))
    # dic_sem_strain_v57.py と同じトリミングを適用（下部TRIM_BOTTOM px除去）
    ref      = ref_raw[:-TRIM_BOTTOM, :]  if TRIM_BOTTOM > 0 else ref_raw
    deformed = deformed_raw[:-TRIM_BOTTOM, :] if TRIM_BOTTOM > 0 else deformed_raw
    print(f"  画像サイズ（トリミング後）: {ref.shape[1]}x{ref.shape[0]} px "
          f"(TRIM_BOTTOM={TRIM_BOTTOM}px)")

    # アライメント補正（sem_alignment.json が指定された場合）
    if alignment_json_path is not None:
        import json as _json
        with open(str(alignment_json_path), encoding='utf-8') as _f:
            _align = _json.load(_f)
        _shifts_raw = _align.get('shifts', {})
        # {filename: (dx, dy)} の辞書に変換
        _shifts = {k: (int(v['dx']), int(v['dy']))
                   for k, v in _shifts_raw.items()}
        def_filename = Path(def_path).name
        deformed = apply_alignment(deformed, def_filename, _shifts)
        dx_used, dy_used = _shifts.get(def_filename, (0, 0))
        print(f"  アライメント補正: {def_filename}  dx={dx_used}, dy={dy_used}")
    else:
        print("  アライメント補正: なし（alignment_json_path 未指定）")

    u_grid, xs, ys = to_grid(x, y, u_cols[label])
    v_grid, _,  _  = to_grid(x, y, v_cols[label])

    disc_grid_run, _, _ = to_grid(x, y, disc_cols_run[label])
    # grain_thr_map が渡されていない場合は disc_threshold を全粒共通で使用
    if grain_thr_map is None:
        grain_thr_map = {gid: disc_threshold for gid in grain_theta_map}

    # 対象粒に属するdisc_subsetを粒ごとのしきい値で抽出
    target_gids = set(grain_theta_map.keys())
    subset_list = []
    n_by_grain = {}
    for gid in target_gids:
        thr = grain_thr_map.get(gid, disc_threshold)
        disc_mask_g = disc_grid_run > thr
        rows_idx, cols_idx = np.where(disc_mask_g)
        n_g = 0
        for r, c in zip(rows_idx, cols_idx):
            cx_i = int(xs[c])
            cy_i = int(ys[r])
            if coord_to_grain.get((cx_i, cy_i)) != gid:
                continue
            theta_c = grain_theta_map[gid]
            u0 = float(u_grid[r, c]) if not np.isnan(u_grid[r, c]) else 0.0
            v0 = float(v_grid[r, c]) if not np.isnan(v_grid[r, c]) else 0.0
            subset_list.append((cx_i, cy_i, u0, v0, theta_c, gid))
            n_g += 1
        n_by_grain[gid] = n_g
        print(f"  Grain {gid:3d}: thr={thr:.4f}  subsets={n_g}")

    print(f"対象サブセット数: {len(subset_list)}  "
          f"({len(target_gids)} 粒: {sorted(target_gids)})")
    print(f"Running Heaviside DIC  OFFSET_MAX={offset_max:.0f}px  MIN_SIDE_PX={min_side_px}px ...")

    if JOBLIB_AVAILABLE and n_jobs != 1:
        results = Parallel(n_jobs=n_jobs, verbose=5)(
            delayed(process_one_subset)(
                ref, deformed, cx, cy, u0, v0, theta_c,
                offset_max=offset_max, min_side_px=min_side_px)
            for cx, cy, u0, v0, theta_c, gid in subset_list)
    else:
        # シリアル実行: Tkinter プログレスウィンドウ付き
        try:
            import tkinter as _tk
            from tkinter import ttk as _ttk
            _prog_root = _tk.Tk()
            _prog_root.withdraw()
            _prog_win = _tk.Toplevel(_prog_root)
            _prog_win.title('H-DIC 計算中...')
            _prog_win.resizable(False, False)
            _prog_win.attributes('-topmost', True)
            _tk.Label(_prog_win, text='Heaviside DIC 計算中',
                      font=('', 11, 'bold'), pady=10).pack(padx=20)
            _prog_lbl = _tk.Label(_prog_win,
                                  text=f'0 / {len(subset_list)} サブセット',
                                  font=('', 9))
            _prog_lbl.pack(padx=20)
            _prog_bar = _ttk.Progressbar(_prog_win, orient='horizontal',
                                         length=320, mode='determinate',
                                         maximum=max(len(subset_list), 1))
            _prog_bar.pack(padx=20, pady=(4, 16))
            _prog_win.update()
            _USE_PROG = True
        except Exception:
            _USE_PROG = False

        # 粒ごとにマスクキャッシュを事前生成（theta_center 別）
        _theta_to_cache = {}
        unique_thetas = set(theta_c for _, _, _, _, theta_c, _ in subset_list)
        for tc in unique_thetas:
            _theta_to_cache[tc] = build_mask_cache(tc)
        print(f"  マスクキャッシュ生成完了: {len(_theta_to_cache)} 粒分")

        results = []
        for i, (cx, cy, u0, v0, theta_c, gid) in enumerate(subset_list):
            if i % 50 == 0:
                print(f"  {i}/{len(subset_list)}")
            results.append(
                process_one_subset(ref, deformed, cx, cy, u0, v0, theta_c,
                                   offset_max=offset_max, min_side_px=min_side_px,
                                   _mask_cache=_theta_to_cache.get(theta_c)))
            if _USE_PROG and i % 5 == 0:
                _prog_bar['value'] = i + 1
                _prog_lbl.config(text=f'{i+1} / {len(subset_list)} サブセット')
                _prog_win.update()

        if _USE_PROG:
            _prog_bar['value'] = len(subset_list)
            _prog_lbl.config(text=f'完了: {len(subset_list)} サブセット')
            _prog_win.update()
            _prog_win.after(800, _prog_win.destroy)
            _prog_root.after(900, _prog_root.destroy)

    # grain_id を結果に付与
    for res, (cx, cy, u0, v0, theta_c, gid) in zip(results, subset_list):
        if res is not None:
            res['grain_id'] = gid

    # スキップされたサブセットを別途保持
    skipped_list = [(subset_list[i], gid)
                    for i, ((cx, cy, u0, v0, theta_c, gid), res)
                    in enumerate(zip(subset_list, results))
                    if res is None]

    results = [r for r in results if r is not None]
    print(f"Valid results: {len(results)}  Skipped: {len(skipped_list)}")
    results = filter_by_neighbors(results)

    # Excel出力（skipped列を追加、スキップサブセットも記載）
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = 'heaviside'
    headers = ['cx', 'cy', 'grain_id', 'theta', 'offset',
               'u_A', 'v_A', 'u_B', 'v_B',
               'delta_u', 'delta_v', 'delta_s', 'delta_n',
               'ncc_A', 'ncc_B', 'skipped']
    ws.append(headers)
    for r in results:
        ws.append([r['cx'], r['cy'], r.get('grain_id'),
                   r['theta'], r['offset'],
                   r['u_A'], r['v_A'], r['u_B'], r['v_B'],
                   r['delta_u'], r['delta_v'], r['delta_s'], r['delta_n'],
                   r['ncc_A'], r['ncc_B'], 0])
    for (cx, cy, u0, v0, theta_c, gid), _ in skipped_list:
        ws.append([cx, cy, gid,
                   None, None, None, None, None, None,
                   None, None, None, None, None, None, 1])
    wb_out.save(out_path)
    print(f"Saved: {out_path}")

    return results, deformed, subset_list



# =============================================================================
# 結果可視化GUI（インタラクティブ表示 + ON/OFF切替 + PNG保存）
# =============================================================================
def _build_disc_line_map(results, H, W):
    """
    results から delta_s オーバーレイ用マップを生成して返す。
    不連続線の描画位置は変形後座標（参照座標 + Side A/B 変位の平均）を使用。
    """
    half = SUBSET_SIZE // 2
    disc_line_map = np.full((H, W), np.nan)
    for r in results:
        # 変形後座標: 参照座標 + Side A/B 変位の平均
        cx_def = int(round(r['cx'] + (r['u_A'] + r['u_B']) / 2.0))
        cy_def = int(round(r['cy'] + (r['v_A'] + r['v_B']) / 2.0))
        theta_rad = np.radians(float(r['theta']))
        offset    = float(r['offset'])
        ds        = float(r['delta_s'])
        cc, rr    = np.meshgrid(np.arange(-half, half + 1),
                                np.arange(-half, half + 1))
        proj    = cc * (-np.sin(theta_rad)) + rr * np.cos(theta_rad)
        on_line = np.abs(proj - offset) < 1.5
        local_rows, local_cols = np.where(on_line)
        py = cy_def - half + local_rows
        px = cx_def - half + local_cols
        valid = (py >= 0) & (py < H) & (px >= 0) & (px < W)
        disc_line_map[py[valid], px[valid]] = ds
    return disc_line_map


def _save_png(results, sem_img, out_xlsx_path, label=''):
    """
    delta_s オーバーレイ図を PNG として保存する。
    出力先: out_xlsx_path と同じディレクトリ、同じ stem に '_delta_s.png' を付与。
    不連続線・サブセットの描画位置は変形後座標（v32対応）。
      不連続線: 参照座標 + (u_A + u_B) / 2
    """
    H, W = sem_img.shape
    disc_line_map = _build_disc_line_map(results, H, W)
    valid_ds = disc_line_map[~np.isnan(disc_line_map)]
    if len(valid_ds) == 0:
        print("Warning: No valid delta_s values, skipping PNG output.")
        return
    p2, p98 = np.percentile(valid_ds, [2, 98])
    vlim = float(max(abs(p2), abs(p98)))
    cmap_rdbu = plt.cm.RdBu_r.copy()
    cmap_rdbu.set_bad(alpha=0.0)
    fig, ax = plt.subplots(figsize=(10, 8))
    ax.imshow(sem_img, cmap='gray',
              extent=[-0.5, W-0.5, H-0.5, -0.5], alpha=0.6)
    im = ax.imshow(disc_line_map, cmap=cmap_rdbu,
                   extent=[-0.5, W-0.5, H-0.5, -0.5],
                   vmin=-vlim, vmax=vlim, alpha=1.0)
    plt.colorbar(im, ax=ax, label='delta_s [px]')
    ax.set_title(f'Heaviside DIC: {label}  delta_s (±{vlim:.2f} px), n={len(results)}')
    ax.set_xlabel('x [px]')
    ax.set_ylabel('y [px]')
    plt.tight_layout()
    png_path = str(Path(out_xlsx_path).with_suffix('')) + '_delta_s.png'
    fig.savefig(png_path, dpi=300)
    plt.close(fig)
    print(f"Saved PNG: {png_path}")


def _show_result_gui(results, sem_img, out_xlsx_path, label,
                     subset_list=None, strain_data=None):
    """
    計算結果をインタラクティブに確認・調整して PNG 保存するGUI。
    左側: matplotlib figure（マップ表示）
    右側: Tk Frame（全コントロール）
    """
    import matplotlib
    matplotlib.use('TkAgg')
    import matplotlib.pyplot as _plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import tkinter as _tk
    from tkinter import ttk as _ttk

    if subset_list is None:
        subset_list = []
    if strain_data is None:
        strain_data = {}

    H, W = sem_img.shape
    extent = [-0.5, W-0.5, H-0.5, -0.5]

    # ---- delta_s / delta_n の座標・値リストを事前生成（scatter用）----
    # 通常DICの変位 (u0, v0) を使った変形後座標で描画（ひずみマップと同じ座標系）
    half = SUBSET_SIZE // 2
    _u0v0_map = {(cx_i, cy_i): (u0, v0)
                 for cx_i, cy_i, u0, v0, _, _ in (subset_list or [])}
    maps = {}
    for key in ('delta_s', 'delta_n'):
        xs_list, ys_list, vals_list = [], [], []
        for r in results:
            cx_i, cy_i = r['cx'], r['cy']
            u0, v0 = _u0v0_map.get((cx_i, cy_i), (0.0, 0.0))
            cx_def = cx_i + u0
            cy_def = cy_i + v0
            theta_rad = np.radians(float(r['theta']))
            offset    = float(r['offset'])
            val       = float(r[key])
            cc, rr = np.meshgrid(np.arange(-half, half + 1),
                                  np.arange(-half, half + 1))
            proj    = cc * (-np.sin(theta_rad)) + rr * np.cos(theta_rad)
            on_line = np.abs(proj - offset) < 1.5
            lrows, lcols = np.where(on_line)
            xs_list.extend((cx_def - half + lcols).tolist())
            ys_list.extend((cy_def - half + lrows).tolist())
            vals_list.extend([val] * int(on_line.sum()))
        maps[key] = {
            'x': np.array(xs_list),
            'y': np.array(ys_list),
            'val': np.array(vals_list),
        }

    # ---- 初期 vmin/vmax ----
    def _strain_vlim(bk):
        if bk == 'SEM像' or bk not in strain_data:
            return 0.0, 1.0
        vals = strain_data[bk]['val']
        valid_v = vals[~np.isnan(vals)]
        if len(valid_v) == 0:
            return 0.0, 1.0
        return float(np.percentile(valid_v, 2)), float(np.percentile(valid_v, 98))

    valid0 = maps['delta_s']['val']
    valid0 = valid0[~np.isnan(valid0)]
    p2, p98 = (np.percentile(valid0, [2, 98]) if len(valid0) else (0.0, 1.0))
    init_vlim = float(max(abs(p2), abs(p98)))

    state = {
        'vmin':      -init_vlim,
        'vmax':       init_vlim,
        'data_key':  'delta_s',
        'base_key':  'SEM像',
        'base_vmin':  0.0,
        'base_vmax':  1.0,
        'disc_cmap':  'RdBu_r',
        'base_cmap':  'hot',
    }
    if strain_data:
        first_sh = list(strain_data.keys())[0]
        state['base_vmin'], state['base_vmax'] = _strain_vlim(first_sh)

    base_keys = ['SEM像'] + list(strain_data.keys())

    # ---- Tk ルートウィンドウ ----
    root = _tk.Tk()
    root.title('H-DIC 結果確認  |  設定を調整して PNG 保存')
    root.configure(bg='#1c1c2e')

    # ---- 左: matplotlib figure ----
    fig, (ax_main, ax_cb_base, ax_cb_disc) = _plt.subplots(
        1, 3,
        gridspec_kw={'width_ratios': [40, 1, 1], 'wspace': 0.05},
        figsize=(12, 8))
    fig.patch.set_facecolor('#1c1c2e')
    ax_main.set_facecolor('#111122')
    for ax in [ax_cb_base, ax_cb_disc]:
        ax.set_facecolor('#1c1c2e')

    # ---- 右: Tk コントロールフレーム（先に pack して領域を確保） ----
    ctrl_outer = _tk.Frame(root, bg='#1c1c2e', width=290)
    ctrl_outer.pack(side=_tk.RIGHT, fill=_tk.Y, padx=8)
    ctrl_outer.pack_propagate(False)

    # 上下のスペーサーで中央寄せ
    _tk.Frame(ctrl_outer, bg='#1c1c2e').pack(side=_tk.TOP, fill=_tk.Y, expand=True)
    ctrl = _tk.Frame(ctrl_outer, bg='#1c1c2e')
    ctrl.pack(side=_tk.TOP, pady=8, fill=_tk.X)
    _tk.Frame(ctrl_outer, bg='#1c1c2e').pack(side=_tk.TOP, fill=_tk.Y, expand=True)

    # ---- 左: matplotlib figure（右パネルの後に pack） ----
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.get_tk_widget().pack(side=_tk.LEFT, fill=_tk.BOTH, expand=True)

    def _on_resize(event):
        _redraw()

    fig.canvas.mpl_connect('resize_event', _on_resize)

    # スタイル定数
    BG   = '#1c1c2e'
    BG2  = '#2d2d44'
    FG   = '#ddddee'
    FG2  = '#aaaacc'
    BLUE = '#2a5caa'
    ORG  = '#7a4a00'
    FONT     = ('', 10)
    FONT_SM  = ('', 9)

    def _sep():
        _tk.Frame(ctrl, height=1, bg='#333355').pack(fill=_tk.X, pady=4)

    def _lbl(text, font=None):
        _tk.Label(ctrl, text=text, bg=BG, fg=FG,
                  font=font or FONT).pack(anchor='w', pady=(6, 1))

    def _btn_row(texts, cmds, colors=None):
        f = _tk.Frame(ctrl, bg=BG)
        f.pack(fill=_tk.X, pady=2)
        btns = []
        for i, (t, c) in enumerate(zip(texts, cmds)):
            col = (colors[i] if colors else BG2)
            b = _tk.Button(f, text=t, command=c, bg=col, fg=FG,
                           font=FONT_SM, relief='flat',
                           activebackground='#444466', activeforeground=FG,
                           padx=4, pady=4)
            b.pack(side=_tk.LEFT, fill=_tk.X, expand=True, padx=2)
            btns.append(b)
        return btns

    def _entry_row(label_text, init_val):
        f = _tk.Frame(ctrl, bg=BG)
        f.pack(fill=_tk.X, pady=1)
        _tk.Label(f, text=label_text, bg=BG, fg=FG2,
                  font=FONT_SM, width=6, anchor='w').pack(side=_tk.LEFT)
        e = _tk.Entry(f, bg='#252538', fg='#eeeeee',
                      insertbackground='white', relief='flat',
                      font=FONT_SM)
        e.insert(0, init_val)
        e.pack(side=_tk.LEFT, fill=_tk.X, expand=True, ipady=3)
        return e

    def _calc_marker_size(step_px):
        """ax_main の現在の表示サイズから scatter マーカーサイズを動的計算"""
        ax_bbox = ax_main.get_window_extent(renderer=fig.canvas.get_renderer())
        ax_w_pt = ax_bbox.width   # ax_main の幅 [points]
        ax_h_pt = ax_bbox.height  # ax_main の高さ [points]
        data_w = W  # data座標の幅
        data_h = H  # data座標の高さ
        pt_per_data_x = ax_w_pt / data_w
        pt_per_data_y = ax_h_pt / data_h
        pt_per_data = min(pt_per_data_x, pt_per_data_y)
        # 1サブセット = step_px data単位 → ポイント換算してマーカー面積
        marker_pt = step_px * pt_per_data
        return max(marker_pt ** 2, 1.0)

    # ---- 描画関数 ----
    _cb_base_obj = [None]
    _cb_disc_obj = [None]

    def _redraw():
        ax_main.clear()
        ax_main.set_facecolor('#111122')
        ax_cb_base.clear()
        ax_cb_disc.clear()

        # まず xlim/ylim/aspect を設定してからマーカーサイズを計算
        ax_main.set_xlim(-0.5, W - 0.5)
        ax_main.set_ylim(H - 0.5, -0.5)
        ax_main.set_aspect('equal')
        fig.canvas.draw()  # ax_main のサイズを確定させる
        ms_base = _calc_marker_size(STEP)
        ms_disc = _calc_marker_size(1) * 0.3   # 不連続線は細め

        bk = state['base_key']
        if bk == 'SEM像':
            ax_main.imshow(sem_img, cmap='gray', extent=extent, alpha=0.6)
            ax_cb_base.set_visible(False)
        else:
            sd = strain_data[bk]
            mask_in = ((sd['x_def'] >= -0.5) & (sd['x_def'] <= W - 0.5) &
                       (sd['y_def'] >= -0.5) & (sd['y_def'] <= H - 0.5))
            cmap_base = state['base_cmap']
            sc = ax_main.scatter(
                sd['x_def'][mask_in], sd['y_def'][mask_in],
                c=sd['val'][mask_in], cmap=cmap_base,
                vmin=state['base_vmin'], vmax=state['base_vmax'],
                s=ms_base, marker='s', linewidths=0, alpha=1.0,
                zorder=1)
            ax_cb_base.set_visible(True)
            fig.colorbar(sc, cax=ax_cb_base)
            ax_cb_base.set_ylabel(bk, color=FG2, fontsize=8)
            ax_cb_base.tick_params(colors=FG2, labelsize=7)
            ax_cb_base.yaxis.set_label_position('left')
            ax_cb_base.yaxis.tick_left()

        # 不連続線を scatter で描画
        md = maps[state['data_key']]
        mask_disc = ((md['x'] >= -0.5) & (md['x'] <= W - 0.5) &
                     (md['y'] >= -0.5) & (md['y'] <= H - 0.5))
        cmap_rdbu = _plt.cm.get_cmap('RdBu_r').copy()
        sc_disc = ax_main.scatter(
            md['x'][mask_disc], md['y'][mask_disc],
            c=md['val'][mask_disc], cmap=cmap_rdbu,
            vmin=state['vmin'], vmax=state['vmax'],
            s=ms_disc, marker='s', linewidths=0, alpha=1.0, zorder=2)
        fig.colorbar(sc_disc, cax=ax_cb_disc)
        ax_cb_disc.set_ylabel(f"{state['data_key']} [px]", color=FG2, fontsize=8)
        ax_cb_disc.tick_params(colors=FG2, labelsize=7)
        ax_cb_disc.yaxis.set_label_position('right')
        ax_cb_disc.yaxis.tick_right()

        ax_main.set_xlim(-0.5, W - 0.5)
        ax_main.set_ylim(H - 0.5, -0.5)
        ax_main.set_aspect('equal')
        ax_main.set_title(f"{bk} + {state['data_key']}  n={len(results)}",
                          color='white', fontsize=10)
        ax_main.set_xlabel('x [px]', color=FG2, fontsize=9)
        ax_main.set_ylabel('y [px]', color=FG2, fontsize=9)
        ax_main.tick_params(colors=FG2)
        for sp in ax_main.spines.values():
            sp.set_edgecolor('#333355')
        canvas.draw_idle()

    _redraw()

    # ---- 表示データ選択 ----
    _lbl('表示データ')
    data_var = _tk.StringVar(value='delta_s')
    data_btn_frame = _tk.Frame(ctrl, bg=BG)
    data_btn_frame.pack(fill=_tk.X, pady=2)
    data_btns = {}
    for dk in ['delta_s', 'delta_n']:
        b = _tk.Button(data_btn_frame, text=dk,
                       bg=BLUE if dk == 'delta_s' else BG2,
                       fg=FG, font=FONT_SM, relief='flat',
                       activebackground='#3a5ccc', activeforeground=FG,
                       padx=4, pady=4)
        b.pack(side=_tk.LEFT, fill=_tk.X, expand=True, padx=2)
        data_btns[dk] = b

    def _on_data(dk):
        state['data_key'] = dk
        for k, b in data_btns.items():
            b.config(bg=BLUE if k == dk else BG2)
        _redraw()

    for dk in data_btns:
        data_btns[dk].config(command=lambda k=dk: _on_data(k))

    # ---- ベース画像選択 ----
    _sep()
    _lbl('ベース画像')
    base_btn_frame = _tk.Frame(ctrl, bg=BG)
    base_btn_frame.pack(fill=_tk.X, pady=2)
    base_btns = {}
    for bk in base_keys:
        b = _tk.Button(base_btn_frame, text=bk,
                       bg=BLUE if bk == 'SEM像' else BG2,
                       fg=FG, font=('', 8), relief='flat',
                       activebackground='#3a5ccc', activeforeground=FG,
                       padx=2, pady=4)
        b.pack(side=_tk.LEFT, fill=_tk.X, expand=True, padx=1)
        base_btns[bk] = b

    def _on_base(bk):
        state['base_key'] = bk
        for k, b in base_btns.items():
            b.config(bg=BLUE if k == bk else BG2)
        vmin_b, vmax_b = _strain_vlim(bk)
        state['base_vmin'] = vmin_b
        state['base_vmax'] = vmax_b
        ent_base_vmin.delete(0, _tk.END)
        ent_base_vmin.insert(0, f'{vmin_b:.4f}')
        ent_base_vmax.delete(0, _tk.END)
        ent_base_vmax.insert(0, f'{vmax_b:.4f}')
        _redraw()

    for bk in base_btns:
        base_btns[bk].config(command=lambda k=bk: _on_base(k))

    # ---- 不連続線カラーマップ ----
    _sep()
    _lbl('不連続線カラーマップ範囲')
    ent_disc_vmin = _entry_row('vmin:', f'{state["vmin"]:.4f}')
    ent_disc_vmax = _entry_row('vmax:', f'{state["vmax"]:.4f}')

    def _on_disc_apply():
        try:
            state['vmin'] = float(ent_disc_vmin.get())
        except ValueError:
            pass
        try:
            state['vmax'] = float(ent_disc_vmax.get())
        except ValueError:
            pass
        _redraw()

    def _on_disc_auto():
        valid = maps[state['data_key']]['val']
        valid = valid[~np.isnan(valid)]
        if len(valid):
            p2, p98 = np.percentile(valid, [2, 98])
            vlim = float(max(abs(p2), abs(p98)))
            state['vmin'] = -vlim
            state['vmax'] =  vlim
            ent_disc_vmin.delete(0, _tk.END)
            ent_disc_vmin.insert(0, f'{state["vmin"]:.4f}')
            ent_disc_vmax.delete(0, _tk.END)
            ent_disc_vmax.insert(0, f'{state["vmax"]:.4f}')
        _redraw()

    _btn_row(['適用', '自動設定'],
             [_on_disc_apply, _on_disc_auto],
             ['#1a4a8a', BG2])

    # ---- ベースカラーマップ ----
    _sep()
    _lbl('ベースカラーマップ')

    BASE_CMAPS = ['RdBu_r', 'hot', 'cividis', 'magma', 'viridis']
    base_cmap_var = _tk.StringVar(value=state['base_cmap'])
    base_cmap_cb = _ttk.Combobox(ctrl, textvariable=base_cmap_var,
                                  values=BASE_CMAPS, state='readonly',
                                  font=FONT_SM)
    base_cmap_cb.pack(fill=_tk.X, pady=2)

    def _on_base_cmap(event=None):
        selected = base_cmap_cb.get()
        state['base_cmap'] = selected
        root.focus_set()  # Combobox からフォーカスを外す
        _redraw()

    base_cmap_cb.bind('<<ComboboxSelected>>', _on_base_cmap)

    _lbl('ベースカラーマップ範囲')
    ent_base_vmin = _entry_row('vmin:', f'{state["base_vmin"]:.4f}')
    ent_base_vmax = _entry_row('vmax:', f'{state["base_vmax"]:.4f}')

    def _on_base_apply():
        try:
            state['base_vmin'] = float(ent_base_vmin.get())
        except ValueError:
            pass
        try:
            state['base_vmax'] = float(ent_base_vmax.get())
        except ValueError:
            pass
        _redraw()

    def _on_base_auto():
        vmin_b, vmax_b = _strain_vlim(state['base_key'])
        state['base_vmin'] = vmin_b
        state['base_vmax'] = vmax_b
        ent_base_vmin.delete(0, _tk.END)
        ent_base_vmin.insert(0, f'{vmin_b:.4f}')
        ent_base_vmax.delete(0, _tk.END)
        ent_base_vmax.insert(0, f'{vmax_b:.4f}')
        _redraw()

    _btn_row(['適用', '自動設定'],
             [_on_base_apply, _on_base_auto],
             ['#1a4a8a', BG2])

    # ---- PNG 保存 ----
    _sep()

    def _on_save():
        md = maps[state['data_key']]
        bk = state['base_key']
        cmap_rdbu = _plt.cm.get_cmap('RdBu_r').copy()
        fig_out, ax_out = _plt.subplots(figsize=(10, 8))

        sc_base_out = None
        mask_in_count = 0
        if bk == 'SEM像':
            ax_out.imshow(sem_img, cmap='gray', extent=extent, alpha=0.6)
        else:
            sd = strain_data[bk]
            mask_in = ((sd['x_def'] >= -0.5) & (sd['x_def'] <= W - 0.5) &
                       (sd['y_def'] >= -0.5) & (sd['y_def'] <= H - 0.5))
            mask_in_count = int(mask_in.sum())
            cmap_base = state['base_cmap']
            sc_base_out = ax_out.scatter(
                sd['x_def'][mask_in], sd['y_def'][mask_in],
                c=sd['val'][mask_in], cmap=cmap_base,
                vmin=state['base_vmin'], vmax=state['base_vmax'],
                s=1, marker='s', linewidths=0, alpha=1.0, zorder=1)
            _plt.colorbar(sc_base_out, ax=ax_out, label=bk)

        mask_disc = ((md['x'] >= -0.5) & (md['x'] <= W - 0.5) &
                     (md['y'] >= -0.5) & (md['y'] <= H - 0.5))
        mask_disc_count = int(mask_disc.sum())
        sc_disc_out = ax_out.scatter(
            md['x'][mask_disc], md['y'][mask_disc],
            c=md['val'][mask_disc], cmap=cmap_rdbu,
            vmin=state['vmin'], vmax=state['vmax'],
            s=1, marker='s', linewidths=0, alpha=1.0, zorder=2)
        _plt.colorbar(sc_disc_out, ax=ax_out, label=f"{state['data_key']} [px]")
        ax_out.set_title(
            f'Heaviside DIC: {label}  {state["data_key"]}  n={len(results)}')
        ax_out.set_xlabel('x [px]')
        ax_out.set_ylabel('y [px]')
        _plt.tight_layout()
        ax_out.set_xlim(-0.5, W - 0.5)
        ax_out.set_ylim(H - 0.5, -0.5)
        ax_out.set_aspect('equal')
        fig_out.canvas.draw()  # アスペクト比を確定させてから軸サイズを取得

        # tight_layout + set_aspect後の実際の軸サイズからマーカーサイズを計算
        _renderer = fig_out.canvas.get_renderer()
        _bbox = ax_out.get_window_extent(renderer=_renderer)
        _px_per_pt = fig_out.get_dpi() / 72.0
        _pt_per_data = min(_bbox.width / _px_per_pt / W,
                           _bbox.height / _px_per_pt / H)
        ms_base_png = max((_pt_per_data * STEP) ** 2 * 1.3, 0.5)
        ms_disc_png = max(ms_base_png * 0.3 * 0.7, 0.5)
        if sc_base_out is not None:
            sc_base_out.set_sizes([ms_base_png] * mask_in_count)
        sc_disc_out.set_sizes([ms_disc_png] * mask_disc_count)

        base_tag = state['base_key'].replace(' ', '_')
        png_path = (str(Path(out_xlsx_path).with_suffix(''))
                    + f'_{state["data_key"]}_on_{base_tag}.png')
        fig_out.savefig(png_path, dpi=300)
        _plt.close(fig_out)
        print(f"Saved PNG: {png_path}")

    _tk.Button(ctrl, text='▶ PNG 保存', command=_on_save,
               bg=ORG, fg='white', font=('', 11, 'bold'),
               relief='flat', pady=8,
               activebackground='#aa6600', activeforeground='white'
               ).pack(fill=_tk.X, pady=(8, 2))

    # ---- ヒント ----
    _sep()
    _tk.Label(ctrl,
              text='使い方:\n'
                   '① 表示データ / ベース画像を選択\n'
                   '② 各カラーマップ範囲を入力して「適用」\n'
                   '   「自動設定」で 2〜98%ile に戻す\n'
                   '③「PNG 保存」で保存して閉じる',
              bg=BG, fg='#888899', font=('', 9),
              justify='left').pack(anchor='w', pady=4)

    root.mainloop()
if __name__ == '__main__':
    import sys as _sys
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import json as _json

    # argv[1] が .json ファイルならウィザードをスキップしてパスを直接受け取る
    _param_json = None
    _initial_dir = None
    if len(_sys.argv) > 1:
        if _sys.argv[1].endswith('.json'):
            _param_json = _sys.argv[1]
        else:
            _initial_dir = _sys.argv[1]

    # ==========================================================================
    # ウィザード形式ファイル選択ランチャー
    # ==========================================================================
    root = tk.Tk()
    root.withdraw()

    class FileWizard:
        """
        全ファイル選択を1ウィンドウにまとめたウィザードダイアログ。
        ステップ:
          1. dic_results_georef.xlsx
          2. dic_config.txt（省略可）
          3. sem_alignment.json（省略可）
          4. 出力先フォルダ
        """
        def __init__(self, parent, initial_dir=None):
            self.win = tk.Toplevel(parent)
            self.win.title('Heaviside DIC  —  ファイル選択')
            self.win.resizable(False, False)
            self.win.grab_set()
            self._initial_dir = initial_dir

            # ---- パス変数 ----
            self.v_xlsx  = tk.StringVar(value='（未選択）')
            self.v_cfg   = tk.StringVar(value='（スキップ）')
            self.v_align = tk.StringVar(value='（スキップ）')
            self.v_out   = tk.StringVar(value='（未選択）')

            self.result = None   # {'xlsx':..., 'config':..., 'align':..., 'out':...}

            # ---- タイトル ----
            hdr = tk.Frame(self.win, bg='#1a1a2e')
            hdr.pack(fill='x')
            tk.Label(hdr, text='Heaviside DIC',
                     bg='#1a1a2e', fg='white',
                     font=('', 13, 'bold'), pady=10).pack(side='left', padx=16)
            tk.Label(hdr, text='ファイル設定ウィザード',
                     bg='#1a1a2e', fg='#aaaacc',
                     font=('', 10), pady=10).pack(side='left')

            # ---- 進行バー ----
            pb_frame = tk.Frame(self.win, padx=16, pady=8)
            pb_frame.pack(fill='x')
            self.step_labels = []
            steps = ['① xlsx', '② config', '③ align', '④ 出力先']
            for i, s in enumerate(steps):
                lbl = tk.Label(pb_frame, text=s, font=('', 9),
                               relief='groove', width=9, padx=4, pady=3)
                lbl.grid(row=0, column=i, padx=3)
                self.step_labels.append(lbl)
            self._highlight_step(0)

            # ---- 各行 ----
            rows_frame = tk.Frame(self.win, padx=20, pady=6)
            rows_frame.pack(fill='both')
            self._make_row(rows_frame, 0,
                           '① dic_results_georef.xlsx *',
                           self.v_xlsx,
                           lambda: self._pick_file(
                               self.v_xlsx, 0,
                               'dic_results_georef.xlsx を選択',
                               [('Excel files', '*.xlsx'), ('All files', '*.*')]))
            self._make_row(rows_frame, 1,
                           '② dic_config.txt（省略可）',
                           self.v_cfg,
                           lambda: self._pick_file(
                               self.v_cfg, 1,
                               'dic_config.txt を選択（キャンセルでスキップ）',
                               [('Text files', '*.txt'), ('All files', '*.*')],
                               optional=True))
            self._make_row(rows_frame, 2,
                           '③ sem_alignment.json（省略可）',
                           self.v_align,
                           lambda: self._pick_file(
                               self.v_align, 2,
                               'sem_alignment.json を選択（キャンセルでスキップ）',
                               [('JSON files', '*.json'), ('All files', '*.*')],
                               optional=True))
            self._make_row(rows_frame, 3,
                           '④ 出力先フォルダ *',
                           self.v_out,
                           lambda: self._pick_dir())

            # ---- OK / キャンセル ----
            btn_frame = tk.Frame(self.win, pady=12)
            btn_frame.pack()
            tk.Button(btn_frame, text='次へ  ▶', width=14, bg='#2a5caa', fg='white',
                      font=('', 10, 'bold'), relief='flat', pady=6,
                      command=self._on_ok).pack(side='left', padx=8)
            tk.Button(btn_frame, text='キャンセル', width=10,
                      font=('', 10), relief='flat', pady=6,
                      command=self._on_cancel).pack(side='left', padx=8)

            self.win.wait_window()

        def _make_row(self, parent, row, label_text, var, cmd):
            tk.Label(parent, text=label_text, font=('', 9, 'bold'),
                     anchor='w').grid(row=row*2, column=0, columnspan=3,
                                      sticky='w', pady=(10, 0))
            lbl = tk.Label(parent, textvariable=var, font=('', 9),
                           fg='#555555', anchor='w', width=52,
                           relief='sunken', padx=6, pady=3)
            lbl.grid(row=row*2+1, column=0, sticky='ew', padx=(0, 6))
            tk.Button(parent, text='参照...', command=cmd,
                      font=('', 9), width=8, relief='groove').grid(
                          row=row*2+1, column=1)

        def _highlight_step(self, step):
            for i, lbl in enumerate(self.step_labels):
                if i == step:
                    lbl.config(bg='#2a5caa', fg='white', font=('', 9, 'bold'))
                elif i < step:
                    lbl.config(bg='#4caf50', fg='white', font=('', 9))
                else:
                    lbl.config(bg='#eeeeee', fg='#888888', font=('', 9))

        def _pick_file(self, var, step, title, ftypes, optional=False):
            path = filedialog.askopenfilename(title=title, filetypes=ftypes,
                                              initialdir=self._initial_dir)
            if path:
                var.set(path)
                self._highlight_step(step + 1)
            elif optional:
                var.set('（スキップ）')

        def _pick_dir(self):
            path = filedialog.askdirectory(title='出力先フォルダを選択',
                                           initialdir=self._initial_dir)
            if path:
                self.v_out.set(path)
                self._highlight_step(4)

        def _on_ok(self):
            if self.v_xlsx.get() == '（未選択）':
                messagebox.showwarning('未選択', 'dic_results_georef.xlsx を選択してください。',
                                       parent=self.win)
                return
            if self.v_out.get() == '（未選択）':
                messagebox.showwarning('未選択', '出力先フォルダを選択してください。',
                                       parent=self.win)
                return
            cfg   = self.v_cfg.get()
            align = self.v_align.get()
            self.result = {
                'xlsx'  : self.v_xlsx.get(),
                'config': cfg   if cfg   != '（スキップ）' else None,
                'align' : align if align != '（スキップ）' else None,
                'out'   : self.v_out.get(),
            }
            self.win.destroy()

        def _on_cancel(self):
            self.result = None
            self.win.destroy()

    # JSONパラメータが渡された場合はウィザードをスキップ
    if _param_json:
        with open(_param_json, encoding='utf-8') as _f:
            _p = _json.load(_f)
        georef_xlsx     = _p['xlsx']
        dic_config_path = _p.get('cfg')
        alignment_json  = _p.get('align')
        out_dir         = Path(_p['out'])
    else:
        wizard = FileWizard(root, _initial_dir)
        if wizard.result is None:
            raise SystemExit('キャンセルされました')
        georef_xlsx     = wizard.result['xlsx']
        dic_config_path = wizard.result['config']
        alignment_json  = wizard.result['align']
        out_dir         = Path(wizard.result['out'])

    # ---- dic_config.txt 読み込み ----
    if dic_config_path:
        applied = load_dic_config(dic_config_path)
        print(f"  -> dic_config.txt 読み込み完了: {dic_config_path}")
        print(f"     適用パラメータ: {applied}")
    else:
        print(f"  -> スキップ。ハードコード値を使用:")
        print(f"     SUBSET_SIZE={SUBSET_SIZE}, STEP={STEP}, "
              f"TRIM_BOTTOM={TRIM_BOTTOM}, NCC_THRESHOLD={NCC_THRESHOLD}, "
              f"OFFSET_MAX={OFFSET_MAX:.1f}")

    if alignment_json:
        print(f"  -> アライメントファイル: {alignment_json}")
    else:
        print("  -> アライメント補正なし")

    # ==========================================================================
    # データ読み込み（GUI起動前に完了させる）
    # ==========================================================================
    print("データ読み込み中...")
    from PIL import Image as _PIL_Image

    _wb = openpyxl.load_workbook(georef_xlsx, read_only=True)
    _hdr = list(_wb['u'].iter_rows(values_only=True, max_row=1))[0]
    _data_start = 3 if _hdr[0] == 'subset_id' else 2
    available_labels = [str(h) for h in _hdr[_data_start + 1:] if h is not None]
    _wb.close()
    print(f"利用可能なラベル: {available_labels}")

    coord_to_grain = load_grain_assignment(georef_xlsx)
    all_grain_ids  = sorted(set(coord_to_grain.values()))
    print(f"有効な結晶粒数: {len(all_grain_ids)}")

    # ==========================================================================
    # ステップ2: 応力段階選択ダイアログ（改良版）
    # ==========================================================================
    def ask_label():
        dialog = tk.Toplevel(root)
        dialog.title('変形段階を選択')
        dialog.grab_set()
        dialog.resizable(False, False)
        dialog.attributes('-topmost', True)
        dialog.lift()
        dialog.focus_force()

        tk.Label(dialog, text='処理する変形段階を選択してください',
                 font=('', 11, 'bold'), pady=12).pack(padx=20, anchor='w')

        frame = tk.Frame(dialog, padx=20)
        frame.pack(fill='both')

        scrollbar = tk.Scrollbar(frame, orient='vertical')
        listbox = tk.Listbox(frame, font=('', 11), height=min(len(available_labels), 8),
                             selectmode='single', activestyle='none',
                             yscrollcommand=scrollbar.set,
                             selectbackground='#2a5caa', selectforeground='white',
                             relief='flat', bd=1)
        scrollbar.config(command=listbox.yview)
        listbox.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        for lbl in available_labels:
            listbox.insert('end', f'  {lbl}')
        listbox.select_set(len(available_labels) - 1)
        listbox.see(len(available_labels) - 1)

        selected = []
        def on_ok():
            idx = listbox.curselection()
            if idx:
                selected.append(available_labels[idx[0]])
            dialog.destroy()

        btn_frame = tk.Frame(dialog, pady=12)
        btn_frame.pack()
        tk.Button(btn_frame, text='OK', width=12, bg='#2a5caa', fg='white',
                  font=('', 10, 'bold'), relief='flat', pady=6,
                  command=on_ok).pack(side='left', padx=8)
        tk.Button(btn_frame, text='キャンセル', width=10,
                  font=('', 10), relief='flat', pady=6,
                  command=dialog.destroy).pack(side='left', padx=8)

        dialog.wait_window()
        return selected[0] if selected else None

    label = ask_label()
    if not label:
        raise SystemExit('キャンセルされました')

    # 選択ラベルのひずみデータ読み込み（全シートを事前ロード）
    print(f"ひずみデータ読み込み中: {label}")
    MAP_SHEETS = ['u', 'v', 'zncc', 'exx', 'eyy', 'exy', 'e1', 'gamma_max', 'omega_xy']
    _wb2 = openpyxl.load_workbook(georef_xlsx, read_only=True)
    x_arr, y_arr, u_cols    = load_sheet(_wb2['u'])
    _,     _,     v_cols    = load_sheet(_wb2['v'])
    _,     _,     init_cols = load_sheet(_wb2['gamma_max'])
    all_sheet_data = {}
    for sh in MAP_SHEETS:
        if sh in _wb2.sheetnames:
            _, _, cols = load_sheet(_wb2[sh])
            if label in cols:
                grid, _, _ = to_grid(x_arr, y_arr, cols[label])
                all_sheet_data[sh] = grid
    _wb2.close()

    u_grid,    xs, ys = to_grid(x_arr, y_arr, u_cols[label])
    v_grid,    _,  _  = to_grid(x_arr, y_arr, v_cols[label])
    init_grid, _,  _  = to_grid(x_arr, y_arr, init_cols[label])

    current_map_grid  = init_grid.copy()
    current_map_sheet = 'gamma_max'

    THR_INIT = 0.02
    disc_mask_all = current_map_grid > THR_INIT

    _xi_map = {float(v): i for i, v in enumerate(xs)}
    _yi_map = {float(v): i for i, v in enumerate(ys)}
    grain_grid = np.zeros(init_grid.shape, dtype=int)
    for (cx_i, cy_i), gid in coord_to_grain.items():
        xi_ = _xi_map.get(float(cx_i))
        yi_ = _yi_map.get(float(cy_i))
        if xi_ is not None and yi_ is not None:
            grain_grid[yi_, xi_] = gid
    gb_mask = np.zeros(init_grid.shape, dtype=bool)
    gb_mask[:, 1:]  |= (grain_grid[:, 1:]  != grain_grid[:, :-1])
    gb_mask[1:, :]  |= (grain_grid[1:, :]  != grain_grid[:-1, :])
    gb_cx = xs[np.where(gb_mask)[1]]
    gb_cy = ys[np.where(gb_mask)[0]]

    # ==========================================================================
    # ステップ3: HoughGUI（改良版）
    # ==========================================================================
    grain_theta_map = {}
    grain_thr_map   = {}   # {grain_id: disc_thr at registration}

    class HoughGUI:
        def __init__(self):
            import matplotlib
            matplotlib.use('TkAgg')
            import matplotlib.pyplot as _plt
            import matplotlib.gridspec as gridspec
            from matplotlib.widgets import Slider, Button as MplButton, TextBox
            self.plt  = _plt

            # ---- フィギュア ----
            self.fig  = _plt.figure(figsize=(17, 11))
            self.fig.patch.set_facecolor('#1c1c2e')
            self.fig.canvas.manager.set_window_title(
                '結晶粒選択 & Hough変換  |  左パネルをクリックして結晶粒を選択')

            # ---- レイアウト: 左3段（マップ/Hough/プレビュー） + 右コントロール列 ----
            gs_left = gridspec.GridSpec(
                3, 1,
                left=0.05, right=0.54,
                top=0.96, bottom=0.05,
                hspace=0.38)
            self.ax_strain  = self.fig.add_subplot(gs_left[0])
            self.ax_hough   = self.fig.add_subplot(gs_left[1])
            self.ax_preview = self.fig.add_subplot(gs_left[2])

            for ax in [self.ax_strain, self.ax_hough, self.ax_preview]:
                ax.set_facecolor('#111122')
                ax.tick_params(colors='#aaaacc')
                for spine in ax.spines.values():
                    spine.set_edgecolor('#333355')
                ax.title.set_color('#ddddee')
                ax.xaxis.label.set_color('#aaaacc')
                ax.yaxis.label.set_color('#aaaacc')

            self.fig.subplots_adjust(left=0.05, right=0.54,
                                     top=0.96, bottom=0.05, hspace=0.38)
            for ax in [self.ax_hough, self.ax_preview]:
                ax.set_aspect('auto')

            # ---- 右コントロール列 (固定 axes) ----
            RX = 0.58   # 右列の開始x
            RW = 0.38   # 幅

            # 登録リスト（上部）
            self.ax_list = self.fig.add_axes([RX, 0.76, RW, 0.19])
            self.ax_list.set_facecolor('#111122')

            # マップ選択ボタン行
            # （_setup_widgets で生成）

            # 状態変数
            self.current_gid   = None
            self.current_theta = None
            self.disc_thr      = THR_INIT
            self.map_grid      = current_map_grid.copy()
            self.map_sheet     = current_map_sheet
            self.disc_mask     = disc_mask_all.copy()
            self._hough_vline  = None
            self._preview_line = None
            self.offset_max    = OFFSET_MAX
            self.min_side_px   = MIN_SIDE_PX

            self._draw_strain_map()
            self._draw_list()
            self._setup_widgets()
            self.fig.canvas.mpl_connect('button_press_event', self._on_click)
            self.fig.canvas.mpl_connect('key_press_event', self._on_key)

        # ------------------------------------------------------------------
        def _on_key(self, event):
            """キーボードショートカット"""
            if event.key == 'd':
                self._on_delete(None)
            elif event.key == 'left':
                new_val = max(self.slider_thr.valmin,
                              self.slider_thr.val - self.thr_step)
                self.slider_thr.set_val(new_val)
            elif event.key == 'right':
                new_val = min(self.slider_thr.valmax,
                              self.slider_thr.val + self.thr_step)
                self.slider_thr.set_val(new_val)

        # ------------------------------------------------------------------
        def _draw_strain_map(self):
            ax = self.ax_strain
            ax.clear()
            ax.set_facecolor('#111122')
            ax.tick_params(colors='#aaaacc')
            for spine in ax.spines.values():
                spine.set_edgecolor('#333355')

            grid = self.map_grid
            valid = grid[~np.isnan(grid)]
            if len(valid) == 0:
                vmin, vmax = 0, 1
            else:
                vmin = float(np.percentile(valid, 2))
                vmax = float(np.percentile(valid, 98))
            cmap = 'hot' if self.map_sheet == 'gamma_max' else 'RdBu_r'
            ax.imshow(grid, cmap=cmap, origin='upper',
                      extent=[xs.min(), xs.max(), ys.max(), ys.min()],
                      vmin=vmin, vmax=vmax, alpha=1.0)
            gb_rgba = np.zeros((*grain_grid.shape, 4), dtype=float)
            gb_rgba[gb_mask, 3] = 0.85
            ax.imshow(gb_rgba, origin='upper',
                      extent=[xs.min(), xs.max(), ys.max(), ys.min()])
            # 登録済み粒: ライムグリーンでハイライト＋grain_id ラベル
            for gid_reg in grain_theta_map:
                gcoords = [(cx_i, cy_i) for (cx_i, cy_i), g
                           in coord_to_grain.items() if g == gid_reg]
                if gcoords:
                    ax.scatter([c[0] for c in gcoords],
                               [c[1] for c in gcoords],
                               s=6, c='lime', alpha=0.8, linewidths=0, zorder=5)
                    cx_mean = float(np.mean([c[0] for c in gcoords]))
                    cy_mean = float(np.mean([c[1] for c in gcoords]))
                    ax.text(cx_mean, cy_mean, str(gid_reg),
                            fontsize=13, color='yellow',
                            ha='center', va='center', zorder=7,
                            bbox=dict(boxstyle='round,pad=0.15',
                                      facecolor='#222244', alpha=0.7,
                                      edgecolor='none'))
            # 現在選択中の粒をシアンでハイライト
            if self.current_gid is not None:
                cur_coords = [(cx_i, cy_i) for (cx_i, cy_i), g
                              in coord_to_grain.items() if g == self.current_gid]
                if cur_coords:
                    ax.scatter([c[0] for c in cur_coords],
                               [c[1] for c in cur_coords],
                               s=10, c='cyan', alpha=0.6, linewidths=0, zorder=6)
            ax.set_xlim(xs.min(), xs.max())
            ax.set_ylim(ys.max(), ys.min())
            n_reg = len(grain_theta_map)
            ax.set_title(
                f'{self.map_sheet} ({label})  登録粒:{n_reg}件  ← クリックで粒選択',
                fontsize=15, color='#ddddee')
            ax.set_xlabel('x [px]', color='#aaaacc', fontsize=14)
            ax.set_ylabel('y [px]', color='#aaaacc', fontsize=14)
            self.fig.canvas.draw_idle()

        # ------------------------------------------------------------------
        def _on_click(self, event):
            if event.inaxes != self.ax_strain:
                return
            cx_click = int(round(event.xdata))
            cy_click = int(round(event.ydata))
            best_gid  = None
            best_dist = np.inf
            for (cx_i, cy_i), gid in coord_to_grain.items():
                d = (cx_i - cx_click)**2 + (cy_i - cy_click)**2
                if d < best_dist:
                    best_dist = d
                    best_gid  = gid
            if best_gid is None:
                return
            self.current_gid = best_gid
            self._update_hough(best_gid)

        # ------------------------------------------------------------------
        def _update_hough(self, gid):
            grain_disc = np.zeros_like(self.disc_mask, dtype=np.uint8)
            for (cx_i, cy_i), g in coord_to_grain.items():
                if g != gid:
                    continue
                xi = np.searchsorted(xs, cx_i)
                yi = np.searchsorted(ys, cy_i)
                if (0 <= yi < self.disc_mask.shape[0] and
                        0 <= xi < self.disc_mask.shape[1] and
                        self.disc_mask[yi, xi]):
                    grain_disc[yi, xi] = 1
            n_disc = int(grain_disc.sum())

            for ax in [self.ax_hough, self.ax_preview]:
                ax.clear()
                ax.set_facecolor('#111122')
                ax.tick_params(colors='#aaaacc')
                for spine in ax.spines.values():
                    spine.set_edgecolor('#333355')
                ax.set_aspect('auto')
            self._hough_vline  = None
            self._preview_line = None

            if n_disc < 2:
                self.ax_hough.text(
                    0.5, 0.5, f'Grain {gid}\ndisc点不足 ({n_disc}点)',
                    ha='center', va='center',
                    transform=self.ax_hough.transAxes,
                    fontsize=14, color='#aaaacc')
                self.ax_hough.set_title(f'Grain {gid}  Hough空間',
                                        fontsize=15, color='#ddddee')
                self.current_theta = 0.0
                self._hough_data = None
            else:
                theta_deg, raw_hough_deg, hspace, theta_arr, d_arr = estimate_theta(
                    grain_disc.astype(bool))
                self.current_theta = theta_deg
                self._raw_hough_deg = raw_hough_deg
                self._hough_data = (hspace, theta_arr, d_arr)
                self.ax_hough.imshow(
                    np.log1p(hspace),
                    extent=[np.degrees(theta_arr[0]), np.degrees(theta_arr[-1]),
                            d_arr[-1], d_arr[0]],
                    cmap='hot', aspect='auto', origin='upper')
                self.ax_hough.axvline(
                    raw_hough_deg, color='yellow',
                    linewidth=1.5, linestyle='-',
                    label=f'走向θ={theta_deg:.1f}°')
                self.ax_hough.legend(
                    fontsize=7, loc='upper right',
                    facecolor='#222233', labelcolor='white')
                self.ax_hough.set_title(
                    f'Grain {gid}  Hough空間', fontsize=15, color='#ddddee')
                self.ax_hough.set_xlabel('θ [deg]', color='#aaaacc', fontsize=14)
                self.ax_hough.set_ylabel('ρ [px]', color='#aaaacc', fontsize=14)

            rows_g, cols_g = np.where(grain_disc)
            self.ax_preview.scatter(xs[cols_g], ys[rows_g],
                                    s=4, c='#ff6666', alpha=0.8, linewidths=0)
            self.ax_preview.set_xlim(xs.min(), xs.max())
            self.ax_preview.set_ylim(ys.max(), ys.min())
            self.ax_preview.set_title(
                f'Grain {gid}  disc={n_disc}点', fontsize=15, color='#ddddee')
            self.ax_preview.set_xlabel('x [px]', color='#aaaacc', fontsize=14)
            self.ax_preview.set_ylabel('y [px]', color='#aaaacc', fontsize=14)

            if hasattr(self, 'slider'):
                self.slider.set_val(self.current_theta)  # これで _on_slider が発火し線を描く
            self.fig.canvas.draw_idle()

        # ------------------------------------------------------------------
        def _setup_widgets(self):
            from matplotlib.widgets import Slider, Button as MplButton, TextBox
            from matplotlib.patches import FancyBboxPatch

            RX = 0.58
            RW = 0.38

            # ---- 閾値スライダー ----
            ax_thr_lbl = self.fig.add_axes([RX, 0.73, RW, 0.022])
            ax_thr_lbl.axis('off')
            ax_thr_lbl.set_facecolor('none')
            ax_thr_lbl.text(0.0, 0.5, 'しきい値',
                            va='center', fontsize=15, color='#ddddee',
                            transform=ax_thr_lbl.transAxes)

            # valtext がウィンドウ右端にはみ出ないよう幅を RW*0.82 に縮小
            ax_thr = self.fig.add_axes([RX, 0.695, RW * 0.82, 0.025])
            ax_thr.set_facecolor('#1c1c2e')

            # ---- θスライダー ----
            ax_sl_lbl = self.fig.add_axes([RX, 0.66, RW, 0.022])
            ax_sl_lbl.axis('off')
            ax_sl_lbl.set_facecolor('none')
            ax_sl_lbl.text(0.0, 0.5, '不連続線の角度 θ [deg]',
                           va='center', fontsize=15, color='#ddddee',
                           transform=ax_sl_lbl.transAxes)

            ax_sl = self.fig.add_axes([RX, 0.625, RW * 0.82, 0.025])
            ax_sl.set_facecolor('#1c1c2e')

            # ---- マップ選択ボタン ----
            map_keys = list(all_sheet_data.keys())
            n = len(map_keys)
            btn_w = RW / n
            btn_gap = 0.003
            self._map_btns = []
            self._map_btn_axes = []
            lbl_ax = self.fig.add_axes([RX, 0.597, RW, 0.020])
            lbl_ax.axis('off')
            lbl_ax.set_facecolor('none')
            lbl_ax.text(0.0, 0.5, 'マップ選択',
                        va='center', fontsize=15, color='#ddddee',
                        transform=lbl_ax.transAxes)
            for i, sh in enumerate(map_keys):
                w_i = btn_w - btn_gap
                ax_b = self.fig.add_axes([RX + i*btn_w, 0.555, w_i, 0.036])
                is_active = (sh == self.map_sheet)
                color     = '#2a5caa' if is_active else '#2d2d44'
                btn = MplButton(ax_b, sh, color=color, hovercolor='#3a3a5c')
                btn.label.set_fontsize(7.5)
                btn.label.set_color('white')
                btn.on_clicked(lambda _, s=sh: self._on_map_btn(s))
                self._map_btns.append(btn)
                self._map_btn_axes.append(ax_b)

            # ---- OFFSET_MAX / MIN_SIDE_PX ----
            param_lbl_ax = self.fig.add_axes([RX, 0.525, RW, 0.020])
            param_lbl_ax.axis('off')
            param_lbl_ax.set_facecolor('none')
            param_lbl_ax.text(0.0, 0.5, 'パラメータ設定',
                              va='center', fontsize=15, color='#ddddee',
                              transform=param_lbl_ax.transAxes)

            # OFFSET_MAX
            ax_off_lbl = self.fig.add_axes([RX, 0.485, RW * 0.55, 0.030])
            ax_off_lbl.axis('off')
            ax_off_lbl.set_facecolor('none')
            ax_off_lbl.text(0.0, 0.5, 'OFFSET_MAX [px]:',
                            va='center', fontsize=15, color='#ccccdd',
                            transform=ax_off_lbl.transAxes)
            ax_off = self.fig.add_axes([RX + RW * 0.57, 0.482, RW * 0.35, 0.034])

            # MIN_SIDE_PX
            ax_msp_lbl = self.fig.add_axes([RX, 0.435, RW * 0.55, 0.030])
            ax_msp_lbl.axis('off')
            ax_msp_lbl.set_facecolor('none')
            ax_msp_lbl.text(0.0, 0.5, 'MIN_SIDE_PX [px]:',
                            va='center', fontsize=15, color='#ccccdd',
                            transform=ax_msp_lbl.transAxes)
            ax_msp = self.fig.add_axes([RX + RW * 0.57, 0.432, RW * 0.35, 0.034])

            # ---- ボタン行（追加 / 削除 / 実行） ----
            BW = RW / 3 - 0.008
            ax_add = self.fig.add_axes([RX,                  0.37, BW, 0.048])
            ax_del = self.fig.add_axes([RX + RW/3 + 0.004,   0.37, BW, 0.048])
            ax_run = self.fig.add_axes([RX + 2*RW/3 + 0.008, 0.37, BW, 0.048])

            # ---- スライダー生成 ----
            grid = self.map_grid
            valid = grid[~np.isnan(grid)]
            thr_max  = float(np.percentile(valid, 99)) if len(valid) else 1.0
            thr_min  = float(np.percentile(valid, 0))  if len(valid) else 0.0
            thr_init = (THR_INIT if self.map_sheet == 'gamma_max'
                        else float(np.percentile(valid, 0)) if len(valid) else 0.0)

            self.slider_thr = Slider(ax_thr, '', thr_min, thr_max,
                                     valinit=thr_init, color='#4488cc')
            self.slider_thr.label.set_fontsize(15)
            self.slider_thr.valtext.set_color('#ddddee')
            self.slider_thr.valtext.set_fontsize(15)
            self.slider_thr.on_changed(self._on_thr_slider)
            self.thr_step = (thr_max - thr_min) / 1000

            self.slider = Slider(ax_sl, '', -90, 90, valinit=0, valstep=0.5,
                                 color='#44aa88')
            self.slider.label.set_fontsize(15)
            self.slider.valtext.set_color('#ddddee')
            self.slider.valtext.set_fontsize(15)
            self.slider.on_changed(self._on_slider)

            # ---- 操作ボタン ----
            self.btn_add = MplButton(ax_add, '＋ 結晶粒を追加',
                                     color='#1a6b2a', hovercolor='#258a38')
            self.btn_add.label.set_fontsize(14)
            self.btn_add.label.set_color('white')
            self.btn_add.on_clicked(self._on_add)

            self.btn_del = MplButton(ax_del, '－ 削除',
                                     color='#3d3d4a', hovercolor='#555566')
            self.btn_del.label.set_fontsize(14)
            self.btn_del.label.set_color('#ccccdd')
            self.btn_del.on_clicked(self._on_delete)

            self.btn_run = MplButton(ax_run, '▶ H-DIC実行',
                                     color='#7a4a00', hovercolor='#aa6600')
            self.btn_run.label.set_fontsize(14)
            self.btn_run.label.set_color('white')
            self.btn_run.on_clicked(self._on_run)

            # ---- TextBox ----
            self.tb_off = TextBox(ax_off, '', initial=str(int(self.offset_max)),
                                  color='#1c1c2e', hovercolor='#2a2a3e')
            self.tb_off.text_disp.set_color('#eeeeee')
            self.tb_msp = TextBox(ax_msp, '', initial=str(int(self.min_side_px)),
                                  color='#1c1c2e', hovercolor='#2a2a3e')
            self.tb_msp.text_disp.set_color('#eeeeee')

            # ---- ヒントテキスト ----
            hint_ax = self.fig.add_axes([RX, 0.28, RW, 0.07])
            hint_ax.axis('off')
            hint_ax.set_facecolor('none')
            hint_ax.text(0.0, 1.0,
                         '使い方:\n'
                         '① 左マップをクリック → 粒選択\n'
                         '② θスライダーで不連続線の角度を微調整\n'
                         '③「結晶粒を追加」で登録\n'
                         '④「H-DIC実行」で計算開始\n\n'
                         'キーボード:\n'
                         '  ←/→ = しきい値微調整',
                         va='top', fontsize=14, color='#888899', linespacing=1.5,
                         transform=hint_ax.transAxes)

        # ------------------------------------------------------------------
        def _on_map_btn(self, sheet_name):
            if sheet_name not in all_sheet_data:
                return
            self.map_sheet = sheet_name
            self.map_grid  = all_sheet_data[sheet_name]
            map_keys = list(all_sheet_data.keys())
            for btn, sh in zip(self._map_btns, map_keys):
                is_active = (sh == sheet_name)
                btn.color = '#2a5caa' if is_active else '#2d2d44'
                btn.ax.set_facecolor(btn.color)
            valid = self.map_grid[~np.isnan(self.map_grid)]
            if len(valid):
                thr_min  = float(np.percentile(valid, 0))
                thr_max  = float(np.percentile(valid, 99))
                thr_init = float(np.percentile(valid, 0))
                # valmin/valmax + 軸xlim を両方更新しないとトラックが古い範囲のまま
                self.slider_thr.valmin = thr_min
                self.slider_thr.valmax = thr_max
                self.slider_thr.ax.set_xlim(thr_min, thr_max)
                self.slider_thr.set_val(thr_init)
            self.disc_mask = self.map_grid > self.slider_thr.val
            self.disc_thr  = self.slider_thr.val
            self._draw_strain_map()
            self.fig.canvas.draw_idle()

        def _on_thr_slider(self, val):
            self.disc_thr  = float(val)
            self.disc_mask = self.map_grid > self.disc_thr
            self._draw_strain_map()
            if self.current_gid is not None:
                self._update_hough(self.current_gid)

        def _on_slider(self, val):
            self.current_theta = float(val)
            ax_h = self.ax_hough
            ax_p = self.ax_preview

            if self.current_gid is not None:
                # Hough 空間の縦線更新
                if self._hough_vline is not None:
                    try:
                        self._hough_vline.remove()
                    except ValueError:
                        pass
                raw_val = val - 90
                if raw_val < -90:
                    raw_val += 180
                self._hough_vline = ax_h.axvline(
                    raw_val, color='cyan', linewidth=1.5,
                    linestyle='--', label=f'走向θ={val:.1f}°')
                ax_h.legend(fontsize=7, loc='upper right',
                            facecolor='#222233', labelcolor='white')

                # プレビュー直線を「現在選択粒のdisc点群の重心」中心に描く
                if self._preview_line is not None:
                    try:
                        self._preview_line.remove()
                    except ValueError:
                        pass
                # 粒のdisc点座標を取得
                grain_disc_pts = [(xs[c_idx], ys[r_idx])
                                  for (cx_i, cy_i), g in coord_to_grain.items()
                                  if g == self.current_gid
                                  for r_idx, c_idx in [
                                      (np.searchsorted(ys, cy_i),
                                       np.searchsorted(xs, cx_i))]
                                  if (0 <= r_idx < self.disc_mask.shape[0] and
                                      0 <= c_idx < self.disc_mask.shape[1] and
                                      self.disc_mask[r_idx, c_idx])]
                if grain_disc_pts:
                    xc = float(np.mean([p[0] for p in grain_disc_pts]))
                    yc = float(np.mean([p[1] for p in grain_disc_pts]))
                    # 粒の広がりに合わせた直線長
                    x_span = (max(p[0] for p in grain_disc_pts)
                              - min(p[0] for p in grain_disc_pts))
                    y_span = (max(p[1] for p in grain_disc_pts)
                              - min(p[1] for p in grain_disc_pts))
                    length = max(x_span, y_span, STEP * 4) * 1.3
                else:
                    xc = (xs.min() + xs.max()) / 2
                    yc = (ys.min() + ys.max()) / 2
                    length = max(xs.max() - xs.min(), ys.max() - ys.min())
                t_rad = np.radians(val)
                dx_l = np.cos(t_rad) * length / 2
                dy_l = np.sin(t_rad) * length / 2
                lines = ax_p.plot(
                    [xc - dx_l, xc + dx_l], [yc - dy_l, yc + dy_l],
                    color='cyan', linewidth=1.8, alpha=0.9,
                    label=f'θ={val:.1f}°')
                self._preview_line = lines[0]
                ax_p.legend(fontsize=7, loc='upper right')

            self.fig.canvas.draw_idle()

        def _on_add(self, event):
            if self.current_gid is None:
                return
            grain_theta_map[self.current_gid] = self.current_theta
            grain_thr_map[self.current_gid]   = self.disc_thr
            print(f"  追加: grain {self.current_gid}  theta={self.current_theta:.1f}deg  thr={self.disc_thr:.4f}")
            self._draw_list()
            self._draw_strain_map()

        def _on_delete(self, event):
            if not grain_theta_map:
                return
            # 現在選択中の粒を優先削除、なければ最後の粒を削除
            target = (self.current_gid
                      if self.current_gid in grain_theta_map
                      else list(grain_theta_map.keys())[-1])
            del grain_theta_map[target]
            grain_thr_map.pop(target, None)
            print(f"  削除: grain {target}")
            self._draw_list()
            self._draw_strain_map()

        def _draw_list(self):
            ax = self.ax_list
            ax.clear()
            ax.axis('off')
            ax.set_facecolor('#111122')
            if not grain_theta_map:
                ax.text(0.04, 0.5, '（結晶粒が登録されていません）',
                        transform=ax.transAxes, fontsize=17,
                        color='#666688', va='center')
            else:
                n = len(grain_theta_map)
                target_gids = set(grain_theta_map.keys())
                # 粒ごとのしきい値で個別にサブセット数を集計
                subset_counts = {gid: 0 for gid in target_gids}
                for (cx_i, cy_i), gid in coord_to_grain.items():
                    if gid not in target_gids:
                        continue
                    thr_g = grain_thr_map.get(gid, self.disc_thr)
                    xi = np.searchsorted(xs, cx_i)
                    yi = np.searchsorted(ys, cy_i)
                    if (0 <= yi < self.map_grid.shape[0] and
                            0 <= xi < self.map_grid.shape[1] and
                            not np.isnan(self.map_grid[yi, xi]) and
                            self.map_grid[yi, xi] > thr_g):
                        subset_counts[gid] += 1

                header = f'登録結晶粒  {n}件'
                lines  = [header, '']
                for gid, th in sorted(grain_theta_map.items()):
                    n_sub = subset_counts.get(gid, 0)
                    thr   = grain_thr_map.get(gid, float('nan'))
                    lines.append(
                        f'Grain {gid:3d}  '
                        f'theta={th:+.1f} [deg]  '
                        f'thr={thr:.4f}  '
                        f'n={n_sub}')
                ax.text(0.04, 0.97, '\n'.join(lines),
                        transform=ax.transAxes,
                        fontsize=15, va='top',
                        color='#ddddee')
            self.fig.canvas.draw_idle()

        def _on_run(self, event):
            if not grain_theta_map:
                messagebox.showwarning('警告', '粒が登録されていません。')
                return
            try:
                self.offset_max  = float(self.tb_off.text)
            except Exception:
                pass
            try:
                self.min_side_px = int(self.tb_msp.text)
            except Exception:
                pass
            self.plt.close(self.fig)

    gui = HoughGUI()
    gui.fig.canvas.draw()
    pos0 = gui.ax_strain.get_position()
    for ax2 in [gui.ax_hough, gui.ax_preview]:
        p = ax2.get_position()
        ax2.set_position([pos0.x0, p.y0, pos0.width, p.height])
    gui.plt.show()

    if not grain_theta_map:
        raise SystemExit('粒が登録されなかったため処理を終了します。')

    OFFSET_MAX  = gui.offset_max
    MIN_SIDE_PX = gui.min_side_px
    print(f"パラメータ確定: OFFSET_MAX={OFFSET_MAX:.0f}px, MIN_SIDE_PX={MIN_SIDE_PX}px")

    # ==========================================================================
    # ステップ4: SEM画像選択 & H-DIC実行
    # ==========================================================================
    print(f"\n登録粒: {dict(sorted(grain_theta_map.items()))}")

    print("参照SEM画像 (Ref) を選択してください...")
    ref_path = filedialog.askopenfilename(
        title='参照SEM画像 (Ref) を選択',
        filetypes=[('Image files', '*.bmp *.png *.tif *.tiff'), ('All files', '*.*')])
    if not ref_path:
        raise SystemExit('キャンセルされました')

    def_paths = {}
    for lbl in [label]:
        path = filedialog.askopenfilename(
            title=f'変形後SEM画像を選択: {lbl}',
            filetypes=[('Image files', '*.bmp *.png *.tif *.tiff'), ('All files', '*.*')])
        if not path:
            raise SystemExit(f'{lbl} の画像選択がキャンセルされました')
        def_paths[lbl] = path

    print("\n" + "="*60)
    print(f"H-DIC実行開始: {len(grain_theta_map)} 粒")
    print("="*60)

    result_dir = out_dir / f'heaviside_dic_{label}'
    result_dir.mkdir(parents=True, exist_ok=True)
    out_path = str(result_dir / f'heaviside_results_{label}.xlsx')

    config_txt_path = result_dir / f'hdic_config_{label}.txt'
    import datetime as _dt
    with open(str(config_txt_path), 'w', encoding='utf-8') as _cf:
        _cf.write(f'Heaviside DIC 設定パラメータ\n')
        _cf.write(f'{"="*40}\n')
        _cf.write(f'実行日時      : {_dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
        _cf.write(f'応力段階      : {label}\n')
        _cf.write(f'\n[通常DIC設定]\n')
        _cf.write(f'SUBSET_SIZE   : {SUBSET_SIZE} px\n')
        _cf.write(f'STEP          : {STEP} px\n')
        _cf.write(f'TRIM_BOTTOM   : {TRIM_BOTTOM} px\n')
        _cf.write(f'NCC_THRESHOLD : {NCC_THRESHOLD}\n')
        _cf.write(f'\n[H-DIC設定]\n')
        _cf.write(f'OFFSET_MAX    : {OFFSET_MAX:.0f} px\n')
        _cf.write(f'MIN_SIDE_PX   : {MIN_SIDE_PX} px\n')
        _cf.write(f'THETA_RANGE   : ±{THETA_RANGE:.1f} deg\n')
        _cf.write(f'THETA_STEP    : {THETA_STEP:.1f} deg\n')
        _cf.write(f'SEARCH_RANGE  : ±{SEARCH_RANGE} px\n')
        _cf.write(f'\n[HoughGUI設定]\n')
        _cf.write(f'マップシート  : {gui.map_sheet}\n')
        _cf.write(f'\n[登録結晶粒  grain_id : theta [deg] / threshold / n_subsets]\n')
        for gid, th in sorted(grain_theta_map.items()):
            thr   = grain_thr_map.get(gid, float('nan'))
            # 粒ごとのしきい値でサブセット数を集計
            n_sub = sum(
                1 for (cx_i, cy_i), g in coord_to_grain.items()
                if g == gid
                for r_idx, c_idx in [
                    (int(np.searchsorted(ys, cy_i)),
                     int(np.searchsorted(xs, cx_i)))]
                if (0 <= r_idx < gui.map_grid.shape[0] and
                    0 <= c_idx < gui.map_grid.shape[1] and
                    not np.isnan(gui.map_grid[r_idx, c_idx]) and
                    gui.map_grid[r_idx, c_idx] > thr)
            )
            _cf.write(
                f'  Grain {gid:3d} : '
                f'theta={th:+.1f} [deg]  '
                f'thr={thr:.4f}  '
                f'n={n_sub}\n'
            )
        _cf.write(f'\n[ファイル]\n')
        _cf.write(f'georef xlsx   : {georef_xlsx}\n')
        _cf.write(f'REF画像       : {ref_path}\n')
        _cf.write(f'DEF画像       : {def_paths[label]}\n')
        if alignment_json:
            _cf.write(f'alignment json: {alignment_json}\n')
    print(f"設定ファイル保存: {config_txt_path}")

    hdic_results, deformed_img, hdic_subset_list = run_heaviside_dic(
        xlsx_path            = georef_xlsx,
        ref_path             = ref_path,
        def_path             = def_paths[label],
        label                = label,
        grain_theta_map      = grain_theta_map,
        coord_to_grain       = coord_to_grain,
        out_path             = out_path,
        n_jobs               = max(1, (os.cpu_count() or 2) - 1),
        alignment_json_path  = alignment_json,
        disc_threshold       = gui.disc_thr,
        offset_max           = OFFSET_MAX,
        min_side_px          = MIN_SIDE_PX,
        disc_sheet           = gui.map_sheet,
        grain_thr_map        = grain_thr_map,
    )

    print("\n" + "="*60)
    print("計算完了! 結果確認GUIを起動します...")
    print("="*60)

    # strain_data: 全サブセットを通常DIC変位で補正した座標にひずみ値をプロット
    # x_def = cx + u0, y_def = cy + v0
    _XX, _YY = np.meshgrid(xs, ys)
    _U = np.where(np.isnan(u_grid), 0.0, u_grid)
    _V = np.where(np.isnan(v_grid), 0.0, v_grid)
    _Xdef = (_XX + _U).ravel()
    _Ydef = (_YY + _V).ravel()

    strain_data = {}
    for sh, grid in all_sheet_data.items():
        flat = grid.ravel()
        valid_mask = ~np.isnan(flat)
        if not valid_mask.any():
            continue
        strain_data[sh] = {
            'x_def': _Xdef[valid_mask],
            'y_def': _Ydef[valid_mask],
            'val':   flat[valid_mask],
        }

    _show_result_gui(hdic_results, deformed_img, out_path, label,
                     subset_list=hdic_subset_list,
                     strain_data=strain_data)

    print(f"出力先フォルダ: {result_dir}")
    messagebox.showinfo('完了', f'H-DIC完了\n出力先: {result_dir}')
