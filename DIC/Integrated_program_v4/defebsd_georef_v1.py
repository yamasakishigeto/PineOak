"""
defebsd_georef_v1.py
====================
Def_nth EBSD データのジオリファレンス（変形後DIC座標系への位置合わせ）。

処理フロー:
  各変形段階について以下を繰り返す:
  1. dic_results_georef.xlsx から変形後DIC Grain IDマップを生成
  2. Def_nth EBSD ファイル（.mat or .txt）から EBSD Grain IDマップを生成
  3. 左（変形後DIC）→ 右（EBSD）でコントロールポイントを指定
  4. 各EBSDグリッド点に最近傍DICサブセットの subset_id を割り当て
  5. 出力:
     - .mat モード: subset_id + 2D全フィールドを抽出して _georef.mat として保存
     - .txt モード: dic_results_georef.xlsx に新シートを追加

座標系:
  DIC :  SEM画像ピクセル座標 (cx, cy) [px]
  EBSD:  物理座標 (xpos, ypos) [μm]、Grain IDマップ画像座標 (ix, iy) [px]

変換:
  H (2×3 アフィン行列):  変形後DICキャンバスピクセル → EBSDキャンバスピクセル
  H_inv               :  EBSDキャンバスピクセル → 変形後DICキャンバスピクセル
"""

import sys
import os
import json
import numpy as np
from pathlib import Path
from collections import deque

# 非日本語環境でのUnicodeEncodeError対策
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

os.environ['MPLBACKEND'] = 'QtAgg'
import matplotlib
matplotlib.use('QtAgg')

# ── 必須ライブラリチェック ────────────────────────────────────────────────
_REQUIRED = {
    'numpy':      'numpy',
    'scipy':      'scipy',
    'matplotlib': 'matplotlib',
    'openpyxl':   'openpyxl',
    'cv2':        'opencv-python',
}
_missing = []
for _mod, _pkg in _REQUIRED.items():
    try:
        __import__(_mod)
    except ImportError:
        _missing.append(_pkg)
if _missing:
    print("=" * 60)
    print("  [ERROR] Required libraries not found.")
    print(f"  pip install {' '.join(_missing)}")
    print("=" * 60)
    sys.exit(1)

import cv2
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.font_manager as _fm
import platform as _platform
from scipy.spatial import cKDTree

# ── 日本語フォント設定 ────────────────────────────────────────────────────
_JP_FONTS = {
    'Windows': ['Meiryo', 'Yu Gothic', 'MS Gothic'],
    'Darwin':  ['Hiragino Sans', 'Hiragino Maru Gothic Pro', 'AppleGothic'],
    'Linux':   ['IPAexGothic', 'IPAPGothic', 'Noto Sans CJK JP'],
}
_available_fonts = {f.name for f in _fm.fontManager.ttflist}
_jp_candidates   = _JP_FONTS.get(_platform.system(), [])
_jp_font = next((f for f in _jp_candidates if f in _available_fonts), None)
if _jp_font:
    matplotlib.rcParams['font.family'] = _jp_font
matplotlib.rcParams['axes.unicode_minus'] = False


# =============================================================================
# dic_results_georef.xlsx から変形後DICデータを読み込む
# =============================================================================
def load_deformed_dic_grain_map(georef_xlsx, label):
    """
    dic_results_georef.xlsx から指定ラベルの変形後 DIC Grain ID マップを読み込む。

    Returns
    -------
    cx_def     : ndarray (N,)  変形後X座標 [px]
    cy_def     : ndarray (N,)  変形後Y座標 [px]
    grain_ids  : ndarray (N,)  Grain ID（>0のみ）
    subset_ids : ndarray (N,)  subset_id
    """
    wb = openpyxl.load_workbook(georef_xlsx, read_only=True)

    # ebsd_georef シート: subset_id → grain_id
    rows_eg = list(wb['ebsd_georef'].iter_rows(values_only=True))
    hdr_eg  = rows_eg[0]
    sid_col_eg = next(i for i, h in enumerate(hdr_eg) if h == 'subset_id')
    gid_col_eg = next(i for i, h in enumerate(hdr_eg) if h == 'grain_id')
    sid_to_gid = {int(r[sid_col_eg]): int(r[gid_col_eg])
                  for r in rows_eg[1:] if r[sid_col_eg] is not None}

    # u シート: subset_id, x [px], u[label]
    rows_u  = list(wb['u'].iter_rows(values_only=True))
    hdr_u   = rows_u[0]
    sid_col_u  = next(i for i, h in enumerate(hdr_u) if h == 'subset_id')
    xref_col   = next(i for i, h in enumerate(hdr_u) if h == 'x [px]')
    try:
        u_col = next(i for i, h in enumerate(hdr_u) if h == label)
    except StopIteration:
        wb.close()
        raise ValueError(f"Label '{label}' not found in 'u' sheet of {georef_xlsx}")

    # v シート: subset_id, y [px], v[label]
    rows_v  = list(wb['v'].iter_rows(values_only=True))
    hdr_v   = rows_v[0]
    sid_col_v  = next(i for i, h in enumerate(hdr_v) if h == 'subset_id')
    yref_col   = next(i for i, h in enumerate(hdr_v) if h == 'y [px]')
    try:
        v_col = next(i for i, h in enumerate(hdr_v) if h == label)
    except StopIteration:
        wb.close()
        raise ValueError(f"Label '{label}' not found in 'v' sheet of {georef_xlsx}")

    wb.close()

    # v シートから sid → y_def を構築
    sid_to_ydef = {}
    for r in rows_v[1:]:
        if r[sid_col_v] is None:
            continue
        sid  = int(r[sid_col_v])
        yref = float(r[yref_col]) if r[yref_col] is not None else None
        v    = float(r[v_col])    if r[v_col]    is not None else 0.0
        if yref is not None:
            sid_to_ydef[sid] = yref + v

    # u シートを走査して cx_def, cy_def, grain_id を収集
    subset_ids_list = []
    cx_def_list     = []
    cy_def_list     = []
    grain_ids_list  = []

    for r in rows_u[1:]:
        if r[sid_col_u] is None:
            continue
        sid  = int(r[sid_col_u])
        gid  = sid_to_gid.get(sid, 0)
        if gid <= 0:
            continue
        xref = float(r[xref_col]) if r[xref_col] is not None else None
        u    = float(r[u_col])    if r[u_col]    is not None else 0.0
        if xref is None or sid not in sid_to_ydef:
            continue
        subset_ids_list.append(sid)
        cx_def_list.append(xref + u)
        cy_def_list.append(sid_to_ydef[sid])
        grain_ids_list.append(gid)

    cx_def     = np.array(cx_def_list,    dtype=float)
    cy_def     = np.array(cy_def_list,    dtype=float)
    grain_ids  = np.array(grain_ids_list, dtype=int)
    subset_ids = np.array(subset_ids_list, dtype=int)

    print(f"  Deformed DIC subsets loaded: {len(subset_ids)}  (label='{label}')")
    return cx_def, cy_def, grain_ids, subset_ids


def get_available_labels(georef_xlsx):
    """dic_results_georef.xlsx の u シートから利用可能な変形段階ラベルを返す。"""
    wb   = openpyxl.load_workbook(georef_xlsx, read_only=True)
    hdr  = list(wb['u'].iter_rows(values_only=True, max_row=1))[0]
    wb.close()
    skip = {None, 'subset_id', 'x [px]', 'y [px]'}
    return [str(h) for h in hdr if h not in skip]


# =============================================================================
# 変形後 DIC Grain ID マップを画像（numpy RGB）に描画
# =============================================================================
def render_deformed_grain_image(cx_def, cy_def, grain_ids, step=None):
    """
    変形後DICサブセット座標 (cx_def, cy_def) をGrain IDで色付けし、
    numpy RGB画像（H×W×3, uint8）として返す。

    画像座標 (col, row) と DIC実座標の対応:
        DIC_x = col + x_offset
        DIC_y = row + y_offset

    Returns
    -------
    canvas   : ndarray (H, W, 3) uint8
    x_offset : int  DIC実座標のX最小値（マージン込み）
    y_offset : int  DIC実座標のY最小値（マージン込み）
    step_est : float  推定DICステップ [px]
    """
    if step is None:
        # 点の密度から推定（大変形時でも安定）
        # 変形後の座標は非等間隔になるため、x 差分の中央値は過小推定になりやすい
        x_range = float(cx_def.max() - cx_def.min())
        y_range = float(cy_def.max() - cy_def.min())
        n = len(cx_def)
        if n > 1 and x_range > 0 and y_range > 0:
            step_est = float(np.sqrt(x_range * y_range / n))
        else:
            step_est = 15.0
    else:
        step_est = float(step)

    half   = max(1, int(step_est // 2))
    margin = int(step_est * 3)

    x_offset = max(0, int(cx_def.min()) - margin)
    y_offset = max(0, int(cy_def.min()) - margin)
    W = int(cx_def.max()) + margin - x_offset + 1
    H = int(cy_def.max()) + margin - y_offset + 1

    # Grain IDごとのカラーパレット
    valid_ids = sorted(set(int(g) for g in grain_ids if g > 0))
    pool = []
    for _name in ('tab20', 'tab20b', 'tab20c'):
        _c = matplotlib.colormaps[_name]
        pool.extend([_c(i) for i in range(20)])
    idxs = np.linspace(0, len(pool) - 1, max(len(valid_ids), 1), dtype=int)
    gid_to_color = {
        gid: (np.array(pool[idxs[i]][:3]) * 255).astype(np.uint8)
        for i, gid in enumerate(valid_ids)
    }

    canvas = np.zeros((H, W, 3), dtype=np.uint8)
    for cx, cy, gid in zip(cx_def, cy_def, grain_ids):
        if gid <= 0:
            continue
        color = gid_to_color.get(int(gid))
        if color is None:
            continue
        col = int(round(cx)) - x_offset
        row = int(round(cy)) - y_offset
        r0, r1 = max(0, row - half), min(H, row + half + 1)
        c0, c1 = max(0, col - half), min(W, col + half + 1)
        canvas[r0:r1, c0:c1] = color

    return canvas, x_offset, y_offset, step_est


# =============================================================================
# EBSD Grain File (.txt) 読み込み（ebsd_georef_v68 から流用）
# =============================================================================
def load_grain_file(path):
    """OIM Analysis Grain File（固定11列）を読み込む。"""
    num_rows   = []
    phase_col  = []
    grain_tol_angle = None

    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            line_s = line.strip()
            if not line_s:
                continue
            if line_s.startswith('#'):
                if 'Grain Tolerance Angle' in line_s:
                    try:
                        grain_tol_angle = float(line_s.split(':')[-1].strip())
                    except ValueError:
                        pass
                continue
            parts = line_s.split()
            if len(parts) < 11:
                continue
            try:
                num_rows.append([float(p) for p in parts[:10]])
                phase_col.append(parts[10])
            except ValueError:
                continue

    if not num_rows:
        raise ValueError(f"No data rows found: {path}")

    data        = np.array(num_rows, dtype=np.float64)
    phase_names = np.array(phase_col)

    x_vals  = np.unique(np.round(data[:, 3], 6))
    y_vals  = np.unique(np.round(data[:, 4], 6))
    nx      = len(x_vals)
    ny      = len(y_vals)
    x_step  = float(np.median(np.diff(x_vals))) if nx > 1 else 1.0
    y_step  = float(np.median(np.diff(y_vals))) if ny > 1 else 1.0
    x_min, x_max = float(x_vals.min()), float(x_vals.max())
    y_min, y_max = float(y_vals.min()), float(y_vals.max())

    unique_phases   = sorted(set(phase_col))
    grain_ids       = data[:, 8].astype(int)
    n_valid_grains  = len(np.unique(grain_ids[grain_ids > 0]))

    print(f"  Grain File loaded: {len(data)} pixels  ({nx} x {ny})")
    print(f"  Step: x={x_step:.4f} um, y={y_step:.4f} um")
    print(f"  Valid grains (GrainID>0): {n_valid_grains}")

    meta = dict(
        nx=nx, ny=ny,
        x_step=x_step, y_step=y_step,
        x_min=x_min, y_min=y_min,
        x_max=x_max, y_max=y_max,
        grain_tol_angle=grain_tol_angle,
        phases=unique_phases,
    )
    return data, phase_names, meta


# =============================================================================
# .mat ファイル読み込み（ジオリファレンス用最小フィールド）
# =============================================================================
def load_mat_minimal(mat_path):
    """
    .mat から grain_number / xpos / ypos / euler角 / metadata のみ読み込む。
    大型配列（3次元以上）は読み込まない。
    """
    import scipy.io

    GEOREF_VARS = [
        'grain_number', 'xpos', 'ypos', 'xstep', 'ystep', 'phasetxt',
        'euler_phi1', 'euler_phi', 'euler_phi2',
        'numcols', 'numrows',
    ]
    mat = scipy.io.loadmat(mat_path, variable_names=GEOREF_VARS)

    ny, nx = mat['grain_number'].shape
    xpos   = mat['xpos'].flatten()
    ypos   = mat['ypos'].flatten()
    x_step = float(mat['xstep'][0, 0])
    y_step = float(mat['ystep'][0, 0])
    x_min, x_max = float(xpos.min()), float(xpos.max())
    y_min, y_max = float(ypos.min()), float(ypos.max())

    try:
        phase_name = str(mat['phasetxt'][0, 0][0])
    except Exception:
        phase_name = 'unknown'

    gid_raw = mat['grain_number'].flatten()
    gid     = np.where(np.isnan(gid_raw), 0.0, gid_raw)
    zeros   = np.zeros(len(gid), dtype=np.float64)

    phi1 = np.radians(mat['euler_phi1'].flatten())
    phi  = np.radians(mat['euler_phi'].flatten())
    phi2 = np.radians(mat['euler_phi2'].flatten())

    data = np.column_stack([phi1, phi, phi2, xpos, ypos,
                            zeros, zeros, zeros, gid, zeros])
    phase_names = np.array([phase_name] * len(data))

    print(f"  .mat loaded (minimal): {len(data)} pixels  ({nx} x {ny})")
    print(f"  Step: x={x_step:.4f} um, y={y_step:.4f} um")

    meta = dict(
        nx=nx, ny=ny,
        x_step=x_step, y_step=y_step,
        x_min=x_min, y_min=y_min,
        x_max=x_max, y_max=y_max,
        grain_tol_angle=None,
        phases=[phase_name],
    )
    return data, phase_names, meta


# =============================================================================
# EBSD Grain ID マップ画像生成（ebsd_georef_v68 から流用）
# =============================================================================
def make_grain_id_image(data, meta):
    """Grain ID からカラー RGB 画像（ny × nx × 3, uint8）を生成する。"""
    nx, ny   = meta['nx'], meta['ny']
    x_min    = meta['x_min']
    y_min    = meta['y_min']
    x_step   = meta['x_step']
    y_step   = meta['y_step']

    grain_id_map = np.zeros((ny, nx), dtype=np.int32)
    for row in data:
        ix = int(round((row[3] - x_min) / x_step))
        iy = int(round((row[4] - y_min) / y_step))
        if 0 <= ix < nx and 0 <= iy < ny:
            grain_id_map[iy, ix] = int(row[8])

    valid_ids = sorted(set(int(v) for v in grain_id_map.flat if v > 0))
    pool = []
    for _name in ('tab20', 'tab20b', 'tab20c'):
        _c = matplotlib.colormaps[_name]
        pool.extend([_c(i) for i in range(20)])
    idxs = np.linspace(0, len(pool) - 1, max(len(valid_ids), 1), dtype=int)
    id_to_color = {
        gid: (np.array(pool[idxs[i]][:3]) * 255).astype(np.uint8)
        for i, gid in enumerate(valid_ids)
    }

    rgb = np.zeros((ny, nx, 3), dtype=np.uint8)
    for iy in range(ny):
        for ix in range(nx):
            gid = grain_id_map[iy, ix]
            if gid > 0:
                rgb[iy, ix] = id_to_color[gid]
            elif gid < 0:
                rgb[iy, ix] = [80, 80, 80]

    return rgb, grain_id_map, id_to_color


# =============================================================================
# コントロールポイント指定 GUI（両パネル RGB 対応版）
# =============================================================================
def pick_control_points(left_img, right_img, left_label='', right_label=''):
    """
    左（変形後DIC Grain IDマップ）と右（EBSD Grain IDマップ）を並べて表示し、
    左→右の順に交互クリックでコントロールポイントを指定する。

    左パネルは RGB 画像（render_deformed_grain_image の出力）を想定。

    Parameters
    ----------
    left_label  : str  左パネルのデータ識別名（DIC ラベル）
    right_label : str  右パネルのデータ識別名（EBSD ファイル名など）

    Returns
    -------
    pts_left  : ndarray (N, 2)  左パネル画像座標 [px]
    pts_right : ndarray (N, 2)  右パネル画像座標 [px]
    """
    from matplotlib.widgets import Button

    _C = {
        'bg': '#0a0c0f', 'surface': '#111318', 'surface2': '#1a1d24',
        'border': '#2a2d35', 'accent': '#00d4ff', 'green': '#00ff88',
        'orange': '#ff6b35', 'text': '#e0e4ec', 'dim': '#6b7280',
    }

    H_L, W_L = left_img.shape[:2]
    H_R, W_R = right_img.shape[:2]

    plt.style.use('dark_background')
    fig, axes = plt.subplots(1, 2, figsize=(22, 10))
    fig.patch.set_facecolor(_C['bg'])
    for ax in axes:
        ax.set_facecolor(_C['surface'])
        for sp in ax.spines.values():
            sp.set_edgecolor(_C['border'])
        ax.tick_params(colors=_C['dim'], labelsize=11)
    fig.subplots_adjust(bottom=0.15, top=0.88, left=0.04, right=0.97, wspace=0.08)

    _win_left  = f'Deformed DIC  [{left_label}]'  if left_label  else 'Deformed DIC Grain ID'
    _win_right = f'EBSD  [{right_label}]'          if right_label else 'EBSD Grain ID'
    fig.canvas.manager.set_window_title(
        f'Def EBSD Georef  |  {_win_left} (L) → {_win_right} (R)'
        '  |  Right-click: undo  |  q: confirm')

    ax_left, ax_right = axes

    # 両パネルとも RGB 表示
    ax_left.imshow(left_img,  origin='upper')
    ax_right.imshow(right_img, origin='upper')

    _title_left  = f'Deformed DIC Grain ID  —  {left_label}\n[ click here first ]'  if left_label  else 'Deformed DIC Grain ID  [ click here first ]'
    _title_right = f'EBSD Grain ID  —  {right_label}\n[ click here second ]'         if right_label else 'EBSD Grain ID map  [ click here second ]'
    ax_left.set_title(_title_left,  fontsize=12, color=_C['accent'], pad=8)
    ax_right.set_title(_title_right, fontsize=12, color=_C['dim'],   pad=8)

    # オーバーレイ（EBSDをワープして左パネルに半透明表示）
    _overlay_im = ax_left.imshow(
        np.zeros((H_L, W_L, 4), dtype=np.uint8),
        origin='upper', alpha=1.0, zorder=2)
    _overlay_im.set_visible(False)
    _OVERLAY_OPACITY = 0.4

    # ボタン共通スタイル
    def _make_btn(rect, label, color):
        ax_b = fig.add_axes(rect)
        ax_b.set_facecolor(_C['surface2'])
        for sp in ax_b.spines.values():
            sp.set_edgecolor(color)
            sp.set_linewidth(1.2)
        btn = Button(ax_b, label, color=_C['surface2'], hovercolor=_C['surface'])
        btn.label.set_fontsize(11)
        btn.label.set_color(color)
        btn.label.set_fontfamily('monospace')
        return btn

    _btn_overlay = _make_btn([0.12, 0.025, 0.14, 0.055], 'Overlay: ON',  _C['accent'])
    _btn_done    = _make_btn([0.68, 0.025, 0.14, 0.055], 'Confirm',      _C['green'])

    pts_left  = []
    pts_right = []
    artists_left  = []
    artists_right = []
    state = {'next': 0, 'overlay_visible': True}

    def _redraw_labels():
        for a_l, a_r in zip(artists_left, artists_right):
            for a in a_l + a_r:
                a.remove()
        artists_left.clear()
        artists_right.clear()
        for i, (pl, pr) in enumerate(zip(pts_left, pts_right), start=1):
            mk_l, = ax_left.plot(pl[0],  pl[1],  'r+', markersize=16, markeredgewidth=2.5, zorder=5)
            lb_l  = ax_left.annotate(str(i),  (pl[0], pl[1]),  color='red', fontsize=14,
                                     fontweight='bold', xytext=(6, 6),
                                     textcoords='offset points', zorder=5)
            mk_r, = ax_right.plot(pr[0], pr[1], 'r+', markersize=16, markeredgewidth=2.5, zorder=5)
            lb_r  = ax_right.annotate(str(i), (pr[0], pr[1]), color='red', fontsize=14,
                                      fontweight='bold', xytext=(6, 6),
                                      textcoords='offset points', zorder=5)
            artists_left.append([mk_l, lb_l])
            artists_right.append([mk_r, lb_r])
        fig.canvas.draw_idle()

    def _update_overlay():
        n = min(len(pts_left), len(pts_right))
        if n < 3 or not state['overlay_visible']:
            _overlay_im.set_visible(False)
            fig.canvas.draw_idle()
            return
        src = np.array(pts_left[:n],  dtype=np.float32)
        dst = np.array(pts_right[:n], dtype=np.float32)
        try:
            if n == 3:
                H_mat = cv2.getAffineTransform(dst[:3], src[:3])
            else:
                H_mat, _ = cv2.estimateAffine2D(dst, src,
                                                method=cv2.RANSAC,
                                                ransacReprojThreshold=10.0)
        except Exception:
            _overlay_im.set_visible(False)
            fig.canvas.draw_idle()
            return
        if H_mat is None:
            _overlay_im.set_visible(False)
            fig.canvas.draw_idle()
            return
        warped = cv2.warpAffine(right_img, H_mat, (W_L, H_L),
                                flags=cv2.INTER_NEAREST,
                                borderMode=cv2.BORDER_CONSTANT, borderValue=(0, 0, 0))
        rgba = np.zeros((H_L, W_L, 4), dtype=np.uint8)
        rgba[:, :, :3] = warped
        is_black = (warped[:, :, 0] == 0) & (warped[:, :, 1] == 0) & (warped[:, :, 2] == 0)
        rgba[:, :, 3] = np.where(is_black, 0, int(_OVERLAY_OPACITY * 255))
        _overlay_im.set_data(rgba)
        _overlay_im.set_visible(True)
        fig.canvas.draw_idle()

    def _update_title():
        n = min(len(pts_left), len(pts_right))
        next_str = 'DIC (left)' if state['next'] == 0 else 'EBSD (right)'
        overlay_str = (f'Overlay: ON (40%)'
                       if n >= 3 and state['overlay_visible']
                       else ('Overlay: OFF' if n >= 3 else 'Overlay: need 3+ pairs'))
        fig.suptitle(
            f'Pairs: {n}  |  Next: click {next_str}  |  4+ pairs recommended  |  {overlay_str}',
            fontsize=12, color=_C['accent'], fontfamily='monospace', y=0.97)
        fig.canvas.draw_idle()

    def on_click(event):
        if event.inaxes is None or event.xdata is None:
            return
        x, y = event.xdata, event.ydata

        if event.button == 3:  # 右クリック: 最近傍ペアを削除
            n = min(len(pts_left), len(pts_right))
            if n == 0:
                return
            pts_search = pts_left[:n] if event.inaxes is ax_left else pts_right[:n]
            dists = [((p[0]-x)**2 + (p[1]-y)**2) for p in pts_search]
            idx = int(np.argmin(dists))
            pts_left.pop(idx)
            pts_right.pop(idx)
            if len(pts_left) > len(pts_right):
                for a in artists_left.pop():
                    a.remove()
                pts_left.pop()
                state['next'] = 0
            _redraw_labels()
            _update_title()
            _update_overlay()
            return

        if event.button != 1:
            return

        n_pair = min(len(pts_left), len(pts_right)) + 1

        if state['next'] == 0 and event.inaxes is ax_left:
            pts_left.append([x, y])
            mk, = ax_left.plot(x, y, 'r+', markersize=16, markeredgewidth=2.5, zorder=5)
            lb  = ax_left.annotate(str(n_pair), (x, y), color='red', fontsize=14,
                                   fontweight='bold', xytext=(6, 6),
                                   textcoords='offset points', zorder=5)
            artists_left.append([mk, lb])
            state['next'] = 1
        elif state['next'] == 1 and event.inaxes is ax_right:
            pts_right.append([x, y])
            mk, = ax_right.plot(x, y, 'r+', markersize=16, markeredgewidth=2.5, zorder=5)
            lb  = ax_right.annotate(str(n_pair), (x, y), color='red', fontsize=14,
                                    fontweight='bold', xytext=(6, 6),
                                    textcoords='offset points', zorder=5)
            artists_right.append([mk, lb])
            state['next'] = 0
        else:
            return

        _update_title()
        _update_overlay()

    def on_key(event):
        if event.key == 'q':
            plt.close(fig)
        elif event.key == 'o':
            state['overlay_visible'] = not state['overlay_visible']
            _btn_overlay.label.set_text(
                'Overlay: ON' if state['overlay_visible'] else 'Overlay: OFF')
            _update_title()
            _update_overlay()

    def _on_btn_overlay(_event):
        state['overlay_visible'] = not state['overlay_visible']
        _btn_overlay.label.set_text(
            'Overlay: ON' if state['overlay_visible'] else 'Overlay: OFF')
        _update_title()
        _update_overlay()

    def _on_btn_done(_event):
        plt.close(fig)

    fig.canvas.mpl_connect('button_press_event', on_click)
    fig.canvas.mpl_connect('key_press_event',    on_key)
    _btn_overlay.on_clicked(_on_btn_overlay)
    _btn_done.on_clicked(_on_btn_done)

    _update_title()
    plt.show()

    n = min(len(pts_left), len(pts_right))
    return (np.array(pts_left[:n],  dtype=np.float32),
            np.array(pts_right[:n], dtype=np.float32))


# =============================================================================
# 変換行列推定（ebsd_georef_v68 から流用）
# =============================================================================
def estimate_transform(pts_left, pts_right):
    """
    左パネル座標 → 右パネル座標 のアフィン変換行列 H (2×3) を推定する。
    """
    n = len(pts_left)
    if n < 3:
        raise ValueError(f"Not enough control points ({n}). At least 3 required.")

    if n == 3:
        H = cv2.getAffineTransform(pts_left[:3], pts_right[:3])
        print(f"  Affine transform estimated (exact, 3 points)")
    else:
        H, inliers = cv2.estimateAffine2D(pts_left, pts_right,
                                          method=cv2.RANSAC,
                                          ransacReprojThreshold=10.0)
        if H is None:
            raise RuntimeError("Affine estimation failed.")
        n_inliers = int(inliers.sum()) if inliers is not None else n
        pts_h  = np.hstack([pts_left, np.ones((n, 1), dtype=np.float32)])
        pts_pred = (H @ pts_h.T).T
        residuals = np.linalg.norm(pts_pred - pts_right, axis=1)
        print(f"  Affine transform estimated (least squares, {n} pts, {n_inliers} inliers)")
        for i, (res, inl) in enumerate(zip(residuals, inliers.flatten())):
            flag = '' if inl else '  [outlier]'
            print(f"    pt{i+1}: residual = {res:.2f} px{flag}")

    return H


# =============================================================================
# EBSDグリッド点への subset_id 割り当て
# =============================================================================
def assign_subset_ids(data, meta, H, cx_def, cy_def, subset_ids,
                      x_offset, y_offset, dic_step):
    """
    各EBSDグリッド点（data の行）に最近傍DICサブセットの subset_id を割り当てる。

    座標変換の流れ:
      EBSD物理座標 (μm)
        → EBSD Grain IDキャンバスピクセル (ix, iy)
        → H_inv → 変形後DICキャンバスピクセル
        → + offset → DIC実座標 (cx_def, cy_def) と比較

    Returns
    -------
    subset_id_map : ndarray (ny × nx, int32)
        各EBSDグリッド点の subset_id。DIC視野外は -1。
    """
    nx, ny   = meta['nx'], meta['ny']
    x_min    = meta['x_min']
    y_min    = meta['y_min']
    x_step   = meta['x_step']
    y_step   = meta['y_step']
    x_max    = meta['x_max']
    y_max    = meta['y_max']

    # EBSD物理座標 → EBSDキャンバスピクセル座標
    xpos = data[:, 3]
    ypos = data[:, 4]
    ebsd_ix = (xpos - x_min) / x_step if x_step > 0 else np.zeros(len(xpos))
    ebsd_iy = (ypos - y_min) / y_step if y_step > 0 else np.zeros(len(ypos))

    # H_inv: EBSDキャンバスピクセル → 変形後DICキャンバスピクセル
    H_full   = np.vstack([H, [0.0, 0.0, 1.0]])
    H_inv    = np.linalg.inv(H_full)[:2]           # (2, 3)

    N = len(ebsd_ix)
    ebsd_pts_h = np.column_stack([ebsd_ix, ebsd_iy, np.ones(N)])  # (N, 3)
    dic_canvas = (H_inv @ ebsd_pts_h.T).T                         # (N, 2)

    # キャンバスピクセル → DIC実座標
    dic_cx_actual = dic_canvas[:, 0] + x_offset
    dic_cy_actual = dic_canvas[:, 1] + y_offset

    # KDTree で最近傍DICサブセットを検索
    max_dist = dic_step * 1.5
    tree = cKDTree(np.column_stack([cx_def, cy_def]))
    dists, idxs = tree.query(np.column_stack([dic_cx_actual, dic_cy_actual]), k=1)

    sid_flat = np.where(dists <= max_dist,
                        np.array(subset_ids)[idxs],
                        -1).astype(np.int32)

    n_assigned = int((sid_flat >= 0).sum())
    n_outside  = N - n_assigned
    print(f"  subset_id assigned: {n_assigned} / {N}")
    if n_outside > 0:
        print(f"  Outside DIC area  : {n_outside}")

    return sid_flat.reshape(ny, nx)


# =============================================================================
# _georef.mat 保存（2D フィールド + subset_id）
# =============================================================================
def save_georef_mat(mat_path, subset_id_map, out_path):
    """
    元の .mat から 2D フィールド（3次元以上は除外）＋ subset_id を
    _georef.mat として保存する。
    """
    import scipy.io

    mat_info = scipy.io.whosmat(mat_path)

    # 3次元以上の大型配列を除外
    include_names = [name for name, shape, _ in mat_info if len(shape) < 3]
    print(f"  Loading {len(include_names)} 2D fields from {Path(mat_path).name} ...")

    mat = scipy.io.loadmat(mat_path, variable_names=include_names)

    # subset_id を追加（float64 で保存）
    mat['subset_id'] = subset_id_map.astype(np.float64)

    # MATLAB 内部変数を除去
    mat_clean = {k: v for k, v in mat.items() if not k.startswith('_')}

    print(f"  Saving to {Path(out_path).name} ...")
    scipy.io.savemat(str(out_path), mat_clean)

    size_mb = Path(out_path).stat().st_size / 1024 / 1024
    print(f"  Saved: {out_path}  ({size_mb:.1f} MB)")


# =============================================================================
# dic_results_georef.xlsx に新シート追加（Grain File モード用）
# =============================================================================
def save_georef_xlsx_sheet(georef_xlsx, sheet_name, data, phase_names, meta,
                           subset_id_map):
    """
    Grain File (.txt) モードの場合、dic_results_georef.xlsx に
    'ebsd_georef_{label}' シートを追加する。
    """
    import zipfile, re

    nx, ny  = meta['nx'], meta['ny']
    x_step  = meta['x_step']
    y_step  = meta['y_step']
    x_min   = meta['x_min']
    y_min   = meta['y_min']

    # シートデータ組み立て
    header = ['subset_id', 'ebsd_ix', 'ebsd_iy', 'xpos_um', 'ypos_um',
              'grain_id', 'phase',
              'phi1_deg', 'PHI_deg', 'phi2_deg', 'IQ', 'CI']
    rows_data = [header]

    for flat_idx, row in enumerate(data):
        iy = flat_idx // nx
        ix = flat_idx %  nx
        sid = int(subset_id_map[iy, ix])
        rows_data.append([
            sid if sid >= 0 else '',
            ix, iy,
            round(float(row[3]), 6), round(float(row[4]), 6),
            int(row[8]),
            phase_names[flat_idx],
            round(float(np.degrees(row[0])), 6),
            round(float(np.degrees(row[1])), 6),
            round(float(np.degrees(row[2])), 6),
            round(float(row[5]), 6) if not np.isnan(row[5]) else '',
            round(float(row[6]), 6) if not np.isnan(row[6]) else '',
        ])

    def _escape(v):
        if v is None or v == '' or (isinstance(v, float) and np.isnan(v)):
            return ''
        s = str(v)
        return (s.replace('&','&amp;').replace('<','&lt;')
                  .replace('>','&gt;').replace('"','&quot;'))

    sheet_xml_lines = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
        '<sheetData>',
    ]
    for ri, row in enumerate(rows_data, start=1):
        sheet_xml_lines.append(f'<row r="{ri}">')
        for ci, val in enumerate(row):
            col_letter = chr(ord('A') + ci)
            ref = f'{col_letter}{ri}'
            if isinstance(val, str) or val == '':
                sheet_xml_lines.append(
                    f'<c r="{ref}" t="inlineStr"><is><t>{_escape(val)}</t></is></c>')
            elif val is None or (isinstance(val, float) and np.isnan(val)):
                sheet_xml_lines.append(f'<c r="{ref}"/>')
            else:
                sheet_xml_lines.append(f'<c r="{ref}"><v>{val}</v></c>')
        sheet_xml_lines.append('</row>')
    sheet_xml_lines += ['</sheetData>', '</worksheet>']
    sheet_xml = '\n'.join(sheet_xml_lines).encode('utf-8')

    SHEET_NAME = sheet_name
    SHEET_PATH = f'xl/worksheets/sheet_{sheet_name}.xml'
    REL_ID     = f'rId_{sheet_name}'

    import tempfile, shutil
    tmp_path = georef_xlsx + '.tmp'

    with zipfile.ZipFile(georef_xlsx, 'r') as zin, \
         zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:

        skip = {SHEET_PATH, 'xl/workbook.xml',
                'xl/_rels/workbook.xml.rels', '[Content_Types].xml'}
        for item in zin.infolist():
            if item.filename in skip:
                continue
            zout.writestr(item, zin.read(item.filename))

        wb_xml = zin.read('xl/workbook.xml').decode('utf-8')
        wb_xml = re.sub(
            rf'<sheet [^/]*name="{re.escape(SHEET_NAME)}"[^/]*/>', '', wb_xml)
        ids = list(map(int, re.findall(r'sheetId="(\d+)"', wb_xml)))
        new_id = max(ids) + 1 if ids else 1
        new_sheet_tag = (f'<sheet name="{SHEET_NAME}" sheetId="{new_id}" '
                         f'r:id="{REL_ID}"/>')
        wb_xml = wb_xml.replace('</sheets>', new_sheet_tag + '</sheets>')
        zout.writestr('xl/workbook.xml', wb_xml.encode('utf-8'))

        rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        rels_xml = re.sub(
            rf'<Relationship [^/]*Id="{re.escape(REL_ID)}"[^/]*/>', '', rels_xml)
        new_rel = (f'<Relationship Id="{REL_ID}" '
                   f'Type="http://schemas.openxmlformats.org/officeDocument/'
                   f'2006/relationships/worksheet" '
                   f'Target="worksheets/sheet_{sheet_name}.xml"/>')
        rels_xml = rels_xml.replace('</Relationships>', new_rel + '</Relationships>')
        zout.writestr('xl/_rels/workbook.xml.rels', rels_xml.encode('utf-8'))

        ct_xml = zin.read('[Content_Types].xml').decode('utf-8')
        ct_xml = re.sub(
            rf'<Override PartName="/xl/worksheets/sheet_{re.escape(sheet_name)}\.xml"[^/]*/>', '',
            ct_xml)
        new_ct = (f'<Override PartName="/xl/worksheets/sheet_{sheet_name}.xml" '
                  f'ContentType="application/vnd.openxmlformats-officedocument'
                  f'.spreadsheetml.worksheet+xml"/>')
        ct_xml = ct_xml.replace('</Types>', new_ct + '</Types>')
        zout.writestr('[Content_Types].xml', ct_xml.encode('utf-8'))

        zout.writestr(SHEET_PATH, sheet_xml)

    shutil.move(tmp_path, georef_xlsx)
    print(f"  Sheet '{SHEET_NAME}' added to {Path(georef_xlsx).name}")


# =============================================================================
# オーバーレイ PNG 保存
# =============================================================================
def save_overlay_png(left_img, right_img, pts_left, pts_right, out_path, opacity=0.45):
    """位置合わせ結果のオーバーレイ画像を PNG として保存する。"""
    H_L, W_L = left_img.shape[:2]
    n = len(pts_left)

    src = pts_left.astype(np.float32)
    dst = pts_right.astype(np.float32)
    if n == 3:
        H_mat = cv2.getAffineTransform(dst[:3], src[:3])
    else:
        H_mat, _ = cv2.estimateAffine2D(dst, src,
                                        method=cv2.RANSAC,
                                        ransacReprojThreshold=10.0)
    warped = cv2.warpAffine(right_img, H_mat, (W_L, H_L),
                            flags=cv2.INTER_NEAREST,
                            borderMode=cv2.BORDER_CONSTANT, borderValue=(0, 0, 0))

    left_f  = left_img.astype(np.float32)
    warp_f  = warped.astype(np.float32)
    is_black = ((warped[:,:,0]==0)&(warped[:,:,1]==0)&(warped[:,:,2]==0))
    alpha   = np.where(is_black, 0.0, opacity)[:,:,np.newaxis]
    blended = (left_f*(1-alpha) + warp_f*alpha).clip(0, 255).astype(np.uint8)

    fig, ax = plt.subplots(figsize=(12, 9))
    fig.patch.set_facecolor('#0a0c0f')
    ax.set_facecolor('#111318')
    ax.imshow(blended, origin='upper')
    for i, pl in enumerate(pts_left, start=1):
        ax.plot(pl[0], pl[1], 'r+', markersize=14, markeredgewidth=2, zorder=5)
        ax.annotate(str(i), (pl[0], pl[1]), color='red', fontsize=9,
                    fontweight='bold', xytext=(5, 5),
                    textcoords='offset points', zorder=5)
    ax.set_title(f'EBSD Grain ID overlay on Deformed DIC  (Affine, {n} pts, opacity={int(opacity*100)}%)',
                 fontsize=11)
    plt.tight_layout()
    fig.savefig(str(out_path), dpi=150)
    plt.close(fig)
    print(f"  Saved overlay PNG: {out_path}")


# =============================================================================
# メイン処理
# =============================================================================
if __name__ == '__main__':
    from PyQt6.QtWidgets import QApplication
    import sys as _sys

    _qt_app = QApplication.instance() or QApplication(_sys.argv)

    if len(_sys.argv) < 2:
        print("Usage: python defebsd_georef_v1.py <param_file.json>")
        sys.exit(1)

    with open(_sys.argv[1], encoding='utf-8') as _f:
        params = json.load(_f)

    georef_xlsx = params['georef_xlsx']
    stages      = params['stages']   # list of {label, ebsd_path, mode}
    out_dir     = Path(params.get('out_dir', Path(georef_xlsx).parent / 'defebsd_georef'))
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"  Def EBSD Georef  —  {len(stages)} stage(s)")
    print(f"  Georef xlsx : {georef_xlsx}")
    print(f"  Output dir  : {out_dir}")
    print(f"{'='*60}")

    for stage_idx, stage in enumerate(stages, 1):
        label     = stage['label']
        ebsd_path = stage['ebsd_path']
        mode      = stage['mode']   # 'mat' or 'grain'

        print(f"\n[Stage {stage_idx}/{len(stages)}]  {label}  ({Path(ebsd_path).name})")
        print('-' * 60)

        # [1] 変形後DIC Grain IDマップ読み込み
        print("[1/5] Loading deformed DIC Grain ID map...")
        cx_def, cy_def, grain_ids_dic, subset_ids = \
            load_deformed_dic_grain_map(georef_xlsx, label)

        # [2] 変形後DIC Grain IDマップを画像化
        print("[2/5] Rendering deformed DIC Grain ID image...")
        dic_grain_img, x_offset, y_offset, dic_step = \
            render_deformed_grain_image(cx_def, cy_def, grain_ids_dic)
        print(f"  Canvas size : {dic_grain_img.shape[1]} x {dic_grain_img.shape[0]} px")
        print(f"  DIC step est: {dic_step:.1f} px")

        # [3] EBSD データ読み込みと Grain IDマップ生成
        print("[3/5] Loading EBSD data...")
        if mode == 'mat':
            data, phase_names, meta = load_mat_minimal(ebsd_path)
        else:
            data, phase_names, meta = load_grain_file(ebsd_path)

        print("[3/5] Generating EBSD Grain ID image...")
        ebsd_grain_img, _, _ = make_grain_id_image(data, meta)

        # [4] コントロールポイント指定
        print("[4/5] Please specify control points in the GUI...")
        print("  Left(Deformed DIC) → Right(EBSD) click alternately")
        print("  Right-click: undo  /  q: confirm")
        pts_left, pts_right = pick_control_points(
            dic_grain_img, ebsd_grain_img,
            left_label=label,
            right_label=Path(ebsd_path).name,
        )

        n_pts = len(pts_left)
        if n_pts < 3:
            print(f"  Skipped: not enough control points ({n_pts} < 3)")
            continue
        print(f"  {n_pts} control point pairs confirmed.")

        H = estimate_transform(pts_left, pts_right)

        # [5] subset_id 割り当て
        print("[5/5] Assigning subset_id to each EBSD point...")
        subset_id_map = assign_subset_ids(
            data, meta, H,
            cx_def, cy_def, subset_ids,
            x_offset, y_offset, dic_step,
        )

        # 出力
        safe_label = label.replace('/', '-').replace(' ', '_')
        overlay_png = out_dir / f'overlay_{safe_label}.png'
        save_overlay_png(dic_grain_img, ebsd_grain_img,
                         pts_left, pts_right, overlay_png)

        if mode == 'mat':
            stem     = Path(ebsd_path).stem
            out_mat  = out_dir / f'{stem}_georef.mat'
            save_georef_mat(ebsd_path, subset_id_map, str(out_mat))
        else:
            sheet_name = f'ebsd_georef_{safe_label}'
            save_georef_xlsx_sheet(georef_xlsx, sheet_name,
                                   data, phase_names, meta, subset_id_map)

        print(f"  Stage '{label}' complete.")

    print(f"\n{'='*60}")
    print(f"  All {len(stages)} stage(s) complete.")
    print(f"  Output: {out_dir}")
    print(f"{'='*60}")
