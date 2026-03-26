"""
ebsd_georef.py  v68
==================
v68での変更点:
  tkinter を PyQt6 に全面置き換え。
  - matplotlib backend: TkAgg → QtAgg
  - pick_control_points(): matplotlib widgets はそのまま、Tk依存を排除
  - show_map_viewer(): tkinter UI を PyQt6 QWidget に全面書き換え（ダークテーマ統一）
  - __main__ のファイル選択ダイアログ: tkinter → PyQt6 QFileDialog
  - 必要パッケージ: pip install PyQt6

v67での変更点:
==================
v3での変更点:
  pick_control_points() にオーバーレイプレビュー機能を追加。
  - ペアが3組以上そろった時点で、Grain IDマップをSEM座標系にワープして
    SEM画像左パネルに半透明で重ねて表示する。
  - 点を追加・削除するたびにオーバーレイが自動更新される。
  - [o] キーで透明度をサイクル切替（20% → 40% → 60%）。
  - 粒界（黒ピクセル）はアルファ=0にして透かし、粒内のみ着色表示。
  その他の処理・出力形式はv1（= v2）と同一。


EBSD Grain File（TSL/EDAX OIM Analysis エクスポート）をSEM画像座標系に
ジオリファレンスし、DICサブセットグリッドに結晶粒情報を割り当てて
ebsd_georef.xlsx に保存する。

入力ファイル:
  Grain_File_*.txt  OIM Analysis の "Grain File" エクスポート
  列構成（固定11列）:
    0:phi1[rad]  1:PHI[rad]  2:phi2[rad]  3:x[μm]  4:y[μm]
    5:IQ  6:CI  7:Fit[deg]  8:GrainID(int)  9:Edge(int)  10:PhaseName(str)
  グリッド数・ステップ・視野サイズ・相数はデータから自動検出（ハードコードなし）。
  GrainID=0: 無効粒（粒界・未解析）、負値: anti-grain

処理フロー:
  ① Grain File を読み込む（グリッド情報自動検出）
  ② IQ マップ画像を生成
  ③ SEM 参照画像と EBSD IQ マップを GUI で左右に並べて表示
     左（SEM）→ 右（EBSD IQ）の順に交互クリックでコントロールポイントを対応付け
     4点以上: 射影変換（Homography）
     3点:     アフィン変換（フォールバック）
  ④ dic_results.xlsx の u シートからサブセット座標を読み込み、
     SEM→EBSD 座標変換 → 最近傍 EBSD データを割り当て
  ⑤ ebsd_georef.xlsx に保存
     列: cx[px], cy[px], grain_id, phase, phi1[deg], PHI[deg], phi2[deg], IQ, CI

座標系:
  SEM画像: x右向き・y下向き [px]
  EBSD:    x右向き・y下向き [μm]（Grain File の x, y と同方向）
"""

import sys
import os
import numpy as np
from pathlib import Path
from collections import deque

# matplotlib backend を最初に設定（import前に必須）
os.environ['MPLBACKEND'] = 'QtAgg'
import matplotlib
matplotlib.use('QtAgg')

# ── 必須ライブラリチェック ────────────────────────────────────────────────
_REQUIRED = {
    'numpy':      'numpy',
    'scipy':      'scipy',
    'matplotlib': 'matplotlib',
    'openpyxl':   'openpyxl',
    'cv2':        'opencv-python',
}
_missing = []
for _mod, _pkg in _REQUIRED.items():
    try:
        __import__(_mod)
    except ImportError:
        _missing.append(_pkg)
if _missing:
    print("=" * 60)
    print("  [ERROR] Required libraries not found.")
    print(f"  pip install {' '.join(_missing)}")
    print("=" * 60)
    sys.exit(1)

import cv2
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.font_manager as _fm
import platform as _platform
from scipy.spatial import cKDTree

# ── 日本語フォント設定 ────────────────────────────────────────────────────
_JP_FONTS = {
    'Windows': ['Meiryo', 'Yu Gothic', 'MS Gothic'],
    'Darwin':  ['Hiragino Sans', 'Hiragino Maru Gothic Pro', 'AppleGothic'],
    'Linux':   ['IPAexGothic', 'IPAPGothic', 'Noto Sans CJK JP'],
}
_available_fonts = {f.name for f in _fm.fontManager.ttflist}
_jp_candidates   = _JP_FONTS.get(_platform.system(), [])
_jp_font = next((f for f in _jp_candidates if f in _available_fonts), None)
if _jp_font:
    matplotlib.rcParams['font.family'] = _jp_font
matplotlib.rcParams['axes.unicode_minus'] = False


# =============================================================================
# Grain File 読み込み
# =============================================================================
def load_grain_file(path):
    """
    OIM Analysis Grain File（固定11列）を読み込む。
    グリッド数・ステップ・視野・相はデータから自動検出する。

    Returns
    -------
    data : ndarray, shape (N, 10)  数値列（phi1〜Edge）
    phase_names : ndarray, shape (N,)  Phase名文字列
    meta : dict
        nx, ny, x_step, y_step, x_min, y_min, x_max, y_max
        grain_tol_angle : OIMヘッダから読んだ粒界閾値（なければNone）
        phases : list of str  ユニークなPhase名
    """
    num_rows = []
    phase_col = []
    grain_tol_angle = None

    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            line_s = line.strip()
            if not line_s:
                continue
            if line_s.startswith('#'):
                # OIM ヘッダから粒界閾値を読む
                if 'Grain Tolerance Angle' in line_s:
                    try:
                        grain_tol_angle = float(line_s.split(':')[-1].strip())
                    except ValueError:
                        pass
                continue
            parts = line_s.split()
            if len(parts) < 11:
                continue
            try:
                num_rows.append([float(p) for p in parts[:10]])
                phase_col.append(parts[10])
            except ValueError:
                continue

    if not num_rows:
        raise ValueError(f"No data rows found: {path}")

    data = np.array(num_rows, dtype=np.float64)
    phase_names = np.array(phase_col)

    # グリッド構造を自動検出
    x_vals = np.unique(np.round(data[:, 3], 6))
    y_vals = np.unique(np.round(data[:, 4], 6))
    nx = len(x_vals)
    ny = len(y_vals)
    x_step = float(np.median(np.diff(x_vals))) if nx > 1 else 1.0
    y_step = float(np.median(np.diff(y_vals))) if ny > 1 else 1.0
    x_min, x_max = float(x_vals.min()), float(x_vals.max())
    y_min, y_max = float(y_vals.min()), float(y_vals.max())

    unique_phases = sorted(set(phase_col))
    grain_ids = data[:, 8].astype(int)
    n_valid_grains = len(np.unique(grain_ids[grain_ids > 0]))

    print(f"  Grain File loaded: {len(data)} pixels  ({nx} x {ny})")
    print(f"  Step: x={x_step:.4f} um, y={y_step:.4f} um")
    print(f"  FOV: x={x_min:.2f}~{x_max:.2f} um, y={y_min:.2f}~{y_max:.2f} um")
    print(f"  Valid grains (GrainID>0): {n_valid_grains}")
    print(f"  Phases: {unique_phases}")
    if grain_tol_angle is not None:
        print(f"  OIM grain tolerance angle: {grain_tol_angle} deg")

    meta = dict(
        nx=nx, ny=ny,
        x_step=x_step, y_step=y_step,
        x_min=x_min, y_min=y_min,
        x_max=x_max, y_max=y_max,
        grain_tol_angle=grain_tol_angle,
        phases=unique_phases,
    )
    return data, phase_names, meta


# =============================================================================
# Grain ID マップ画像生成（RGB）
# =============================================================================
def make_grain_id_image(data, meta):
    """
    Grain ID（列8）からカラーRGB画像（ny × nx × 3, uint8）を生成する。
    - GrainID > 0: 粒ごとに固定色（tab20 カラーマップをサイクル）
    - GrainID = 0: 黒（粒界・未解析）
    - GrainID < 0: 濃灰（anti-grain）
    """
    nx, ny = meta['nx'], meta['ny']
    x_min, y_min = meta['x_min'], meta['y_min']
    x_step, y_step = meta['x_step'], meta['y_step']

    grain_id_map = np.zeros((ny, nx), dtype=np.int32)
    for row in data:
        ix = int(round((row[3] - x_min) / x_step))
        iy = int(round((row[4] - y_min) / y_step))
        if 0 <= ix < nx and 0 <= iy < ny:
            grain_id_map[iy, ix] = int(row[8])

    # 有効 Grain ID（>0）に対してカラーを割り当て
    # tab20/tab20b/tab20c を合成した60色パレットから均等に選ぶ
    valid_ids = sorted(set(int(v) for v in grain_id_map.flat if v > 0))
    import matplotlib.cm as _cm
    _pool = []
    for _name in ('tab20', 'tab20b', 'tab20c'):
        _c = matplotlib.colormaps[_name]
        _pool.extend([_c(i) for i in range(20)])
    _idxs = np.linspace(0, len(_pool) - 1, max(len(valid_ids), 1), dtype=int)
    id_to_color = {gid: (np.array(_pool[_idxs[i]][:3]) * 255).astype(np.uint8)
                   for i, gid in enumerate(valid_ids)}

    rgb = np.zeros((ny, nx, 3), dtype=np.uint8)
    for iy in range(ny):
        for ix in range(nx):
            gid = grain_id_map[iy, ix]
            if gid > 0:
                rgb[iy, ix] = id_to_color[gid]
            elif gid < 0:
                rgb[iy, ix] = [80, 80, 80]   # anti-grain: 濃灰
            # gid == 0: 黒のまま（粒界・未解析）

    return rgb, grain_id_map, id_to_color


# =============================================================================
# コントロールポイント指定 GUI（v3: オーバーレイプレビュー付き）
# =============================================================================
def pick_control_points(sem_img, grain_id_img, cx_arr=None, cy_arr=None):
    """
    SEM画像（左）と EBSD Grain ID マップ（右）を並べて表示し、
    左→右の順に交互クリックでコントロールポイントを対応付ける。

    v3での追加機能:
      ペアが3組以上そろった時点で、現在のコントロールポイントから
      変換行列を推定し、SEM画像上に半透明のGrain IDマップを重ねて表示する。
      点を追加・削除するたびにオーバーレイが自動更新される。

    操作:
      左クリック : 点を追加（左→右→左→右…の順）
      右クリック : 直前の点を取消
      q キー     : 確定してウィンドウを閉じる
      o キー     : オーバーレイの透明度を変更（0.2 → 0.4 → 0.6 → 0.2 …）

    Returns
    -------
    pts_sem  : ndarray (N, 2)  SEM画像座標 [px]
    pts_ebsd : ndarray (N, 2)  EBSD Grain ID画像座標 [px]
    """
    from matplotlib.widgets import RangeSlider, Button
    import matplotlib
    matplotlib.use('QtAgg')   # 念押し（既にロード済みでも上書き試行）

    # ── デザイントークン（sem_align_tool_v3.html 準拠）──────────────────
    _C = {
        'bg':       '#0a0c0f',
        'surface':  '#111318',
        'surface2': '#1a1d24',
        'border':   '#2a2d35',
        'accent':   '#00d4ff',
        'green':    '#00ff88',
        'orange':   '#ff6b35',
        'text':     '#e0e4ec',
        'dim':      '#6b7280',
    }

    H_sem, W_sem = sem_img.shape[:2]
    H_ebsd, W_ebsd = grain_id_img.shape[:2]

    plt.style.use('dark_background')
    fig, axes = plt.subplots(1, 2, figsize=(22, 10))
    fig.patch.set_facecolor(_C['bg'])
    for ax in axes if hasattr(axes, '__iter__') else [axes]:
        ax.set_facecolor(_C['surface'])
        for sp in ax.spines.values():
            sp.set_edgecolor(_C['border'])
        ax.tick_params(colors=_C['dim'], labelsize=11)
    fig.subplots_adjust(bottom=0.22, top=0.88, left=0.05, right=0.97,
                        wspace=0.08)
    fig.canvas.manager.set_window_title(
        'EBSD Georef  |  Control Points  —  SEM(L) → EBSD(R)  |  Right-click: undo  |  q: confirm')

    ax_sem, ax_ebsd = axes

    # --- ベース画像を描画 ---
    _sem_vmin0 = int(sem_img.min())
    _sem_vmax0 = int(sem_img.max())
    _sem_im = ax_sem.imshow(sem_img, cmap='gray', origin='upper',
                            vmin=_sem_vmin0, vmax=_sem_vmax0)
    ax_ebsd.imshow(grain_id_img, origin='upper')
    ax_sem.set_title('SEM ref image  [ click here first ]',
                     fontsize=13, color=_C['accent'], pad=8)
    ax_ebsd.set_title('EBSD Grain ID map  [ click here second ]',
                      fontsize=13, color=_C['dim'], pad=8)
    ax_sem.tick_params(labelsize=11, colors=_C['dim'])
    ax_ebsd.tick_params(labelsize=11, colors=_C['dim'])

    # DICサブセット点をSEM像上に描画（デフォルトOFF）
    _subset_sc = None
    if cx_arr is not None and cy_arr is not None:
        _subset_sc = ax_sem.scatter(cx_arr, cy_arr, s=2, c='cyan', alpha=0.5,
                                    linewidths=0, zorder=3)
        _subset_sc.set_visible(False)   # 最初は非表示

    # オーバーレイ用 imshow（初期は非表示）
    # SEM側: grain_id_img をワープしてSEM座標系に重ねる
    _overlay_im = ax_sem.imshow(
        np.zeros((H_sem, W_sem, 4), dtype=np.uint8),
        origin='upper', alpha=1.0, zorder=2)
    _overlay_im.set_visible(False)

    # --- コントラストスライダー ---
    ax_slider = fig.add_axes([0.05, 0.10, 0.44, 0.030])
    ax_slider.set_facecolor(_C['surface2'])
    for sp in ax_slider.spines.values():
        sp.set_edgecolor(_C['border'])
    _slider = RangeSlider(ax_slider, '', _sem_vmin0, _sem_vmax0,
                          valinit=(_sem_vmin0, _sem_vmax0))
    _slider.label.set_color(_C['dim'])
    _slider.valtext.set_color(_C['accent'])
    _slider.poly.set_facecolor(_C['accent'])
    _slider.poly.set_alpha(0.4)
    fig.text(0.05, 0.142, 'SEM CONTRAST', fontsize=9,
             va='bottom', ha='left', color=_C['accent'],
             fontfamily='monospace', fontweight='bold', linespacing=1.2)

    def _on_slider(val):
        lo, hi = _slider.val
        _sem_im.set_clim(lo, hi)
        fig.canvas.draw_idle()

    _slider.on_changed(_on_slider)

    # --- ボタン共通スタイル ---
    def _make_btn(rect, label, color):
        ax_b = fig.add_axes(rect)
        ax_b.set_facecolor(_C['surface2'])
        for sp in ax_b.spines.values():
            sp.set_edgecolor(color)
            sp.set_linewidth(1.2)
        btn = Button(ax_b, label, color=_C['surface2'], hovercolor=_C['surface'])
        btn.label.set_fontsize(11)
        btn.label.set_color(color)
        btn.label.set_fontfamily('monospace')
        return btn

    # --- Overlay ON/OFFボタン ---
    _btn_overlay = _make_btn([0.12, 0.025, 0.14, 0.055], 'Overlay: ON', _C['accent'])

    # --- Subsets ON/OFFボタン ---
    _btn_subsets = _make_btn([0.28, 0.025, 0.14, 0.055], 'Subsets: OFF', _C['dim'])

    # --- 完了ボタン ---
    _btn_done = _make_btn([0.68, 0.025, 0.14, 0.055], 'Confirm', _C['green'])

    # 状態
    pts_sem  = []
    pts_ebsd = []
    # artists_sem / artists_ebsd: ペアごとの [marker, label] リスト
    # インデックスがペア番号と対応する（ペア単位で削除するため分離管理）
    artists_sem  = []
    artists_ebsd = []
    state    = {'next': 0,              # 0=SEM待ち, 1=EBSD待ち
                'overlay_visible': True, # オーバーレイ表示ON/OFF
                'subset_visible': False} # サブセット点表示ON/OFF（デフォルトOFF）
    _OVERLAY_OPACITY = 0.4               # 透過度は40%固定

    # -----------------------------------------------------------------
    # 番号ラベルの全再描画（ペア削除後に番号を振り直す）
    # -----------------------------------------------------------------
    def _redraw_labels():
        """全ペアのラベルを削除して番号を振り直す。"""
        for sem_arts, ebsd_arts in zip(artists_sem, artists_ebsd):
            for a in sem_arts + ebsd_arts:
                a.remove()
        artists_sem.clear()
        artists_ebsd.clear()
        for i, (ps, pe) in enumerate(zip(pts_sem, pts_ebsd), start=1):
            mk_s, = ax_sem.plot(ps[0], ps[1], 'r+', markersize=16,
                                markeredgewidth=2.5, zorder=5)
            lb_s  = ax_sem.annotate(str(i), (ps[0], ps[1]), color='red',
                                    fontsize=14, fontweight='bold',
                                    xytext=(6, 6), textcoords='offset points',
                                    zorder=5)
            mk_e, = ax_ebsd.plot(pe[0], pe[1], 'r+', markersize=16,
                                 markeredgewidth=2.5, zorder=5)
            lb_e  = ax_ebsd.annotate(str(i), (pe[0], pe[1]), color='red',
                                     fontsize=14, fontweight='bold',
                                     xytext=(6, 6), textcoords='offset points',
                                     zorder=5)
            artists_sem.append([mk_s, lb_s])
            artists_ebsd.append([mk_e, lb_e])
        fig.canvas.draw_idle()

    # -----------------------------------------------------------------
    # オーバーレイ更新
    # -----------------------------------------------------------------
    def _update_overlay():
        """ペアが3組以上あればワープしてオーバーレイを更新する。"""
        n = min(len(pts_sem), len(pts_ebsd))
        if n < 3 or not state['overlay_visible']:
            _overlay_im.set_visible(False)
            fig.canvas.draw_idle()
            return

        src = np.array(pts_sem[:n],  dtype=np.float32)
        dst = np.array(pts_ebsd[:n], dtype=np.float32)

        # 変換行列推定（EBSD→SEM方向、warpAffine用に逆方向）
        try:
            if n == 3:
                H_mat = cv2.getAffineTransform(dst[:3], src[:3])
            else:
                H_mat, _ = cv2.estimateAffine2D(dst, src,
                                                method=cv2.RANSAC,
                                                ransacReprojThreshold=10.0)
        except Exception:
            _overlay_im.set_visible(False)
            fig.canvas.draw_idle()
            return

        if H_mat is None:
            _overlay_im.set_visible(False)
            fig.canvas.draw_idle()
            return

        # grain_id_img (RGB) を SEM座標系にワープ（常にアフィン）
        ebsd_rgb = grain_id_img  # (H_ebsd, W_ebsd, 3) uint8
        warped = cv2.warpAffine(
            ebsd_rgb, H_mat, (W_sem, H_sem),
            flags=cv2.INTER_NEAREST,
            borderMode=cv2.BORDER_CONSTANT, borderValue=(0, 0, 0))

        # アルファチャンネル付きRGBAに変換
        # 黒ピクセル（粒界・ワープ外）はほぼ透明にする
        rgba = np.zeros((H_sem, W_sem, 4), dtype=np.uint8)
        rgba[:, :, :3] = warped
        is_black = (warped[:, :, 0] == 0) & \
                   (warped[:, :, 1] == 0) & \
                   (warped[:, :, 2] == 0)
        alpha_val = int(_OVERLAY_OPACITY * 255)
        rgba[:, :, 3] = np.where(is_black, 0, alpha_val)

        _overlay_im.set_data(rgba)
        _overlay_im.set_visible(True)
        fig.canvas.draw_idle()

    # -----------------------------------------------------------------
    # タイトル更新
    # -----------------------------------------------------------------
    def _update_title():
        n = min(len(pts_sem), len(pts_ebsd))
        next_str = 'SEM (left)' if state['next'] == 0 else 'EBSD (right)'
        subset_str  = 'Subsets: ON' if state['subset_visible'] else 'Subsets: OFF'
        overlay_str = (f'Overlay: ON (40%)'
                       if n >= 3 and state['overlay_visible']
                       else ('Overlay: OFF' if n >= 3 else 'Overlay: need 3+ pairs'))
        fig.suptitle(
            f'Pairs: {n}  |  Next: click {next_str}  |  4+ pairs recommended  '
            f'|  {overlay_str}  |  {subset_str}',
            fontsize=12, color=_C['accent'], fontfamily='monospace',
            y=0.97)
        fig.canvas.draw_idle()

    # -----------------------------------------------------------------
    # クリックイベント
    # -----------------------------------------------------------------
    def on_click(event):
        if event.inaxes is None or event.xdata is None:
            return
        x, y = event.xdata, event.ydata

        # 右クリック: クリック位置に最も近いペアを両側同時に削除
        if event.button == 3:
            n = min(len(pts_sem), len(pts_ebsd))
            if n == 0:
                return
            # クリックした側の点群から最近傍ペアを探す
            if event.inaxes is ax_sem:
                pts_search = pts_sem[:n]
            else:
                pts_search = pts_ebsd[:n]
            dists = [((p[0]-x)**2 + (p[1]-y)**2) for p in pts_search]
            idx = int(np.argmin(dists))
            # ペアを削除
            pts_sem.pop(idx)
            pts_ebsd.pop(idx)
            # 半完成ペア（SEM側だけ追加済みの場合）の処理
            if len(pts_sem) > len(pts_ebsd):
                # SEM側の未完成点も削除
                for a in artists_sem.pop():
                    a.remove()
                pts_sem.pop()
                state['next'] = 0
            # 番号を振り直して再描画
            _redraw_labels()
            _update_title()
            _update_overlay()
            return

        if event.button != 1:
            return

        n_pair = min(len(pts_sem), len(pts_ebsd)) + 1

        if state['next'] == 0 and event.inaxes is ax_sem:
            pts_sem.append([x, y])
            mk, = ax_sem.plot(x, y, 'r+', markersize=16, markeredgewidth=2.5, zorder=5)
            lb  = ax_sem.annotate(str(n_pair), (x, y), color='red', fontsize=14,
                                  fontweight='bold',
                                  xytext=(6, 6), textcoords='offset points', zorder=5)
            artists_sem.append([mk, lb])
            state['next'] = 1

        elif state['next'] == 1 and event.inaxes is ax_ebsd:
            pts_ebsd.append([x, y])
            mk, = ax_ebsd.plot(x, y, 'r+', markersize=16, markeredgewidth=2.5, zorder=5)
            lb  = ax_ebsd.annotate(str(n_pair), (x, y), color='red', fontsize=14,
                                   fontweight='bold',
                                   xytext=(6, 6), textcoords='offset points', zorder=5)
            artists_ebsd.append([mk, lb])
            state['next'] = 0

        else:
            return  # 順番違いのクリックは無視

        _update_title()
        _update_overlay()

    # -----------------------------------------------------------------
    # キーイベント
    # -----------------------------------------------------------------
    def on_key(event):
        if event.key == 'q':
            plt.close(fig)
        elif event.key == 's':
            # サブセット点表示ON/OFF
            state['subset_visible'] = not state['subset_visible']
            if _subset_sc is not None:
                _subset_sc.set_visible(state['subset_visible'])
            _btn_subsets.label.set_text(
                'Subsets: ON' if state['subset_visible'] else 'Subsets: OFF')
            _update_title()
            fig.canvas.draw_idle()
        elif event.key == 'o':
            # オーバーレイON/OFF
            state['overlay_visible'] = not state['overlay_visible']
            _btn_overlay.label.set_text(
                'Overlay: ON' if state['overlay_visible'] else 'Overlay: OFF')
            _update_title()
            _update_overlay()

    fig.canvas.mpl_connect('button_press_event', on_click)
    fig.canvas.mpl_connect('key_press_event',    on_key)

    # --- ボタンのコールバックを接続 ---
    def _on_btn_overlay(_event):
        state['overlay_visible'] = not state['overlay_visible']
        _btn_overlay.label.set_text(
            'Overlay: ON' if state['overlay_visible'] else 'Overlay: OFF')
        _update_title()
        _update_overlay()

    def _on_btn_subsets(_event):
        state['subset_visible'] = not state['subset_visible']
        if _subset_sc is not None:
            _subset_sc.set_visible(state['subset_visible'])
        _btn_subsets.label.set_text(
            'Subsets: ON' if state['subset_visible'] else 'Subsets: OFF')
        _btn_subsets.label.set_color(
            _C['accent'] if state['subset_visible'] else _C['dim'])
        for sp in _btn_subsets.ax.spines.values():
            sp.set_edgecolor(_C['accent'] if state['subset_visible'] else _C['dim'])
        _update_title()
        fig.canvas.draw_idle()

    def _on_btn_done(_event):
        plt.close(fig)

    _btn_overlay.on_clicked(_on_btn_overlay)
    _btn_subsets.on_clicked(_on_btn_subsets)
    _btn_done.on_clicked(_on_btn_done)

    _update_title()
    plt.show()

    n = min(len(pts_sem), len(pts_ebsd))
    return (np.array(pts_sem[:n],  dtype=np.float32),
            np.array(pts_ebsd[:n], dtype=np.float32),
            state['subset_visible'])


# =============================================================================
# 変換行列推定
# =============================================================================
def estimate_transform(pts_sem, pts_ebsd):
    """
    SEM画像座標 → EBSD IQ画像座標 のアフィン変換行列を推定する。

    点数にかかわらず常にアフィン変換（2×3）を使用する。
    3点: cv2.getAffineTransform（厳密解）
    4点以上: cv2.estimateAffine2D（全点を使った最小二乗、RANSAC付き）

    射影変換（Homography）は自由度が高すぎてSEM-EBSD間のような
    スケール・回転・シフトが主体のズレには過剰適合しやすいため使用しない。

    Returns
    -------
    H    : ndarray (2×3)
    mode : str  'affine'
    """
    n = len(pts_sem)
    if n < 3:
        raise ValueError(f"Not enough control points ({n}). At least 3 required.")

    if n == 3:
        H = cv2.getAffineTransform(pts_sem[:3], pts_ebsd[:3])
        print(f"  Affine transform estimated (exact, 3 points)")
    else:
        H, inliers = cv2.estimateAffine2D(pts_sem, pts_ebsd,
                                          method=cv2.RANSAC,
                                          ransacReprojThreshold=10.0)
        if H is None:
            raise RuntimeError("Affine estimation failed. Check control point placement.")
        n_inliers = int(inliers.sum()) if inliers is not None else n
        # 残差を計算して表示
        pts_h = np.hstack([pts_sem, np.ones((n,1), dtype=np.float32)])
        pts_pred = (H @ pts_h.T).T
        residuals = np.linalg.norm(pts_pred - pts_ebsd, axis=1)
        print(f"  Affine transform estimated (least squares, {n} points, "
              f"{n_inliers} inliers)")
        for i, (res, inl) in enumerate(zip(residuals, inliers.flatten())):
            flag = '' if inl else '  [outlier]'
            print(f"    pt{i+1}: residual = {res:.2f} px{flag}")

    return H, 'affine'


# =============================================================================
# SEM座標 [px] → EBSD物理座標 [μm] 変換
# =============================================================================
def sem_px_to_ebsd_um(cx_arr, cy_arr, H, mode, meta):
    """
    SEM画像座標 [px] → EBSD IQ画像座標 [px] → EBSD物理座標 [μm] に変換する。

    EBSD IQ画像は (ny × nx) ピクセル = 物理視野 (x_max-x_min) × (y_max-y_min) μm に対応。
    """
    pts = np.column_stack([cx_arr, cy_arr]).astype(np.float32)

    # SEM [px] → EBSD IQ画像 [px]（常にアフィン変換）
    ones = np.ones((len(pts), 1), dtype=np.float32)
    pts_h = np.hstack([pts, ones]).T        # (3, N)
    pts_ebsd_px = (H @ pts_h).T             # (N, 2)

    # EBSD IQ画像 [px] → 物理座標 [μm]
    nx, ny = meta['nx'], meta['ny']
    x_min, y_min = meta['x_min'], meta['y_min']
    x_max, y_max = meta['x_max'], meta['y_max']

    x_um = pts_ebsd_px[:, 0] / (nx - 1) * (x_max - x_min) + x_min
    y_um = pts_ebsd_px[:, 1] / (ny - 1) * (y_max - y_min) + y_min

    return x_um, y_um


# =============================================================================
# 最近傍割り当て
# =============================================================================
def assign_ebsd_to_subsets(cx_arr, cy_arr, x_um, y_um,
                           data, phase_names, meta):
    """
    各 DIC サブセット座標 (x_um, y_um) に最近傍の EBSD データを割り当てる。
    EBSD 視野外（距離 > ステップの2倍）のサブセットは grain_id=-1 で保存。

    Returns
    -------
    assigned   : list of dict
    n_outside  : int  視野外サブセット数
    """
    # KDTree で最近傍検索
    ebsd_xy = data[:, 3:5]  # x, y [μm]
    tree = cKDTree(ebsd_xy)
    max_dist = 2.0 * max(meta['x_step'], meta['y_step'])
    dists, idxs = tree.query(np.column_stack([x_um, y_um]), k=1)

    assigned = []
    n_outside = 0
    for i, (cx, cy) in enumerate(zip(cx_arr, cy_arr)):
        if dists[i] > max_dist:
            n_outside += 1
            assigned.append(dict(
                subset_id=i + 1,
                cx=int(cx), cy=int(cy),
                grain_id=-1, phase='',
                phi1_deg=np.nan, PHI_deg=np.nan, phi2_deg=np.nan,
                IQ=np.nan, CI=np.nan))
        else:
            row = data[idxs[i]]
            assigned.append(dict(
                subset_id=i + 1,
                cx=int(cx), cy=int(cy),
                grain_id=int(row[8]),
                phase=phase_names[idxs[i]],
                phi1_deg=float(np.degrees(row[0])),
                PHI_deg=float(np.degrees(row[1])),
                phi2_deg=float(np.degrees(row[2])),
                IQ=float(row[5]),
                CI=float(row[6]),
            ))

    if n_outside > 0:
        print(f"  Warning: {n_outside}/{len(cx_arr)} subsets outside EBSD FOV -> saved as grain_id=-1")

    return assigned, n_outside


# =============================================================================
# Excel 出力（既存ファイルにシート追加）
# =============================================================================
def save_xlsx(assigned, dic_xlsx_path, out_path):
    """
    dic_results.xlsx に 'ebsd_georef' シートを追加して out_path に保存する。

    高速化のため zipfile を直接操作する:
      ① 元 xlsx の zip エントリをそのままコピー（全シートを openpyxl で読み直さない）
      ② ebsd_georef シートの XML を新規生成して追加
      ③ [Content_Types].xml / workbook.xml / workbook.xml.rels のみ書き換え

    Parameters
    ----------
    assigned      : list of dict  割り当て結果
    dic_xlsx_path : str           元の dic_results.xlsx パス
    out_path      : str           出力先パス
    """
    import zipfile, shutil, re, xml.etree.ElementTree as ET

    # --- ebsd_georef シートの XML を生成 ---
    def _escape(v):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return ''
        s = str(v)
        return (s.replace('&','&amp;').replace('<','&lt;')
                  .replace('>','&gt;').replace('"','&quot;'))

    rows_data = [['subset_id', 'cx [px]', 'cy [px]', 'grain_id', 'phase',
                  'phi1 [deg]', 'PHI [deg]', 'phi2 [deg]', 'IQ', 'CI']]
    for r in assigned:
        rows_data.append([r['subset_id'], r['cx'], r['cy'], r['grain_id'], r['phase'],
                          r['phi1_deg'], r['PHI_deg'], r['phi2_deg'],
                          r['IQ'], r['CI']])

    sheet_xml_lines = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
        '<sheetData>',
    ]
    for ri, row in enumerate(rows_data, start=1):
        sheet_xml_lines.append(f'<row r="{ri}">')
        for ci, val in enumerate(row):
            col_letter = chr(ord('A') + ci)
            ref = f'{col_letter}{ri}'
            if isinstance(val, str) or val == '':
                sheet_xml_lines.append(
                    f'<c r="{ref}" t="inlineStr"><is><t>{_escape(val)}</t></is></c>')
            elif val is None or (isinstance(val, float) and np.isnan(val)):
                sheet_xml_lines.append(f'<c r="{ref}"/>')
            else:
                sheet_xml_lines.append(f'<c r="{ref}"><v>{val}</v></c>')
        sheet_xml_lines.append('</row>')
    sheet_xml_lines += ['</sheetData>', '</worksheet>']
    sheet_xml = '\n'.join(sheet_xml_lines).encode('utf-8')

    SHEET_NAME = 'ebsd_georef'
    SHEET_PATH = 'xl/worksheets/sheet_ebsd_georef.xml'
    REL_ID     = 'rId_ebsd_georef'

    # --- zipfile直接操作でコピー＋シート追加 ---
    with zipfile.ZipFile(dic_xlsx_path, 'r') as zin, \
         zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as zout:

        # 既存エントリをコピー（書き換え対象・ebsd_georef の残骸は除外）
        skip = {
            SHEET_PATH,
            'xl/worksheets/sheet_ebsd_georef.xml',
            'xl/workbook.xml',
            'xl/_rels/workbook.xml.rels',
            '[Content_Types].xml',
        }
        for item in zin.infolist():
            if item.filename in skip:
                continue
            zout.writestr(item, zin.read(item.filename))

        # workbook.xml を読んでシート登録を追加
        wb_xml = zin.read('xl/workbook.xml').decode('utf-8')
        # 既存 ebsd_georef エントリを除去
        wb_xml = re.sub(
            r'<sheet [^/]*name="ebsd_georef"[^/]*/>', '', wb_xml)
        # 最後の </sheets> の前に追加
        # sheetId は既存の最大+1
        ids = list(map(int, re.findall(r'sheetId="(\d+)"', wb_xml)))
        new_id = max(ids) + 1 if ids else 1
        new_sheet_tag = (f'<sheet name="{SHEET_NAME}" sheetId="{new_id}" '
                         f'r:id="{REL_ID}"/>')
        wb_xml = wb_xml.replace('</sheets>', new_sheet_tag + '</sheets>')
        zout.writestr('xl/workbook.xml', wb_xml.encode('utf-8'))

        # xl/_rels/workbook.xml.rels にリレーション追加
        rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        rels_xml = re.sub(
            r'<Relationship [^/]*Id="' + REL_ID + r'"[^/]*/>', '', rels_xml)
        new_rel = (f'<Relationship Id="{REL_ID}" '
                   f'Type="http://schemas.openxmlformats.org/officeDocument/'
                   f'2006/relationships/worksheet" '
                   f'Target="worksheets/sheet_ebsd_georef.xml"/>')
        rels_xml = rels_xml.replace('</Relationships>',
                                    new_rel + '</Relationships>')
        zout.writestr('xl/_rels/workbook.xml.rels', rels_xml.encode('utf-8'))

        # [Content_Types].xml に Override 追加
        ct_xml = zin.read('[Content_Types].xml').decode('utf-8')
        ct_xml = re.sub(
            r'<Override PartName="/xl/worksheets/sheet_ebsd_georef\.xml"[^/]*/>', '', ct_xml)
        new_ct = ('<Override PartName="/xl/worksheets/sheet_ebsd_georef.xml" '
                  'ContentType="application/vnd.openxmlformats-officedocument'
                  '.spreadsheetml.worksheet+xml"/>')
        ct_xml = ct_xml.replace('</Types>', new_ct + '</Types>')
        zout.writestr('[Content_Types].xml', ct_xml.encode('utf-8'))

        # 新シートの XML を追加
        zout.writestr(SHEET_PATH, sheet_xml)

    print(f"  Saved: {out_path}  (sheet '{SHEET_NAME}' added)")


# =============================================================================
# メイン
# =============================================================================
# =============================================================================
# オーバーレイ画像保存
# =============================================================================
def save_overlay_png(sem_img, grain_id_img, pts_sem, pts_ebsd, out_path,
                     opacity=0.45, show_subsets=False, cx_arr=None, cy_arr=None):
    """
    位置合わせ結果のオーバーレイ画像をPNGとして保存する。
    SEM画像上にGrain IDマップをワープして半透明で重ね、
    コントロールポイントを番号付きで表示する。
    show_subsets=True のとき DIC サブセット点も描画する。
    """
    import matplotlib.pyplot as _plt

    H_sem, W_sem = sem_img.shape[:2]
    n = len(pts_sem)

    # --- ワープ（EBSD→SEM方向）---
    src = pts_sem.astype(np.float32)
    dst = pts_ebsd.astype(np.float32)
    if n == 3:
        H_mat = cv2.getAffineTransform(dst[:3], src[:3])
    else:
        H_mat, _ = cv2.estimateAffine2D(dst, src,
                                        method=cv2.RANSAC,
                                        ransacReprojThreshold=10.0)
    warped = cv2.warpAffine(
        grain_id_img, H_mat, (W_sem, H_sem),
        flags=cv2.INTER_NEAREST,
        borderMode=cv2.BORDER_CONSTANT, borderValue=(0, 0, 0))

    # --- SEM + Grain IDを合成 ---
    sem_rgb  = np.stack([sem_img] * 3, axis=-1).astype(np.float32)
    warp_f   = warped.astype(np.float32)
    is_black = ((warped[:, :, 0] == 0) &
                (warped[:, :, 1] == 0) &
                (warped[:, :, 2] == 0))
    alpha    = np.where(is_black, 0.0, opacity)[:, :, np.newaxis]
    blended  = (sem_rgb * (1 - alpha) + warp_f * alpha).clip(0, 255).astype(np.uint8)

    # --- 描画・保存 ---
    _plt.style.use('dark_background')
    fig, ax = _plt.subplots(figsize=(12, 9))
    fig.patch.set_facecolor('#0a0c0f')
    ax.set_facecolor('#111318')
    ax.imshow(blended, origin='upper')
    # DICサブセット点
    if show_subsets and cx_arr is not None and cy_arr is not None:
        ax.scatter(cx_arr, cy_arr, s=8, c='cyan', alpha=0.7,
                   linewidths=0, zorder=3)
    # コントロールポイント
    for i, ps in enumerate(pts_sem, start=1):
        ax.plot(ps[0], ps[1], 'r+', markersize=14, markeredgewidth=2, zorder=5)
        ax.annotate(str(i), (ps[0], ps[1]), color='red', fontsize=9,
                    fontweight='bold', xytext=(5, 5),
                    textcoords='offset points', zorder=5)
    mode_str = 'Affine'
    subset_str = '  + subsets' if show_subsets else ''
    ax.set_title(
        f'EBSD Grain ID overlay on SEM  ({mode_str}, {n} pts, opacity={int(opacity*100)}%){subset_str}',
        fontsize=11)
    ax.set_xlabel('x [px]')
    ax.set_ylabel('y [px]')
    _plt.tight_layout()
    fig.savefig(str(out_path), dpi=150)
    _plt.close(fig)
    print(f"  Saved overlay PNG: {out_path}")



# =============================================================================
# マップビューワー（ジオリファレンス後）
# =============================================================================
def show_map_viewer(georef_xlsx_path, sem_img, grain_id_img, pts_sem, pts_ebsd):
    """
    ジオリファレンス完了後に立ち上げるインタラクティブマップビューワー。

    操作:
      - シート名ドロップダウン: dic_results の列シート（u/v/exx/…）を自動検出
      - 応力段階ドロップダウン: シートのヘッダ列名を自動検出
      - 「表示」ボタン: matplotlib figure で新規表示
        左: SEM + Grain IDオーバーレイ + サブセット散布（選択データ）
        右: Grain IDカラーマップ
    """
    from PyQt6.QtWidgets import (
        QApplication, QWidget, QVBoxLayout, QHBoxLayout,
        QLabel, QComboBox, QCheckBox, QPushButton,
        QScrollArea, QFrame, QGridLayout, QSizePolicy,
        QFileDialog,
    )
    from PyQt6.QtCore import Qt
    import matplotlib.pyplot as _plt
    import matplotlib.colors as _mcolors
    import sys as _sys

    # QApplication が未起動なら起動
    _app = QApplication.instance() or QApplication(_sys.argv)

    # =================================================================
    # georef xlsx を読み込んで subset_id ベースのデータ構造を構築する
    #
    # データソース:
    #   ebsd_georef シート: subset_id → grain_id の対応
    #   u シート: subset_id → x_ref [px], u値（変形ステップごと）
    #   v シート: subset_id → y_ref [px], v値（変形ステップごと）
    #   その他シート: subset_id → 物理量値
    #
    # 変形後座標:
    #   cx_deformed = x_ref + u[label]
    #   cy_deformed = y_ref + v[label]
    # =================================================================
    wb_e = openpyxl.load_workbook(georef_xlsx_path, read_only=True)

    # --- ebsd_georef から subset_id → grain_id を取得 ---
    rows_em = list(wb_e['ebsd_georef'].iter_rows(values_only=True))[1:]
    # ヘッダ確認: subset_id, cx, cy, grain_id, ...
    hdr_em = list(wb_e['ebsd_georef'].iter_rows(values_only=True, max_row=1))[0]
    sid_em_col   = next((i for i, h in enumerate(hdr_em) if h == 'subset_id'), 0)
    gid_em_col   = next((i for i, h in enumerate(hdr_em) if h == 'grain_id'),  3)
    sid_to_gid   = {int(r[sid_em_col]): int(r[gid_em_col])
                    for r in rows_em if r[sid_em_col] is not None}
    all_sids     = sorted(sid_to_gid.keys())
    sid_view     = np.array(all_sids, dtype=int)
    gid_view     = np.array([sid_to_gid[s] for s in all_sids], dtype=float)

    # --- uシートから x_ref（DIC参照x座標）を取得 ---
    rows_u    = list(wb_e['u'].iter_rows(values_only=True))
    hdr_u     = rows_u[0]
    sid_u_col = next((i for i, h in enumerate(hdr_u) if h == 'subset_id'), None)
    xref_col  = next((i for i, h in enumerate(hdr_u) if h == 'x [px]'), 1)
    sid_to_xref = {int(r[sid_u_col]): float(r[xref_col])
                   for r in rows_u[1:] if r[sid_u_col] is not None}

    # --- vシートから y_ref（DIC参照y座標）を取得 ---
    rows_v    = list(wb_e['v'].iter_rows(values_only=True))
    hdr_v     = rows_v[0]
    sid_v_col = next((i for i, h in enumerate(hdr_v) if h == 'subset_id'), None)
    yref_col  = next((i for i, h in enumerate(hdr_v) if h == 'y [px]'), 2)
    sid_to_yref = {int(r[sid_v_col]): float(r[yref_col])
                   for r in rows_v[1:] if r[sid_v_col] is not None}

    cx_ref = np.array([sid_to_xref.get(s, np.nan) for s in all_sids], dtype=float)
    cy_ref = np.array([sid_to_yref.get(s, np.nan) for s in all_sids], dtype=float)

    # --- 全データシートから物理量をキャッシュ ---
    # _val_cache[(sheet_name, label)] = ndarray (sid_view順)
    _val_cache = {}
    for _sh in wb_e.sheetnames:
        _rows = list(wb_e[_sh].iter_rows(values_only=True))
        _hdr  = _rows[0]
        _sid_col = next((i for i, h in enumerate(_hdr) if h == 'subset_id'), None)
        if _sid_col is None:
            continue
        _data_start = _sid_col + 3  # subset_id, x [px], y [px] の次から
        for _ci, _h in enumerate(_hdr[_data_start:], start=_data_start):
            if _h is None:
                continue
            _lbl = str(_h)
            _sid_to_val = {int(r[_sid_col]): (r[_ci] if r[_ci] is not None else np.nan)
                           for r in _rows[1:] if r[_sid_col] is not None}
            # 数値に変換できない列（phase等の文字列列）はスキップ
            try:
                _val_cache[(_sh, _lbl)] = np.array(
                    [_sid_to_val.get(s, np.nan) for s in all_sids], dtype=float)
            except (ValueError, TypeError):
                continue
    wb_e.close()

    # 表示可能なシート・ラベルを列挙（ebsd_georef は除外）
    sheet_labels = {}
    for (_sh, _lbl), _ in _val_cache.items():
        if _sh != 'ebsd_georef':
            sheet_labels.setdefault(_sh, []).append(_lbl)
    data_sheets = list(sheet_labels.keys())

    if not data_sheets:
        print("  No valid DIC data sheets found in georef xlsx.")
        return

    # Grain IDカラーマップ
    # tab20 / tab20b / tab20c を合成して最大60色の定性カラーパレットを生成
    valid_ids = sorted(set(int(g) for g in gid_view if g > 0))
    n_grains  = len(valid_ids)
    def _make_qualitative_colors(n):
        import matplotlib.cm as _cm
        pool = []
        for _name in ('tab20', 'tab20b', 'tab20c'):
            _c = matplotlib.colormaps[_name]
            pool.extend([_c(i) for i in range(20)])
        # pool(60色)から n 色を均等間隔で選ぶ
        idxs = np.linspace(0, len(pool) - 1, max(n, 1), dtype=int)
        return [pool[i] for i in idxs]
    _colors    = _make_qualitative_colors(n_grains)
    gid_to_rgba = {gid: _colors[i] for i, gid in enumerate(valid_ids)}

    # --- 粒界セグメント計算 ---
    # 参照グリッドのインデックスで隣接判定し、変形後座標の中点をセル境界として使う
    def _calc_grain_boundary_segments(cx_arr, cy_arr, gid_arr):
        """
        隣接サブセット（右・下）でgrain_idが異なる場合に粒界線分を生成する。
        pcolormesh のセル境界（隣接2点の中点）と一致するよう端点を計算する。
        戻り値: list of ((x0,y0),(x1,y1))
        """
        xs_u = np.unique(cx_ref)
        ys_u = np.unique(cy_ref)
        nx_g = len(xs_u)
        ny_g = len(ys_u)
        xi_map = {int(round(v)): i for i, v in enumerate(xs_u)}
        yi_map = {int(round(v)): i for i, v in enumerate(ys_u)}

        grid_cx  = np.full((ny_g, nx_g), np.nan)
        grid_cy  = np.full((ny_g, nx_g), np.nan)
        grid_gid = np.full((ny_g, nx_g), -999, dtype=int)
        for i in range(len(cx_ref)):
            ix = xi_map.get(int(round(cx_ref[i])))
            iy = yi_map.get(int(round(cy_ref[i])))
            if ix is None or iy is None:
                continue
            grid_cx[iy, ix]  = cx_arr[i]
            grid_cy[iy, ix]  = cy_arr[i]
            grid_gid[iy, ix] = int(gid_arr[i])

        segs = []
        for iy in range(ny_g):
            for ix in range(nx_g):
                g0 = grid_gid[iy, ix]
                if g0 == -999:
                    continue
                cx0, cy0 = grid_cx[iy, ix], grid_cy[iy, ix]

                # 右隣との境界（縦線：上下隣との中点を端点とする）
                if ix + 1 < nx_g:
                    g_r = grid_gid[iy, ix + 1]
                    if g_r != -999 and g_r != g0:
                        cx1, cy1 = grid_cx[iy, ix + 1], grid_cy[iy, ix + 1]
                        # 境界の中心（pcolormeshのセル境界と一致）
                        bx = (cx0 + cx1) * 0.5
                        by = (cy0 + cy1) * 0.5
                        # 上下の端点：上(iy-1)との中点 ～ 下(iy+1)との中点
                        if iy > 0 and grid_gid[iy-1, ix] != -999:
                            top_y = (grid_cy[iy-1, ix] + grid_cy[iy-1, ix+1]) * 0.5 if ix + 1 < nx_g and grid_gid[iy-1, ix+1] != -999 else (cy0 + grid_cy[iy-1, ix]) * 0.5
                            top_x = (grid_cx[iy-1, ix] + grid_cx[iy-1, ix+1]) * 0.5 if ix + 1 < nx_g and grid_gid[iy-1, ix+1] != -999 else (cx0 + grid_cx[iy-1, ix]) * 0.5
                            top_pt = ((bx + top_x) * 0.5, (by + top_y) * 0.5)
                        else:
                            top_pt = (bx, by - abs(cy1 - cy0) * 0.5) if abs(cy1 - cy0) > 1e-6 else (bx, by - abs(cx1 - cx0) * 0.5)
                        if iy + 1 < ny_g and grid_gid[iy+1, ix] != -999:
                            bot_y = (grid_cy[iy+1, ix] + grid_cy[iy+1, ix+1]) * 0.5 if ix + 1 < nx_g and grid_gid[iy+1, ix+1] != -999 else (cy0 + grid_cy[iy+1, ix]) * 0.5
                            bot_x = (grid_cx[iy+1, ix] + grid_cx[iy+1, ix+1]) * 0.5 if ix + 1 < nx_g and grid_gid[iy+1, ix+1] != -999 else (cx0 + grid_cx[iy+1, ix]) * 0.5
                            bot_pt = ((bx + bot_x) * 0.5, (by + bot_y) * 0.5)
                        else:
                            bot_pt = (bx, by + abs(cy1 - cy0) * 0.5) if abs(cy1 - cy0) > 1e-6 else (bx, by + abs(cx1 - cx0) * 0.5)
                        segs.append((top_pt, bot_pt))

                # 下隣との境界（横線）
                if iy + 1 < ny_g:
                    g_d = grid_gid[iy + 1, ix]
                    if g_d != -999 and g_d != g0:
                        cx1, cy1 = grid_cx[iy + 1, ix], grid_cy[iy + 1, ix]
                        bx = (cx0 + cx1) * 0.5
                        by = (cy0 + cy1) * 0.5
                        # 左右の端点
                        if ix > 0 and grid_gid[iy, ix-1] != -999:
                            lx = (grid_cx[iy, ix-1] + grid_cx[iy+1, ix-1]) * 0.5 if grid_gid[iy+1, ix-1] != -999 else (cx0 + grid_cx[iy, ix-1]) * 0.5
                            ly = (grid_cy[iy, ix-1] + grid_cy[iy+1, ix-1]) * 0.5 if grid_gid[iy+1, ix-1] != -999 else (cy0 + grid_cy[iy, ix-1]) * 0.5
                            left_pt = ((bx + lx) * 0.5, (by + ly) * 0.5)
                        else:
                            left_pt = (bx - abs(cx1 - cx0) * 0.5, by) if abs(cx1 - cx0) > 1e-6 else (bx - abs(cy1 - cy0) * 0.5, by)
                        if ix + 1 < nx_g and grid_gid[iy, ix+1] != -999:
                            rx = (grid_cx[iy, ix+1] + grid_cx[iy+1, ix+1]) * 0.5 if grid_gid[iy+1, ix+1] != -999 else (cx0 + grid_cx[iy, ix+1]) * 0.5
                            ry = (grid_cy[iy, ix+1] + grid_cy[iy+1, ix+1]) * 0.5 if grid_gid[iy+1, ix+1] != -999 else (cy0 + grid_cy[iy, ix+1]) * 0.5
                            right_pt = ((bx + rx) * 0.5, (by + ry) * 0.5)
                        else:
                            right_pt = (bx + abs(cx1 - cx0) * 0.5, by) if abs(cx1 - cx0) > 1e-6 else (bx + abs(cy1 - cy0) * 0.5, by)
                        segs.append((left_pt, right_pt))
        return segs

    _gb_segs_ref = _calc_grain_boundary_segments(cx_ref, cy_ref, gid_view)
    print(f"  Grain boundary segments (ref): {len(_gb_segs_ref)}")
    H_sem, W_sem = sem_img.shape[:2]
    n_cp = len(pts_sem)
    src = pts_sem.astype(np.float32)
    dst = pts_ebsd.astype(np.float32)
    if n_cp == 3:
        H_warp = cv2.getAffineTransform(dst[:3], src[:3])
    else:
        H_warp, _ = cv2.estimateAffine2D(dst, src,
                                         method=cv2.RANSAC,
                                         ransacReprojThreshold=10.0)
    warped_base = cv2.warpAffine(
        grain_id_img, H_warp, (W_sem, H_sem),
        flags=cv2.INTER_NEAREST,
        borderMode=cv2.BORDER_CONSTANT, borderValue=(0, 0, 0))

    # ================================================================
    # PyQt6 ダークテーマ定数
    # ================================================================
    _DARK_QSS = """
        QWidget {
            background-color: #111318;
            color: #e0e4ec;
            font-family: 'Noto Sans JP', 'Yu Gothic', sans-serif;
            font-size: 12px;
        }
        QLabel {
            color: #6b7280;
            font-size: 11px;
        }
        QLabel#section_label {
            color: #00d4ff;
            font-size: 10px;
            font-family: 'JetBrains Mono', monospace;
            letter-spacing: 1px;
        }
        QComboBox {
            background-color: #1a1d24;
            border: 1px solid #2a2d35;
            border-radius: 3px;
            padding: 4px 8px;
            color: #e0e4ec;
            font-size: 12px;
            min-width: 90px;
        }
        QComboBox:hover { border-color: #00d4ff; }
        QComboBox::drop-down { border: none; width: 20px; }
        QComboBox QAbstractItemView {
            background-color: #1a1d24;
            border: 1px solid #2a2d35;
            selection-background-color: #00d4ff;
            selection-color: #0a0c0f;
            color: #e0e4ec;
        }
        QCheckBox {
            color: #e0e4ec;
            font-size: 11px;
            spacing: 6px;
        }
        QCheckBox::indicator {
            width: 14px; height: 14px;
            border: 1px solid #2a2d35;
            border-radius: 2px;
            background: #1a1d24;
        }
        QCheckBox::indicator:checked {
            background: #00d4ff;
            border-color: #00d4ff;
        }
        QPushButton {
            background-color: #1a1d24;
            border: 1px solid #2a2d35;
            border-radius: 4px;
            color: #e0e4ec;
            padding: 6px 18px;
            font-size: 12px;
        }
        QPushButton:hover { border-color: #00d4ff; color: #00d4ff; background-color: rgba(0,212,255,0.07); }
        QPushButton#btn_show {
            border-color: #00d4ff; color: #00d4ff;
            background-color: rgba(0,212,255,0.10);
            font-weight: bold;
        }
        QPushButton#btn_show:hover { background-color: rgba(0,212,255,0.20); }
        QPushButton#btn_save {
            border-color: #00ff88; color: #00ff88;
            background-color: rgba(0,255,136,0.08);
        }
        QPushButton#btn_save:hover { background-color: rgba(0,255,136,0.18); }
        QPushButton#btn_close {
            border-color: #ff6b35; color: #ff6b35;
            background-color: rgba(255,107,53,0.08);
        }
        QPushButton#btn_close:hover { background-color: rgba(255,107,53,0.18); }
        QFrame#sep {
            background-color: #2a2d35;
            max-height: 1px;
        }
        QScrollArea { border: none; }
    """

    # grain_id リスト（"All" + 各ID）
    gid_options = ['All'] + [str(g) for g in valid_ids]

    # ================================================================
    # PyQt6 Map Viewer ウィンドウ
    # ================================================================
    win = QWidget()
    win.setWindowTitle('Map Viewer  |  EBSD Georef')
    win.setStyleSheet(_DARK_QSS)
    win.setWindowFlags(win.windowFlags() | Qt.WindowType.WindowStaysOnTopHint)

    root_layout = QVBoxLayout(win)
    root_layout.setContentsMargins(16, 14, 16, 14)
    root_layout.setSpacing(10)

    def _sep():
        f = QFrame()
        f.setObjectName('sep')
        f.setFixedHeight(1)
        root_layout.addWidget(f)

    def _section(text):
        lb = QLabel(text.upper())
        lb.setObjectName('section_label')
        root_layout.addWidget(lb)

    # --- 物理量・Grain ID ---
    _section('物理量 / Grain ID')
    row0 = QHBoxLayout()
    row0.setSpacing(8)
    row0.addWidget(QLabel('物理量'))
    sheet_cb = QComboBox()
    sheet_cb.addItems(data_sheets)
    sheet_cb.setCurrentIndex(0)
    row0.addWidget(sheet_cb)
    row0.addSpacing(12)
    row0.addWidget(QLabel('Grain ID'))
    gid_cb = QComboBox()
    gid_cb.addItems(gid_options)
    gid_cb.setCurrentIndex(0)
    row0.addWidget(gid_cb)
    row0.addStretch()
    root_layout.addLayout(row0)

    _sep()

    # --- 表示内容チェックボックス ---
    _section('表示内容')
    row1 = QHBoxLayout()
    row1.setSpacing(16)
    show_dic_cb = QCheckBox('物理量マップ')
    show_dic_cb.setChecked(True)
    show_gid_cb = QCheckBox('Grain IDマップ')
    show_gid_cb.setChecked(True)
    row1.addWidget(show_dic_cb)
    row1.addWidget(show_gid_cb)
    row1.addStretch()
    root_layout.addLayout(row1)

    _sep()

    # --- 変形ステップ（スクロール可チェックボックス群）---
    _section('変形ステップ')

    _step_scroll_area = QScrollArea()
    _step_scroll_area.setWidgetResizable(True)
    _step_scroll_area.setFixedHeight(110)
    _step_scroll_area.setStyleSheet(
        'QScrollArea { border: 1px solid #2a2d35; border-radius: 3px; background: #0a0c0f; }')
    _step_container = QWidget()
    _step_container.setStyleSheet('background: #0a0c0f;')
    _step_grid = QGridLayout(_step_container)
    _step_grid.setContentsMargins(8, 6, 8, 6)
    _step_grid.setSpacing(4)
    _step_scroll_area.setWidget(_step_container)
    root_layout.addWidget(_step_scroll_area)

    _step_checks = {}   # label -> QCheckBox

    def _rebuild_step_checks(labels):
        # 既存ウィジェットを削除
        for i in reversed(range(_step_grid.count())):
            w = _step_grid.itemAt(i).widget()
            if w:
                w.setParent(None)
        _step_checks.clear()
        cols = 6
        for idx, lbl in enumerate(labels):
            cb = QCheckBox(lbl)
            cb.setChecked(True)
            cb.setStyleSheet('font-size: 11px; color: #e0e4ec;')
            _step_grid.addWidget(cb, idx // cols, idx % cols)
            _step_checks[lbl] = cb

    def _on_sheet_change():
        sh = sheet_cb.currentText()
        labels = sheet_labels.get(sh, [])
        _rebuild_step_checks(labels)

    sheet_cb.currentIndexChanged.connect(_on_sheet_change)
    _on_sheet_change()

    _sep()

    # 開いているfigureを追跡（一括保存用）
    _open_figs = []

    def _on_show():
        sh       = sheet_cb.currentText()
        show_dic = show_dic_cb.isChecked()
        show_gid = show_gid_cb.isChecked()
        sel_gid  = gid_cb.currentText()

        # 選択されたステップを取得
        selected_labels = [lbl for lbl, cb in _step_checks.items() if cb.isChecked()]
        if not selected_labels:
            return
        if not show_dic and not show_gid:
            return

        import matplotlib.colors as _mcolors_local
        from matplotlib.gridspec import GridSpec

        gid_mask = (np.ones(len(gid_view), dtype=bool) if sel_gid == 'All'
                    else (gid_view == int(sel_gid)))
        gid_str  = f'  Grain {sel_gid}' if sel_gid != 'All' else ''

        # --- 軸範囲・パネルサイズ ---
        x_lim   = (cx_ref.min() - 20, cx_ref.max() + 20)
        y_lim   = (cy_ref.max() + 20, cy_ref.min() - 20)
        x_span  = x_lim[1] - x_lim[0]
        y_span  = y_lim[0] - y_lim[1]
        panel_h = 7.0
        panel_w = panel_h * x_span / y_span
        n_map_types = sum([show_dic, show_gid])

        def _setup_ax(ax, title=''):
            ax.set_facecolor('white')
            for sp in ax.spines.values():
                sp.set_edgecolor('#cccccc')
                sp.set_linewidth(0.8)
            ax.set_xlim(*x_lim)
            ax.set_ylim(*y_lim)
            ax.set_aspect('equal')
            ax.set_xlabel('x [px]', fontsize=9, color='#444444')
            ax.set_ylabel('y [px]', fontsize=9, color='#444444')
            ax.tick_params(colors='#444444', labelsize=8)
            if title:
                ax.set_title(title, fontsize=10, color='#222222', pad=6)

        def _draw_gb(ax, segs):
            from matplotlib.collections import LineCollection
            if segs:
                lc_bg = LineCollection(segs, colors='white',
                                       linewidths=2.5, alpha=1.0, zorder=5)
                lc_fg = LineCollection(segs, colors='#111111',
                                       linewidths=1.0, alpha=0.8, zorder=6)
                ax.add_collection(lc_bg)
                ax.add_collection(lc_fg)

        # --- カラーマップ・スケールを全選択ステップ横断で決定 ---
        KEYS_SYM  = ('exx', 'eyy', 'exy')
        KEYS_ASYM = ('e1', 'gamma_max')
        KEYS_DISP = ('u', 'v')
        all_valid_vals = []
        for lbl in selected_labels:
            cached = _val_cache.get((sh, lbl))
            if cached is None:
                continue
            u_arr = _val_cache.get(('u', lbl))
            v_arr = _val_cache.get(('v', lbl))
            valid_uv = (~np.isnan(u_arr) & ~np.isnan(v_arr)
                        if u_arr is not None and v_arr is not None
                        else np.ones(len(cx_ref), dtype=bool))
            bmask = gid_mask & valid_uv
            v = cached[bmask & ~np.isnan(cached)]
            all_valid_vals.append(v)
        all_valid_vals = np.concatenate(all_valid_vals) if all_valid_vals else np.array([])

        if len(all_valid_vals) == 0:
            vmin_use, vmax_use, cmap_use = 0, 1, 'RdBu_r'
        elif sh == 'ncc':
            vmin_use, vmax_use, cmap_use = 0.2, 1.0, 'plasma'
        elif sh == 'omega_xy':
            vabs = max(float(np.percentile(np.abs(all_valid_vals), 98)), 1e-6)
            vmin_use, vmax_use, cmap_use = -vabs, vabs, 'PiYG'
        elif sh in KEYS_DISP:
            vabs = max(float(np.percentile(np.abs(all_valid_vals), 95)), 1.0)
            vmin_use, vmax_use, cmap_use = -vabs, vabs, 'RdBu_r'
        elif sh in KEYS_SYM:
            vabs = max(float(np.percentile(np.abs(all_valid_vals), 98)), 1e-6)
            vmin_use, vmax_use, cmap_use = -vabs, vabs, 'RdBu_r'
        elif sh in KEYS_ASYM:
            vmin_use = float(np.percentile(all_valid_vals, 2))
            vmax_use = float(np.percentile(all_valid_vals, 98))
            cmap_use = 'hot_r'
        else:
            vabs = max(float(np.percentile(np.abs(all_valid_vals), 98)), 1e-6)
            vmin_use, vmax_use, cmap_use = -vabs, vabs, 'RdBu_r'
        xs_u   = np.unique(cx_ref)
        ys_u   = np.unique(cy_ref)
        nx_g   = len(xs_u)
        ny_g   = len(ys_u)
        xi_map_l = {int(round(v)): i for i, v in enumerate(xs_u)}
        yi_map_l = {int(round(v)): i for i, v in enumerate(ys_u)}

        def _build_grids(cx_plot, cy_plot, values, mask):
            """
            変形後座標 cx_plot/cy_plot を (ny_g, nx_g) グリッドに配置。
            pcolormesh 用のセル中心座標グリッド X,Y と値グリッド V を返す。
            """
            X = np.full((ny_g, nx_g), np.nan)
            Y = np.full((ny_g, nx_g), np.nan)
            V = np.full((ny_g, nx_g), np.nan)
            for i in range(len(cx_ref)):
                ix = xi_map_l.get(int(round(cx_ref[i])))
                iy = yi_map_l.get(int(round(cy_ref[i])))
                if ix is None or iy is None:
                    continue
                X[iy, ix] = cx_plot[i]
                Y[iy, ix] = cy_plot[i]
                if mask[i] and values is not None:
                    V[iy, ix] = values[i]
            return X, Y, V

        def _build_gid_index_grid(cx_plot, cy_plot, mask):
            """
            grain_id → 整数インデックス (ny_g, nx_g) グリッドを生成。
            pcolormesh + ListedColormap で色付けするためのインデックス値。
            -1: 無効（非表示）
            """
            grid = np.full((ny_g, nx_g), -1, dtype=int)
            for i in range(len(cx_ref)):
                ix = xi_map_l.get(int(round(cx_ref[i])))
                iy = yi_map_l.get(int(round(cy_ref[i])))
                if ix is None or iy is None:
                    continue
                g = int(gid_view[i])
                if not mask[i]:
                    grid[iy, ix] = -1
                elif g in gid_to_rgba:
                    grid[iy, ix] = valid_ids.index(g)
                elif g == 0:
                    grid[iy, ix] = len(valid_ids)      # 粒界色用インデックス
                else:
                    grid[iy, ix] = len(valid_ids) + 1  # anti-grain用インデックス
            return grid

        def _pcmesh_edges(X, Y):
            """
            pcolormesh に渡すセル境界座標 (ny_g+1, nx_g+1) を
            セル中心座標グリッド X,Y から生成する。
            """
            # 内部境界: 隣接セル中点
            # 外部境界: 端セルから同じ幅だけ外側に延長
            def _edge(arr):
                # arr: (n,) → edges: (n+1,)
                mid = (arr[:-1] + arr[1:]) * 0.5
                left  = arr[0]  - (arr[1]  - arr[0])  * 0.5
                right = arr[-1] + (arr[-1] - arr[-2]) * 0.5
                return np.concatenate([[left], mid, [right]])
            # X,Y は (ny_g, nx_g)。行・列ごとに1D edgeを計算して2Dに展開
            # 列方向(x)のedge
            ex = np.apply_along_axis(_edge, 1, X)   # (ny_g, nx_g+1)
            # 行方向(y)のedge
            ey = np.apply_along_axis(_edge, 0, Y)   # (ny_g+1, nx_g)
            # 2D edge grid (ny_g+1, nx_g+1)
            Xe = np.zeros((ny_g + 1, nx_g + 1))
            Ye = np.zeros((ny_g + 1, nx_g + 1))
            # xエッジを行方向に補間
            ex_full = np.vstack([ex, ex[-1:] + (ex[-1:] - ex[-2:-1]) * 0.5])
            ex_full[0] = ex_full[1] - (ex_full[2] - ex_full[1])
            Xe = ex_full
            # yエッジを列方向に補間
            ey_t = ey.T  # (nx_g, ny_g+1)
            ey_full = np.vstack([ey_t, ey_t[-1:] + (ey_t[-1:] - ey_t[-2:-1]) * 0.5]).T
            ey_full[:, 0] = ey_full[:, 1] - (ey_full[:, 2] - ey_full[:, 1])
            Ye = ey_full
            return Xe, Ye

        # --- style・DPI をループ外で1度だけ設定 ---
        _plt.style.use('default')
        _FIXED_DPI = 100  # 全figureで統一するDPI
        FIG_W = 7.0 * n_map_types + 1.5 * n_map_types
        FIG_H = 7.0

        # --- ステップごとに1 figure ずつ生成 ---
        for lbl in selected_labels:
            cached = _val_cache.get((sh, lbl))
            u_arr  = _val_cache.get(('u', lbl))
            v_arr  = _val_cache.get(('v', lbl))

            # 変形後座標（最重要の値 — 一切変更しない）
            if u_arr is not None and v_arr is not None:
                cx_plot = np.where(np.isnan(u_arr), cx_ref, cx_ref + u_arr)
                cy_plot = np.where(np.isnan(v_arr), cy_ref, cy_ref + v_arr)
                valid_uv = ~np.isnan(u_arr) & ~np.isnan(v_arr)
            else:
                cx_plot  = cx_ref
                cy_plot  = cy_ref
                valid_uv = np.ones(len(cx_ref), dtype=bool)
            base_mask = gid_mask & valid_uv

            # 粒界セグメントを変形後座標で計算（subset IDとGrain IDの対応は維持）
            gb_segs = _calc_grain_boundary_segments(cx_plot, cy_plot, gid_view)

            # pcolormesh 用グリッド構築
            X, Y, V  = _build_grids(cx_plot, cy_plot, cached, base_mask)
            GID_grid = _build_gid_index_grid(cx_plot, cy_plot, base_mask)
            Xe, Ye   = _pcmesh_edges(X, Y)

            # GridSpec
            col_ratios = []
            for pi in range(n_map_types):
                col_ratios += [1, 0.06]
                if pi < n_map_types - 1:
                    col_ratios += [0.13]  # マップ間スペーサー（80px分追加）
            n_cols = len(col_ratios)

            # figureサイズ・DPI固定
            fig = _plt.figure(figsize=(FIG_W, FIG_H), dpi=_FIXED_DPI)
            fig.set_size_inches(FIG_W, FIG_H)
            fig.patch.set_facecolor('white')
            fig.suptitle(f'{sh}  /  {lbl}{gid_str}',
                         fontsize=11, color='#222222', y=0.98)
            gs = GridSpec(1, n_cols, figure=fig,
                          width_ratios=col_ratios,
                          wspace=0.0, left=0.07, right=0.97,
                          top=0.91, bottom=0.09)

            col_idx = 0
            # --- 物理量マップ（pcolormesh）---
            if show_dic:
                ax    = fig.add_subplot(gs[0, col_idx])
                ax_cb = fig.add_subplot(gs[0, col_idx + 1])
                col_idx += 2 + (1 if show_gid else 0)
                Vm = np.ma.masked_invalid(V)
                if not Vm.mask.all():
                    pc = ax.pcolormesh(Xe, Ye, Vm,
                                       cmap=cmap_use,
                                       vmin=vmin_use, vmax=vmax_use,
                                       shading='flat', zorder=2)
                    cb = _plt.colorbar(pc, cax=ax_cb)
                    cb.ax.tick_params(colors='#444444', labelsize=7)
                    ax_cb.set_facecolor('white')
                    for sp in ax_cb.spines.values():
                        sp.set_edgecolor('#cccccc')
                else:
                    ax_cb.axis('off')
                _draw_gb(ax, gb_segs)
                _setup_ax(ax, title=f'{sh}  /  {lbl}{gid_str}')

            # --- Grain IDマップ（pcolormesh + ListedColormap）---
            if show_gid:
                ax    = fig.add_subplot(gs[0, col_idx])
                ax_cb = fig.add_subplot(gs[0, col_idx + 1])

                # カラーパレット: valid_ids分 + 粒界色 + anti-grain色 + 無効色
                palette = [gid_to_rgba[g] for g in valid_ids]
                palette += [(0, 0, 0, 0.4),    # GrainID=0（粒界・未解析）
                            (0.3, 0.3, 0.3, 0.4),  # anti-grain
                            (0, 0, 0, 0)]       # 無効セル（透明）
                n_colors = len(palette)
                from matplotlib.colors import ListedColormap, BoundaryNorm
                cmap_gid  = ListedColormap(palette)
                # -1（無効）を末尾インデックスに丸めるため vmin/vmax を設定
                GID_disp  = np.where(GID_grid < 0, n_colors - 1, GID_grid)
                norm_gid  = BoundaryNorm(
                    np.arange(-0.5, n_colors, 1), ncolors=n_colors)
                ax.pcolormesh(Xe, Ye, GID_disp,
                              cmap=cmap_gid, norm=norm_gid,
                              shading='flat', zorder=2)
                ax_cb.axis('off')
                _draw_gb(ax, gb_segs)
                _setup_ax(ax, title=f'Grain ID map{gid_str}')

            _open_figs.append((lbl, fig))
            _plt.show(block=False)
            # ウィンドウサイズをピクセル単位で固定（全ステップ統一）
            try:
                _mgr = fig.canvas.manager
                _px_w = int(FIG_W * _FIXED_DPI)
                _px_h = int(FIG_H * _FIXED_DPI)
                _mgr.window.resize(_px_w, _px_h)
            except Exception:
                pass

        # 全figure表示後、Map ViewerをFigureの下に配置
        if _open_figs:
            try:
                last_fig = _open_figs[-1][1]
                fig_win  = last_fig.canvas.manager.window
                fig_win.adjustSize() if hasattr(fig_win, 'adjustSize') else None
                fy = fig_win.y() if hasattr(fig_win, 'y') else 0
                fh = fig_win.height() if hasattr(fig_win, 'height') else 600
                win.move(0, fy + fh + 10)
            except Exception:
                pass

    def _on_save():
        if not _open_figs:
            return
        save_dir = QFileDialog.getExistingDirectory(win, '保存先フォルダを選択')
        if not save_dir:
            return
        sh = sheet_cb.currentText()
        for lbl, fig in _open_figs:
            fname = f'{sh}_{lbl}.png'.replace('/', '-').replace(' ', '_')
            fpath = Path(save_dir) / fname
            fig.savefig(str(fpath), dpi=150, bbox_inches='tight')
            print(f'  Saved: {fpath}')
        print(f'  {len(_open_figs)} figure(s) saved to {save_dir}')

    # --- ボタン行 ---
    btn_row = QHBoxLayout()
    btn_row.setSpacing(8)
    btn_show  = QPushButton('表示')
    btn_show.setObjectName('btn_show')
    btn_save  = QPushButton('保存')
    btn_save.setObjectName('btn_save')
    btn_close = QPushButton('閉じる')
    btn_close.setObjectName('btn_close')
    btn_show.clicked.connect(_on_show)
    btn_save.clicked.connect(_on_save)

    def _on_close():
        win.close()
        import sys as _sys2
        _sys2.exit(0)

    btn_close.clicked.connect(_on_close)
    btn_row.addWidget(btn_show)
    btn_row.addWidget(btn_save)
    btn_row.addStretch()
    btn_row.addWidget(btn_close)
    root_layout.addLayout(btn_row)

    win.resize(680, 320)
    win.show()
    _app.exec()


if __name__ == '__main__':
    import sys as _sys
    from PyQt6.QtWidgets import QApplication, QFileDialog
    import json as _json

    _qt_app = QApplication.instance() or QApplication(_sys.argv)

    # argv[1] が .json ファイルならウィザードをスキップ
    _param_json = None
    _initial_dir = None
    if len(_sys.argv) > 1:
        if _sys.argv[1].endswith('.json'):
            _param_json = _sys.argv[1]
        else:
            _initial_dir = _sys.argv[1]

    if _param_json:
        # ウィザードスキップ: JSONからパスを取得
        with open(_param_json, encoding='utf-8') as _f:
            _p = _json.load(_f)
        grain_path    = _p['grain']
        sem_path      = _p['sem']
        dic_xlsx_path = _p['xlsx']
        out_dir       = _p['out']
    else:
        # ① Grain File 選択
        print("Select EBSD Grain File...")
        grain_path, _ = QFileDialog.getOpenFileName(
            None, 'Select EBSD Grain File (Grain_File_*.txt)',
            _initial_dir or '',
            'Grain File (*.txt);;All files (*.*)')
        if not grain_path:
            raise SystemExit('Cancelled')

        # ② SEM 参照画像選択
        print("Select SEM reference image (Ref_0th)...")
        sem_path, _ = QFileDialog.getOpenFileName(
            None, 'Select SEM reference image',
            _initial_dir or '',
            'Image files (*.bmp *.png *.tif *.tiff);;All files (*.*)')
        if not sem_path:
            raise SystemExit('Cancelled')

        # ③ dic_results.xlsx 選択
        print("Select dic_results.xlsx...")
        dic_xlsx_path, _ = QFileDialog.getOpenFileName(
            None, 'Select dic_results.xlsx',
            _initial_dir or '',
            'Excel files (*.xlsx);;All files (*.*)')
        if not dic_xlsx_path:
            raise SystemExit('Cancelled')

        # ④ 出力先フォルダ選択
        print("Select output folder...")
        out_dir = QFileDialog.getExistingDirectory(
            None, 'Select output folder (will create ebsd_georef/ subfolder)',
            _initial_dir or '')
        if not out_dir:
            raise SystemExit('Cancelled')
    out_dir = Path(out_dir) / 'ebsd_georef'
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"  Output folder: {out_dir}")

    print("\n" + "=" * 60)

    # [1] Grain File 読み込み
    print("[1/4] Loading Grain File...")
    data, phase_names, meta = load_grain_file(grain_path)

    # [2] Grain ID マップ生成
    print("[2/4] Generating Grain ID map...")
    grain_id_img, grain_id_map, _ = make_grain_id_image(data, meta)

    # SEM 画像読み込み（グレースケール）
    # cv2.imread は日本語パスを読めないため numpy + imdecode で対応
    with open(sem_path, 'rb') as _f:
        _buf = np.frombuffer(_f.read(), dtype=np.uint8)
    sem_img = cv2.imdecode(_buf, cv2.IMREAD_GRAYSCALE)
    if sem_img is None:
        raise FileNotFoundError(f"SEM image not found: {sem_path}")
    print(f"  SEM image size: {sem_img.shape[1]} x {sem_img.shape[0]} px")

    # DIC サブセット座標を先読み（コントロールポイントGUI表示用）
    wb_dic_pre = openpyxl.load_workbook(dic_xlsx_path, read_only=True)
    rows_u_pre = list(wb_dic_pre['u'].iter_rows(values_only=True))
    wb_dic_pre.close()
    hdr_u_pre = rows_u_pre[0]
    if hdr_u_pre[0] == 'subset_id':
        cx_arr = np.array([r[1] for r in rows_u_pre[1:] if r[1] is not None], dtype=float)
        cy_arr = np.array([r[2] for r in rows_u_pre[1:] if r[2] is not None], dtype=float)
    else:
        cx_arr = np.array([r[0] for r in rows_u_pre[1:] if r[0] is not None], dtype=float)
        cy_arr = np.array([r[1] for r in rows_u_pre[1:] if r[1] is not None], dtype=float)
    print(f"  DIC subsets (pre-load): {len(cx_arr)}")

    # [3] コントロールポイント指定 GUI
    print("[3/4] Please specify control points...")
    print("  Left(SEM) -> Right(EBSD Grain ID) click alternately")
    print("  Right-click: undo last point  /  q key: confirm")
    pts_sem, pts_ebsd, subset_visible = pick_control_points(sem_img, grain_id_img, cx_arr, cy_arr)

    n_pts = len(pts_sem)
    if n_pts < 3:
        raise SystemExit(f'Not enough control points ({n_pts}). At least 3 required.')
    print(f"  {n_pts} control point pairs confirmed.")

    # 変換行列推定
    H, mode = estimate_transform(pts_sem, pts_ebsd)

    # [4] DIC サブセット座標の読み込み・割り当て
    print("[4/4] Loading DIC subset coordinates and assigning EBSD data...")
    wb_dic = openpyxl.load_workbook(dic_xlsx_path, read_only=True)
    rows_u = list(wb_dic['u'].iter_rows(values_only=True))
    wb_dic.close()
    hdr_u = rows_u[0]
    # subset_id列があれば列1がsubset_id、列2がx、列3がy
    # なければ（旧形式）列1がx、列2がy
    if hdr_u[0] == 'subset_id':
        cx_arr = np.array([r[1] for r in rows_u[1:] if r[1] is not None], dtype=float)
        cy_arr = np.array([r[2] for r in rows_u[1:] if r[2] is not None], dtype=float)
    else:
        cx_arr = np.array([r[0] for r in rows_u[1:] if r[0] is not None], dtype=float)
        cy_arr = np.array([r[1] for r in rows_u[1:] if r[1] is not None], dtype=float)
    print(f"  DIC subsets: {len(cx_arr)}")

    # SEM [px] → EBSD [μm] 変換
    x_um, y_um = sem_px_to_ebsd_um(cx_arr, cy_arr, H, mode, meta)

    # 最近傍割り当て
    assigned, n_outside = assign_ebsd_to_subsets(
        cx_arr, cy_arr, x_um, y_um, data, phase_names, meta)

    # Excel 保存（dic_results.xlsx に ebsd_georef シートを追加）
    print("[5/5] Saving Excel and overlay PNG...")
    dic_stem = Path(dic_xlsx_path).stem
    xlsx_path_out = out_dir / f'{dic_stem}_georef.xlsx'
    save_xlsx(assigned, dic_xlsx_path, str(xlsx_path_out))

    # オーバーレイPNG保存
    png_path_out = out_dir / 'ebsd_overlay.png'
    save_overlay_png(sem_img, grain_id_img, pts_sem, pts_ebsd, png_path_out,
                     show_subsets=subset_visible, cx_arr=cx_arr, cy_arr=cy_arr)

    # サマリ
    valid = [r for r in assigned if r['grain_id'] > 0]
    n_grains_found = len(set(r['grain_id'] for r in valid))
    n_total_grains = len(np.unique(data[:, 8].astype(int)[data[:, 8] > 0]))
    print(f"\n  Assignment summary:")
    print(f"    Valid subsets : {len(valid)} / {len(assigned)}")
    print(f"    Grains found  : {n_grains_found} / {n_total_grains}")
    print(f"    Outside EBSD  : {n_outside}")

    print("\n" + "=" * 60)
    print("Done.")
    print(f"  Output folder: {out_dir}")
    print(f"  {dic_stem}_georef.xlsx  (+ sheet: ebsd_georef)")
    print(f"  ebsd_overlay.png")
    print(f"  Valid subsets : {len(valid)}/{len(assigned)}")
    print(f"  Grains found  : {n_grains_found}")
    print(f"  Outside EBSD  : {n_outside}")

    # マップビューワーを起動（ジオリファレンス後の確認用）
    print("\nLaunching Map Viewer...")
    show_map_viewer(str(xlsx_path_out), sem_img, grain_id_img, pts_sem, pts_ebsd)
